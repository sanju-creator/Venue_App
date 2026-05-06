"""
AI-Powered Call Log Classifier
================================
Reads raw call log Excel files, consolidates key columns into a single
text block, then uses Google Gemini FREE API to classify each row into
a predefined category hierarchy.

Columns consolidated  : CATEGORY | SUBCATEGORY | ITEM | SUBJECT | DESCRIPTION | RESOLUTION
Columns preserved     : REQUEST_ID, CREATED_TIME, REGION, STATE, CITY, PROJECT_NAME, VENUE_NUMBER, CENTER_CODE
New columns added     : CONSOLIDATED, MAIN_CATEGORY, SUB_CATEGORY, HIERARCHY_PATH, CONFIDENCE

Requirements:
    pip install google-generativeai pandas openpyxl

Usage:
    python call_log_classifier.py
    python call_log_classifier.py --batch-size 10   (test on first 10 rows only)
"""

import os
import sys
import json
import time
import argparse
import pandas as pd
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import google.generativeai as genai

# ===============================================================
#  CONFIGURE YOUR PATHS HERE
# ===============================================================
DEFAULT_INPUT_DIR  = r"D:\Venue Performance\Help_Desk"
DEFAULT_OUTPUT_DIR = r"D:\Venue Performance\Help_Desk"

DEFAULT_INPUT_FILE  = "Call Logs.xlsx"
DEFAULT_OUTPUT_FILE = "Call_Logs_Classified.xlsx"

# ===============================================================
#  COLUMNS TO CONSOLIDATE
# ===============================================================
CONSOLIDATE_COLS = ["CATEGORY", "SUBCATEGORY", "ITEM", "SUBJECT", "DESCRIPTION", "RESOLUTION"]

# ===============================================================
#  METADATA COLUMNS TO CARRY THROUGH
# ===============================================================
META_COLS = ["REQUEST_ID", "CREATED_TIME", "REGION", "STATE", "CITY",
             "PROJECT_NAME", "VENUE_NUMBER", "CENTER_CODE"]

# ===============================================================
#  CATEGORY HIERARCHY
#  Edit this dict freely to add / remove / rename categories
# ===============================================================
CATEGORY_HIERARCHY = {
    "Power Related": [
        "Power Cut - Exam on UPS/Genset",
        "Power Cut - Exam Halted - DG/UPS Failure/Not Available",
        "Power Cut - No Power Available - Exam on Genset Only",
        "Power Cut - Continuous Voltage Issue/Electricity Not Available",
        "Power Cut - Exam Delayed - MCB/Switch/Short Circuit",
        "Power Cut - Maintenance on Exam Day",
        "Power Related - Short Circuit",
        "Power Related - MCB Short/High Voltage",
        "CCTV Recording Issue - Power Cut",
        "Other Power Issues",
    ],
    "PC Changed": [
        "PC Change - System Failure/Network Failure/Other",
        "PC Changed",
    ],
    "QB Pull Issue": [
        "QB Pull/Push/Cache Issue",
        "QB Pull Issue",
    ],
    "Auto Seat Allocation": [
        "Exam - Unable to Auto Seat Allocation",
        "Exam System Issue - Auto Seat Allocation",
    ],
    "D Mapping": [
        "IT Infra - DMapping Required",
        "D Mapping",
    ],
    "Venue is Not PWD Friendly": [
        "Venue - Not PWD Friendly",
        "PWD Candidate - Allowed Scribe/Required Scribe/Need Extra Time",
    ],
    "White Screen Related": [
        "Exam - Blank/White Screen",
        "White Screen Related",
    ],
    "SPS Node": [
        "SPS Node Issue",
        "IT Infra - SPS Related",
        "Node Issue",
    ],
    "Candidate Debarred": [
        "Candidate Debarred - Dummy Candidate/Using Mobile/Carried Cheats",
        "Candidate Debarred",
        "Candidate - Caught While Stealing",
        "Candidate - Misbehaviour",
    ],
    "IT Asset - Damaged/Misplaced/Missing": [
        "Asset IT - Damaged/Misplaced/Missing",
        "Asset IT - Damaged by Candidate",
        "IT Infra - Laptop Related",
        "IT Infra - Printer",
        "IT Asset Related",
        "Laptop Related",
    ],
    "Venue Closed/Other Exam Running": [
        "Venue - Other Exam Going in Venue",
        "Venue - Exam Delayed - Venue Closed",
        "Venue - Denied to Open",
        "Venue Not Opened/Another Exam Running",
        "Venue - Snake Issue",
        "Venue - Cleaning Not Done",
        "Venue - Water Filling in Rain",
        "Venue - Other Issues",
    ],
    "Batch - Start batch/Stop batch delay": [
        "Batch - Start Batch/Stop Batch Delay",
        "Batch - Reappeared",
        "Batch - Password Issue",
        "Start Batch Password Related",
        "CP Application - App Failure - Batch Reappear",
        "Batch Related",
    ],
    "Venue Payment Related Issue": [
        "Venue - Payment Issue",
        "Candidate - Disallowed Due to Payment Issue",
    ],
    "Other low impact issues": [
        "Exam Application Issue",
        "Exam System Issue",
        "Exam - Application Login Issue/URL/Portal Not Started",
        "Exam - Application/Button Issue",
        "Exam - Dashboard Not Updating/Not Functional",
        "Exam - Result Pull/Push Issue",
        "Exam - Database Related Issue",
        "Exam Module Issue",
        "IT Infra - MAC Binding Error",
        "IT Infra - VPN",
        "IT Infra - Server Issues",
        "IT Infra - CAARC Not Functional/Not Configured",
        "IT Infra - FTP Issue",
        "IT Infra - SBT Issues",
        "IT Infra - Firewall",
        "IT Infra - JJT Portal Not Working",
        "IT Infra - Error E001/E002/E003",
        "IT Infra - SSID/NTS/Spark Related",
        "IT Infra - Invalid Credential CAC ID",
        "IT Infra - Laptop Backup Issue",
        "Network Issue - Exam Delayed",
        "Network Issue - Exam Halted",
        "Network Issue - CCTV Not Working",
        "PC Shortage for Exams and Mocks",
        "PC Buffer Quantity Shortage",
        "PC Related - System/Network Failure/Restart",
        "CP Application - Log In Issue",
        "CP Application - Wrong Document or Data Uploaded",
        "CP Application - Unable to Login",
        "CP Application - Project Related Issue",
        "CP Application - Venue Name Issue",
        "CP Application - App Using by Another Person/Venue",
        "CP Application - Invigilator Photo Mismatch",
        "CP Application - SSID/Wrong QP Uploaded",
        "CP Application - Unable to Use on iOS",
        "CP Application - PAT Safe Application Issue",
        "CP Application - Other",
        "DCVD/Verification - Details Mismatch",
        "DCVD/Verification - Exception Password Required",
        "DCVD/Verification - Unauthorised Activity/Wrong Verification",
        "DCVD/Verification - Dashboard Not Updating",
        "DCVD/Verification - PC Shortage",
        "DCVD/Verification - Link Issue",
        "DCVD/Verification - DCVD Failure",
        "Test Schedule - Candidate Details/Attendance Issue",
        "Candidate Related Issues",
        "Candidate - Detail Mismatch/Photo/ID/Document Issue",
        "Candidate - Reported Late",
        "Candidate - Wrongly Allowed/Extra Allowed",
        "Candidate - Allowed in Wrong Batch/Module",
        "Candidate - Chosen Wrong Module",
        "Candidate - Language Issues",
        "Candidate - Bio Break/Medical/Emergency",
        "Candidate - Paper Submitted Late",
        "Observer Related Issues",
        "Observer Not Available/Not Reported",
        "Observer Feedbacks",
        "Observer Link Not Working",
        "Invigilator - Not Reported",
        "Invigilator - Selected Wrong Module/Incorrect Language",
        "Venue - Staff Issue/Not Reported",
        "Venue - Jammer Not Available/Not Working",
        "Venue - Police Validation Form Not Accepted",
        "Venue - Asset Damaged",
        "Venue - Undertaking Form Not Signed",
        "Venue - Exam Delay by Centre",
        "Venue - Other",
        "IT Technician Not Reported at Venue",
        "CP Not Available - Exam Delayed",
        "Translation/Content/Font/Symbol Error",
        "Question Error",
        "Help Desk Queries",
        "CCTV Related",
        "Courier Issue",
        "Security Guard Related",
        "Operation - Exam Over-scheduled/Rake Issue",
        "Primary/Secondary Server Issue",
        "Postgress Related",
        "SSID Related",
        "Other IT Issues",
        "Unclassified",
    ],
}


# ===============================================================
#  STEP 1 — CONSOLIDATION
# ===============================================================
def consolidate_row(row, cols):
    parts = []
    for col in cols:
        val = row.get(col, "")
        val = str(val).strip() if pd.notna(val) else ""
        if val and val.lower() not in ("nan", "none", "na", "n/a", "-", ""):
            parts.append(val)
    return " | ".join(parts)


def consolidate_dataframe(df):
    available = [c for c in CONSOLIDATE_COLS if c in df.columns]
    missing   = [c for c in CONSOLIDATE_COLS if c not in df.columns]
    if missing:
        print(f"  Warning - columns not found (skipped): {missing}")
    df["CONSOLIDATED"] = df.apply(lambda row: consolidate_row(row, available), axis=1)
    print(f"  Consolidated: {available}")
    return df


# ===============================================================
#  STEP 2 — AI CLASSIFICATION (Google Gemini)
# ===============================================================
def build_prompt(consolidated_text):
    hierarchy_text = json.dumps(CATEGORY_HIERARCHY, indent=2)
    return f"""You are an expert call-log classifier for an examination management system in India.

Classify the following call log entry into a MAIN CATEGORY and SUB CATEGORY.

Available categories:
{hierarchy_text}

Call log entry:
{consolidated_text}

Rules:
- Read ALL parts of the text carefully.
- Choose the MOST SPECIFIC and ACCURATE match.
- main_category must exactly match one of the top-level keys.
- sub_category must come from the list under the chosen main_category.
- If nothing fits, use "Other low impact issues" > "Unclassified".
- confidence: "High" if clearly matches, "Medium" if likely, "Low" if uncertain.
- Respond ONLY with a valid JSON object. No explanation, no markdown fences.

Response format:
{{
  "main_category": "<exact main category name>",
  "sub_category": "<exact sub-category name>",
  "confidence": "<High | Medium | Low>"
}}"""


def prompt_for_api_key():
    """Interactively ask the user for their Google Gemini API key at runtime."""
    import getpass
    print("\n" + "=" * 65)
    print("  GOOGLE GEMINI API KEY REQUIRED")
    print("=" * 65)
    print("  Get FREE key : https://aistudio.google.com/apikey")
    print("  Sign in with your Gmail — no credit card needed!")
    print("  Your key starts with : AIza...")
    print("  (Input is hidden for security)\n")
    while True:
        key = getpass.getpass("AIzaSyCauT4ciktuwNQFEe18Sp2aosS9pEczSpY").strip()
        if key == "":
            print("  No key entered. Exiting.")
            sys.exit(0)
        elif key.startswith("AIza") and len(key) > 20:
            print("  API key accepted.\n")
            return key
        else:
            print("  Invalid key format — it should start with 'AIza'. Please try again.\n")


class CallLogClassifier:
    def __init__(self, api_key=None):
        # Priority: 1) CLI argument  2) Environment variable  3) Interactive prompt
        key = api_key or os.environ.get("GEMINI_API_KEY")
        if not key:
            key = prompt_for_api_key()
        genai.configure(api_key=key)
        self.model = genai.GenerativeModel("gemini-1.5-flash")
        self.fallback = {"main_category": "Other low impact issues",
                         "sub_category": "Unclassified", "confidence": "Low"}

    def classify_single(self, consolidated_text, retries=3):
        if not consolidated_text or str(consolidated_text).strip() == "":
            return self.fallback

        prompt = build_prompt(str(consolidated_text).strip())

        for attempt in range(1, retries + 1):
            try:
                response = self.model.generate_content(prompt)
                raw = response.text.strip().replace("```json", "").replace("```", "").strip()
                result = json.loads(raw)
                if result.get("main_category") not in CATEGORY_HIERARCHY:
                    return self.fallback
                return result

            except json.JSONDecodeError as e:
                print(f"    JSON error attempt {attempt}: {e}")
                if attempt == retries:
                    return self.fallback
                time.sleep(1)

            except Exception as e:
                err = str(e).lower()
                # Handle rate limit (free tier = 15 requests/min)
                if "quota" in err or "rate" in err or "429" in err:
                    wait = 60
                    print(f"\n    Free tier rate limit hit — waiting {wait}s before retrying...")
                    time.sleep(wait)
                else:
                    print(f"    Error attempt {attempt}: {e}")
                    if attempt == retries:
                        return self.fallback
                    time.sleep(2)

        return self.fallback

    def classify_batch(self, texts, delay=1.5):
        """
        delay=1.5s between calls keeps us safely under the
        free tier limit of 15 requests/minute.
        """
        results = []
        total = len(texts)
        for i, text in enumerate(texts, 1):
            print(f"  [{i:>5}/{total}] Processing row {i}...", end="\r")
            results.append(self.classify_single(text))
            if i < total:
                time.sleep(delay)
        print(f"  [{total}/{total}] Classification complete!          ")
        return results


# ===============================================================
#  STEP 3 — EXCEL FORMATTING
# ===============================================================
CONFIDENCE_COLORS = {"High": "C6EFCE", "Medium": "FFEB9C", "Low": "FFC7CE"}
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT    = Font(name="Arial", size=10)
ALT_ROW_FILL = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

OUTPUT_COLS = (
    META_COLS
    + ["CONSOLIDATED"]
    + ["MAIN_CATEGORY", "SUB_CATEGORY", "HIERARCHY_PATH", "CONFIDENCE"]
    + CONSOLIDATE_COLS
)

OUTPUT_COL_WIDTHS = {
    "REQUEST_ID": 15, "CREATED_TIME": 22, "REGION": 15, "STATE": 15,
    "CITY": 18, "PROJECT_NAME": 25, "VENUE_NUMBER": 15, "CENTER_CODE": 15,
    "CONSOLIDATED": 90, "MAIN_CATEGORY": 35, "SUB_CATEGORY": 50,
    "HIERARCHY_PATH": 72, "CONFIDENCE": 12,
    "CATEGORY": 25, "SUBCATEGORY": 30, "ITEM": 25,
    "SUBJECT": 35, "DESCRIPTION": 50, "RESOLUTION": 50,
}


def style_excel_output(output_path, df):
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Classified Results"
    actual_cols = df.columns.tolist()

    for col_idx, col_name in enumerate(actual_cols, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = OUTPUT_COL_WIDTHS.get(col_name, 20)

    ws.row_dimensions[1].height = 32
    conf_col_idx = (actual_cols.index("CONFIDENCE") + 1) if "CONFIDENCE" in actual_cols else None

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(actual_cols)), start=2
    ):
        conf_val = row[conf_col_idx - 1].value if conf_col_idx else "Low"
        is_alt = row_idx % 2 == 0
        for cell in row:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if conf_col_idx and cell.column == conf_col_idx:
                hex_color = CONFIDENCE_COLORS.get(conf_val, "FFFFFF")
                cell.fill = PatternFill("solid", start_color=hex_color, end_color=hex_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=10, bold=True)
            elif is_alt:
                cell.fill = ALT_ROW_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Category Summary Sheet
    ws2 = wb.create_sheet("Category Summary")
    for col_idx, (h, w) in enumerate(zip(["MAIN CATEGORY", "COUNT", "% OF TOTAL"], [45, 10, 14]), 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    counts = Counter(df["MAIN_CATEGORY"].dropna())
    total = sum(counts.values())
    for i, (cat, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), start=2):
        pct = f"{cnt / total * 100:.1f}%"
        for col_idx, val in enumerate([cat, cnt, pct], 1):
            c = ws2.cell(row=i, column=col_idx, value=val)
            c.font = DATA_FONT; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            if i % 2 == 0:
                c.fill = ALT_ROW_FILL
    ws2.freeze_panes = "A2"

    # SubCategory Breakdown Sheet
    ws3 = wb.create_sheet("SubCategory Breakdown")
    for col_idx, (h, w) in enumerate(
        zip(["MAIN CATEGORY", "SUB CATEGORY", "COUNT", "% OF TOTAL"], [35, 52, 10, 14]), 1
    ):
        c = ws3.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = THIN_BORDER
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    sub_counts = Counter(zip(df["MAIN_CATEGORY"].fillna(""), df["SUB_CATEGORY"].fillna("")))
    for i, ((main, sub), cnt) in enumerate(
        sorted(sub_counts.items(), key=lambda x: (-x[1], x[0][0])), start=2
    ):
        pct = f"{cnt / total * 100:.1f}%"
        for col_idx, val in enumerate([main, sub, cnt, pct], 1):
            c = ws3.cell(row=i, column=col_idx, value=val)
            c.font = DATA_FONT; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
            if i % 2 == 0:
                c.fill = ALT_ROW_FILL
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  Formatted Excel saved -> {output_path}")


# ===============================================================
#  MAIN PIPELINE
# ===============================================================
def run_pipeline(input_path, output_path, api_key=None, batch_size=None):
    print("\n" + "=" * 65)
    print("   AI-Powered Call Log Classifier  |  Venue Performance")
    print("   Powered by Google Gemini (Free API)")
    print("=" * 65)

    if not Path(input_path).exists():
        print(f"\nERROR: Input file not found: {input_path}")
        print("  Please update DEFAULT_INPUT_DIR and DEFAULT_INPUT_FILE at the top of the script.\n")
        sys.exit(1)

    print(f"\n  Input  : {input_path}")
    print(f"  Output : {output_path}")

    df = pd.read_excel(input_path)
    print(f"\n  Rows loaded : {len(df)}")
    print(f"  Columns     : {df.columns.tolist()}")

    if batch_size:
        df = df.head(batch_size)
        print(f"  Batch mode  : first {batch_size} rows only")

    # Estimate time
    total_rows = len(df)
    est_minutes = round((total_rows * 1.5) / 60, 1)
    print(f"  Est. time   : ~{est_minutes} minutes (free tier pace)\n")

    print("Step 1/3 - Consolidating columns...")
    df = consolidate_dataframe(df)

    print(f"\nStep 2/3 - Classifying {total_rows} rows via Google Gemini AI...")
    classifier = CallLogClassifier(api_key=api_key)
    classifications = classifier.classify_batch(df["CONSOLIDATED"].tolist())

    df["MAIN_CATEGORY"]  = [c["main_category"] for c in classifications]
    df["SUB_CATEGORY"]   = [c["sub_category"]   for c in classifications]
    df["HIERARCHY_PATH"] = [f"{c['main_category']} > {c['sub_category']}" for c in classifications]
    df["CONFIDENCE"]     = [c["confidence"]      for c in classifications]

    ordered = [c for c in OUTPUT_COLS if c in df.columns]
    extra   = [c for c in df.columns  if c not in ordered]
    df = df[ordered + extra]

    print("\nStep 3/3 - Saving & formatting Excel...")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    style_excel_output(output_path, df)

    print("\n" + "-" * 65)
    print("CLASSIFICATION SUMMARY")
    print("-" * 65)
    for cat, cnt in df["MAIN_CATEGORY"].value_counts().items():
        bar = "#" * int(cnt / len(df) * 35)
        print(f"  {cat:<45} {cnt:>4}  {bar}")
    print("-" * 65)
    print(f"  Total rows : {len(df)}")
    for conf, cnt in df["CONFIDENCE"].value_counts().items():
        print(f"  Confidence {conf:<8}: {cnt}")
    print(f"\nDone! Output saved to: {output_path}\n")
    return df


# ===============================================================
#  ENTRY POINT
# ===============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="AI-Powered Call Log Classifier (Google Gemini)")
    parser.add_argument("--input",      default=None, help="Full path to input Excel file")
    parser.add_argument("--output",     default=None, help="Full path to output Excel file")
    parser.add_argument("--api-key",    default=None, help="Google Gemini API key (optional — will prompt if not provided)")
    parser.add_argument("--batch-size", type=int, default=None,
                        help="Process only first N rows (for testing, e.g. --batch-size 10)")
    args = parser.parse_args()

    input_path  = args.input  or str(Path(DEFAULT_INPUT_DIR)  / DEFAULT_INPUT_FILE)
    output_path = args.output or str(Path(DEFAULT_OUTPUT_DIR) / DEFAULT_OUTPUT_FILE)

    run_pipeline(
        input_path=input_path,
        output_path=output_path,
        api_key=args.api_key,
        batch_size=args.batch_size,
    )