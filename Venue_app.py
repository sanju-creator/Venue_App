from __future__ import annotations

# import streamlit as st
# import pandas as pd
# import os
# import numpy as np
# import plotly.express as px
# import plotly.graph_objects as go
# import html
# import json
# import re
# import zipfile
# from datetime import datetime
# from streamlit_autorefresh import st_autorefresh
# import io
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment


# # ---------------- PAGE CONFIG ----------------
# st.set_page_config(layout="wide", page_title="Venue Inventory Master", initial_sidebar_state="expanded")

# # ---------------- STYLE ----------------
# STYLE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ui", "styles.css")
# try:
#     with open(STYLE_PATH, "r", encoding="utf-8") as f:
#         st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
# except FileNotFoundError:
#     # Styling is optional, don't break if CSS file not found
#     pass
# except Exception as e:
#     # Log styling loading errors but don't crash
#     print(f"Warning: Could not load styles from {STYLE_PATH}: {e}")


# def metric_card(label, value, tooltip: str | None = None, pct: float | None = None, trend: str | None = None):
#     safe_label = html.escape(str(label))
#     safe_value = html.escape(str(value))

#     details_html = ""
#     graph_html = ""
#     trend_html = ""

#     if trend:
#         t_str = str(trend).strip()
#         if "+" in t_str:
#             t_class = "trend-up"
#         elif "-" in t_str:
#             t_class = "trend-down"
#         else:
#             t_class = "trend-neutral"
#         trend_html = f'<span class="{t_class}">{html.escape(t_str)}</span>'

#     if pct is not None:
#         safe_pct = min(max(pct, 0), 100)
#         graph_html = f'<div class="kpi-graph"><div class="kpi-graph-fill" style="width: {safe_pct}%;"></div></div>'

#     if tooltip:
#         parts = str(tooltip).split("\n")
#         formatted_parts = []
#         for p in parts:
#             if ":" in p:
#                 k, v = p.split(":", 1)
#                 formatted_parts.append(
#                     "<span class='kpi-detail-item'>"
#                     f"<span style='color:var(--text);font-weight:700'>{html.escape(k)}:</span>"
#                     f"<span>{html.escape(v.strip())}</span>"
#                     "</span>"
#                 )
#             else:
#                 formatted_parts.append(f"<span class='kpi-detail-item'>{html.escape(p)}</span>")
#         safe_details = ' <span style="margin:0 6px; color:var(--border-dark)">|</span> '.join(formatted_parts)

#         details_html = f'<div class="kpi-subtext">{trend_html} {safe_details}</div>'
#     elif trend_html:
#         details_html = f'<div class="kpi-subtext">{trend_html}</div>'

#     final_html = f"""<div class="kpi-static">
# <div class="kpi-label">{safe_label}</div>
# <div class="kpi-value">{safe_value}</div>
# {graph_html}
# {details_html}
# </div>"""
#     st.markdown(final_html, unsafe_allow_html=True)

# def metric_card_compact(label: str, value: int, accent: str = "#81CDEC"):
#     """Compact KPI card for inline executive summaries."""
#     safe_label = html.escape(str(label))
#     safe_value = html.escape(f"{int(value):,}" if isinstance(value, (int, np.integer, float, np.floating)) else str(value))
#     st.markdown(
#         f"""
#         <div class="kpi-compact" style="--kpi-accent:{accent};">
#             <div class="kpi-compact-label">{safe_label}</div>
#             <div class="kpi-compact-value">{safe_value}</div>
#         </div>
#         """,
#         unsafe_allow_html=True,
#     )

# def render_brand_logo():
#     """Render company logo inline — call inside a st.columns block for positioning."""
#     import base64 as _b64
#     logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
#     if not os.path.exists(logo_path):
#         logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.jpg")
#     if os.path.exists(logo_path):
#         with open(logo_path, "rb") as _lf:
#             _logo_b64 = _b64.b64encode(_lf.read()).decode("utf-8")
#         _ext = os.path.splitext(logo_path)[1][1:].lower()
#         _mime = "jpeg" if _ext == "jpg" else _ext
#         st.markdown(
#             f"""
#             <div style="display:flex; justify-content:flex-end; align-items:flex-start; margin:-14px 0 8px 0; padding:0;">
#                 <img src="data:image/{_mime};base64,{_logo_b64}"
#                      alt="DEXIT Global"
#                      style="width:205px; height:auto; display:block; margin:0; padding:0; object-fit:contain;" />
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#     else:
#         st.markdown(
#             """
#             <div style="display:flex; justify-content:flex-end; margin:-4px 0 0 0; padding:0;
#                         font-size:15px; font-weight:800; color:#1A202C; letter-spacing:0.5px;">VMS AUTOMATION</div>
#             """,
#             unsafe_allow_html=True,
#         )

# def sidebar_checkbox_filter(
#     label: str,
#     options: list,
#     key_prefix: str,
#     *,
#     expanded: bool = False,
#     columns: int = 1,
# ) -> list:
#     """Render an executive-style checkbox filter block and return selected items."""
#     clean_options = [opt for opt in options if str(opt).strip() != ""]
#     if not clean_options:
#         return []

#     for idx in range(len(clean_options)):
#         state_key = f"{key_prefix}_{idx}"
#         if state_key not in st.session_state:
#             st.session_state[state_key] = True

#     selected_before = sum(
#         1 for idx in range(len(clean_options)) if st.session_state.get(f"{key_prefix}_{idx}", True)
#     )

#     with st.sidebar.expander(f"{label}  ({selected_before}/{len(clean_options)})", expanded=expanded):
#         search_query = st.text_input(
#             f"Search {label}",
#             key=f"{key_prefix}_search",
#             placeholder="Search options...",
#             label_visibility="collapsed",
#         ).strip().lower()

#         ctl_col1, ctl_col2 = st.columns(2)
#         if ctl_col1.button("Select All", key=f"{key_prefix}_all", use_container_width=True):
#             for idx in range(len(clean_options)):
#                 st.session_state[f"{key_prefix}_{idx}"] = True
#         if ctl_col2.button("Clear All", key=f"{key_prefix}_none", use_container_width=True):
#             for idx in range(len(clean_options)):
#                 st.session_state[f"{key_prefix}_{idx}"] = False

#         visible_options = []
#         for idx, opt in enumerate(clean_options):
#             opt_str = str(opt)
#             if not search_query or search_query in opt_str.lower():
#                 visible_options.append((idx, opt_str))

#         if not visible_options:
#             st.caption("No matching options.")
#         else:
#             check_cols = st.columns(max(1, columns))
#             for pos, (idx, opt_text) in enumerate(visible_options):
#                 with check_cols[pos % max(1, columns)]:
#                     st.checkbox(opt_text, key=f"{key_prefix}_{idx}")

#     selected = [opt for idx, opt in enumerate(clean_options) if st.session_state.get(f"{key_prefix}_{idx}", True)]
#     return selected

# # ---------------- PATHS ----------------
# SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# BASE_PATH = os.path.join(SCRIPT_DIR, "Data")
# OUTPUT_PATH = os.path.join(BASE_PATH, "Output")
# PHOTOS_PATH = os.path.join(BASE_PATH, "Venue_Photos")
# INDIA_GEOJSON_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "india.geojson")
# STATE_LIST_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "state-list.json")
# STATE_GEOJSON_DIR = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "states")
# BIHAR_REFERENCE_PPTX_PATH = os.path.join("C:\\Users", "DELL", "Downloads", "Bihar district maps.pptx")
# INDIA_MAP_CENTER = {"lat": 21.92, "lon": 82.74}
# INDIA_MAP_BOUNDS = {"west": 65.5, "east": 100.0, "south": 4.8, "north": 39.8}
# if not os.path.exists(OUTPUT_PATH): os.makedirs(OUTPUT_PATH)
# if not os.path.exists(PHOTOS_PATH): os.makedirs(PHOTOS_PATH)

# # ---------------- CONFIGURATION ----------------
# FILES = {
#     "Venue Inventory": {"file": "VMS Master.xlsx", "users": ["Nishant", "Prafull", "Admin"]},
#     "Blacklisted": {"file": "Blacklisted Venues.xlsx", "users": ["Prafull", "Admin"]},
#     "FFA Data": {"file": "FFA Data.xlsx", "users": ["Anil", "Admin"]},
#     "Suspicious": {"file": "Suspicious Activities.xlsx", "users": ["Prafull", "Admin"]},
#     "Day End": {"file": "Day End Observations.xlsx", "users": ["Prafull", "Admin"]},
#     "Market Reviews": {"file": "Google review and ratings.xlsx", "users": ["Prafull", "Admin"]},
#     "Call Logs": {"file": "Master Call Logs.xlsx", "users": ["Prafull", "Admin"]},
#     "Project Master": {"file": "Project Master.xlsx", "users": ["Prafull", "Admin"]}
# }

# USERS = {"Prafull": "Prafull@123", "Nishant": "Nishant@123", "Mayuresh": "Mayuresh@123", "Anil": "Anil@123", "Admin": "Admin@123"}

# # --- SORTING DEFINITIONS ---
# CAT_ORDER = ["A", "B", "C", "BL-C", "BL"]
# REGION_ORDER = ["North One", "North Two", "East", "West", "South", "Foreign"]

# # ---------------- SESSION STATE ----------------
# if "logged_in" not in st.session_state: st.session_state.logged_in = False
# if "user" not in st.session_state: st.session_state.user = None
# if "search_main_box" not in st.session_state: st.session_state.search_main_box = ""
# if "venue_search_query" not in st.session_state: st.session_state.venue_search_query = ""
# if "venue_detail_code" not in st.session_state: st.session_state.venue_detail_code = None
# if "map_selected_state" not in st.session_state: st.session_state.map_selected_state = None
# if "map_selected_district" not in st.session_state: st.session_state.map_selected_district = None
# if "map_selected_district_state" not in st.session_state: st.session_state.map_selected_district_state = None

# VALID_PAGES = {"dashboard", "login", "main", "edit", "process", "venue_detail", "india_map_detail", "manpower_dashboard", "project_wise_details"}
# VALID_DATASETS = set(FILES.keys())

# qp_page = st.query_params.get("page", None)
# if isinstance(qp_page, list):
#     qp_page = qp_page[0] if qp_page else None

# qp_dataset = st.query_params.get("dataset", None)
# if isinstance(qp_dataset, list):
#     qp_dataset = qp_dataset[0] if qp_dataset else None

# if "page" not in st.session_state:
#     st.session_state.page = qp_page if qp_page in VALID_PAGES else "dashboard"
# elif st.session_state.page not in VALID_PAGES:
#     st.session_state.page = "dashboard"

# if "dataset_key" not in st.session_state:
#     st.session_state.dataset_key = qp_dataset if qp_dataset in VALID_DATASETS else "Venue Inventory"
# elif st.session_state.dataset_key not in VALID_DATASETS:
#     st.session_state.dataset_key = "Venue Inventory"
# elif qp_dataset in VALID_DATASETS and qp_dataset != st.session_state.dataset_key:
#     st.session_state.dataset_key = qp_dataset

# if qp_page != st.session_state.page:
#     st.query_params["page"] = st.session_state.page
# if qp_dataset != st.session_state.dataset_key:
#     st.query_params["dataset"] = st.session_state.dataset_key

# if st.session_state.page == "dashboard":
#     st.session_state["_auto_refresh_count"] = st_autorefresh(interval=60000, key="auto_refresh_main")
# else:
#     st.session_state["_auto_refresh_count"] = st.session_state.get("_auto_refresh_count", 0)

# # ---------------- SMART DATA LOADER ----------------
# def load_smart_data(file_path):
#     if not os.path.exists(file_path): return pd.DataFrame()
#     try:
#         if file_path.lower().endswith(".xlsx"):
#             preview = pd.read_excel(file_path, header=None, nrows=10)
#             header_idx = 0
#             for i, row in preview.iterrows():
#                 row_str = " ".join([str(x) for x in row.values])
#                 if any(key in row_str.upper() for key in ["ROW LABELS", "VENUE_TYPE", "CODE", "ACTIVE", "DMS_CODE", "STATUS"]):
#                     header_idx = i; break
#             return pd.read_excel(file_path, header=header_idx)
#         return pd.read_csv(file_path)
#     except: return pd.DataFrame()

# # ---------------- HELPER FUNCTIONS FOR DATA PROCESSING ----------------
# def clean_project_name(series: pd.Series) -> pd.Series:
#     """Clean project names by removing special chars, normalizing spacing, and filtering blanks."""
#     s = series.fillna("").astype(str).str.strip()
#     s_upper = s.str.upper()
#     is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
#     s = s.mask(is_blank, "")
#     s = s.str.replace(r"[^A-Za-z ]+", " ", regex=True)
#     s = s.str.replace(r"\s+", " ", regex=True).str.strip()
#     return s.str.upper()

# def _blankify_text(series: pd.Series) -> pd.Series:
#     """Convert blank-like values (-, NAN, NONE) to empty strings."""
#     s = series.fillna("").astype(str).str.strip()
#     s_upper = s.str.upper()
#     is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
#     return s.mask(is_blank, "")

# def _col_or_empty(df_in: pd.DataFrame, col_names: list) -> pd.Series:
#     """Return column from dataframe or empty series if not present. Case-insensitive."""
#     col_names_lower = [c.lower() for c in col_names]
#     for c in df_in.columns:
#         if str(c).lower() in col_names_lower:
#             return df_in[c]
#     return pd.Series([""] * len(df_in), index=df_in.index)

# def build_concat_address(df_in: pd.DataFrame) -> pd.Series:
#     """Build concatenated full address in format: name, address, landmark, city, district, state - pincode"""
#     name_s = _blankify_text(_col_or_empty(df_in, ["name", "venue_name", "updated_venue_name"]))
#     address_s = _blankify_text(_col_or_empty(df_in, ["address"]))
#     landmark_s = _blankify_text(_col_or_empty(df_in, ["landmark"]))
#     city_s = _blankify_text(_col_or_empty(df_in, ["city"]))
#     district_s = _blankify_text(_col_or_empty(df_in, ["district"]))
#     state_s = _blankify_text(_col_or_empty(df_in, ["state"]))

#     parts_df = pd.concat([name_s, address_s, landmark_s, city_s, district_s, state_s], axis=1)
#     joined_address = parts_df.apply(lambda row: ", ".join([v for v in row.tolist() if v]), axis=1)

#     pincode_s = _blankify_text(_col_or_empty(df_in, ["pincode", "pin_code"]))
#     pincode_s = pincode_s.str.replace(r"\.0$", "", regex=True)

#     # Format: name, address, landmark, city, district, state - pincode
#     result = np.where(
#         joined_address != "",
#         joined_address + np.where(pincode_s != "", " - " + pincode_s, ""),
#         ""
#     )
#     return pd.Series(result, index=df_in.index)

# def find_dms_col(df: pd.DataFrame):
#     """Find the DMS/Venue code column in a dataframe."""
#     return next((c for c in df.columns if c.lower().replace(" ", "_") in ['dms_code', 'venue_code']), None)

# STATE_NAME_ALIASES = {
#     "NCT OF DELHI": "DELHI",
#     "NCT DELHI": "DELHI",
#     "ORISSA": "ODISHA",
#     "UTTARANCHAL": "UTTARAKHAND",
#     "JAMMU KASHMIR": "JAMMU AND KASHMIR",
#     "DADRA AND NAGAR HAVELI AND DAMAN AND DIU": "DADRA AND NAGAR HAVELI DAMAN AND DIU",
#     "ANDAMAN NICOBAR": "ANDAMAN AND NICOBAR ISLANDS",
#     "PONDICHERRY": "PUDUCHERRY",
# }


# @st.cache_data(show_spinner=False)
# def load_india_geojson(path: str):
#     if not os.path.exists(path):
#         return None
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             gj = json.load(f)
#         if isinstance(gj, dict) and isinstance(gj.get("features"), list):
#             return gj
#     except Exception:
#         return None
#     return None


# def normalize_state_name(value: str) -> str:
#     name = str(value or "").upper()
#     name = name.replace("&", " AND ")
#     name = re.sub(r"[^A-Z0-9]+", " ", name)
#     name = re.sub(r"\s+", " ", name).strip()
#     return STATE_NAME_ALIASES.get(name, name)


# def build_india_map_dataframe(source_df: pd.DataFrame | None):
#     geojson_data = load_india_geojson(INDIA_GEOJSON_PATH)
#     if geojson_data is None:
#         return pd.DataFrame(), None, 0, 0

#     state_count_map: dict[str, int] = {}
#     if source_df is not None and not source_df.empty and "state" in source_df.columns:
#         state_series = source_df["state"].fillna("").astype(str).str.strip()
#         state_series = state_series[state_series != ""]
#         for state_name, count in state_series.value_counts().items():
#             normalized = normalize_state_name(state_name)
#             if normalized:
#                 state_count_map[normalized] = state_count_map.get(normalized, 0) + int(count)

#     rows = []
#     matched_states = 0
#     mapped_venues = 0
#     features = geojson_data.get("features", [])

#     for idx, feature in enumerate(features):
#         properties = feature.get("properties", {}) or {}
#         state_name = (
#             properties.get("st_nm")
#             or properties.get("STATE")
#             or properties.get("name")
#             or f"State {idx + 1}"
#         )
#         district_code = properties.get("dt_code") or properties.get("district_code")
#         state_code = properties.get("st_code") or properties.get("state_code")
#         feature_id = str(
#             feature.get("id")
#             or district_code
#             or state_code
#             or properties.get("ID_1")
#             or f"{normalize_state_name(state_name)}_{idx + 1}"
#         )
#         feature["id"] = feature_id

#         count = int(state_count_map.get(normalize_state_name(state_name), 0))
#         if count > 0:
#             matched_states += 1
#             mapped_venues += count

#         rows.append(
#             {
#                 "id": feature_id,
#                 "state": state_name,
#                 "venue_count": count,
#             }
#         )

#     return pd.DataFrame(rows), geojson_data, matched_states, mapped_venues

# def _get_feature_centroid(feature: dict) -> tuple[float, float]:
#     geom = feature.get("geometry", {})
#     if not geom:
#         return 0.0, 0.0
#     coords = geom.get("coordinates", [])
#     flat_coords = []

#     def extract_coords(c):
#         if not c:
#             return
#         if isinstance(c[0], (int, float)):
#             flat_coords.append(c)
#         else:
#             for item in c:
#                 extract_coords(item)
#     extract_coords(coords)
    
#     if not flat_coords:
#         return 0.0, 0.0
#     arr = np.array(flat_coords)
#     return float(np.mean(arr[:, 0])), float(np.mean(arr[:, 1]))


# def create_india_map_figure(
#     map_df: pd.DataFrame,
#     geojson_data: dict,
#     *,
#     height: int = 320,
#     show_data: bool = True,
#     zoom: float = 3.5,
#     center: dict | None = None,
#     opacity: float = 0.78,
#     map_style: str = "white-bg",
#     color_power: float = 1.0,
#     line_color: str = "#94A3B8",
#     line_width: float = 0.9,
#     restrict_to_india: bool = True,
# ):
#     if map_df.empty or not geojson_data:
#         return None

#     if show_data:
#         plot_df = map_df.copy()
#         counts = pd.to_numeric(plot_df.get("venue_count", 0), errors="coerce").fillna(0).astype(float)
#         if color_power != 1.0:
#             plot_df["_color_metric"] = np.power(counts, color_power)
#         else:
#             plot_df["_color_metric"] = counts
#         max_metric = float(plot_df["_color_metric"].max()) if "_color_metric" in plot_df.columns else 0.0
#         fig_map = px.choropleth(
#             plot_df,
#             geojson=geojson_data,
#             locations="id",
#             featureidkey="id",
#             color="_color_metric",
#             hover_name="state",
#             custom_data=["state", "id", "venue_count"],
#             hover_data={"venue_count": False, "id": False, "_color_metric": False},
#             color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
#             range_color=(0, max(1.0, max_metric)),
#         )
#     else:
#         preview_df = map_df.copy()
#         preview_df["map_fill"] = 1
#         plot_df = preview_df
#         fig_map = px.choropleth(
#             preview_df,
#             geojson=geojson_data,
#             locations="id",
#             featureidkey="id",
#             color="map_fill",
#             hover_name="state",
#             custom_data=["state", "id", "venue_count"],
#             hover_data={"map_fill": False, "id": False},
#             color_continuous_scale=["#C9E1F2", "#C9E1F2"],
#             range_color=(0, 1),
#         )

#     fig_map.update_traces(
#         marker_line_color=line_color,
#         marker_line_width=line_width,
#         hovertemplate="<b>%{hovertext}</b><extra></extra>" if not show_data else "<b>%{hovertext}</b><br>Venues: %{customdata[2]:,}<extra></extra>",
#     )
    
#     id_to_name = dict(zip(plot_df["id"], plot_df["state"]))
#     state_to_coords = {}
#     if geojson_data and "features" in geojson_data:
#         for f in geojson_data["features"]:
#             fid = f.get("id")
#             if fid in id_to_name:
#                 state_name = id_to_name[fid]
#                 geom = f.get("geometry", {})
#                 if geom:
#                     if state_name not in state_to_coords:
#                         state_to_coords[state_name] = []
#                     state_to_coords[state_name].append(geom.get("coordinates", []))

#     HIDE_LABELS = {
#         "Dadra and Nagar Haveli and Daman and Diu", 
#         "Chandigarh", 
#         "Puducherry", 
#         "Lakshadweep",
#         "Goa",
#         "Delhi",
#         "Sikkim",
#         "Tripura",
#         "Mizoram",
#         "Manipur",
#         "Nagaland",
#         "Meghalaya",
#         "Haryana"
#     }
#     LABEL_ALIASES = {
#         "Andaman and Nicobar Islands": "A & N",
#         "Jammu and Kashmir": "J & K",
#         "Himachal Pradesh": "HP",
#         "Uttarakhand": "UK",
#         "Punjab": "PB",
#         "West Bengal": "WB",
#         "Jharkhand": "JH",
#         "Arunachal Pradesh": "Arunachal",
#         "Uttar Pradesh": "Uttar<br>Pradesh",
#         "Madhya Pradesh": "Madhya<br>Pradesh",
#         "Andhra Pradesh": "Andhra<br>Pradesh",
#         "Tamil Nadu": "Tamil<br>Nadu"
#     }
    
#     lons, lats, texts = [], [], []
#     for state_name, coords_list in state_to_coords.items():
#         if state_name in HIDE_LABELS:
#             continue
#         cx, cy = _get_feature_centroid({"geometry": {"coordinates": coords_list}})
#         if cx != 0.0 and cy != 0.0:
#             lons.append(cx)
#             lats.append(cy)
#             texts.append(LABEL_ALIASES.get(state_name, state_name))

#     if show_data and lons:
#         fig_map.add_trace(go.Scattergeo(
#             lon=lons,
#             lat=lats,
#             text=texts,
#             mode="text",
#             textfont=dict(color="#1E293B", size=8.5, family="Inter, sans-serif"),
#             showlegend=False,
#             hoverinfo="skip"
#         ))
        
#     fig_map.update_geos(fitbounds="locations", visible=False)
    
#     fig_map.update_layout(
#         clickmode="event+select",
#         dragmode="pan",
#         margin=dict(t=0, b=0, l=0, r=0),
#         paper_bgcolor="rgba(0,0,0,0)",
#         plot_bgcolor="rgba(0,0,0,0)",
#         height=height,
#         coloraxis_showscale=show_data,
#         coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.78),
#     )
#     return fig_map


# def extract_selected_points(selection_event) -> list:
#     if selection_event is None:
#         return []
#     if isinstance(selection_event, dict):
#         return selection_event.get("selection", {}).get("points", []) or []
#     selection_obj = getattr(selection_event, "selection", None)
#     if selection_obj is None:
#         return []
#     return getattr(selection_obj, "points", []) or []


# def get_last_point_signature(points: list) -> str | None:
#     if not isinstance(points, (list, tuple)) or not points:
#         return None
#     point = points[-1]
#     if point is None:
#         return None
#     customdata = _point_get(point, "customdata")
#     if isinstance(customdata, (list, tuple)):
#         custom_sig = "|".join("" if x is None else str(x) for x in customdata[:3])
#     else:
#         custom_sig = ""
#     sig_parts = [
#         str(_point_get(point, "location") or ""),
#         str(_point_get(point, "id") or ""),
#         str(_point_get(point, "hovertext") or ""),
#         str(_point_get(point, "point_index") or _point_get(point, "pointIndex") or ""),
#         custom_sig,
#     ]
#     signature = "||".join(sig_parts).strip()
#     return signature if signature else None


# def _point_get(point, key: str):
#     if isinstance(point, dict):
#         return point.get(key)
#     return getattr(point, key, None)


# @st.cache_data(show_spinner=False)
# def load_ppt_reference_images(pptx_path: str, max_images: int = 6):
#     if not os.path.exists(pptx_path):
#         return []
#     images = []
#     try:
#         with zipfile.ZipFile(pptx_path, "r") as zf:
#             media_files = [
#                 n for n in zf.namelist()
#                 if n.startswith("ppt/media/") and n.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
#             ]
#             media_files = sorted(media_files)
#             for media_path in media_files[:max_images]:
#                 try:
#                     images.append((os.path.basename(media_path), zf.read(media_path)))
#                 except Exception:
#                     continue
#     except Exception:
#         return []
#     return images


# def resolve_state_from_map_points(points: list, id_to_state_map: dict, ordered_names: list | None = None):
#     if not points:
#         return None
#     point = points[-1] if isinstance(points, (list, tuple)) else None
#     if point is None:
#         return None

#     # 0) customdata fallback using known payload order:
#     #    [name, id, ...] for both state and district map traces.
#     cd = _point_get(point, "customdata")
#     if isinstance(cd, (list, tuple)):
#         if len(cd) >= 2:
#             cand_id = str(cd[1]).strip() if cd[1] is not None else ""
#             if cand_id:
#                 resolved = id_to_state_map.get(cand_id)
#                 if resolved:
#                     return resolved
#         if len(cd) >= 1:
#             cand_name = str(cd[0]).strip() if cd[0] is not None else ""
#             if cand_name:
#                 for mapped_name in id_to_state_map.values():
#                     if str(mapped_name).strip().lower() == cand_name.lower():
#                         return mapped_name

#         # legacy fallback for unexpected payload shape
#         for item in cd:
#             item_s = str(item).strip() if item is not None else ""
#             if not item_s:
#                 continue
#             resolved = id_to_state_map.get(item_s)
#             if resolved:
#                 return resolved
#             for mapped_name in id_to_state_map.values():
#                 if str(mapped_name).strip().lower() == item_s.lower():
#                     return mapped_name

#     # 1) direct id/location mapping (common for choropleth)
#     for key in ("location", "id"):
#         loc = _point_get(point, key)
#         if loc is not None:
#             resolved = id_to_state_map.get(str(loc))
#             if resolved:
#                 return resolved

#     # 2) index-based fallback (point_index / point_number)
#     if ordered_names:
#         for idx_key in ("point_index", "point_number", "pointIndex", "pointNumber"):
#             idx_val = _point_get(point, idx_key)
#             try:
#                 idx_int = int(idx_val)
#             except Exception:
#                 idx_int = None
#             if idx_int is not None and 0 <= idx_int < len(ordered_names):
#                 return ordered_names[idx_int]

#     # 3) text-based fallback from hover payload
#     for txt_key in ("hovertext", "label", "text", "name"):
#         txt = _point_get(point, txt_key)
#         if txt is None:
#             continue
#         txt_clean = str(txt).strip()
#         if not txt_clean:
#             continue
#         # First try direct exact map value
#         for mapped_name in id_to_state_map.values():
#             if str(mapped_name).strip().lower() == txt_clean.lower():
#                 return mapped_name

#     return None


# def build_state_detail_payload(df: pd.DataFrame, state_name: str):
#     if df is None or df.empty or "state" not in df.columns or not state_name:
#         return pd.DataFrame()
#     target = normalize_state_name(state_name)
#     matched = df[df["state"].fillna("").astype(str).apply(normalize_state_name) == target].copy()
#     return matched


# def normalize_district_name(value: str) -> str:
#     name = str(value or "").upper()
#     name = re.sub(r"[^A-Z0-9]+", " ", name)
#     return re.sub(r"\s+", " ", name).strip()


# @st.cache_data(show_spinner=False)
# def load_state_slug_mapping(path: str):
#     mapping = {}
#     if not os.path.exists(path):
#         return mapping
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             rows = json.load(f)
#         for row in rows:
#             state_name = row.get("name")
#             slug = row.get("slug")
#             if state_name and slug:
#                 mapping[normalize_state_name(state_name)] = slug
#     except Exception:
#         return {}
#     return mapping


# @st.cache_data(show_spinner=False)
# def load_state_geojson_by_slug(slug: str):
#     if not slug:
#         return None
#     path = os.path.join(STATE_GEOJSON_DIR, f"{slug}.geojson")
#     if not os.path.exists(path):
#         return None
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             gj = json.load(f)
#         if isinstance(gj, dict) and isinstance(gj.get("features"), list):
#             return gj
#     except Exception:
#         return None
#     return None


# def resolve_state_slug(state_name: str):
#     if not state_name:
#         return None
#     state_map = load_state_slug_mapping(STATE_LIST_PATH)
#     return state_map.get(normalize_state_name(state_name))


# def _flatten_coordinates(coords):
#     if isinstance(coords, (list, tuple)) and coords:
#         if isinstance(coords[0], (int, float)) and len(coords) >= 2:
#             yield float(coords[0]), float(coords[1])
#         else:
#             for item in coords:
#                 yield from _flatten_coordinates(item)


# def get_geojson_center_zoom(geojson_data: dict):
#     lons, lats = [], []
#     for feature in geojson_data.get("features", []):
#         geometry = feature.get("geometry") or {}
#         coords = geometry.get("coordinates")
#         if coords is None:
#             continue
#         for lon, lat in _flatten_coordinates(coords):
#             lons.append(lon)
#             lats.append(lat)

#     if not lons or not lats:
#         return {"lat": 22.7, "lon": 79.2}, 5.4

#     min_lon, max_lon = min(lons), max(lons)
#     min_lat, max_lat = min(lats), max(lats)
#     span = max(max_lon - min_lon, max_lat - min_lat)

#     if span > 14:
#         zoom = 4.9
#     elif span > 10:
#         zoom = 5.3
#     elif span > 6:
#         zoom = 5.8
#     elif span > 3.5:
#         zoom = 6.3
#     elif span > 2:
#         zoom = 6.9
#     else:
#         zoom = 7.5

#     center = {"lat": (min_lat + max_lat) / 2, "lon": (min_lon + max_lon) / 2}
#     return center, zoom


# def build_district_map_dataframe(state_df: pd.DataFrame, state_geojson: dict):
#     if state_geojson is None:
#         return pd.DataFrame(), {}

#     district_count_map = {}
#     datc_map = {}
#     dotc_map = {}
#     active_map = {}
#     inactive_map = {}
#     black_map = {}

#     if state_df is not None and not state_df.empty:
#         district_series = state_df["district"].fillna("").astype(str).str.strip() if "district" in state_df.columns else pd.Series([], dtype=str)
#         venue_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([], dtype=str)
#         status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([], dtype=str)

#         for idx, district_name in district_series.items():
#             norm = normalize_district_name(district_name)
#             if not norm:
#                 continue
#             district_count_map[norm] = district_count_map.get(norm, 0) + 1

#             vtype = venue_series.get(idx, "")
#             if vtype == "DATC":
#                 datc_map[norm] = datc_map.get(norm, 0) + 1
#             elif vtype == "DOTC":
#                 dotc_map[norm] = dotc_map.get(norm, 0) + 1

#             status_val = status_series.get(idx, "")
#             if status_val == "ACTIVE":
#                 active_map[norm] = active_map.get(norm, 0) + 1
#             elif status_val == "INACTIVE":
#                 inactive_map[norm] = inactive_map.get(norm, 0) + 1
#             elif "BLACK" in status_val:
#                 black_map[norm] = black_map.get(norm, 0) + 1

#     rows = []
#     id_to_district = {}
#     for i, feature in enumerate(state_geojson.get("features", [])):
#         props = feature.get("properties", {}) or {}
#         district_name = props.get("district") or props.get("DISTRICT") or props.get("name") or f"District {i + 1}"
#         norm = normalize_district_name(district_name)
#         feature_id = str(feature.get("id") or props.get("dt_code") or props.get("district_code") or district_name)
#         feature["id"] = feature_id
#         id_to_district[feature_id] = district_name

#         rows.append(
#             {
#                 "id": feature_id,
#                 "district": district_name,
#                 "venue_count": int(district_count_map.get(norm, 0)),
#                 "datc_count": int(datc_map.get(norm, 0)),
#                 "dotc_count": int(dotc_map.get(norm, 0)),
#                 "active_count": int(active_map.get(norm, 0)),
#                 "inactive_count": int(inactive_map.get(norm, 0)),
#                 "black_count": int(black_map.get(norm, 0)),
#             }
#         )

#     return pd.DataFrame(rows), id_to_district


# def compute_dotc_datc_summary(df_in: pd.DataFrame):
#     if df_in is None or df_in.empty:
#         return {
#             "dotc_centres": 0, "datc_centres": 0, "total_centres": 0,
#             "dotc_seats": 0, "datc_seats": 0, "total_seats": 0,
#         }

#     vtype = df_in["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in df_in.columns else pd.Series([""] * len(df_in))
#     seat_series = pd.to_numeric(df_in.get("venue_max_capacity", 0), errors="coerce").fillna(0)

#     dotc_mask = vtype == "DOTC"
#     datc_mask = vtype == "DATC"

#     dotc_centres = int(dotc_mask.sum())
#     datc_centres = int(datc_mask.sum())
#     total_centres = int(len(df_in))

#     dotc_seats = int(seat_series[dotc_mask].sum())
#     datc_seats = int(seat_series[datc_mask].sum())
#     total_seats = int(seat_series.sum())

#     return {
#         "dotc_centres": dotc_centres,
#         "datc_centres": datc_centres,
#         "total_centres": total_centres,
#         "dotc_seats": dotc_seats,
#         "datc_seats": datc_seats,
#         "total_seats": total_seats,
#     }


# def render_dotc_datc_summary(title: str, summary: dict):
#     st.markdown(
#         f"""
#         <div style="border:1px solid #A8BBC8; border-radius:10px; overflow:hidden; background:#F8FBFD;">
#             <div style="padding:8px 12px; border-bottom:1px solid #A8BBC8; font-weight:800; color:#2A3B4C; text-transform:uppercase; letter-spacing:0.4px;">
#                 {html.escape(title)}
#             </div>
#             <table style="width:100%; border-collapse:collapse; font-size:13px;">
#                 <thead>
#                     <tr style="background:#EAF2F7; color:#334155;">
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:left;">Metric</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DOTC</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DATC</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">Total</th>
#                     </tr>
#                 </thead>
#                 <tbody>
#                     <tr>
#                         <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Centre</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_centres", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_centres", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_centres", 0):,}</td>
#                     </tr>
#                     <tr>
#                         <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Seat Capacity</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_seats", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_seats", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_seats", 0):,}</td>
#                     </tr>
#                 </tbody>
#             </table>
#         </div>
#         """,
#         unsafe_allow_html=True,
#     )


# # If protected pages are refreshed without an active session, route to login cleanly.
# if st.session_state.page in {"main", "edit", "process"} and not st.session_state.logged_in:
#     st.session_state.page = "login"
#     st.query_params["page"] = "login"

# # ================= DASHBOARD PAGE =================
# if st.session_state.page == "dashboard":
#     _title_col, _logo_col = st.columns([3, 1], vertical_alignment="top")
#     with _title_col:
#         st.title("Venue Analysis Dashboard")
#     with _logo_col:
#         render_brand_logo()
#     st.markdown('<div style="height: 8px;"></div>', unsafe_allow_html=True)
    
#     if st.session_state.logged_in:
#         st.sidebar.success(f"User: {st.session_state.user}")
#         if st.sidebar.button("Logout"):
#             st.session_state.logged_in, st.session_state.user = False, None
#             st.rerun()
#     else:
#         if st.sidebar.button("Admin Login"):
#             st.session_state.page = "login"
#             st.rerun()

#     all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH) 
#                  if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]
    
#     if not all_files:
#         st.warning("No processed reports found. Admin must run the Analysis Engine.")
#     else:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         data = pd.read_excel(latest_file_path, sheet_name=0)
#         data.columns = data.columns.str.lower().str.replace(" ", "_")
        
#         st.info(f"Dashboard Data: **{os.path.basename(latest_file_path)}**")

#         # --- VENUE SEARCH BAR ---
#         st.markdown("""
#         <div style="background: var(--surface);
#                     border-radius: 12px; padding: 22px 26px; margin: 12px 0 20px 0;
#                     border: 1px solid var(--border); border-left: 5px solid var(--primary); 
#                     box-shadow: var(--shadow-sm);">
#             <div style="color:var(--text); font-size:18px; font-weight:700; letter-spacing:0.5px; margin-bottom:6px; text-transform:uppercase;">
#                 Venue Search
#             </div>
#             <p style="color:var(--text-muted); font-size:13px; margin:0; font-weight:500;">
#                 Search by Venue Name, DMS Code, City, or any keyword to view complete venue profile and metrics
#             </p>
#         </div>
#         """, unsafe_allow_html=True)

#         vs_col1, vs_col2 = st.columns([4, 1], vertical_alignment="center")
#         with vs_col1:
#             venue_search_input = st.text_input(
#                 "Search Venue",
#                 value=st.session_state.venue_search_query,
#                 placeholder="e.g. AP COMPUTER POINT or EST-AR-1161 or ITANAGAR...",
#                 label_visibility="collapsed"
#             )
#         with vs_col2:
#             search_btn = st.button("Search Venue", use_container_width=True, type="primary")

#         # Update session state if button is clicked or if they type and press enter
#         if search_btn or venue_search_input != st.session_state.venue_search_query:
#             st.session_state.venue_search_query = venue_search_input

#         query = st.session_state.venue_search_query.strip()
#         if query:
#             # Search across all text columns
#             search_mask = data.astype(str).apply(
#                 lambda col: col.str.contains(query, case=False, na=False)
#             ).any(axis=1)
#             results = data[search_mask]

#             is_authorized = st.session_state.get("logged_in") and st.session_state.get("user") in ["Admin", "Prafull"]

#             if results.empty:
#                 st.warning(f"No venues found matching **\"{query}\"**. Try a different keyword.")
#                 if st.button("Clear Search", key="clear_search_empty"):
#                     st.session_state.venue_search_query = ""
#                     st.rerun()
#             elif len(results) == 1 and search_btn and is_authorized:
#                 # Single result -> go directly to detail page (only auto-redirect if search button was explicitly clicked)
#                 st.session_state.venue_detail_code = str(results.iloc[0].get('dms_code', ''))
#                 st.session_state.page = "venue_detail"
#                 st.session_state.venue_search_query = "" # Clear search after found
#                 st.rerun()
#             else:
#                 # Multiple results (or single result persistent view) -> show selection list
#                 head_col, clear_col = st.columns([5, 1], vertical_alignment="center")
#                 with head_col:
#                     st.markdown(f"##### Found {len(results)} venue(s) matching **\"{query}\"**")
#                 with clear_col:
#                     if st.button("Clear Search", key="clear_search_results", use_container_width=True):
#                         st.session_state.venue_search_query = ""
#                         st.rerun()
#                 for idx, row in results.iterrows():
#                     dms = str(row.get('dms_code', 'N/A'))
#                     name = str(row.get('venue_name', row.get('name', 'N/A')))
#                     city = str(row.get('city', ''))
#                     region = str(row.get('region', ''))
#                     status = str(row.get('status', ''))
#                     category = str(row.get('category', ''))
                    
#                     # Color coding for category
#                     cat_colors = {"A": "#10B981", "B": "#3B82F6", "C": "#F59E0B", "BL": "#EF4444", "BL-C": "#8B5CF6"}
#                     cat_color = cat_colors.get(category, "#6B7280")
                    
#                     st.markdown(f"""
#                     <div style="background:#fff; border:1px solid #E2E8F0; border-left:4px solid {cat_color};
#                                 border-radius:12px; padding:14px 18px; margin:6px 0;
#                                 display:flex; justify-content:space-between; align-items:center;
#                                 transition:all 0.2s; box-shadow:0 1px 3px rgba(0,0,0,0.05);">
#                         <div>
#                             <div style="font-weight:700; font-size:15px; color:#1A202C;">{html.escape(name)}</div>
#                             <div style="font-size:12px; color:#64748B; margin-top:4px; font-weight:600; letter-spacing:0.3px;">
#                                 <span>CODE: {html.escape(dms)}</span>
#                                 <span style="margin:0 8px; color:#CBD5E0;">|</span>
#                                 <span>CITY: {html.escape(city)}</span>
#                                 <span style="margin:0 8px; color:#CBD5E0;">|</span>
#                                 <span>REGION: {html.escape(region)}</span>
#                             </div>
#                         </div>
#                         <div style="display:flex; align-items:center; gap:8px;">
#                             <span style="background:{cat_color}22; color:{cat_color}; font-weight:700;
#                                         padding:4px 12px; border-radius:999px; font-size:12px;">
#                                 Cat {html.escape(category)}
#                             </span>
#                             <span style="background:#F0F4F8; color:#4A5568; font-weight:600;
#                                         padding:4px 10px; border-radius:999px; font-size:12px;">
#                                 {html.escape(status)}
#                             </span>
#                         </div>
#                     </div>
#                     """, unsafe_allow_html=True)
                    
#                     if st.button(f"View Details \u2192 {dms}", key=f"view_{dms}_{idx}", use_container_width=True, disabled=not is_authorized):
#                         st.session_state.venue_detail_code = dms
#                         st.session_state.page = "venue_detail"
#                         st.session_state.venue_search_query = "" # Clear search upon click
#                         st.rerun()

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)

#         if 'region' in data.columns:
#             data['region'] = data['region'].astype(str).str.title().str.strip()
#             data['region'] = pd.Categorical(data['region'], categories=REGION_ORDER, ordered=True)
#         if 'category' in data.columns:
#             data['category'] = pd.Categorical(data['category'], categories=CAT_ORDER, ordered=True)

#         # Executive sidebar filter header
#         st.sidebar.markdown(
#             """
#             <div class="filters-shell">
#                 <div class="filters-title-row">
#                     <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
#                         <polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"></polygon>
#                     </svg>
#                     <span class="filters-title-text">Dashboard Filters</span>
#                 </div>
#                 <p class="filters-subtitle">Refine live metrics by operational segments</p>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         if st.sidebar.button("Reset All Filters", key="dash_reset_all", use_container_width=True):
#             for k in list(st.session_state.keys()):
#                 if k.startswith("dash_vtype_") or k.startswith("dash_region_") or k.startswith("dash_status_") or k.startswith("dash_abc_"):
#                     if k.endswith("_search"):
#                         st.session_state[k] = ""
#                     elif k.endswith("_all") or k.endswith("_none"):
#                         continue
#                     else:
#                         st.session_state[k] = True
#             st.rerun()
#         st.sidebar.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        
#         v_type_list = sorted(data['venue_type'].dropna().unique().tolist())
#         sel_vtype = sidebar_checkbox_filter("Venue Type", v_type_list, "dash_vtype", expanded=True, columns=1)

#         avail_regs = sorted(data['region'].dropna().unique().tolist())
#         sel_reg = sidebar_checkbox_filter("Region", avail_regs, "dash_region", expanded=False, columns=1)

#         stat_list = sorted(data['status'].dropna().unique().tolist())
#         sel_stat = sidebar_checkbox_filter("Status", stat_list, "dash_status", expanded=False, columns=1)

#         avail_cats = sorted(data['category'].dropna().unique().tolist()) if 'category' in data.columns else []
#         if avail_cats:
#             sel_cat = sidebar_checkbox_filter("ABC Category", avail_cats, "dash_abc", expanded=False, columns=2)
#         else:
#             sel_cat = []

#         st.sidebar.markdown(
#             f"""
#             <div class="filters-summary-card">
#                 <div class="filters-summary-title">Current Selection</div>
#                 <div class="filters-summary-item">Venue Type: <b>{len(sel_vtype)}</b></div>
#                 <div class="filters-summary-item">Region: <b>{len(sel_reg)}</b></div>
#                 <div class="filters-summary-item">Status: <b>{len(sel_stat)}</b></div>
#                 <div class="filters-summary-item">ABC Category: <b>{len(sel_cat)}</b></div>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         st.sidebar.markdown(
#             """
#             <div class="sidebar-actions-shell">
#                 <div class="sidebar-actions-title">Quick Actions</div>
#                 <div class="sidebar-actions-subtitle">Navigate to admin tools instantly</div>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         if st.sidebar.button("Manage Master Data", key="dash_manage_master_sidebar", use_container_width=True):
#             st.session_state.page = "main" if st.session_state.logged_in else "login"
#             st.rerun()
#         if st.sidebar.button("Process Analysis Engine", key="dash_process_engine_sidebar", use_container_width=True):
#             if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
#                 st.session_state.page = "process"
#             elif not st.session_state.logged_in:
#                 st.session_state.page = "login"
#             else:
#                 st.error("Restricted to Admin/Prafull.")
#             st.rerun()
        
#         mask = (data['venue_type'].isin(sel_vtype)) & (data['region'].isin(sel_reg)) & (data['status'].isin(sel_stat))
#         if 'category' in data.columns:
#             mask = mask & (data['category'].isin(sel_cat))
#         df_dash = data[mask].copy()

#         def fmt_pct(part: int, total: int) -> str:
#             if total <= 0:
#                 return "N/A"
#             return f"{(part / total) * 100:.1f}%"

#         def yes_count(df_in, col_name: str) -> int:
#             if col_name not in df_in.columns:
#                 return 0
#             s = df_in[col_name].astype(str).str.strip().str.upper()
#             return int((s == "YES").sum())

#         def numeric_sum(df_in, col_name: str) -> float:
#             if col_name not in df_in.columns:
#                 return 0.0
#             return float(pd.to_numeric(df_in[col_name], errors="coerce").fillna(0).sum())

#         def numeric_nonnull(df_in, col_name: str) -> int:
#             if col_name not in df_in.columns:
#                 return 0
#             return int(pd.to_numeric(df_in[col_name], errors="coerce").notna().sum())

#         total_filtered = int(len(df_dash))
#         total_overall = int(len(data))

#         active_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'ACTIVE'])) if 'status' in df_dash.columns else 0
#         inactive_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'INACTIVE'])) if 'status' in df_dash.columns else 0
#         black_filtered = int(len(df_dash[df_dash['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in df_dash.columns else 0

#         active_overall = int(len(data[data['status'].str.upper() == 'ACTIVE'])) if 'status' in data.columns else 0
#         inactive_overall = int(len(data[data['status'].str.upper() == 'INACTIVE'])) if 'status' in data.columns else 0
#         black_overall = int(len(data[data['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in data.columns else 0

#         if 'category' in data.columns:
#             cat_series_overall = data['category'].astype(str).str.strip().str.upper()
#             cat_counts_overall = cat_series_overall.value_counts()
#         else:
#             cat_counts_overall = {}

#         # --- Row 1 ---
#         m1, m2, m3, m4 = st.columns(4)

#         with m4:
#             pct = (total_filtered/total_overall)*100 if total_overall > 0 else 0
#             metric_card(
#                 "Total Venues",
#                 total_filtered,
#                 pct=pct,
#                 trend=f"Coverage: {fmt_pct(total_filtered, total_overall)}",
#                 tooltip=(f"Overall: {total_overall:,}"),
#             )

#         with m1:
#             pct = (active_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Active Venues",
#                 active_filtered,
#                 pct=pct,
#                 trend=f"Share: {fmt_pct(active_filtered, total_filtered)}",
#                 tooltip=(f"Others: {(total_filtered - active_filtered):,}"),
#             )

#         with m2:
#             pct = (black_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Blacklisted Venues",
#                 black_filtered,
#                 pct=pct,
#                 trend=f"Share: {fmt_pct(black_filtered, total_filtered)}",
#                 tooltip=(f"Others: {(total_filtered - black_filtered):,}"),
#             )

#         with m3:
#             pct = (inactive_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Inactive Venues",
#                 inactive_filtered,
#                 pct=pct,
#                 trend=f"Share: {fmt_pct(inactive_filtered, total_filtered)}",
#                 tooltip=(f"Others: {(total_filtered - inactive_filtered):,}"),
#             )


#         # Insert a custom small gap (e.g., 10 pixels)
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)


#         # --- Interactive Executive Dashboards ---
#         st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#         graph_col1, graph_col2, graph_col3 = st.columns([1, 1, 1.05], gap="large")

#         with graph_col1:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Category Distribution</h3>", unsafe_allow_html=True)
#             if 'category' in df_dash.columns:
#                 cat_series = df_dash['category'].astype(str).str.strip().str.upper()
#                 cat_counts = cat_series.value_counts()

#                 labels = ["A", "B", "C", "BL", "BL-C"]
#                 values = [cat_counts.get(l, 0) for l in labels]
#                 colors = ["#C8F6C2", "#AEEBFE", "#FFF6A3", "#FECCCE", "#E9B6FA"]

#                 chart_rows = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0]
#                 if chart_rows:
#                     pie_labels = [r[0] for r in chart_rows]
#                     pie_values = [r[1] for r in chart_rows]
#                     pie_colors = [r[2] for r in chart_rows]
#                     total_pie = sum(pie_values)
#                     pie_text = [
#                         f"<b>{lab}</b><br>{(val/total_pie)*100:.1f}%"
#                         if total_pie and (val / total_pie) >= 0.06 else ""
#                         for lab, val in zip(pie_labels, pie_values)
#                     ]

#                     fig_pie = go.Figure(data=[go.Pie(
#                         labels=pie_labels,
#                         values=pie_values,
#                         sort=False,
#                         direction="clockwise",
#                         marker=dict(colors=pie_colors, line=dict(color='#FFFFFF', width=2)),
#                         text=pie_text,
#                         textinfo='text',
#                         textposition='inside',
#                         insidetextfont=dict(size=12, color='#243447', family='sans-serif'),
#                         hovertemplate="<b>%{label}</b><br>Count: %{value:,}<br>Share: %{percent}<extra></extra>"
#                     )])
#                     fig_pie.update_layout(
#                         showlegend=True,
#                         legend=dict(
#                             orientation="h",
#                             yanchor="bottom",
#                             y=-0.2,
#                             xanchor="center",
#                             x=0.5,
#                             font=dict(size=11, color="#334155")
#                         ),
#                         margin=dict(t=8, b=26, l=8, r=8),
#                         paper_bgcolor="rgba(0,0,0,0)",
#                         plot_bgcolor="rgba(0,0,0,0)",
#                         height=305
#                     )
#                     st.plotly_chart(fig_pie, use_container_width=True)

#         with graph_col2:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Infrastructure Readiness Index</h3>", unsafe_allow_html=True)
#             cctv_c = yes_count(df_dash, "cctv")
#             pwd_c = yes_count(df_dash, "pwd_venue")
#             genset_c = yes_count(df_dash, "genset_availability")
#             ups_c = yes_count(df_dash, "ups_availability")

#             total_active = total_filtered if total_filtered > 0 else 1

#             r_values = [
#                 (cctv_c/total_active)*100,
#                 (pwd_c/total_active)*100,
#                 (ups_c/total_active)*100,
#                 (genset_c/total_active)*100
#             ]
#             r_labels = ['CCTV', 'PWD Ready', 'UPS', 'Genset']

#             r_values_loop = r_values + [r_values[0]]
#             r_labels_loop = r_labels + [r_labels[0]]

#             fig_radar = go.Figure()
#             fig_radar.add_trace(go.Scatterpolar(
#                 r=r_values_loop,
#                 theta=r_labels_loop,
#                 mode='lines+markers',
#                 fill='toself',
#                 fillcolor='rgba(129, 205, 236, 0.30)',
#                 line=dict(color='#236AA5', width=2.6),
#                 marker=dict(size=6, color='#236AA5', line=dict(color='#FFFFFF', width=1.1)),
#                 hoverinfo="theta+r",
#                 hovertemplate="<b>%{theta}</b><br>Readiness: %{r:.1f}%<extra></extra>"
#             ))
#             fig_radar.update_layout(
#                 polar=dict(
#                     bgcolor='rgba(255,255,255,0.45)',
#                     radialaxis=dict(
#                         visible=True,
#                         range=[0, 100],
#                         tickmode='linear',
#                         tick0=0,
#                         dtick=20,
#                         showline=True,
#                         linewidth=1,
#                         linecolor='#9AAFC3',
#                         gridcolor='#C6D5E5',
#                         gridwidth=1.2,
#                         tickfont=dict(size=10, color='#5F7285')
#                     ),
#                     angularaxis=dict(
#                         showline=True,
#                         linecolor='#AABDD0',
#                         gridcolor='#DCE7F2',
#                         gridwidth=1,
#                         tickfont=dict(size=12, color='#2D3748', family='sans-serif')
#                     )
#                 ),
#                 showlegend=False,
#                 margin=dict(t=20, b=5, l=28, r=28),
#                 paper_bgcolor="rgba(0,0,0,0)",
#                 plot_bgcolor="rgba(0,0,0,0)",
#                 height=305
#             )
#             st.plotly_chart(fig_radar, use_container_width=True)

#         with graph_col3:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>India Coverage Map</h3>", unsafe_allow_html=True)
#             map_df, india_geojson, _, _ = build_india_map_dataframe(df_dash)
#             if india_geojson is None:
#                 st.warning(f"GeoJSON not found at: {INDIA_GEOJSON_PATH}")
#             else:
#                 fig_map_preview = create_india_map_figure(
#                     map_df,
#                     india_geojson,
#                     height=250,
#                     show_data=False,
#                     zoom=2.35,
#                     center=INDIA_MAP_CENTER,
#                     opacity=0.9,
#                     line_color="#7C96B5",
#                     line_width=0.85,
#                 )
#                 if fig_map_preview is not None:
#                     id_to_state_map_preview = {
#                         str(row["id"]): str(row["state"])
#                         for _, row in map_df.iterrows()
#                     }
#                     selected_state_from_preview_click = None
#                     try:
#                         map_event = st.plotly_chart(
#                             fig_map_preview,
#                             use_container_width=True,
#                             key="india_map_preview_v5",
#                             on_select="rerun",
#                             selection_mode=("points",),
#                             config={"displayModeBar": False},
#                         )
#                         selected_state_from_preview_click = resolve_state_from_map_points(
#                             extract_selected_points(map_event),
#                             id_to_state_map_preview,
#                             ordered_names=map_df["state"].astype(str).tolist(),
#                         )
#                     except TypeError:
#                         st.plotly_chart(
#                             fig_map_preview,
#                             use_container_width=True,
#                             key="india_map_preview_static_v5",
#                             config={"displayModeBar": False},
#                         )

#                     if selected_state_from_preview_click:
#                         st.session_state.map_selected_state = selected_state_from_preview_click
#                         st.session_state.page = "india_map_detail"
#                         st.query_params["page"] = "india_map_detail"
#                         st.rerun()

#                     if st.button("Open Detailed Map", key="open_detailed_india_map", use_container_width=True):
#                         st.session_state.page = "india_map_detail"
#                         st.query_params["page"] = "india_map_detail"
#                         st.rerun()
             
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)


#         if not df_dash.empty:
#             st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Status Distribution</h3>", unsafe_allow_html=True)
#             status_col1, status_col2 = st.columns([1.5, 1.1], gap="medium")

#             c_data = df_dash.groupby(['region', 'status'], observed=False).size().reset_index(name='Count')
#             status_color_map = {"ACTIVE": "#81CDEC", "INACTIVE": "#CBD5E1", "BLACK LISTED": "#FECCCE"}
#             fig_stat = px.bar(c_data, x='region', y='Count', color='status', barmode='group', text='Count', color_discrete_map=status_color_map, category_orders={"region": REGION_ORDER})
#             fig_stat.update_traces(textposition='outside', textfont=dict(color='#334155', size=12), cliponaxis=False, marker_line_width=0)
#             fig_stat.update_layout(
#                 paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
#                 font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
#                 legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
#                 margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
#             )
#             with status_col1:
#                 st.plotly_chart(fig_stat, use_container_width=True)

#             pivot_status = df_dash.pivot_table(index='region', columns='status', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
#             status_row_order = [r for r in REGION_ORDER if r in pivot_status.index] + ["Grand Total"]
#             pivot_status = pivot_status.reindex(status_row_order)
            
#             headers_stat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_status.columns]
#             cells_stat = [pivot_status.index.tolist()] + [pivot_status[col].tolist() for col in pivot_status.columns]
            
#             fig_tab_stat = go.Figure(data=[go.Table(
#                 header=dict(values=headers_stat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
#                 cells=dict(values=cells_stat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_status)//2 + 1)][0][:len(pivot_status)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
#             )])
#             fig_tab_stat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")
            
#             with status_col2:
#                 st.plotly_chart(fig_tab_stat, use_container_width=True)


#             if 'category' in df_dash.columns:
#                 st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#                 st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Category Breakdown</h3>", unsafe_allow_html=True)
#                 cat_col1, cat_col2 = st.columns([1.5, 1.1], gap="medium")

#                 cat_data = df_dash.groupby(['region', 'category'], observed=False).size().reset_index(name='Count')
#                 color_map = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
#                 fig_cat = px.bar(cat_data, x='region', y='Count', color='category', barmode='group', text='Count', color_discrete_map=color_map, category_orders={"region": REGION_ORDER, "category": CAT_ORDER})
#                 fig_cat.update_traces(textposition='outside', textfont=dict(color='#334155', size=11), cliponaxis=False, marker_line_width=0)
#                 fig_cat.update_layout(
#                     paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
#                     font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
#                     legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
#                     margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
#                 )

#                 with cat_col1:
#                     st.plotly_chart(fig_cat, use_container_width=True)

#                 pivot_cat = df_dash.pivot_table(index='region', columns='category', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
#                 cat_col_order = [c for c in CAT_ORDER if c in pivot_cat.columns] + ["Grand Total"]
#                 pivot_cat = pivot_cat.reindex(index=status_row_order, columns=cat_col_order)
                
#                 headers_cat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_cat.columns]
#                 cells_cat = [pivot_cat.index.tolist()] + [pivot_cat[col].tolist() for col in pivot_cat.columns]
                
#                 fig_tab_cat = go.Figure(data=[go.Table(
#                     header=dict(values=headers_cat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
#                     cells=dict(values=cells_cat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_cat)//2 + 1)][0][:len(pivot_cat)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
#                 )])
#                 fig_tab_cat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")

#                 with cat_col2:
#                     st.plotly_chart(fig_tab_cat, use_container_width=True)

# # ================= INDIA MAP DETAIL PAGE =================
# elif st.session_state.page == "india_map_detail":
#     map_head_col, map_back_col = st.columns([4, 1], vertical_alignment="center")
#     with map_head_col:
#         st.markdown(
#             """
#             <div style="margin-bottom:8px;">
#                 <h2 style="margin:0; color:#1E293B; font-weight:800;">India Venue Coverage Map</h2>
#                 <p style="margin:4px 0 0 0; color:#64748B; font-size:13px;">Detailed state-wise view (click/hover for counts)</p>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#     with map_back_col:
#         if st.button("\u2190 Back to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.query_params["page"] = "dashboard"
#             st.rerun()

#     all_files = [
#         os.path.join(OUTPUT_PATH, f)
#         for f in os.listdir(OUTPUT_PATH)
#         if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")
#     ]

#     detail_df = pd.DataFrame()
#     if all_files:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         detail_df = pd.read_excel(latest_file_path, sheet_name=0)
#         detail_df.columns = detail_df.columns.str.lower().str.replace(" ", "_")
#         st.caption(f"Data Source: {os.path.basename(latest_file_path)}")
#     else:
#         st.warning("No processed reports found. Showing map boundaries only.")

#     map_df, india_geojson, mapped_states, mapped_venues = build_india_map_dataframe(detail_df)
#     if india_geojson is None:
#         st.error(f"India GeoJSON file not found: {INDIA_GEOJSON_PATH}")
#     else:
#         map_total_states = int(len(map_df))
#         data_total_venues = int(len(detail_df)) if not detail_df.empty else 0

#         mk1, mk2, mk3 = st.columns(3)
#         with mk1:
#             metric_card_compact("States in Map", map_total_states, accent="#7CAFD2")
#         with mk2:
#             metric_card_compact("States with Data", mapped_states, accent="#4B89B8")
#         with mk3:
#             metric_card_compact("Mapped Venues", mapped_venues if mapped_venues else data_total_venues, accent="#236AA5")

#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         fig_map_detail = create_india_map_figure(
#             map_df,
#             india_geojson,
#             height=760,
#             show_data=True,
#             zoom=2.45,
#             center=INDIA_MAP_CENTER,
#             opacity=0.92,
#             map_style="white-bg",
#             color_power=0.55,
#             line_color="#334155",
#             line_width=1.05,
#             restrict_to_india=True,
#         )
#         selected_state_from_click = None
#         id_to_state_map = {
#             str(row["id"]): str(row["state"])
#             for _, row in map_df.iterrows()
#         }

#         if fig_map_detail is not None:
#             try:
#                 detail_map_event = st.plotly_chart(
#                     fig_map_detail,
#                     use_container_width=True,
#                     key="india_map_detail_chart_v5",
#                     on_select="rerun",
#                     selection_mode=("points",),
#                     config={"displayModeBar": False, "scrollZoom": False},
#                 )
#                 selected_points = extract_selected_points(detail_map_event)
#                 selected_state_from_click = resolve_state_from_map_points(
#                     selected_points,
#                     id_to_state_map,
#                     ordered_names=map_df["state"].astype(str).tolist(),
#                 )
#                 selected_click_signature = get_last_point_signature(selected_points)
#             except TypeError:
#                 st.plotly_chart(
#                     fig_map_detail,
#                     use_container_width=True,
#                     key="india_map_detail_chart_static_v5",
#                     config={"displayModeBar": False, "scrollZoom": False},
#                 )
#                 selected_click_signature = None
#         else:
#             selected_click_signature = None

#         if selected_state_from_click and selected_click_signature:
#             prev_sig = st.session_state.get("_india_map_detail_last_click_sig")
#             if selected_click_signature != prev_sig:
#                 st.session_state["_india_map_detail_last_click_sig"] = selected_click_signature
#                 st.session_state.map_selected_state = selected_state_from_click
#         elif selected_click_signature is None:
#             st.session_state["_india_map_detail_last_click_sig"] = st.session_state.get("_india_map_detail_last_click_sig", None)

#         if selected_state_from_click and not selected_click_signature:
#             st.session_state.map_selected_state = selected_state_from_click

#         state_totals = (
#             map_df.groupby("state", as_index=False)["venue_count"]
#             .max()
#             .sort_values(["venue_count", "state"], ascending=[False, True], kind="stable")
#         )
#         available_states = state_totals["state"].astype(str).tolist()

#         if available_states:
#             if st.session_state.map_selected_state not in available_states:
#                 st.session_state.map_selected_state = available_states[0]

#             selected_state = st.selectbox(
#                 "Selected State",
#                 options=available_states,
#                 index=available_states.index(st.session_state.map_selected_state),
#                 key="india_detail_state_picker",
#             )
#             if selected_state != st.session_state.map_selected_state:
#                 st.session_state.map_selected_state = selected_state
#                 st.session_state["_india_map_detail_last_click_sig"] = None
#             selected_state = st.session_state.map_selected_state
#             st.caption(f"Current State: {selected_state}")
#         else:
#             selected_state = None

#         if selected_state and not detail_df.empty:
#             state_df = build_state_detail_payload(detail_df, selected_state)
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             st.markdown(
#                 f"<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>{html.escape(selected_state)} Detailed Snapshot</h3>",
#                 unsafe_allow_html=True,
#             )
#             if state_df.empty:
#                 st.info(f"No records available for {selected_state} in current output.")
#             else:
#                 status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([""] * len(state_df))
#                 venue_type_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([""] * len(state_df))
#                 total_state = int(len(state_df))
#                 active_state = int((status_series == "ACTIVE").sum())
#                 inactive_state = int((status_series == "INACTIVE").sum())
#                 black_state = int(status_series.str.contains("BLACK", na=False).sum())
#                 datc_state = int((venue_type_series == "DATC").sum())
#                 dotc_state = int((venue_type_series == "DOTC").sum())

#                 sk1, sk2, sk3, sk4, sk5, sk6 = st.columns(6)
#                 with sk1:
#                     metric_card_compact("State Total", total_state, accent="#236AA5")
#                 with sk2:
#                     metric_card_compact("DATC", datc_state, accent="#0EA5E9")
#                 with sk3:
#                     metric_card_compact("DOTC", dotc_state, accent="#6366F1")
#                 with sk4:
#                     metric_card_compact("Active", active_state, accent="#10B981")
#                 with sk5:
#                     metric_card_compact("Inactive", inactive_state, accent="#94A3B8")
#                 with sk6:
#                     metric_card_compact("Blacklisted", black_state, accent="#EF4444")

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 render_dotc_datc_summary(f"{selected_state} Summary", compute_dotc_datc_summary(state_df))

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 state_slug = resolve_state_slug(selected_state)
#                 state_geojson = load_state_geojson_by_slug(state_slug) if state_slug else None

#                 if st.session_state.map_selected_district_state != selected_state:
#                     st.session_state.map_selected_district = None
#                     st.session_state.map_selected_district_state = selected_state

#                 district_map_df = pd.DataFrame()
#                 district_id_map = {}

#                 if state_geojson is not None:
#                     district_map_df, district_id_map = build_district_map_dataframe(state_df, state_geojson)
#                     if not district_map_df.empty:
#                         dplot = district_map_df.copy()
#                         dplot["_color_metric"] = np.power(pd.to_numeric(dplot["venue_count"], errors="coerce").fillna(0).astype(float), 0.55)
#                         fig_district_map = px.choropleth(
#                             dplot,
#                             geojson=state_geojson,
#                             locations="id",
#                             featureidkey="id",
#                             color="_color_metric",
#                             hover_name="district",
#                             custom_data=["district", "id", "venue_count", "datc_count", "dotc_count", "active_count", "inactive_count", "black_count"],
#                             color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
#                             range_color=(0, max(1.0, float(dplot["_color_metric"].max()))),
#                         )
#                         fig_district_map.update_traces(
#                             marker_line_color="#334155",
#                             marker_line_width=1.0,
#                             hovertemplate=(
#                                 "<b>%{hovertext}</b><br>"
#                                 "Total: %{customdata[2]:,}<br>"
#                                 "DATC: %{customdata[3]:,}<br>"
#                                 "DOTC: %{customdata[4]:,}<br>"
#                                 "Active: %{customdata[5]:,}<br>"
#                                 "Inactive: %{customdata[6]:,}<br>"
#                                 "Blacklisted: %{customdata[7]:,}<extra></extra>"
#                             ),
#                         )
                        
#                         id_to_dname = dict(zip(dplot["id"], dplot["district"]))
#                         dlons, dlats, dtexts = [], [], []
#                         if state_geojson and "features" in state_geojson:
#                             for f in state_geojson["features"]:
#                                 fid = f.get("id")
#                                 if fid in id_to_dname:
#                                     cx, cy = _get_feature_centroid(f)
#                                     dlons.append(cx)
#                                     dlats.append(cy)
#                                     dtexts.append(id_to_dname[fid])
                        
#                         if dlons:
#                             fig_district_map.add_trace(go.Scattergeo(
#                                 lon=dlons,
#                                 lat=dlats,
#                                 text=dtexts,
#                                 mode="text",
#                                 textfont=dict(color="#1E293B", size=9, family="Inter, sans-serif"),
#                                 showlegend=False,
#                                 hoverinfo="skip"
#                             ))
                            
#                         fig_district_map.update_geos(fitbounds="locations", visible=False)
#                         fig_district_map.update_layout(
#                             clickmode="event+select",
#                             dragmode="pan",
#                             margin=dict(t=0, b=0, l=0, r=0),
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             height=560,
#                             coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.75),
#                         )

#                         selected_district_from_click = None
#                         try:
#                             district_map_event = st.plotly_chart(
#                                 fig_district_map,
#                                 use_container_width=True,
#                                 key=f"district_map_{state_slug}",
#                                 on_select="rerun",
#                                 selection_mode=("points",),
#                             )
#                             selected_district_from_click = resolve_state_from_map_points(
#                                 extract_selected_points(district_map_event),
#                                 district_id_map,
#                                 ordered_names=district_map_df["district"].astype(str).tolist(),
#                             )
#                         except TypeError:
#                             st.plotly_chart(fig_district_map, use_container_width=True, key=f"district_map_{state_slug}_static")

#                         if selected_district_from_click:
#                             st.session_state.map_selected_district = selected_district_from_click

#                         district_options = district_map_df["district"].astype(str).tolist()
#                         if district_options:
#                             if st.session_state.map_selected_district not in district_options:
#                                 district_options_sorted = (
#                                     district_map_df.sort_values("venue_count", ascending=False)["district"].astype(str).tolist()
#                                 )
#                                 st.session_state.map_selected_district = district_options_sorted[0]

#                             picked_district = st.selectbox(
#                                 f"Selected District in {selected_state}",
#                                 options=district_options,
#                                 index=district_options.index(st.session_state.map_selected_district),
#                                 key=f"district_picker_{state_slug}",
#                             )
#                             if picked_district != st.session_state.map_selected_district:
#                                 st.session_state.map_selected_district = picked_district

#                             selected_district = st.session_state.map_selected_district
#                             st.caption(f"Current District: {selected_district}")
#                             district_df = state_df[
#                                 state_df["district"].fillna("").astype(str).apply(normalize_district_name) == normalize_district_name(selected_district)
#                             ].copy()

#                             st.markdown(
#                                 f"<h4 style='color:#1E293B; margin-top:6px; margin-bottom:8px;'>{html.escape(selected_district)} District Stats</h4>",
#                                 unsafe_allow_html=True,
#                             )

#                             if district_df.empty:
#                                 st.info(f"No data available for {selected_district}.")
#                             else:
#                                 d_status_series = district_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in district_df.columns else pd.Series([""] * len(district_df))
#                                 d_vtype_series = district_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in district_df.columns else pd.Series([""] * len(district_df))
#                                 d_total = int(len(district_df))
#                                 d_datc = int((d_vtype_series == "DATC").sum())
#                                 d_dotc = int((d_vtype_series == "DOTC").sum())
#                                 d_active = int((d_status_series == "ACTIVE").sum())
#                                 d_inactive = int((d_status_series == "INACTIVE").sum())
#                                 d_black = int(d_status_series.str.contains("BLACK", na=False).sum())

#                                 dk1, dk2, dk3, dk4, dk5, dk6 = st.columns(6)
#                                 with dk1:
#                                     metric_card_compact("District Total", d_total, accent="#1D4ED8")
#                                 with dk2:
#                                     metric_card_compact("DATC", d_datc, accent="#0EA5E9")
#                                 with dk3:
#                                     metric_card_compact("DOTC", d_dotc, accent="#6366F1")
#                                 with dk4:
#                                     metric_card_compact("Active", d_active, accent="#10B981")
#                                 with dk5:
#                                     metric_card_compact("Inactive", d_inactive, accent="#94A3B8")
#                                 with dk6:
#                                     metric_card_compact("Blacklisted", d_black, accent="#EF4444")

#                                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                                 render_dotc_datc_summary(f"{selected_district} Summary", compute_dotc_datc_summary(district_df))

#                                 dc1, dc2 = st.columns([1.2, 1], gap="large")
#                                 with dc1:
#                                     d_cat_counts = (
#                                         district_df["category"].fillna("NA").astype(str).str.strip().replace("", "NA").value_counts().reset_index()
#                                         if "category" in district_df.columns else pd.DataFrame(columns=["index", "category"])
#                                     )
#                                     if not d_cat_counts.empty:
#                                         d_cat_counts.columns = ["Category", "Count"]
#                                         fig_d_cat = px.bar(
#                                             d_cat_counts,
#                                             x="Category",
#                                             y="Count",
#                                             text="Count",
#                                             color="Category",
#                                             color_discrete_map={"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"},
#                                         )
#                                         fig_d_cat.update_traces(textposition="outside", cliponaxis=False)
#                                         fig_d_cat.update_layout(
#                                             paper_bgcolor="rgba(0,0,0,0)",
#                                             plot_bgcolor="rgba(0,0,0,0)",
#                                             margin=dict(t=20, b=10, l=10, r=10),
#                                             xaxis=dict(title=""),
#                                             yaxis=dict(title="Venue Count", gridcolor="#E2E8F0"),
#                                             showlegend=False,
#                                             height=300,
#                                         )
#                                         st.plotly_chart(fig_d_cat, use_container_width=True, key=f"district_category_{state_slug}_{normalize_district_name(selected_district)}")
#                                     else:
#                                         st.info("Category data unavailable for this district.")
#                                 with dc2:
#                                     infra_fields = [
#                                         ("CCTV Ready", "cctv", "#3B82F6"),
#                                         ("PWD Ready", "pwd_venue", "#14B8A6"),
#                                         ("UPS Ready", "ups_availability", "#6366F1"),
#                                         ("Genset Ready", "genset_availability", "#0EA5E9"),
#                                     ]
#                                     for label, col_name, color_code in infra_fields:
#                                         yes_count = 0
#                                         if col_name in district_df.columns:
#                                             yes_count = int((district_df[col_name].fillna("").astype(str).str.strip().str.upper() == "YES").sum())
#                                         pct_val = (yes_count / d_total) * 100 if d_total > 0 else 0
#                                         metric_card_compact(label, f"{pct_val:.1f}%", accent=color_code)
#                         else:
#                             st.info("District list not available for this state map.")
#                 else:
#                     st.warning(f"District map file not available for {selected_state}.")

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 sd_col1, sd_col2 = st.columns([1.25, 1], gap="large")
#                 with sd_col1:
#                     if "district" in state_df.columns:
#                         district_counts = (
#                             state_df["district"]
#                             .fillna("Unknown")
#                             .astype(str)
#                             .str.strip()
#                             .replace("", "Unknown")
#                             .value_counts()
#                             .head(15)
#                             .reset_index()
#                         )
#                         district_counts.columns = ["District", "Venues"]
#                         fig_dist = px.bar(
#                             district_counts,
#                             x="District",
#                             y="Venues",
#                             text="Venues",
#                             color_discrete_sequence=["#4299E1"],
#                         )
#                         fig_dist.update_traces(textposition="outside", cliponaxis=False)
#                         fig_dist.update_layout(
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             xaxis=dict(title="", tickangle=-30, showgrid=False),
#                             yaxis=dict(title="Venue Count", showgrid=True, gridcolor="#E2E8F0"),
#                             margin=dict(t=20, b=10, l=10, r=10),
#                             height=320,
#                             showlegend=False,
#                         )
#                         st.plotly_chart(fig_dist, use_container_width=True, key=f"district_breakdown_{normalize_state_name(selected_state)}")
#                     else:
#                         st.info("District-wise details are not available in this dataset.")
#                 with sd_col2:
#                     if "category" in state_df.columns:
#                         cat_counts = state_df["category"].fillna("NA").astype(str).str.strip().replace("", "NA").value_counts().reset_index()
#                         cat_counts.columns = ["Category", "Count"]
#                         fig_cat_state = px.pie(
#                             cat_counts,
#                             names="Category",
#                             values="Count",
#                             color="Category",
#                             color_discrete_map={"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"},
#                             hole=0.35,
#                         )
#                         fig_cat_state.update_layout(
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             margin=dict(t=20, b=10, l=10, r=10),
#                             height=320,
#                             legend=dict(orientation="h", y=-0.12),
#                         )
#                         st.plotly_chart(fig_cat_state, use_container_width=True, key=f"category_breakdown_{normalize_state_name(selected_state)}")
#                     else:
#                         st.info("Category-wise details are not available in this dataset.")

#                 if normalize_state_name(selected_state) == "BIHAR":
#                     st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                     st.markdown("##### Bihar District Reference (Attached PPT)")
#                     ref_images = load_ppt_reference_images(BIHAR_REFERENCE_PPTX_PATH, max_images=6)
#                     if ref_images:
#                         img_col1, img_col2 = st.columns(2, gap="medium")
#                         for idx, (img_name, img_bytes) in enumerate(ref_images):
#                             with (img_col1 if idx % 2 == 0 else img_col2):
#                                 st.image(img_bytes, caption=f"Reference {idx + 1}: {img_name}", use_container_width=True)
#                     else:
#                         st.info("Reference PPT found, but image previews were not extractable.")
#                     if os.path.exists(BIHAR_REFERENCE_PPTX_PATH):
#                         with open(BIHAR_REFERENCE_PPTX_PATH, "rb") as ref_file:
#                             st.download_button(
#                                 "Download Bihar District Reference PPT",
#                                 data=ref_file.read(),
#                                 file_name=os.path.basename(BIHAR_REFERENCE_PPTX_PATH),
#                                 mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
#                                 use_container_width=False,
#                             )

#         state_summary = state_totals[state_totals["venue_count"] > 0].head(12)
#         if not state_summary.empty:
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             with st.expander("Top States by Venue Count", expanded=False):
#                 st.dataframe(
#                     state_summary.rename(columns={"state": "State", "venue_count": "Venue Count"})[["State", "Venue Count"]],
#                     use_container_width=True,
#                     hide_index=True,
#                 )

# # ================= LOGIN PAGE =================
# elif st.session_state.page == "login":
#     # Inject large top spacing to vertically center the card mentally
#     st.markdown('<div class="gap-lg"></div><div class="gap-lg"></div>', unsafe_allow_html=True)
    
#     col_l, col_center, col_r = st.columns([1, 1.2, 1])
    
#     with col_center:
#         # High-End Security Icon & Header
#         st.markdown(f"""
#         <div style="text-align: center; margin-bottom: 30px;">
#             <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect>
#                 <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
#             </svg>
#             <h2 style="color:var(--primary-dark); font-weight:800; text-transform:uppercase; letter-spacing:1px; margin-top:16px;">Secure Executive Portal</h2>
#             <p style="color:#64748B; font-size:14px; margin-top:-8px;">Authorized Administrative Personnel Only</p>
#         </div>
#         """, unsafe_allow_html=True)
        
#         with st.form("login_form", border=True):
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             u = st.text_input("Username", placeholder="e.g. Admin")
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             p = st.text_input("Password", type="password", placeholder="\u2022\u2022\u2022\u2022\u2022\u2022\u2022\u2022")
#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
            
#             # Using primary type for maximum focus
#             if st.form_submit_button("Authenticate Access", type="primary", use_container_width=True):
#                 if u in USERS and USERS[u] == p:
#                     st.session_state.logged_in, st.session_state.user, st.session_state.page = True, u, "dashboard"
#                     st.rerun()
#                 else: 
#                     st.error("Authentication Failed: Invalid Credentials")
                
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()

# # ================= MASTER DATA VIEW =================
# elif st.session_state.page == "main":
#     allowed = [k for k, v in FILES.items() if st.session_state.user in v["users"]]
#     st.session_state.dataset_key = st.sidebar.radio("ACTIVE MASTER", allowed)
#     if st.query_params.get("dataset", None) != st.session_state.dataset_key:
#         st.query_params["dataset"] = st.session_state.dataset_key
#     current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
#     current_df = load_smart_data(current_path)

#     # --- CEO Level Header Layout ---
#     head_col1, head_col2 = st.columns([3, 1])
#     with head_col1:
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:14px; margin-bottom: 10px;">
#             <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <ellipse cx="12" cy="5" rx="9" ry="3"></ellipse>
#                 <path d="M21 12c0 1.66-4 3-9 3s-9-1.34-9-3"></path>
#                 <path d="M3 5v14c0 1.66 4 3 9 3s9-1.34 9-3V5"></path>
#             </svg>
#             <div>
#                 <h2 style="margin:0; font-weight:800; color:#1E293B;">Master Data Center</h2>
#                 <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Active Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
#             </div>
#         </div>
#         """, unsafe_allow_html=True)

#     with head_col2:
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()

#     if st.session_state.dataset_key == "Venue Inventory" and not current_df.empty:
#         status_series_main = (
#             current_df["status"].astype(str).str.strip().str.upper()
#             if "status" in current_df.columns else pd.Series([""] * len(current_df))
#         )
#         main_active = int((status_series_main == "ACTIVE").sum())
#         main_inactive = int((status_series_main == "INACTIVE").sum())
#         main_black = int(status_series_main.str.contains("BLACK", na=False).sum())
#         main_total = int(len(current_df))

#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         mk1, mk2, mk3, mk4 = st.columns(4, gap="small")
#         with mk1:
#             metric_card_compact("Active Venues", main_active, accent="#86d7a8")
#         with mk2:
#             metric_card_compact("Blacklisted", main_black, accent="#f4a3b5")
#         with mk3:
#             metric_card_compact("Inactive", main_inactive, accent="#b8c6d9")
#         with mk4:
#             metric_card_compact("Total Venues", main_total, accent="#81cdec")

#     st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
    
#     # --- Action Panel & Metrics ---
#     act_col1, act_col2, ext_col = st.columns([1.5, 3, 1.5])
#     with act_col1:
#         if st.button("\u2699 Open Advanced Editor", type="primary", use_container_width=True): 
#             st.session_state.page = "edit"
#             st.rerun()
            
#     with act_col2:
#         # Display dataset health quickly as premium pill badges
#         st.markdown(f"""
#         <div style="display:flex; gap:15px; padding-top:6px;">
#             <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Rows Scanned: {len(current_df):,}</span>
#             <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Columns Mapped: {len(current_df.columns)}</span>
#         </div>
#         """, unsafe_allow_html=True)
    
#     st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#     st.dataframe(current_df, use_container_width=True, height=600)


# # ================= EDIT PAGE =================
# elif st.session_state.page == "edit":
#     st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
#     st.markdown(f"""
#     <div style="display:flex; align-items:center; gap:14px; margin-bottom: 25px;">
#         <svg width="36" height="36" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
#             <polyline points="16 18 22 12 16 6"></polyline>
#             <polyline points="8 6 2 12 8 18"></polyline>
#         </svg>
#         <div>
#             <h2 style="margin:0; font-weight:800; color:#1E293B;">Advanced Data Controller</h2>
#             <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Target Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
#         </div>
#     </div>
#     """, unsafe_allow_html=True)

#     current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
#     current_df = load_smart_data(current_path)

#     # --- TOP ACTIONS & FILTER ROW ---
#     st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#     col_search, col_save, col_clear, col_exit = st.columns([3.5, 1.5, 1.2, 1.2], gap="small")
    
#     with col_search:
#         search = st.text_input("Search", key="search_main_box", placeholder="\U0001F50D Search live database by keywords, target names...", label_visibility="collapsed")
#     with col_save: 
#         save_trigger = st.button("\U0001F4BE Commit Changes", type="primary", use_container_width=True)
#     with col_clear:
#         if st.button("\U0001F9F9 Clear", use_container_width=True):
#             st.session_state.search_main_box = ""
#             st.rerun()
#     with col_exit: 
#         if st.button("\u2190 Back", use_container_width=True):
#             st.session_state.page = "main"
#             st.rerun()
#     filtered = current_df.copy()
#     if search:
#         filtered = filtered[filtered.astype(str).apply(lambda x: x.str.contains(search, case=False).any(), axis=1)]

#     # --- BULK TOOLS (CLEANING & DUPLICATE CHECK) ---
#     if st.session_state.dataset_key != "Call Logs":
#         st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
#         with st.container(border=True):
#             st.subheader("Bulk Injection & Extraction Tools")
#             tool1, tool2 = st.columns([1, 2], gap="large")
            
#             with tool1:
#                 st.caption("Download Blank Template")
#                 blank_df = pd.DataFrame(columns=current_df.columns)
#                 buffer = io.BytesIO()
#                 with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#                     blank_df.to_excel(writer, index=False)
#                 st.download_button(label="\U0001F4E5 Download Format", data=buffer.getvalue(), 
#                                    file_name=f"Format_{FILES[st.session_state.dataset_key]['file']}", mime="application/vnd.ms-excel", use_container_width=True)
            
#             with tool2:
#                 st.caption("Bulk Upload Data (Clean & Update)")
#                 uploaded_file = st.file_uploader("Upload Excel", type=["xlsx"], label_visibility="collapsed")
                
#                 if uploaded_file:
#                     try:
#                         new_data = pd.read_excel(uploaded_file)

#                         d_master = find_dms_col(current_df)
#                         d_new = find_dms_col(new_data)

#                         if d_master and d_new:
#                             # Pre-clean the codes for matching
#                             new_data[d_new] = new_data[d_new].astype(str).str.replace(r'\s+', '', regex=True).str.upper()
#                             current_df[d_master] = current_df[d_master].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

#                             # Check length (informational)
#                             invalid_count = len(new_data[new_data[d_new].str.len() != 11])
#                             if invalid_count > 0:
#                                 st.warning(f"{invalid_count} codes in upload are not 11 digits.")

#                             if st.button("Confirm Bulk Update & Clean", use_container_width=True):
#                                 # 1. Set index to DMS Code for both dataframes
#                                 current_df.set_index(d_master, inplace=True)
#                                 new_data.set_index(d_new, inplace=True)

#                                 # 2. Update existing rows with new data
#                                 current_df.update(new_data)

#                                 # 3. Identify and append completely NEW rows
#                                 new_ids = new_data.index.difference(current_df.index)
#                                 new_rows = new_data.loc[new_ids]
                                
#                                 # 4. Combine and reset index back to original
#                                 final_df = pd.concat([current_df, new_rows]).reset_index()
                                
#                                 # Rename the index column back to the original Master column name if it changed
#                                 final_df.rename(columns={'index': d_master}, inplace=True)

#                                 # 5. Save to File
#                                 final_df.to_excel(current_path, index=False)
#                                 st.success(f"Done! {len(new_data)} rows processed (Updates + New entries).")
#                                 st.rerun()
#                         else:
#                             st.error(f"Missing ID Column! Master needs '{d_master}' and Upload needs a DMS/Venue Code column.")
                    
#                     except Exception as e:
#                         st.error(f"Error processing file: {e}")

#     # --- ML AUTO-TAGGING FOR CALL LOGS ---
#     if st.session_state.dataset_key == "Call Logs":
#         st.markdown("<hr style='margin: 30px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:10px; margin-bottom: 12px;">
#             <svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
#                 <rect x="2" y="2" width="20" height="8" rx="2" ry="2"></rect>
#                 <rect x="2" y="14" width="20" height="8" rx="2" ry="2"></rect>
#                 <line x1="6" y1="6" x2="6.01" y2="6"></line>
#                 <line x1="6" y1="18" x2="6.01" y2="18"></line>
#             </svg>
#             <h3 style="margin:0; color:#1E293B; font-size:18px;">AI Neural Auto-Classification Engine</h3>
#         </div>
#         """, unsafe_allow_html=True)
        
#         ml_col1, ml_col2 = st.columns(2, gap="large")
#         with ml_col1:
#             with st.container(border=True):
#                 st.subheader("1. Model Training Module")
#                 st.caption("Upload historical Call Logs data so the AI can learn human categorizations.")
    
#                 uploaded_train = st.file_uploader(
#                     "Upload Training Batch",
#                     type=["xlsx", "xls", "csv"],
#                     key="ml_train_uploader",
#                     label_visibility="collapsed",
#                     help="File must contain FINAL_ISSUE_REMARK_BY_PRAFULL (and other target columns if available).",
#                 )

#             if st.button("Train/Update ML Model", key="ml_train_btn", use_container_width=True):
#                 with st.spinner("Training Model... This may take a minute..."):
#                     import ml_engine

#                     if not uploaded_train:
#                         st.error("Please upload a training Excel file first.")
#                     else:
#                         train_ext = os.path.splitext(uploaded_train.name)[1].lower()
#                         if train_ext not in [".xlsx", ".xls", ".csv"]:
#                             train_ext = ".xlsx"
#                         train_path = os.path.join(BASE_PATH, f"temp_train_upload{train_ext}")
#                         with open(train_path, "wb") as f:
#                             f.write(uploaded_train.getbuffer())

#                         success, msg = ml_engine.train_ml_model(train_path)
#                         if success:
#                             st.success(f"{msg} (Source: uploaded file)")
#                         else:
#                             st.error(msg)
#         with ml_col2:
#             with st.container(border=True):
#                 st.subheader("2. Live Processing Module")
#                 st.caption("Upload newly downloaded raw Call Logs to classify and format them automatically.")
                
#                 uploaded_raw = st.file_uploader("Upload Raw Data Stream", type=["xlsx", "xls", "csv"], key="ml_uploader", label_visibility="collapsed")
#                 if uploaded_raw and st.button("\u26A1 Process with AI Engine", type="primary", use_container_width=True):
#                     with st.spinner("Processing..."):
#                         import ml_engine
#                         raw_ext = os.path.splitext(uploaded_raw.name)[1].lower()
#                         if raw_ext not in [".xlsx", ".xls", ".csv"]:
#                             raw_ext = ".xlsx"
#                         temp_path = os.path.join(BASE_PATH, f"temp_raw_upload{raw_ext}")
                        
#                         # --- AUTO-COMBINE Gateway/Core/Mock sheets if present ---
#                         try:
#                             if raw_ext in [".xlsx", ".xls"]:
#                                 xl = pd.ExcelFile(io.BytesIO(uploaded_raw.getvalue()))
#                                 sheet_names = xl.sheet_names
                                
#                                 # Check for Gateway/Core/Mock sheets
#                                 gwcm_sheets = [s for s in sheet_names if any(k in s.lower() for k in ["gateway", "core", "mock"])]
                                
#                                 if len(gwcm_sheets) > 1:
#                                     # Auto-combine multiple Gateway/Core/Mock sheets
#                                     st.info(f"Auto-combining {len(gwcm_sheets)} Gateway/Core/Mock sheets: {', '.join(gwcm_sheets)}")
#                                     frames = []
#                                     for s in gwcm_sheets:
#                                         df_part = xl.parse(s)
#                                         s_low = str(s).lower()
#                                         if "gateway" in s_low:
#                                             data_source = "GATEWAY"
#                                         elif "core" in s_low:
#                                             data_source = "CORE"
#                                         elif "mock" in s_low:
#                                             data_source = "MOCK"
#                                         else:
#                                             data_source = str(s).upper()
#                                         df_part.insert(0, "DATA_SOURCE", data_source)
#                                         df_part.insert(1, "SOURCE_SHEET", s)
#                                         frames.append(df_part)
                                    
#                                     combined_df = pd.concat(frames, ignore_index=True, sort=False)
#                                     # Save combined to single Excel file
#                                     with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
#                                         combined_df.to_excel(writer, sheet_name="Sheet1", index=False)
#                                     st.success(f"Combined {len(gwcm_sheets)} sheets: {len(combined_df):,} total rows")
#                                 else:
#                                     # Single sheet or no Gateway/Core/Mock sheets - save as is
#                                     with open(temp_path, "wb") as f:
#                                         f.write(uploaded_raw.getbuffer())
#                             else:
#                                 # CSV file - save as is
#                                 with open(temp_path, "wb") as f:
#                                     f.write(uploaded_raw.getbuffer())
#                         except Exception as e:
#                             st.warning(f"Could not auto-combine sheets: {e}")
#                             with open(temp_path, "wb") as f:
#                                 f.write(uploaded_raw.getbuffer())
                        
#                         # --- Process with ML ---
#                         res = ml_engine.process_call_logs(temp_path)
#                         if res.get("success"):
#                             btn1_file = res.get("output_file_btn1", res.get("output_file"))
#                             btn1_path = res.get("output_path_btn1", res.get("output_path"))
#                             btn2_file = res.get("output_file_btn2")
#                             btn2_path = res.get("output_path_btn2")

#                             if btn2_file:
#                                 st.success(
#                                     f"Processing Complete! Generated files: {btn1_file} and {btn2_file}"
#                                 )
#                             else:
#                                 st.success(f"Processing Complete! Saved to: {btn1_file}")
#                             ml_output = None
#                             try:
#                                 if btn1_path:
#                                     ml_output = pd.read_excel(btn1_path)
#                             except Exception:
#                                 pass
                            
#                             if "auto_approved" in res and "needs_review" in res and "auto_approved_threshold" in res:
#                                 st.write(
#                                     f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}"
#                                     f" | Auto-approved (\u2265{res['auto_approved_threshold']:.2f}): {res['auto_approved']}"
#                                     f" | Needs review: {res['needs_review']}"
#                                 )
#                             else:
#                                 st.write(f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}")
                            
#                             # Show preview if output was loaded successfully
#                             if ml_output is not None and not ml_output.empty:
#                                 st.info("Processed Data Preview (AI-tagged output)")
#                                 st.dataframe(ml_output.head(100), use_container_width=True)
                            
#                             dl_col1, dl_col2 = st.columns(2)
#                             with dl_col1:
#                                 if btn1_path and btn1_file:
#                                     with open(btn1_path, "rb") as f:
#                                         st.download_button(
#                                             label="Download Combined (New + Training)",
#                                             data=f,
#                                             file_name=btn1_file,
#                                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                                             use_container_width=True
#                                         )
#                             with dl_col2:
#                                 if btn2_path and btn2_file:
#                                     with open(btn2_path, "rb") as f:
#                                         st.download_button(
#                                             label="Download Combined (Training + New)",
#                                             data=f,
#                                             file_name=btn2_file,
#                                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                                             use_container_width=True
#                                         )
#                         else:
#                             st.error(res.get("msg", "Unknown error occurred."))

#     # --- DATA EDITOR ---
#     edited = st.data_editor(filtered, use_container_width=True, height=500, num_rows="dynamic", key="data_editor_tool")

#     if save_trigger:
#         # Manual Edit Cleaning
#         d_edit = find_dms_col(edited)
#         if d_edit:
#             edited[d_edit] = edited[d_edit].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

#         if search:
#             current_df.update(edited)
#             current_df.to_excel(current_path, index=False)
#         else:
#             edited.to_excel(current_path, index=False)
#         st.success("Changes Saved & Cleaned Successfully!")
#         st.rerun()
        

# # ================= PROCESS PAGE =================
# elif st.session_state.page == "process":
#     # --- CEO Level Header Layout ---
#     head_col1, head_col2 = st.columns([3, 1])
#     with head_col1:
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:14px; margin-bottom: 5px;">
#             <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"></polygon>
#             </svg>
#             <div>
#                 <h2 style="margin:0; font-weight:800; color:#1E293B;">Central Analysis Engine</h2>
#                 <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Core Processing & Consolidation Node</p>
#             </div>
#         </div>
#         """, unsafe_allow_html=True)

#     with head_col2:
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()
    
#     st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)

#     with st.container(border=True):
#         st.markdown(f"""
#         <div style="margin-bottom: 20px;">
#             <h3 style="margin:0; color:#1E293B; font-size:18px;">Full Ecosystem Compilation</h3>
#             <p style="margin-top:4px; color:#64748B; font-size:14px;">Running the core engine will securely sequence Blacklisted Venues, FFA Data, Daily Observations, Market Reviews, and Machine-Learning Auto-Tagging into a singular unified Master Database.</p>
#         </div>
#         """, unsafe_allow_html=True)
        
#         col_run, col_space = st.columns([2, 3])
#         with col_run:
#             if st.button("\u26A1 Execute Core Analysis Consolidation", type="primary", use_container_width=True):
#                 status = st.empty()
#                 with st.spinner("Initializing Heavy Processing Sequence..."):
#                     try:
#                         # 1. LOAD
#                         vms = load_smart_data(os.path.join(BASE_PATH, "VMS Master.xlsx"))
#                         blacklisted = load_smart_data(os.path.join(BASE_PATH, "Blacklisted Venues.xlsx"))
#                         ffa = load_smart_data(os.path.join(BASE_PATH, "FFA Data.xlsx"))
#                         deo = load_smart_data(os.path.join(BASE_PATH, "Day End Observations.xlsx"))
#                         suspicious = load_smart_data(os.path.join(BASE_PATH, "Suspicious Activities.xlsx"))
#                         market_review = load_smart_data(os.path.join(BASE_PATH, "Google review and ratings.xlsx"))
#                         call_logs = load_smart_data(os.path.join(BASE_PATH, "Master Call Logs.xlsx"))
    
#                         # 2. ALIGN
#                         mapping = {"VENUE_CODE": "dms_code", "DMS_CODE": "dms_code", "DMS_Code": "dms_code", "DMS Code": "dms_code", "UPDATED_DMS_CODE": "dms_code", "DMS CODE": "dms_code", "Row Labels": "dms_code"}
#                         def align_and_lower(df_map):
#                             df_map.columns = df_map.columns.str.strip()
#                             for old, new in mapping.items():
#                                 if old in df_map.columns: df_map = df_map.rename(columns={old: new})
#                             df_map.columns = df_map.columns.str.lower().str.replace(" ", "_")
#                             return df_map
    
#                         vms, blacklisted, ffa = align_and_lower(vms), align_and_lower(blacklisted), align_and_lower(ffa)
#                         deo, suspicious, market_review, call_logs = align_and_lower(deo), align_and_lower(suspicious), align_and_lower(market_review), align_and_lower(call_logs)
    
#                         # 3. CLEANING
#                         def clean_dms(df_c, col):
#                             if col in df_c.columns:
#                                 df_c[col] = df_c[col].astype(str).str.replace(r'\.0$', '', regex=True).str.replace(r'\s+', '', regex=True).str.strip().str.upper()
#                             return df_c
#                         for d in [vms, blacklisted, ffa, deo, suspicious, market_review, call_logs]: clean_dms(d, "dms_code")
#                         vms = vms[vms["dms_code"].notna() & (~vms["dms_code"].isin(["NAN", "", "GRANDTOTAL"]))]
    
#                         # 4. MERGING
#                         rem_bl = 'remark' if 'remark' in blacklisted.columns else 'remarks'
#                         blacklisted["blacklist_info"] = blacklisted[rem_bl].fillna("").astype(str) + " by " + blacklisted["blacklisted_by"].fillna("").astype(str)
#                         merged = vms.merge(blacklisted[["dms_code", "blacklist_info"]], on="dms_code", how="left").fillna({"blacklist_info": ""})
#                         ffa_count = ffa.groupby("dms_code").size().reset_index(name="ffa_frequency")
#                         merged = merged.merge(ffa_count, on="dms_code", how="left").fillna({"ffa_frequency": 0})
    
#                         rem_deo = 'remark' if 'remark' in deo.columns else 'remarks'
#                         deo[rem_deo] = deo[rem_deo].fillna("").astype(str).str.strip()
#                         deo_summary = deo.groupby(["dms_code", rem_deo]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x[rem_deo], x["count"])), include_groups=False).reset_index(name="deo_remarks_summary")
#                         merged = merged.merge(deo_summary, on="dms_code", how="left").fillna({"deo_remarks_summary": ""})
    
#                         suspicious["issue_type"] = suspicious["issue_type"].fillna("").astype(str).str.strip()
#                         susp_summary = suspicious.groupby(["dms_code", "issue_type"]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x["issue_type"], x["count"])), include_groups=False).reset_index(name="suspicious_summary")
#                         merged = merged.merge(susp_summary, on="dms_code", how="left").fillna({"suspicious_summary": ""})

#                         # 4A. Call Logs Enrichment (moved from Call Logs AI module to Central Analysis Engine)
#                         def _first_non_blank(series: pd.Series) -> str:
#                             s = series.fillna("").astype(str).str.strip()
#                             s = s[~s.str.upper().isin(["", "NAN", "NONE", "NOT ASSIGNED", "NA", "-"])]
#                             return s.iloc[0] if not s.empty else ""

#                         if not call_logs.empty:
#                             if "project_name" in call_logs.columns:
#                                 call_logs["updated_project_name"] = clean_project_name(call_logs["project_name"])

#                             cl_dms_source = (
#                                 "updated_dms_code" if "updated_dms_code" in call_logs.columns else
#                                 "center_code" if "center_code" in call_logs.columns else
#                                 "dms_code" if "dms_code" in call_logs.columns else None
#                             )
#                             if cl_dms_source:
#                                 call_logs["updated_dms_code"] = (
#                                     call_logs[cl_dms_source]
#                                     .astype(str)
#                                     .str.replace(r"\.0$", "", regex=True)
#                                     .str.replace(r"\s+", "", regex=True)
#                                     .str.strip()
#                                     .str.upper()
#                                 )
#                                 if "dms_code" not in call_logs.columns:
#                                     call_logs["dms_code"] = call_logs["updated_dms_code"]
#                                 else:
#                                     call_logs["dms_code"] = call_logs["dms_code"].astype(str).str.strip()
#                                     blank_mask = call_logs["dms_code"].str.upper().isin(["", "NAN", "NONE"])
#                                     call_logs.loc[blank_mask, "dms_code"] = call_logs.loc[blank_mask, "updated_dms_code"]

#                             clean_dms(call_logs, "dms_code")
#                             cl_valid = call_logs[call_logs["dms_code"].notna()].copy()
#                             cl_valid["dms_code"] = cl_valid["dms_code"].astype(str).str.strip().str.upper()
#                             cl_valid = cl_valid[~cl_valid["dms_code"].isin(["", "NAN", "NONE", "GRANDTOTAL"])]

#                             if not cl_valid.empty and "dms_code" in vms.columns:
#                                 vms_cols_needed = ["dms_code", "name", "address", "landmark", "city", "district", "state", "pincode"]
#                                 vms_ref = vms.copy()
#                                 for vc in vms_cols_needed:
#                                     if vc not in vms_ref.columns:
#                                         vms_ref[vc] = ""
#                                 vms_ref = vms_ref[vms_cols_needed].drop_duplicates(subset="dms_code", keep="first")
#                                 vms_map_name = dict(zip(vms_ref["dms_code"], vms_ref.get("name", "")))
#                                 vms_map_addr = dict(zip(vms_ref["dms_code"], vms_ref.get("address", "")))
#                                 vms_map_land = dict(zip(vms_ref["dms_code"], vms_ref.get("landmark", "")))
#                                 vms_map_city = dict(zip(vms_ref["dms_code"], vms_ref.get("city", "")))
#                                 vms_map_dist = dict(zip(vms_ref["dms_code"], vms_ref.get("district", "")))
#                                 vms_map_state = dict(zip(vms_ref["dms_code"], vms_ref.get("state", "")))
#                                 vms_map_pin = dict(zip(vms_ref["dms_code"], vms_ref.get("pincode", "")))

#                                 mapped_name = cl_valid["dms_code"].map(vms_map_name).fillna("")
#                                 venue_fallback = (
#                                     cl_valid["venue_name"].fillna("").astype(str).str.strip().str.upper()
#                                     if "venue_name" in cl_valid.columns else pd.Series([""] * len(cl_valid), index=cl_valid.index)
#                                 )
#                                 cl_valid["updated_venue_name"] = mapped_name.where(mapped_name.astype(str).str.strip() != "", venue_fallback)

#                                 cl_valid["address"] = cl_valid["dms_code"].map(vms_map_addr).fillna(cl_valid.get("address", ""))
#                                 cl_valid["landmark"] = cl_valid["dms_code"].map(vms_map_land).fillna(cl_valid.get("landmark", ""))
#                                 cl_valid["city"] = cl_valid["dms_code"].map(vms_map_city).fillna(cl_valid.get("city", ""))
#                                 cl_valid["district"] = cl_valid["dms_code"].map(vms_map_dist).fillna(cl_valid.get("district", ""))
#                                 cl_valid["state"] = cl_valid["dms_code"].map(vms_map_state).fillna(cl_valid.get("state", ""))
#                                 cl_valid["pincode"] = cl_valid["dms_code"].map(vms_map_pin).fillna(cl_valid.get("pincode", ""))
#                                 cl_valid["concat_address"] = build_concat_address(cl_valid)

#                                 profile_aggs = {}
#                                 if "updated_project_name" in cl_valid.columns:
#                                     profile_aggs["call_logs_project_name"] = ("updated_project_name", _first_non_blank)
#                                 if "updated_venue_name" in cl_valid.columns:
#                                     profile_aggs["call_logs_updated_venue_name"] = ("updated_venue_name", _first_non_blank)
#                                 if "concat_address" in cl_valid.columns:
#                                     profile_aggs["call_logs_concat_address"] = ("concat_address", _first_non_blank)

#                                 if profile_aggs:
#                                     call_logs_profile = cl_valid.groupby("dms_code", as_index=False).agg(**profile_aggs)
#                                     merged = merged.merge(call_logs_profile, on="dms_code", how="left")
#                                     for col_name in profile_aggs.keys():
#                                         merged[col_name] = merged[col_name].fillna("")
    
#                         if "final_issue_remark_by_prafull" in call_logs.columns:
                            
#                             # Remove 'Other low impact issues' directly from remark column
#                             filtered_logs = call_logs[
#                                 ~call_logs["final_issue_remark_by_prafull"]
#                                 .fillna("")
#                                 .str.strip()
#                                 .str.lower()
#                                 .str.contains("other low impact issues", na=False)
#                             ]
                            
#                             cl_summary = (
#                                 filtered_logs
#                                 .groupby(["dms_code", "final_issue_remark_by_prafull"])
#                                 .size()
#                                 .reset_index(name="count")
#                                 .groupby("dms_code")
#                                 .apply(
#                                     lambda x: ", ".join(
#                                         f"{r}({c})" for r, c in zip(
#                                             x["final_issue_remark_by_prafull"], x["count"]
#                                         )
#                                     ),
#                                     include_groups=False
#                                 )
#                                 .reset_index(name="call_log_summary")
#                             )
    
#                             merged = merged.merge(cl_summary, on="dms_code", how="left").fillna({"call_log_summary": ""})
                            
                            
#                         market_review["google_ratings_num"] = pd.to_numeric(market_review["google_ratings"], errors="coerce")
#                         market_review["google_review_count_num"] = pd.to_numeric(market_review["google_review_count"], errors="coerce")
#                         m_cond = [market_review["google_review_count_num"] < 10, (market_review["google_ratings_num"] >= 4) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] <= 2) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] < 4) & (market_review["google_review_count_num"] >= 10)]
#                         m_choice = ["Doubtful (Reviews < 10)", "Good (Rating >=4 & Reviews >=10)", "Poor (Rating <=2 & Reviews >=10)", "Average (Rating <4 & Reviews >=10)"]
#                         market_review["market_remark"] = np.select(m_cond, m_choice, default="")
#                         merged = merged.merge(market_review[["dms_code", "market_remark"]], on="dms_code", how="left").fillna({"market_remark": ""})
    
    
#                         # 5. ANALYSIS RULES
#                         merged["final_analysis"] = "Good"; merged["reason"] = "Good"
#                         mask_comp = merged["blacklist_info"].str.lower().str.startswith("completely blacklisted", na=False)
#                         mask_cust = merged["blacklist_info"].str.lower().str.startswith("customer specific blacklisted", na=False)
                        
#                         # Assigning individually to avoid ValueError
#                         merged.loc[mask_comp, "final_analysis"] = "Completely Blacklisted"
#                         merged.loc[mask_comp, "reason"] = merged.loc[mask_comp, "blacklist_info"]
#                         merged.loc[mask_cust, "final_analysis"] = "Customer Specific Blacklisted"
#                         merged.loc[mask_cust, "reason"] = merged.loc[mask_cust, "blacklist_info"]
    
#                         nb = merged[~(mask_comp | mask_cust)].copy()
                        
#                         mask_ffa = nb["ffa_frequency"] >= 1
#                         nb.loc[mask_ffa, "final_analysis"] = np.where(nb.loc[mask_ffa, "ffa_frequency"] < 5, "Average", "Sensitive")
#                         nb.loc[mask_ffa, "reason"] = "Found in FFA | Count: " + nb.loc[mask_ffa, "ffa_frequency"].astype(int).astype(str)
    
#                         if "deo_remarks_summary" in nb.columns:
#                             mask_deo = (nb["deo_remarks_summary"] != "") & (nb["final_analysis"] == "Good")
#                             deo_counts = deo.groupby("dms_code").size()
#                             nb["deo_count"] = nb["dms_code"].map(deo_counts).fillna(0)
                            
#                             # Force Sensitive Venue if any comment exists
#                             nb.loc[mask_deo, "final_analysis"] = "Sensitive"
#                             nb.loc[mask_deo, "reason"] = nb.loc[mask_deo, "deo_remarks_summary"]
    
#                         if "suspicious_summary" in nb.columns:
#                             mask_susp = (nb["suspicious_summary"] != "") & (nb["final_analysis"] == "Good")
#                             susp_counts = suspicious.groupby("dms_code").size()
#                             nb["susp_count"] = nb["dms_code"].map(susp_counts).fillna(0)
                            
#                             # Force Sensitive Venue if any comment exists
#                             nb.loc[mask_susp, "final_analysis"] = "Sensitive"
#                             nb.loc[mask_susp, "reason"] = nb.loc[mask_susp, "suspicious_summary"]
    
#                         if "call_log_summary" in nb.columns:
#                             remark = nb["call_log_summary"].fillna("").str.strip()
                            
#                             mask_call = (remark != "") & (nb["final_analysis"] == "Good")
                            
#                             # Exclude anything containing 'Other low impact issues'
#                             mask_valid_call = ~remark.str.lower().str.contains("other low impact issues", na=False)
                            
#                             # Final mask
#                             mask_final = mask_call & mask_valid_call
    
#                             nb.loc[mask_final, "final_analysis"] = "Average"
#                             nb.loc[mask_final, "reason"] = "Call Logs | " + nb.loc[mask_final, "call_log_summary"]
                        
#                         if "market_remark" in nb.columns:
#                             # Normalize text
#                             remark = nb["market_remark"].fillna("").str.strip()
#                             call_remark = nb["call_log_summary"].fillna("").str.strip()
                            
#                             # Mask for rows where final_analysis is still Good
#                             mask_base = nb["final_analysis"] == "Good"
                            
#                             # [OK] Apply market logic ONLY if call_log_summary is blank OR only 'Other low impact issues'
#                             mask_call_valid = (
#                                 (call_remark == "") |
#                                 (call_remark.str.lower().str.contains("other low impact issues", na=False))
#                             )
                            
#                             # Final base mask
#                             mask_main = mask_base & mask_call_valid
                            
#                             # Check if remark contains doubtful/poor/average
#                             mask_issue = remark.str.lower().str.contains("doubtful|poor|average", na=False)
                            
#                             # --- Case 1: Issue found -> Average ---
#                             mask_avg = mask_main & mask_issue
#                             nb.loc[mask_avg, "final_analysis"] = "Average"
#                             nb.loc[mask_avg, "reason"] = "Market: " + nb.loc[mask_avg, "market_remark"]
                            
#                             # --- Case 2: Other remarks -> Good ---
#                             mask_good = mask_main & (remark != "") & (~mask_issue)
#                             nb.loc[mask_good, "final_analysis"] = "Good"
#                             nb.loc[mask_good, "reason"] = "Market: " + nb.loc[mask_good, "market_remark"]
                            
#                             # --- Case 3: Blank -> Good Venue ---
#                             mask_blank = mask_main & (remark == "")
#                             nb.loc[mask_blank, "final_analysis"] = "Good"
#                             nb.loc[mask_blank, "reason"] = "Good Venue"
                            
    
#                         final_result = pd.concat([merged[mask_comp | mask_cust], nb], ignore_index=True)
    
#                         # Rename output column: NAME -> VENUE_NAME
#                         final_result = final_result.rename(columns={"name": "venue_name"})
                        
#                         # --- 6. COLUMN WISE ANALYSIS ---
#                         analysis_results = []
#                         for col in final_result.columns:
#                             if col in ['deo_count', 'susp_count']: continue
#                             null_count = (final_result[col] == "").sum() + final_result[col].isna().sum()
#                             analysis_results.append({
#                                 "Column Name": col,
#                                 "Fill Rate (%)": round(((len(final_result) - null_count) / len(final_result)) * 100, 2),
#                                 "Unique Values": final_result[col].nunique(),
#                                 "Top Frequent": str(final_result[col].mode()[0]) if not final_result[col].empty else "N/A"
#                             })
#                         audit_df = pd.DataFrame(analysis_results)
    
#                         # Create Category column based on final_analysis
#                         category_map = {
#                             "Good": "A",
#                             "Average": "B",
#                             "Sensitive": "C",
#                             "Completely Blacklisted": "BL",
#                             "Customer Specific Blacklisted": "BL-C"
#                         }
    
#                         final_result["category"] = final_result["final_analysis"].map(category_map).fillna("")

#                         # DOTC default override: keep DATC untouched
#                         if "venue_type" in final_result.columns:
#                             dotc_mask = final_result["venue_type"].astype(str).str.strip().str.upper() == "DOTC"
#                             final_result.loc[dotc_mask, "final_analysis"] = "Good"
#                             final_result.loc[dotc_mask, "reason"] = "Good Venue"
#                             final_result.loc[dotc_mask, "category"] = "A"
                        
                        
#                         # Cleanup
#                         final_result = final_result.drop(columns=['deo_count', 'susp_count'], errors='ignore')
                        
    
#                         # 7. SAVE
#                         ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
#                         # Convert column names to uppercase for main file
#                         final_result.columns = [col.upper() for col in final_result.columns]
#                         audit_df.columns = [col.upper() for col in audit_df.columns]
    
#                         out_path = os.path.join(OUTPUT_PATH, f"VMS_Full_Analysis_{ts}.xlsx")
    
#                         # If available, also include the latest previous output in the new workbook
#                         prev_out_path = None
#                         try:
#                             prev_candidates = [
#                                 os.path.join(OUTPUT_PATH, f)
#                                 for f in os.listdir(OUTPUT_PATH)
#                                 if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and (not f.startswith("~$"))
#                             ]
#                             prev_candidates = [p for p in prev_candidates if os.path.abspath(p) != os.path.abspath(out_path)]
#                             if prev_candidates:
#                                 prev_out_path = max(prev_candidates, key=os.path.getmtime)
#                         except:
#                             prev_out_path = None
    
#                         with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
#                             final_result.to_excel(writer, sheet_name='Main_Analysis', index=False)
#                             audit_df.to_excel(writer, sheet_name='Column_Audit', index=False)
    
#                             run_info = pd.DataFrame([{
#                                 "CURRENT_OUTPUT_FILE": os.path.basename(out_path),
#                                 "PREVIOUS_OUTPUT_FILE": os.path.basename(prev_out_path) if prev_out_path else "",
#                                 "GENERATED_ON": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#                             }])
#                             run_info.to_excel(writer, sheet_name="Run_Info", index=False)
    
#                             if prev_out_path:
#                                 try:
#                                     old_main = pd.read_excel(prev_out_path, sheet_name="Main_Analysis")
#                                     old_main.to_excel(writer, sheet_name="Old_Main_Analysis", index=False)
#                                 except:
#                                     pass
#                                 try:
#                                     old_audit = pd.read_excel(prev_out_path, sheet_name="Column_Audit")
#                                     old_audit.to_excel(writer, sheet_name="Old_Column_Audit", index=False)
#                                 except:
#                                     pass
    
#                         # --- Inventory File ---
#                         inventory_cols = [
#                             "venue_type", "region", "state", "city", "district", "dms_code",
#                             "venue_name", "address", "pincode", "landmark", "complete_address",
#                             "venue_max_capacity", "pwd_venue", "status", "final_analysis", "reason", "category"
#                         ]
    
#                         inventory_df = final_result[[col.upper() for col in inventory_cols if col.upper() in final_result.columns]].copy()
    
#                         # Folder
#                         inventory_folder = os.path.join(OUTPUT_PATH, "Venue Inventory")
#                         os.makedirs(inventory_folder, exist_ok=True)
    
#                         inventory_path = os.path.join(inventory_folder, f"Venue Inventory as on {ts}.xlsx")
#                         inventory_df.to_excel(inventory_path, index=False)
    
#                         # --- APPLY FORMATTING FUNCTION ---
#                         def format_excel(file_path):
#                             wb = load_workbook(file_path)
                            
#                             header_font = Font(color="FFFFFF", bold=True)
#                             header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
#                             header_alignment = Alignment(horizontal="left")
    
#                             for sheet in wb.sheetnames:
#                                 ws = wb[sheet]
                                
#                                 for cell in ws[1]:  # First row (headers)
#                                     cell.font = header_font
#                                     cell.fill = header_fill
#                                     cell.alignment = header_alignment
    
#                             wb.save(file_path)
    
#                         # Apply formatting to both files
#                         format_excel(out_path)
#                         format_excel(inventory_path)
    
#                         # --- Status ---
#                         status.success(f"Generated Analysis Reports.")
#                         st.dataframe(final_result, use_container_width=True)
    
#                     except Exception as e: st.error(f"Error: {e}"); st.exception(e)


# # ================= VENUE DETAIL PAGE =================
# elif st.session_state.page == "venue_detail":
#     # Load data
#     all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH)
#                  if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]

#     if not all_files:
#         st.error("No analysis data found.")
#         if st.button("\u2190 Back to Dashboard"):
#             st.session_state.page = "dashboard"
#             st.rerun()
#     else:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         data = pd.read_excel(latest_file_path, sheet_name=0)
#         data.columns = data.columns.str.lower().str.replace(" ", "_")

#         venue_code = st.session_state.venue_detail_code
#         venue_row = data[data['dms_code'].astype(str).str.strip().str.upper() == str(venue_code).strip().upper()]

#         if venue_row.empty:
#             st.error(f"Venue with code **{venue_code}** not found.")
#             if st.button("\u2190 Back to Dashboard"):
#                 st.session_state.page = "dashboard"
#                 st.rerun()
#         else:
#             v = venue_row.iloc[0]

#             # Helper to safely get value
#             def _val(col, default="N/A"):
#                 val = v.get(col, default)
#                 if pd.isna(val) or str(val).strip().upper() in ["NAN", "NONE", ""]:
#                     return default
#                 res = str(val).strip()
#                 if res.endswith(".0") and res[:-2].replace('-', '').isdigit():
#                     res = res[:-2]
#                 return res

#             venue_name = _val("venue_name", _val("name", "Unknown Venue"))
#             dms_code = _val("dms_code")
#             category = _val("category")
#             final_analysis = _val("final_analysis")
#             status_val = _val("status")
#             region = _val("region")
#             city = _val("city")
#             state = _val("state")
#             district = _val("district")
#             venue_type = _val("venue_type")

#             # Category color
#             cat_colors = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
#             cat_color = cat_colors.get(category, "#DDE8F0")

#             # Analysis color
#             analysis_colors = {
#                 "Good": "#C8F6C2", "Average": "#FFF6A3", "Sensitive": "#FECCCE",
#                 "Completely Blacklisted": "#FDBED2", "Customer Specific Blacklisted": "#E9B6FA"
#             }
#             analysis_color = analysis_colors.get(final_analysis, "#DDE8F0")

#             # --- TOP BAR ---
#             top_left, top_right = st.columns([1, 4])
#             with top_left:
#                 if st.button("\u2190 Back to Dashboard", use_container_width=True):
#                     st.session_state.page = "dashboard"
#                     st.rerun()

#             # --- HERO HEADER ---
#             st.markdown(f"""
#             <div style="background: linear-gradient(135deg, white 0%, #EBF3F8 40%, #CBEAFC 100%);
#                         border-radius: 16px; padding: 20px 24px; margin: 8px 0 20px 0;
#                         box-shadow: 0 4px 16px rgba(0,0,0,0.05); position:relative; overflow:hidden;
#                         border: 1px solid #DDE8F0;">
#                 <div style="position:absolute; top:-20px; right:-20px; width:140px; height:140px;
#                             background:rgba(174,235,254,0.3); border-radius:50%;"></div>
#                 <div style="position:absolute; bottom:-30px; right:60px; width:90px; height:90px;
#                             background:rgba(174,235,254,0.2); border-radius:50%;"></div>
#                 <div>
#                     <div style="color:#2A3B4C; font-size:28px; font-weight:800; letter-spacing:-0.5px;
#                                 line-height:1.2;">{html.escape(venue_name)}</div>
#                     <div style="color:#566D82; font-size:14px; margin-top:4px; font-weight:600; letter-spacing:0.5px; text-transform:uppercase;">
#                         {html.escape(dms_code)} <span style="margin:0 6px; color:#839BB2;">|</span> {html.escape(venue_type)}
#                     </div>
#                 </div>
#                 <!-- Badges removed as per user request -->
#             </div>
#             """, unsafe_allow_html=True)

#             # --- VENUE PHOTO ---
#             venue_photo_url = _val("venue_photo_url", _val("image_url", ""))
            
#             # Check for local photo if URL is not provided
#             if not venue_photo_url or venue_photo_url in ["N/A", ""]:
#                 import glob
#                 import base64
#                 # Search for any image matching the exact DMS code (e.g. EST-AR-1161.jpg)
#                 photo_matches = glob.glob(os.path.join(PHOTOS_PATH, f"{dms_code}.*"))
#                 if photo_matches:
#                     # Filter for actual images
#                     img_files = [p for p in photo_matches if p.lower().endswith(('.png', '.jpg', '.jpeg', '.webp'))]
#                     if img_files:
#                         with open(img_files[0], "rb") as image_file:
#                             b64_str = base64.b64encode(image_file.read()).decode("utf-8")
#                             ext = os.path.splitext(img_files[0])[1][1:].lower()
#                             ext = "jpeg" if ext == "jpg" else ext
#                             venue_photo_url = f"data:image/{ext};base64,{b64_str}"
            
#             img_col, info_col, metrics_col = st.columns([1.2, 1.8, 1])

#             with img_col:
#                 photo_content = ""
#                 if venue_photo_url and venue_photo_url not in ["N/A", ""]:
#                     photo_content = f"""<div style="border-radius:12px; overflow:hidden; flex:1; min-height:0; background:var(--surface-muted); display:flex; justify-content:center; align-items:center;">
#     <img src="{venue_photo_url}" style="width:100%; height:100%; object-fit:cover;" />
# </div>"""
#                 else:
#                     photo_content = """<div style="border-radius:12px; flex:1; min-height:0; background:var(--surface-muted); border: 1px solid var(--border); display:flex; flex-direction:column; justify-content:center; align-items:center;">
#     <div style="width:48px; height:48px; border-radius:50%; background:var(--border-dark); margin-bottom:12px; display:flex; justify-content:center; align-items:center;">
#         <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="var(--surface)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"></rect><circle cx="8.5" cy="8.5" r="1.5"></circle><polyline points="21 15 16 10 5 21"></polyline></svg>
#     </div>
#     <span style="color:var(--text-muted); font-size:13px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Photo Not Available</span>
# </div>"""

#                 # Render entire block at 0 indentation to prevent Markdown falling into <pre> logic
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto; display:flex; flex-direction:column;">
#     <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Photo</div>
# {photo_content}
# </div>""", unsafe_allow_html=True)

#             with info_col:
#                 address = _val("address", "")
#                 landmark = _val("landmark", "")
#                 pincode = _val("pincode", "")
#                 complete_address = _val("complete_address", "")

#                 loc_items = [
#                     ("Region", region),
#                     ("State", state),
#                     ("City", city),
#                     ("District", district),
#                     ("Pincode", pincode),
#                     ("Landmark", landmark),
#                 ]

#                 loc_html = ""
#                 for label, val in loc_items:
#                     if val != "N/A":
#                         loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border);">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">{html.escape(val)}</span>
# </div>"""

#                 if complete_address != "N/A":
#                     loc_html += f"""<div style="margin-top:12px; padding:10px 14px; background:var(--surface-muted); border-radius:10px; border:1px solid var(--border);">
#     <div style="color:var(--text-muted); font-size:12px; font-weight:600; margin-bottom:4px;">FULL ADDRESS</div>
#     <div style="color:var(--text); font-size:13px; line-height:1.5;">{html.escape(complete_address)}</div>
# </div>"""

#                 google_coords = _val("google_cordinates", "")
#                 if google_coords != "N/A" and google_coords != "":
#                     loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border); margin-top:4px;">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">Google Location</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">
#         <a href="https://maps.google.com/?q={html.escape(google_coords)}" target="_blank" style="color:#2563EB; text-decoration:none;">{html.escape(google_coords)}</a>
#     </span>
# </div>"""

#                 added_on = _val("added_on-date", "")
#                 if added_on != "N/A" and added_on != "":
#                     loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border); margin-top:4px;">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">Added On</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">{html.escape(added_on)}</span>
# </div>"""

#                 header_text = "Location Profile"
#                 if added_on != "N/A" and added_on != "":
#                     header_text = f"Location Profile with added on date"
                    
#                 # Render entire profile block in a single markdown wrapper at 0 indentation
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto;">
#     <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">{header_text}</div>
# {loc_html}
# </div>""", unsafe_allow_html=True)

#             with metrics_col:
#                 def _mini_card(label, value, color="#2563EB"):
#                     return f"""<div style="flex:1; background:var(--surface-muted); border:1px solid var(--border); border-radius:8px; padding:6px; text-align:center; border-left:3px solid {color}; box-sizing:border-box; display:flex; flex-direction:column; justify-content:center; align-items:center;">
# <div style="font-size:17px; font-weight:800; color:{color}; line-height:1;">{html.escape(str(value))}</div>
# <div style="font-size:12px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.4px; margin-top:2px;">{label}</div>
# </div>"""

#                 cctv = _val("cctv")
#                 cctv_color = "#10B981" if cctv.upper() == "YES" else "#EF4444"
#                 pwd = _val("pwd_venue")
#                 pwd_color = "#10B981" if pwd.upper() == "YES" else "#EF4444"
#                 genset = _val("genset_availability")
#                 genset_color = "#10B981" if genset.upper() == "YES" else "#EF4444"
#                 ups = _val("ups_availability")
#                 ups_color = "#10B981" if ups.upper() == "YES" else "#EF4444"

#                 used_tested = _val("used_and_tested_venue")
#                 used_color = "#10B981" if used_tested.upper() == "YES" else "#EF4444"

#                 cards_html = (
#                     _mini_card("Max Capacity", _val("venue_max_capacity"), "#2563EB") +
#                     _mini_card("CCTV", cctv, cctv_color) +
#                     _mini_card("PWD Venue", pwd, pwd_color) +
#                     _mini_card("Genset", genset, genset_color) +
#                     _mini_card("UPS", ups, ups_color) +
#                     _mini_card("Used & Tested", used_tested.title(), used_color)
#                 )

#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:16px 14px; box-shadow:var(--shadow-sm); height:410px; box-sizing:border-box; display:flex; flex-direction:column;">
# <div style="font-size:12px; font-weight:700; color:var(--text-muted); margin-bottom:8px; text-transform:uppercase; letter-spacing:0.5px; flex-shrink:0;">Key Metrics</div>
# <div style="display:flex; flex-direction:column; flex:1; gap:5px; margin-top:0;">
# {cards_html}
# </div>
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             # --- VENUE INFO & CONTACT INFO ---
#             venue_info_col, contact_col = st.columns(2)

#             with venue_info_col:
#                 v_info_items = [
#                     ("Venue SPOC", _val("venue_spoc_name")),
#                     ("SPOC Contact", _val("spoc_contact_number")),
#                     ("SPOC Email", _val("spoc_email_id"))
#                 ]
#                 v_info_html = ""
#                 for label, val in v_info_items:
#                     if val != "N/A":
#                         v_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
# <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
# <span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
# </div>"""
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
# <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Information</div>
# {v_info_html}
# </div>""", unsafe_allow_html=True)

#             with contact_col:
#                 c_info_items = [
#                     ("Channel Manager", _val("channel_manager_name")),
#                     ("CM Contact", _val("channel_manager_contact_no.")),
#                     ("CM Email", _val("channel_manager_email_id")),
#                 ]
#                 c_info_html = ""
#                 for label, val in c_info_items:
#                     if val != "N/A":
#                         c_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
# <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
# <span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
# </div>"""
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
# <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">DEXIT SPOC CONTACT INFORMATION</div>
# {c_info_html}
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             is_dotc = str(venue_type).strip().upper() == "DOTC"
#             suppress_analysis_sections = is_dotc

#             # --- ANALYSIS & REMARKS ---
#             if not suppress_analysis_sections:
#                 st.markdown("""
#                 <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px;
#                             text-transform:uppercase; letter-spacing:0.5px;">Analysis & Remarks</div>
#                 """, unsafe_allow_html=True)

#             cat_val = _val("category", "N/A")
#             subcat_val = _val("subcategory", "N/A")

#             fields = []
#             if (not suppress_analysis_sections) and cat_val != "N/A" and cat_val != "" and cat_val != "0" and str(cat_val).lower() != "nan":
#                 fields.append(("Category", cat_val))
#             if subcat_val != "N/A" and subcat_val != "" and subcat_val != "0" and str(subcat_val).lower() != "nan":
#                 fields.append(("Subcategory", subcat_val))
#             if (not suppress_analysis_sections) and final_analysis != "N/A" and final_analysis != "" and final_analysis != "0" and final_analysis != "0.0":
#                 fields.append(("Venue Intel", final_analysis))

#             if fields:
#                 inner_html = ""
#                 for lbl, field_val in fields:
#                     inner_html += f"""
# <div style="margin-right:48px; margin-bottom:8px;">
#     <div style="font-size:11px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:4px;">{lbl}</div>
#     <div style="font-size:14px; color:#1A202C; line-height:1.6; font-weight:500;">{html.escape(str(field_val))}</div>
# </div>"""

#                 st.markdown(f"""
# <div style="background:#fff; border:1px solid #E2E8F0; border-left:4px solid {analysis_color};
#             border-radius:12px; padding:16px 18px 8px 18px; margin:8px 0;
#             box-shadow:0 1px 4px rgba(0,0,0,0.03); display:flex; flex-wrap:wrap;">
#     {inner_html}
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             # --- FINAL ISSUE REMARKS ---
#             final_issue = _val("call_log_summary", "")
#             if (not suppress_analysis_sections) and final_issue and final_issue != "N/A":
#                 st.markdown("""
#                 <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Major Call Log Issues Identified</div>
#                 """, unsafe_allow_html=True)

#                 st.markdown(f"""
#                 <div style="background:var(--surface); border:1px solid var(--border); border-left:4px solid #F59E0B; border-radius:12px; padding:14px 18px; margin:8px 0; box-shadow:var(--shadow-sm);">
#                     <div style="font-size:14px; color:var(--text); line-height:1.6; white-space:pre-wrap;">{html.escape(final_issue)}</div>
#                 </div>
#                 """, unsafe_allow_html=True)

#             st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)


# from __future__ import annotations

# import streamlit as st
# import pandas as pd
# import os
# import numpy as np
# import plotly.express as px
# import plotly.graph_objects as go
# import html
# import json
# import re
# import zipfile
# from datetime import datetime
# from streamlit_autorefresh import st_autorefresh
# import io
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment


# # ---------------- PAGE CONFIG ----------------
# st.set_page_config(layout="wide", page_title="Venue Inventory Master", initial_sidebar_state="expanded")

# # ---------------- STYLE ----------------
# STYLE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ui", "styles.css")
# try:
#     with open(STYLE_PATH, "r", encoding="utf-8") as f:
#         st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
# except FileNotFoundError:
#     # Styling is optional, don't break if CSS file not found
#     pass
# except Exception as e:
#     # Log styling loading errors but don't crash
#     print(f"Warning: Could not load styles from {STYLE_PATH}: {e}")


# def metric_card(label, value, tooltip: str | None = None, pct: float | None = None, trend: str | None = None):
#     safe_label = html.escape(str(label))
#     safe_value = html.escape(str(value))

#     details_html = ""
#     graph_html = ""
#     trend_html = ""

#     if trend:
#         t_str = str(trend).strip()
#         if "+" in t_str:
#             t_class = "trend-up"
#         elif "-" in t_str:
#             t_class = "trend-down"
#         else:
#             t_class = "trend-neutral"
#         trend_html = f'<span class="{t_class}">{html.escape(t_str)}</span>'

#     if pct is not None:
#         safe_pct = min(max(pct, 0), 100)
#         graph_html = f'<div class="kpi-graph"><div class="kpi-graph-fill" style="width: {safe_pct}%;"></div></div>'

#     if tooltip:
#         parts = str(tooltip).split("\n")
#         formatted_parts = []
#         for p in parts:
#             if ":" in p:
#                 k, v = p.split(":", 1)
#                 formatted_parts.append(
#                     "<span class='kpi-detail-item'>"
#                     f"<span style='color:var(--text);font-weight:700'>{html.escape(k)}:</span>"
#                     f"<span>{html.escape(v.strip())}</span>"
#                     "</span>"
#                 )
#             else:
#                 formatted_parts.append(f"<span class='kpi-detail-item'>{html.escape(p)}</span>")
#         safe_details = ' <span style="margin:0 6px; color:var(--border-dark)">|</span> '.join(formatted_parts)

#         details_html = f'<div class="kpi-subtext">{trend_html} {safe_details}</div>'
#     elif trend_html:
#         details_html = f'<div class="kpi-subtext">{trend_html}</div>'

#     final_html = f"""<div class="kpi-static">
# <div class="kpi-label">{safe_label}</div>
# <div class="kpi-value">{safe_value}</div>
# {graph_html}
# {details_html}
# </div>"""
#     st.markdown(final_html, unsafe_allow_html=True)

# def metric_card_compact(label: str, value: int, accent: str = "#81CDEC"):
#     """Compact KPI card for inline executive summaries."""
#     safe_label = html.escape(str(label))
#     safe_value = html.escape(f"{int(value):,}" if isinstance(value, (int, np.integer, float, np.floating)) else str(value))
#     st.markdown(
#         f"""
#         <div class="kpi-compact" style="--kpi-accent:{accent};">
#             <div class="kpi-compact-label">{safe_label}</div>
#             <div class="kpi-compact-value">{safe_value}</div>
#         </div>
#         """,
#         unsafe_allow_html=True,
#     )

# def render_brand_logo():
#     """Render company logo inline — call inside a st.columns block for positioning."""
#     import base64 as _b64
#     logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
#     if not os.path.exists(logo_path):
#         logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.jpg")
#     if os.path.exists(logo_path):
#         with open(logo_path, "rb") as _lf:
#             _logo_b64 = _b64.b64encode(_lf.read()).decode("utf-8")
#         _ext = os.path.splitext(logo_path)[1][1:].lower()
#         _mime = "jpeg" if _ext == "jpg" else _ext
#         st.markdown(
#             f"""
#             <div style="display:flex; justify-content:flex-end; align-items:flex-start; margin:-14px 0 8px 0; padding:0;">
#                 <img src="data:image/{_mime};base64,{_logo_b64}"
#                      alt="DEXIT Global"
#                      style="width:205px; height:auto; display:block; margin:0; padding:0; object-fit:contain;" />
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#     else:
#         st.markdown(
#             """
#             <div style="display:flex; justify-content:flex-end; margin:-4px 0 0 0; padding:0;
#                         font-size:15px; font-weight:800; color:#1A202C; letter-spacing:0.5px;">VMS AUTOMATION</div>
#             """,
#             unsafe_allow_html=True,
#         )

# def sidebar_checkbox_filter(
#     label: str,
#     options: list,
#     key_prefix: str,
#     *,
#     expanded: bool = False,
#     columns: int = 1,
# ) -> list:
#     """Render an executive-style checkbox filter block and return selected items."""
#     clean_options = [opt for opt in options if str(opt).strip() != ""]
#     if not clean_options:
#         return []

#     for idx in range(len(clean_options)):
#         state_key = f"{key_prefix}_{idx}"
#         if state_key not in st.session_state:
#             st.session_state[state_key] = True

#     selected_before = sum(
#         1 for idx in range(len(clean_options)) if st.session_state.get(f"{key_prefix}_{idx}", True)
#     )

#     with st.sidebar.expander(f"{label}  ({selected_before}/{len(clean_options)})", expanded=expanded):
#         search_query = st.text_input(
#             f"Search {label}",
#             key=f"{key_prefix}_search",
#             placeholder="Search options...",
#             label_visibility="collapsed",
#         ).strip().lower()

#         ctl_col1, ctl_col2 = st.columns(2)
#         if ctl_col1.button("Select All", key=f"{key_prefix}_all", use_container_width=True):
#             for idx in range(len(clean_options)):
#                 st.session_state[f"{key_prefix}_{idx}"] = True
#         if ctl_col2.button("Clear All", key=f"{key_prefix}_none", use_container_width=True):
#             for idx in range(len(clean_options)):
#                 st.session_state[f"{key_prefix}_{idx}"] = False

#         visible_options = []
#         for idx, opt in enumerate(clean_options):
#             opt_str = str(opt)
#             if not search_query or search_query in opt_str.lower():
#                 visible_options.append((idx, opt_str))

#         if not visible_options:
#             st.caption("No matching options.")
#         else:
#             check_cols = st.columns(max(1, columns))
#             for pos, (idx, opt_text) in enumerate(visible_options):
#                 with check_cols[pos % max(1, columns)]:
#                     st.checkbox(opt_text, key=f"{key_prefix}_{idx}")

#     selected = [opt for idx, opt in enumerate(clean_options) if st.session_state.get(f"{key_prefix}_{idx}", True)]
#     return selected

# # ---------------- PATHS ----------------
# SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# BASE_PATH = os.path.join(SCRIPT_DIR, "Data")
# OUTPUT_PATH = os.path.join(BASE_PATH, "Output")
# PHOTOS_PATH = os.path.join(BASE_PATH, "Venue_Photos")
# INDIA_GEOJSON_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "india.geojson")
# STATE_LIST_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "state-list.json")
# STATE_GEOJSON_DIR = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "states")
# BIHAR_REFERENCE_PPTX_PATH = os.path.join("C:\\Users", "DELL", "Downloads", "Bihar district maps.pptx")
# INDIA_MAP_CENTER = {"lat": 21.92, "lon": 82.74}
# INDIA_MAP_BOUNDS = {"west": 65.5, "east": 100.0, "south": 4.8, "north": 39.8}
# if not os.path.exists(OUTPUT_PATH): os.makedirs(OUTPUT_PATH)
# if not os.path.exists(PHOTOS_PATH): os.makedirs(PHOTOS_PATH)

# # ---------------- CONFIGURATION ----------------
# FILES = {
#     "Venue Inventory": {"file": "VMS Master.xlsx", "users": ["Nishant", "Prafull", "Admin"]},
#     "Blacklisted": {"file": "Blacklisted Venues.xlsx", "users": ["Prafull", "Admin"]},
#     "FFA Data": {"file": "FFA Data.xlsx", "users": ["Anil", "Admin"]},
#     "Suspicious": {"file": "Suspicious Activities.xlsx", "users": ["Prafull", "Admin"]},
#     "Day End": {"file": "Day End Observations.xlsx", "users": ["Prafull", "Admin"]},
#     "Market Reviews": {"file": "Google review and ratings.xlsx", "users": ["Prafull", "Admin"]},
#     "Call Logs": {"file": "Master Call Logs.xlsx", "users": ["Prafull", "Admin"]},
#     "Project Master": {"file": "Project Master.xlsx", "users": ["Prafull", "Admin"]}
# }

# USERS = {"Prafull": "Prafull@123", "Nishant": "Nishant@123", "Mayuresh": "Mayuresh@123", "Anil": "Anil@123", "Admin": "Admin@123"}

# # --- SORTING DEFINITIONS ---
# CAT_ORDER = ["A", "B", "C", "BL-C", "BL"]
# REGION_ORDER = ["North One", "North Two", "East", "West", "South", "Foreign"]

# # ---------------- SESSION STATE ----------------
# if "logged_in" not in st.session_state: st.session_state.logged_in = False
# if "user" not in st.session_state: st.session_state.user = None
# if "search_main_box" not in st.session_state: st.session_state.search_main_box = ""
# if "venue_search_query" not in st.session_state: st.session_state.venue_search_query = ""
# if "venue_detail_code" not in st.session_state: st.session_state.venue_detail_code = None
# if "map_selected_state" not in st.session_state: st.session_state.map_selected_state = None
# if "map_selected_district" not in st.session_state: st.session_state.map_selected_district = None
# if "map_selected_district_state" not in st.session_state: st.session_state.map_selected_district_state = None

# VALID_PAGES = {"dashboard", "login", "main", "edit", "process", "venue_detail", "india_map_detail", "manpower_dashboard", "project_wise_details"}
# VALID_DATASETS = set(FILES.keys())

# qp_page = st.query_params.get("page", None)
# if isinstance(qp_page, list):
#     qp_page = qp_page[0] if qp_page else None

# qp_dataset = st.query_params.get("dataset", None)
# if isinstance(qp_dataset, list):
#     qp_dataset = qp_dataset[0] if qp_dataset else None

# if "page" not in st.session_state:
#     st.session_state.page = qp_page if qp_page in VALID_PAGES else "dashboard"
# elif st.session_state.page not in VALID_PAGES:
#     st.session_state.page = "dashboard"

# if "dataset_key" not in st.session_state:
#     st.session_state.dataset_key = qp_dataset if qp_dataset in VALID_DATASETS else "Venue Inventory"
# elif st.session_state.dataset_key not in VALID_DATASETS:
#     st.session_state.dataset_key = "Venue Inventory"
# elif qp_dataset in VALID_DATASETS and qp_dataset != st.session_state.dataset_key:
#     st.session_state.dataset_key = qp_dataset

# if qp_page != st.session_state.page:
#     st.query_params["page"] = st.session_state.page
# if qp_dataset != st.session_state.dataset_key:
#     st.query_params["dataset"] = st.session_state.dataset_key

# if st.session_state.page == "dashboard":
#     st.session_state["_auto_refresh_count"] = st_autorefresh(interval=60000, key="auto_refresh_main")
# else:
#     st.session_state["_auto_refresh_count"] = st.session_state.get("_auto_refresh_count", 0)

# # ---------------- SMART DATA LOADER ----------------
# def load_smart_data(file_path):
#     if not os.path.exists(file_path): return pd.DataFrame()
#     try:
#         if file_path.lower().endswith(".xlsx"):
#             preview = pd.read_excel(file_path, header=None, nrows=10)
#             header_idx = 0
#             for i, row in preview.iterrows():
#                 row_str = " ".join([str(x) for x in row.values])
#                 if any(key in row_str.upper() for key in ["ROW LABELS", "VENUE_TYPE", "CODE", "ACTIVE", "DMS_CODE", "STATUS"]):
#                     header_idx = i; break
#             return pd.read_excel(file_path, header=header_idx)
#         return pd.read_csv(file_path)
#     except: return pd.DataFrame()

# # ---------------- HELPER FUNCTIONS FOR DATA PROCESSING ----------------
# def clean_project_name(series: pd.Series) -> pd.Series:
#     """Clean project names by removing special chars, normalizing spacing, and filtering blanks."""
#     s = series.fillna("").astype(str).str.strip()
#     s_upper = s.str.upper()
#     is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
#     s = s.mask(is_blank, "")
#     s = s.str.replace(r"[^A-Za-z ]+", " ", regex=True)
#     s = s.str.replace(r"\s+", " ", regex=True).str.strip()
#     return s.str.upper()

# def _blankify_text(series: pd.Series) -> pd.Series:
#     """Convert blank-like values (-, NAN, NONE) to empty strings."""
#     s = series.fillna("").astype(str).str.strip()
#     s_upper = s.str.upper()
#     is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
#     return s.mask(is_blank, "")

# def _col_or_empty(df_in: pd.DataFrame, col_names: list) -> pd.Series:
#     """Return column from dataframe or empty series if not present. Case-insensitive."""
#     col_names_lower = [c.lower() for c in col_names]
#     for c in df_in.columns:
#         if str(c).lower() in col_names_lower:
#             return df_in[c]
#     return pd.Series([""] * len(df_in), index=df_in.index)

# def build_concat_address(df_in: pd.DataFrame) -> pd.Series:
#     """Build concatenated full address in format: name, address, landmark, city, district, state - pincode"""
#     name_s = _blankify_text(_col_or_empty(df_in, ["name", "venue_name", "updated_venue_name"]))
#     address_s = _blankify_text(_col_or_empty(df_in, ["address"]))
#     landmark_s = _blankify_text(_col_or_empty(df_in, ["landmark"]))
#     city_s = _blankify_text(_col_or_empty(df_in, ["city"]))
#     district_s = _blankify_text(_col_or_empty(df_in, ["district"]))
#     state_s = _blankify_text(_col_or_empty(df_in, ["state"]))

#     parts_df = pd.concat([name_s, address_s, landmark_s, city_s, district_s, state_s], axis=1)
#     joined_address = parts_df.apply(lambda row: ", ".join([v for v in row.tolist() if v]), axis=1)

#     pincode_s = _blankify_text(_col_or_empty(df_in, ["pincode", "pin_code"]))
#     pincode_s = pincode_s.str.replace(r"\.0$", "", regex=True)

#     # Format: name, address, landmark, city, district, state - pincode
#     result = np.where(
#         joined_address != "",
#         joined_address + np.where(pincode_s != "", " - " + pincode_s, ""),
#         ""
#     )
#     return pd.Series(result, index=df_in.index)

# def find_dms_col(df: pd.DataFrame):
#     """Find the DMS/Venue code column in a dataframe."""
#     return next((c for c in df.columns if c.lower().replace(" ", "_") in ['dms_code', 'venue_code']), None)

# STATE_NAME_ALIASES = {
#     "NCT OF DELHI": "DELHI",
#     "NCT DELHI": "DELHI",
#     "ORISSA": "ODISHA",
#     "UTTARANCHAL": "UTTARAKHAND",
#     "JAMMU KASHMIR": "JAMMU AND KASHMIR",
#     "DADRA AND NAGAR HAVELI AND DAMAN AND DIU": "DADRA AND NAGAR HAVELI DAMAN AND DIU",
#     "ANDAMAN NICOBAR": "ANDAMAN AND NICOBAR ISLANDS",
#     "PONDICHERRY": "PUDUCHERRY",
# }


# @st.cache_data(show_spinner=False)
# def load_india_geojson(path: str):
#     if not os.path.exists(path):
#         return None
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             gj = json.load(f)
#         if isinstance(gj, dict) and isinstance(gj.get("features"), list):
#             return gj
#     except Exception:
#         return None
#     return None


# def normalize_state_name(value: str) -> str:
#     name = str(value or "").upper()
#     name = name.replace("&", " AND ")
#     name = re.sub(r"[^A-Z0-9]+", " ", name)
#     name = re.sub(r"\s+", " ", name).strip()
#     return STATE_NAME_ALIASES.get(name, name)


# def build_india_map_dataframe(source_df: pd.DataFrame | None):
#     geojson_data = load_india_geojson(INDIA_GEOJSON_PATH)
#     if geojson_data is None:
#         return pd.DataFrame(), None, 0, 0

#     state_count_map: dict[str, int] = {}
#     if source_df is not None and not source_df.empty and "state" in source_df.columns:
#         state_series = source_df["state"].fillna("").astype(str).str.strip()
#         state_series = state_series[state_series != ""]
#         for state_name, count in state_series.value_counts().items():
#             normalized = normalize_state_name(state_name)
#             if normalized:
#                 state_count_map[normalized] = state_count_map.get(normalized, 0) + int(count)

#     rows = []
#     matched_states = 0
#     mapped_venues = 0
#     features = geojson_data.get("features", [])

#     for idx, feature in enumerate(features):
#         properties = feature.get("properties", {}) or {}
#         state_name = (
#             properties.get("st_nm")
#             or properties.get("STATE")
#             or properties.get("name")
#             or f"State {idx + 1}"
#         )
#         district_code = properties.get("dt_code") or properties.get("district_code")
#         state_code = properties.get("st_code") or properties.get("state_code")
#         feature_id = str(
#             feature.get("id")
#             or district_code
#             or state_code
#             or properties.get("ID_1")
#             or f"{normalize_state_name(state_name)}_{idx + 1}"
#         )
#         feature["id"] = feature_id

#         count = int(state_count_map.get(normalize_state_name(state_name), 0))
#         if count > 0:
#             matched_states += 1
#             mapped_venues += count

#         rows.append(
#             {
#                 "id": feature_id,
#                 "state": state_name,
#                 "venue_count": count,
#             }
#         )

#     return pd.DataFrame(rows), geojson_data, matched_states, mapped_venues

# def _get_feature_centroid(feature: dict) -> tuple[float, float]:
#     geom = feature.get("geometry", {})
#     if not geom:
#         return 0.0, 0.0
#     coords = geom.get("coordinates", [])
#     flat_coords = []

#     def extract_coords(c):
#         if not c:
#             return
#         if isinstance(c[0], (int, float)):
#             flat_coords.append(c)
#         else:
#             for item in c:
#                 extract_coords(item)
#     extract_coords(coords)
    
#     if not flat_coords:
#         return 0.0, 0.0
#     arr = np.array(flat_coords)
#     return float(np.mean(arr[:, 0])), float(np.mean(arr[:, 1]))


# def create_india_map_figure(
#     map_df: pd.DataFrame,
#     geojson_data: dict,
#     *,
#     height: int = 320,
#     show_data: bool = True,
#     zoom: float = 3.5,
#     center: dict | None = None,
#     opacity: float = 0.78,
#     map_style: str = "white-bg",
#     color_power: float = 1.0,
#     line_color: str = "#94A3B8",
#     line_width: float = 0.9,
#     restrict_to_india: bool = True,
# ):
#     if map_df.empty or not geojson_data:
#         return None

#     if show_data:
#         plot_df = map_df.copy()
#         counts = pd.to_numeric(plot_df.get("venue_count", 0), errors="coerce").fillna(0).astype(float)
#         if color_power != 1.0:
#             plot_df["_color_metric"] = np.power(counts, color_power)
#         else:
#             plot_df["_color_metric"] = counts
#         max_metric = float(plot_df["_color_metric"].max()) if "_color_metric" in plot_df.columns else 0.0
#         fig_map = px.choropleth(
#             plot_df,
#             geojson=geojson_data,
#             locations="id",
#             featureidkey="id",
#             color="_color_metric",
#             hover_name="state",
#             custom_data=["state", "id", "venue_count"],
#             hover_data={"venue_count": False, "id": False, "_color_metric": False},
#             color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
#             range_color=(0, max(1.0, max_metric)),
#         )
#     else:
#         preview_df = map_df.copy()
#         preview_df["map_fill"] = 1
#         plot_df = preview_df
#         fig_map = px.choropleth(
#             preview_df,
#             geojson=geojson_data,
#             locations="id",
#             featureidkey="id",
#             color="map_fill",
#             hover_name="state",
#             custom_data=["state", "id", "venue_count"],
#             hover_data={"map_fill": False, "id": False},
#             color_continuous_scale=["#C9E1F2", "#C9E1F2"],
#             range_color=(0, 1),
#         )

#     fig_map.update_traces(
#         marker_line_color=line_color,
#         marker_line_width=line_width,
#         hovertemplate="<b>%{hovertext}</b><extra></extra>" if not show_data else "<b>%{hovertext}</b><br>Venues: %{customdata[2]:,}<extra></extra>",
#     )
    
#     id_to_name = dict(zip(plot_df["id"], plot_df["state"]))
#     state_to_coords = {}
#     if geojson_data and "features" in geojson_data:
#         for f in geojson_data["features"]:
#             fid = f.get("id")
#             if fid in id_to_name:
#                 state_name = id_to_name[fid]
#                 geom = f.get("geometry", {})
#                 if geom:
#                     if state_name not in state_to_coords:
#                         state_to_coords[state_name] = []
#                     state_to_coords[state_name].append(geom.get("coordinates", []))

#     HIDE_LABELS = {
#         "Dadra and Nagar Haveli and Daman and Diu", 
#         "Chandigarh", 
#         "Puducherry", 
#         "Lakshadweep",
#         "Goa",
#         "Delhi",
#         "Sikkim",
#         "Tripura",
#         "Mizoram",
#         "Manipur",
#         "Nagaland",
#         "Meghalaya",
#         "Haryana"
#     }
#     LABEL_ALIASES = {
#         "Andaman and Nicobar Islands": "A & N",
#         "Jammu and Kashmir": "J & K",
#         "Himachal Pradesh": "HP",
#         "Uttarakhand": "UK",
#         "Punjab": "PB",
#         "West Bengal": "WB",
#         "Jharkhand": "JH",
#         "Arunachal Pradesh": "Arunachal",
#         "Uttar Pradesh": "Uttar<br>Pradesh",
#         "Madhya Pradesh": "Madhya<br>Pradesh",
#         "Andhra Pradesh": "Andhra<br>Pradesh",
#         "Tamil Nadu": "Tamil<br>Nadu"
#     }
    
#     lons, lats, texts = [], [], []
#     for state_name, coords_list in state_to_coords.items():
#         if state_name in HIDE_LABELS:
#             continue
#         cx, cy = _get_feature_centroid({"geometry": {"coordinates": coords_list}})
#         if cx != 0.0 and cy != 0.0:
#             lons.append(cx)
#             lats.append(cy)
#             texts.append(LABEL_ALIASES.get(state_name, state_name))

#     if show_data and lons:
#         fig_map.add_trace(go.Scattergeo(
#             lon=lons,
#             lat=lats,
#             text=texts,
#             mode="text",
#             textfont=dict(color="#1E293B", size=8.5, family="Inter, sans-serif"),
#             showlegend=False,
#             hoverinfo="skip"
#         ))
        
#     fig_map.update_geos(fitbounds="locations", visible=False)
    
#     fig_map.update_layout(
#         clickmode="event+select",
#         dragmode="pan",
#         margin=dict(t=0, b=0, l=0, r=0),
#         paper_bgcolor="rgba(0,0,0,0)",
#         plot_bgcolor="rgba(0,0,0,0)",
#         height=height,
#         coloraxis_showscale=show_data,
#         coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.78),
#     )
#     return fig_map


# def extract_selected_points(selection_event) -> list:
#     if selection_event is None:
#         return []
#     if isinstance(selection_event, dict):
#         return selection_event.get("selection", {}).get("points", []) or []
#     selection_obj = getattr(selection_event, "selection", None)
#     if selection_obj is None:
#         return []
#     return getattr(selection_obj, "points", []) or []


# def get_last_point_signature(points: list) -> str | None:
#     if not isinstance(points, (list, tuple)) or not points:
#         return None
#     point = points[-1]
#     if point is None:
#         return None
#     customdata = _point_get(point, "customdata")
#     if isinstance(customdata, (list, tuple)):
#         custom_sig = "|".join("" if x is None else str(x) for x in customdata[:3])
#     else:
#         custom_sig = ""
#     sig_parts = [
#         str(_point_get(point, "location") or ""),
#         str(_point_get(point, "id") or ""),
#         str(_point_get(point, "hovertext") or ""),
#         str(_point_get(point, "point_index") or _point_get(point, "pointIndex") or ""),
#         custom_sig,
#     ]
#     signature = "||".join(sig_parts).strip()
#     return signature if signature else None


# def _point_get(point, key: str):
#     if isinstance(point, dict):
#         return point.get(key)
#     return getattr(point, key, None)


# @st.cache_data(show_spinner=False)
# def load_ppt_reference_images(pptx_path: str, max_images: int = 6):
#     if not os.path.exists(pptx_path):
#         return []
#     images = []
#     try:
#         with zipfile.ZipFile(pptx_path, "r") as zf:
#             media_files = [
#                 n for n in zf.namelist()
#                 if n.startswith("ppt/media/") and n.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
#             ]
#             media_files = sorted(media_files)
#             for media_path in media_files[:max_images]:
#                 try:
#                     images.append((os.path.basename(media_path), zf.read(media_path)))
#                 except Exception:
#                     continue
#     except Exception:
#         return []
#     return images


# def resolve_state_from_map_points(points: list, id_to_state_map: dict, ordered_names: list | None = None):
#     if not points:
#         return None
#     point = points[-1] if isinstance(points, (list, tuple)) else None
#     if point is None:
#         return None

#     # 0) customdata fallback using known payload order:
#     #    [name, id, ...] for both state and district map traces.
#     cd = _point_get(point, "customdata")
#     if isinstance(cd, (list, tuple)):
#         if len(cd) >= 2:
#             cand_id = str(cd[1]).strip() if cd[1] is not None else ""
#             if cand_id:
#                 resolved = id_to_state_map.get(cand_id)
#                 if resolved:
#                     return resolved
#         if len(cd) >= 1:
#             cand_name = str(cd[0]).strip() if cd[0] is not None else ""
#             if cand_name:
#                 for mapped_name in id_to_state_map.values():
#                     if str(mapped_name).strip().lower() == cand_name.lower():
#                         return mapped_name

#         # legacy fallback for unexpected payload shape
#         for item in cd:
#             item_s = str(item).strip() if item is not None else ""
#             if not item_s:
#                 continue
#             resolved = id_to_state_map.get(item_s)
#             if resolved:
#                 return resolved
#             for mapped_name in id_to_state_map.values():
#                 if str(mapped_name).strip().lower() == item_s.lower():
#                     return mapped_name

#     # 1) direct id/location mapping (common for choropleth)
#     for key in ("location", "id"):
#         loc = _point_get(point, key)
#         if loc is not None:
#             resolved = id_to_state_map.get(str(loc))
#             if resolved:
#                 return resolved

#     # 2) index-based fallback (point_index / point_number)
#     if ordered_names:
#         for idx_key in ("point_index", "point_number", "pointIndex", "pointNumber"):
#             idx_val = _point_get(point, idx_key)
#             try:
#                 idx_int = int(idx_val)
#             except Exception:
#                 idx_int = None
#             if idx_int is not None and 0 <= idx_int < len(ordered_names):
#                 return ordered_names[idx_int]

#     # 3) text-based fallback from hover payload
#     for txt_key in ("hovertext", "label", "text", "name"):
#         txt = _point_get(point, txt_key)
#         if txt is None:
#             continue
#         txt_clean = str(txt).strip()
#         if not txt_clean:
#             continue
#         # First try direct exact map value
#         for mapped_name in id_to_state_map.values():
#             if str(mapped_name).strip().lower() == txt_clean.lower():
#                 return mapped_name

#     return None


# def build_state_detail_payload(df: pd.DataFrame, state_name: str):
#     if df is None or df.empty or "state" not in df.columns or not state_name:
#         return pd.DataFrame()
#     target = normalize_state_name(state_name)
#     matched = df[df["state"].fillna("").astype(str).apply(normalize_state_name) == target].copy()
#     return matched


# def normalize_district_name(value: str) -> str:
#     name = str(value or "").upper()
#     name = re.sub(r"[^A-Z0-9]+", " ", name)
#     return re.sub(r"\s+", " ", name).strip()


# @st.cache_data(show_spinner=False)
# def load_state_slug_mapping(path: str):
#     mapping = {}
#     if not os.path.exists(path):
#         return mapping
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             rows = json.load(f)
#         for row in rows:
#             state_name = row.get("name")
#             slug = row.get("slug")
#             if state_name and slug:
#                 mapping[normalize_state_name(state_name)] = slug
#     except Exception:
#         return {}
#     return mapping


# @st.cache_data(show_spinner=False)
# def load_state_geojson_by_slug(slug: str):
#     if not slug:
#         return None
#     path = os.path.join(STATE_GEOJSON_DIR, f"{slug}.geojson")
#     if not os.path.exists(path):
#         return None
#     try:
#         with open(path, "r", encoding="utf-8") as f:
#             gj = json.load(f)
#         if isinstance(gj, dict) and isinstance(gj.get("features"), list):
#             return gj
#     except Exception:
#         return None
#     return None


# def resolve_state_slug(state_name: str):
#     if not state_name:
#         return None
#     state_map = load_state_slug_mapping(STATE_LIST_PATH)
#     return state_map.get(normalize_state_name(state_name))


# def _flatten_coordinates(coords):
#     if isinstance(coords, (list, tuple)) and coords:
#         if isinstance(coords[0], (int, float)) and len(coords) >= 2:
#             yield float(coords[0]), float(coords[1])
#         else:
#             for item in coords:
#                 yield from _flatten_coordinates(item)


# def get_geojson_center_zoom(geojson_data: dict):
#     lons, lats = [], []
#     for feature in geojson_data.get("features", []):
#         geometry = feature.get("geometry") or {}
#         coords = geometry.get("coordinates")
#         if coords is None:
#             continue
#         for lon, lat in _flatten_coordinates(coords):
#             lons.append(lon)
#             lats.append(lat)

#     if not lons or not lats:
#         return {"lat": 22.7, "lon": 79.2}, 5.4

#     min_lon, max_lon = min(lons), max(lons)
#     min_lat, max_lat = min(lats), max(lats)
#     span = max(max_lon - min_lon, max_lat - min_lat)

#     if span > 14:
#         zoom = 4.9
#     elif span > 10:
#         zoom = 5.3
#     elif span > 6:
#         zoom = 5.8
#     elif span > 3.5:
#         zoom = 6.3
#     elif span > 2:
#         zoom = 6.9
#     else:
#         zoom = 7.5

#     center = {"lat": (min_lat + max_lat) / 2, "lon": (min_lon + max_lon) / 2}
#     return center, zoom


# def build_district_map_dataframe(state_df: pd.DataFrame, state_geojson: dict):
#     if state_geojson is None:
#         return pd.DataFrame(), {}

#     district_count_map = {}
#     datc_map = {}
#     dotc_map = {}
#     active_map = {}
#     inactive_map = {}
#     black_map = {}

#     if state_df is not None and not state_df.empty:
#         district_series = state_df["district"].fillna("").astype(str).str.strip() if "district" in state_df.columns else pd.Series([], dtype=str)
#         venue_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([], dtype=str)
#         status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([], dtype=str)

#         for idx, district_name in district_series.items():
#             norm = normalize_district_name(district_name)
#             if not norm:
#                 continue
#             district_count_map[norm] = district_count_map.get(norm, 0) + 1

#             vtype = venue_series.get(idx, "")
#             if vtype == "DATC":
#                 datc_map[norm] = datc_map.get(norm, 0) + 1
#             elif vtype == "DOTC":
#                 dotc_map[norm] = dotc_map.get(norm, 0) + 1

#             status_val = status_series.get(idx, "")
#             if status_val == "ACTIVE":
#                 active_map[norm] = active_map.get(norm, 0) + 1
#             elif status_val == "INACTIVE":
#                 inactive_map[norm] = inactive_map.get(norm, 0) + 1
#             elif "BLACK" in status_val:
#                 black_map[norm] = black_map.get(norm, 0) + 1

#     rows = []
#     id_to_district = {}
#     for i, feature in enumerate(state_geojson.get("features", [])):
#         props = feature.get("properties", {}) or {}
#         district_name = props.get("district") or props.get("DISTRICT") or props.get("name") or f"District {i + 1}"
#         norm = normalize_district_name(district_name)
#         feature_id = str(feature.get("id") or props.get("dt_code") or props.get("district_code") or district_name)
#         feature["id"] = feature_id
#         id_to_district[feature_id] = district_name

#         rows.append(
#             {
#                 "id": feature_id,
#                 "district": district_name,
#                 "venue_count": int(district_count_map.get(norm, 0)),
#                 "datc_count": int(datc_map.get(norm, 0)),
#                 "dotc_count": int(dotc_map.get(norm, 0)),
#                 "active_count": int(active_map.get(norm, 0)),
#                 "inactive_count": int(inactive_map.get(norm, 0)),
#                 "black_count": int(black_map.get(norm, 0)),
#             }
#         )

#     return pd.DataFrame(rows), id_to_district


# def compute_dotc_datc_summary(df_in: pd.DataFrame):
#     if df_in is None or df_in.empty:
#         return {
#             "dotc_centres": 0, "datc_centres": 0, "total_centres": 0,
#             "dotc_seats": 0, "datc_seats": 0, "total_seats": 0,
#         }

#     vtype = df_in["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in df_in.columns else pd.Series([""] * len(df_in))
#     seat_series = pd.to_numeric(df_in.get("venue_max_capacity", 0), errors="coerce").fillna(0)

#     dotc_mask = vtype == "DOTC"
#     datc_mask = vtype == "DATC"

#     dotc_centres = int(dotc_mask.sum())
#     datc_centres = int(datc_mask.sum())
#     total_centres = int(len(df_in))

#     dotc_seats = int(seat_series[dotc_mask].sum())
#     datc_seats = int(seat_series[datc_mask].sum())
#     total_seats = int(seat_series.sum())

#     return {
#         "dotc_centres": dotc_centres,
#         "datc_centres": datc_centres,
#         "total_centres": total_centres,
#         "dotc_seats": dotc_seats,
#         "datc_seats": datc_seats,
#         "total_seats": total_seats,
#     }


# def render_dotc_datc_summary(title: str, summary: dict):
#     st.markdown(
#         f"""
#         <div style="border:1px solid #A8BBC8; border-radius:10px; overflow:hidden; background:#F8FBFD;">
#             <div style="padding:8px 12px; border-bottom:1px solid #A8BBC8; font-weight:800; color:#2A3B4C; text-transform:uppercase; letter-spacing:0.4px;">
#                 {html.escape(title)}
#             </div>
#             <table style="width:100%; border-collapse:collapse; font-size:13px;">
#                 <thead>
#                     <tr style="background:#EAF2F7; color:#334155;">
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:left;">Metric</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DOTC</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DATC</th>
#                         <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">Total</th>
#                     </tr>
#                 </thead>
#                 <tbody>
#                     <tr>
#                         <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Centre</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_centres", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_centres", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_centres", 0):,}</td>
#                     </tr>
#                     <tr>
#                         <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Seat Capacity</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_seats", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_seats", 0):,}</td>
#                         <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_seats", 0):,}</td>
#                     </tr>
#                 </tbody>
#             </table>
#         </div>
#         """,
#         unsafe_allow_html=True,
#     )


# # If protected pages are refreshed without an active session, route to login cleanly.
# if st.session_state.page in {"main", "edit", "process"} and not st.session_state.logged_in:
#     st.session_state.page = "login"
#     st.query_params["page"] = "login"

# # ================= DASHBOARD PAGE =================
# if st.session_state.page == "dashboard":
#     _title_col, _logo_col = st.columns([3, 1], vertical_alignment="top")
#     with _title_col:
#         st.title("Venue Analysis Dashboard")
#     with _logo_col:
#         render_brand_logo()
#     st.markdown('<div style="height: 8px;"></div>', unsafe_allow_html=True)
    
#     if st.session_state.logged_in:
#         st.sidebar.success(f"User: {st.session_state.user}")
#         if st.sidebar.button("Logout"):
#             st.session_state.logged_in, st.session_state.user = False, None
#             st.rerun()
#     else:
#         if st.sidebar.button("Admin Login"):
#             st.session_state.page = "login"
#             st.rerun()

#     all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH) 
#                  if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]
    
#     if not all_files:
#         st.warning("No processed reports found. Admin must run the Analysis Engine.")
#     else:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         data = pd.read_excel(latest_file_path, sheet_name=0)
#         data.columns = data.columns.str.lower().str.replace(" ", "_")
        
#         st.info(f"Dashboard Data: **{os.path.basename(latest_file_path)}**")

#         # --- VENUE SEARCH BAR ---
#         st.markdown("""
#         <div style="background: var(--surface);
#                     border-radius: 12px; padding: 22px 26px; margin: 12px 0 20px 0;
#                     border: 1px solid var(--border); border-left: 5px solid var(--primary); 
#                     box-shadow: var(--shadow-sm);">
#             <div style="color:var(--text); font-size:18px; font-weight:700; letter-spacing:0.5px; margin-bottom:6px; text-transform:uppercase;">
#                 Venue Search
#             </div>
#             <p style="color:var(--text-muted); font-size:13px; margin:0; font-weight:500;">
#                 Search by Venue Name, DMS Code, City, or any keyword to view complete venue profile and metrics
#             </p>
#         </div>
#         """, unsafe_allow_html=True)

#         vs_col1, vs_col2 = st.columns([4, 1], vertical_alignment="center")
#         with vs_col1:
#             venue_search_input = st.text_input(
#                 "Search Venue",
#                 value=st.session_state.venue_search_query,
#                 placeholder="e.g. AP COMPUTER POINT or EST-AR-1161 or ITANAGAR...",
#                 label_visibility="collapsed"
#             )
#         with vs_col2:
#             search_btn = st.button("Search Venue", use_container_width=True, type="primary")

#         # Update session state if button is clicked or if they type and press enter
#         if search_btn or venue_search_input != st.session_state.venue_search_query:
#             st.session_state.venue_search_query = venue_search_input

#         query = st.session_state.venue_search_query.strip()
#         if query:
#             # Search across all text columns
#             search_mask = data.astype(str).apply(
#                 lambda col: col.str.contains(query, case=False, na=False)
#             ).any(axis=1)
#             results = data[search_mask]

#             is_authorized = st.session_state.get("logged_in") and st.session_state.get("user") in ["Admin", "Prafull"]

#             if results.empty:
#                 st.warning(f"No venues found matching **\"{query}\"**. Try a different keyword.")
#                 if st.button("Clear Search", key="clear_search_empty"):
#                     st.session_state.venue_search_query = ""
#                     st.rerun()
#             elif len(results) == 1 and search_btn and is_authorized:
#                 # Single result -> go directly to detail page (only auto-redirect if search button was explicitly clicked)
#                 st.session_state.venue_detail_code = str(results.iloc[0].get('dms_code', ''))
#                 st.session_state.page = "venue_detail"
#                 st.session_state.venue_search_query = "" # Clear search after found
#                 st.rerun()
#             else:
#                 # Multiple results (or single result persistent view) -> show selection list
#                 head_col, clear_col = st.columns([5, 1], vertical_alignment="center")
#                 with head_col:
#                     st.markdown(f"##### Found {len(results)} venue(s) matching **\"{query}\"**")
#                 with clear_col:
#                     if st.button("Clear Search", key="clear_search_results", use_container_width=True):
#                         st.session_state.venue_search_query = ""
#                         st.rerun()
#                 for idx, row in results.iterrows():
#                     dms = str(row.get('dms_code', 'N/A'))
#                     name = str(row.get('venue_name', row.get('name', 'N/A')))
#                     city = str(row.get('city', ''))
#                     region = str(row.get('region', ''))
#                     status = str(row.get('status', ''))
#                     category = str(row.get('category', ''))
                    
#                     # Color coding for category
#                     cat_colors = {"A": "#10B981", "B": "#3B82F6", "C": "#F59E0B", "BL": "#EF4444", "BL-C": "#8B5CF6"}
#                     cat_color = cat_colors.get(category, "#6B7280")
                    
#                     st.markdown(f"""
#                     <div style="background:#fff; border:1px solid #E2E8F0; border-left:4px solid {cat_color};
#                                 border-radius:12px; padding:14px 18px; margin:6px 0;
#                                 display:flex; justify-content:space-between; align-items:center;
#                                 transition:all 0.2s; box-shadow:0 1px 3px rgba(0,0,0,0.05);">
#                         <div>
#                             <div style="font-weight:700; font-size:15px; color:#1A202C;">{html.escape(name)}</div>
#                             <div style="font-size:12px; color:#64748B; margin-top:4px; font-weight:600; letter-spacing:0.3px;">
#                                 <span>CODE: {html.escape(dms)}</span>
#                                 <span style="margin:0 8px; color:#CBD5E0;">|</span>
#                                 <span>CITY: {html.escape(city)}</span>
#                                 <span style="margin:0 8px; color:#CBD5E0;">|</span>
#                                 <span>REGION: {html.escape(region)}</span>
#                             </div>
#                         </div>
#                         <div style="display:flex; align-items:center; gap:8px;">
#                             <span style="background:{cat_color}22; color:{cat_color}; font-weight:700;
#                                         padding:4px 12px; border-radius:999px; font-size:12px;">
#                                 Cat {html.escape(category)}
#                             </span>
#                             <span style="background:#F0F4F8; color:#4A5568; font-weight:600;
#                                         padding:4px 10px; border-radius:999px; font-size:12px;">
#                                 {html.escape(status)}
#                             </span>
#                         </div>
#                     </div>
#                     """, unsafe_allow_html=True)
                    
#                     if st.button(f"View Details \u2192 {dms}", key=f"view_{dms}_{idx}", use_container_width=True, disabled=not is_authorized):
#                         st.session_state.venue_detail_code = dms
#                         st.session_state.page = "venue_detail"
#                         st.session_state.venue_search_query = "" # Clear search upon click
#                         st.rerun()

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)

#         if 'region' in data.columns:
#             data['region'] = data['region'].astype(str).str.title().str.strip()
#             data['region'] = pd.Categorical(data['region'], categories=REGION_ORDER, ordered=True)
#         if 'category' in data.columns:
#             data['category'] = pd.Categorical(data['category'], categories=CAT_ORDER, ordered=True)

#         # Executive sidebar filter header
#         st.sidebar.markdown(
#             """
#             <div class="filters-shell">
#                 <div class="filters-title-row">
#                     <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
#                         <polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"></polygon>
#                     </svg>
#                     <span class="filters-title-text">Dashboard Filters</span>
#                 </div>
#                 <p class="filters-subtitle">Refine live metrics by operational segments</p>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         if st.sidebar.button("Reset All Filters", key="dash_reset_all", use_container_width=True):
#             for k in list(st.session_state.keys()):
#                 if k.startswith("dash_vtype_") or k.startswith("dash_region_") or k.startswith("dash_status_") or k.startswith("dash_abc_"):
#                     if k.endswith("_search"):
#                         st.session_state[k] = ""
#                     elif k.endswith("_all") or k.endswith("_none"):
#                         continue
#                     else:
#                         st.session_state[k] = True
#             st.rerun()
#         st.sidebar.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        
#         v_type_list = sorted(data['venue_type'].dropna().unique().tolist())
#         sel_vtype = sidebar_checkbox_filter("Venue Type", v_type_list, "dash_vtype", expanded=True, columns=1)

#         avail_regs = sorted(data['region'].dropna().unique().tolist())
#         sel_reg = sidebar_checkbox_filter("Region", avail_regs, "dash_region", expanded=False, columns=1)

#         stat_list = sorted(data['status'].dropna().unique().tolist())
#         sel_stat = sidebar_checkbox_filter("Status", stat_list, "dash_status", expanded=False, columns=1)

#         avail_cats = sorted(data['category'].dropna().unique().tolist()) if 'category' in data.columns else []
#         if avail_cats:
#             sel_cat = sidebar_checkbox_filter("ABC Category", avail_cats, "dash_abc", expanded=False, columns=2)
#         else:
#             sel_cat = []

#         st.sidebar.markdown(
#             f"""
#             <div class="filters-summary-card">
#                 <div class="filters-summary-title">Current Selection</div>
#                 <div class="filters-summary-item">Venue Type: <b>{len(sel_vtype)}</b></div>
#                 <div class="filters-summary-item">Region: <b>{len(sel_reg)}</b></div>
#                 <div class="filters-summary-item">Status: <b>{len(sel_stat)}</b></div>
#                 <div class="filters-summary-item">ABC Category: <b>{len(sel_cat)}</b></div>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         st.sidebar.markdown(
#             """
#             <div class="sidebar-actions-shell">
#                 <div class="sidebar-actions-title">Quick Actions</div>
#                 <div class="sidebar-actions-subtitle">Navigate to admin tools instantly</div>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#         if st.sidebar.button("Manage Master Data", key="dash_manage_master_sidebar", use_container_width=True):
#             st.session_state.page = "main" if st.session_state.logged_in else "login"
#             st.rerun()
#         if st.sidebar.button("Process Analysis Engine", key="dash_process_engine_sidebar", use_container_width=True):
#             if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
#                 st.session_state.page = "process"
#             elif not st.session_state.logged_in:
#                 st.session_state.page = "login"
#             else:
#                 st.error("Restricted to Admin/Prafull.")
#             st.rerun()
        
#         mask = (data['venue_type'].isin(sel_vtype)) & (data['region'].isin(sel_reg)) & (data['status'].isin(sel_stat))
#         if 'category' in data.columns:
#             mask = mask & (data['category'].isin(sel_cat))
#         df_dash = data[mask].copy()

#         def fmt_pct(part: int, total: int) -> str:
#             if total <= 0:
#                 return "N/A"
#             return f"{(part / total) * 100:.1f}%"

#         def yes_count(df_in, col_name: str) -> int:
#             if col_name not in df_in.columns:
#                 return 0
#             s = df_in[col_name].astype(str).str.strip().str.upper()
#             return int((s == "YES").sum())

#         def numeric_sum(df_in, col_name: str) -> float:
#             if col_name not in df_in.columns:
#                 return 0.0
#             return float(pd.to_numeric(df_in[col_name], errors="coerce").fillna(0).sum())

#         def numeric_nonnull(df_in, col_name: str) -> int:
#             if col_name not in df_in.columns:
#                 return 0
#             return int(pd.to_numeric(df_in[col_name], errors="coerce").notna().sum())

#         total_filtered = int(len(df_dash))
#         total_overall = int(len(data))

#         active_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'ACTIVE'])) if 'status' in df_dash.columns else 0
#         inactive_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'INACTIVE'])) if 'status' in df_dash.columns else 0
#         black_filtered = int(len(df_dash[df_dash['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in df_dash.columns else 0

#         active_overall = int(len(data[data['status'].str.upper() == 'ACTIVE'])) if 'status' in data.columns else 0
#         inactive_overall = int(len(data[data['status'].str.upper() == 'INACTIVE'])) if 'status' in data.columns else 0
#         black_overall = int(len(data[data['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in data.columns else 0

#         if 'category' in data.columns:
#             cat_series_overall = data['category'].astype(str).str.strip().str.upper()
#             cat_counts_overall = cat_series_overall.value_counts()
#         else:
#             cat_counts_overall = {}

#         # --- Row 1 ---
#         m1, m2, m3, m4 = st.columns(4)

#         with m4:
#             pct = (total_filtered/total_overall)*100 if total_overall > 0 else 0
#             metric_card(
#                 "Total Venues",
#                 total_filtered,
#                 pct=pct,
#             )



#         with m1:
#             pct = (active_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Active Venues",
#                 active_filtered,
#                 pct=pct,
#             )

#         with m2:
#             pct = (black_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Blacklisted Venues",
#                 black_filtered,
#                 pct=pct,
#             )

#         with m3:
#             pct = (inactive_filtered/total_filtered)*100 if total_filtered > 0 else 0
#             metric_card(
#                 "Inactive Venues",
#                 inactive_filtered,
#                 pct=pct,
#             )



#         # Insert a custom small gap (e.g., 10 pixels)
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)


#         # --- Interactive Executive Dashboards ---
#         st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#         graph_col1, graph_col2, graph_col3 = st.columns([1, 1, 1.05], gap="large")

#         with graph_col1:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Category Distribution</h3>", unsafe_allow_html=True)
#             if 'category' in df_dash.columns:
#                 cat_series = df_dash['category'].astype(str).str.strip().str.upper()
#                 cat_counts = cat_series.value_counts()

#                 labels = ["A", "B", "C", "BL", "BL-C"]
#                 values = [cat_counts.get(l, 0) for l in labels]
#                 colors = ["#C8F6C2", "#AEEBFE", "#FFF6A3", "#FECCCE", "#E9B6FA"]

#                 chart_rows = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0]
#                 if chart_rows:
#                     pie_labels = [r[0] for r in chart_rows]
#                     pie_values = [r[1] for r in chart_rows]
#                     pie_colors = [r[2] for r in chart_rows]
#                     total_pie = sum(pie_values)
#                     pie_text = [
#                         f"<b>{lab}</b><br>{(val/total_pie)*100:.1f}%"
#                         if total_pie and (val / total_pie) >= 0.06 else ""
#                         for lab, val in zip(pie_labels, pie_values)
#                     ]

#                     fig_pie = go.Figure(data=[go.Pie(
#                         labels=pie_labels,
#                         values=pie_values,
#                         sort=False,
#                         direction="clockwise",
#                         marker=dict(colors=pie_colors, line=dict(color='#FFFFFF', width=2)),
#                         text=pie_text,
#                         textinfo='text',
#                         textposition='inside',
#                         insidetextfont=dict(size=12, color='#243447', family='sans-serif'),
#                         hovertemplate="<b>%{label}</b><br>Count: %{value:,}<br>Share: %{percent}<extra></extra>"
#                     )])
#                     fig_pie.update_layout(
#                         showlegend=True,
#                         legend=dict(
#                             orientation="h",
#                             yanchor="bottom",
#                             y=-0.2,
#                             xanchor="center",
#                             x=0.5,
#                             font=dict(size=11, color="#334155")
#                         ),
#                         margin=dict(t=8, b=26, l=8, r=8),
#                         paper_bgcolor="rgba(0,0,0,0)",
#                         plot_bgcolor="rgba(0,0,0,0)",
#                         height=305
#                     )
#                     st.plotly_chart(fig_pie, use_container_width=True)

#         with graph_col2:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Infrastructure Readiness Index</h3>", unsafe_allow_html=True)
#             cctv_c = yes_count(df_dash, "cctv")
#             pwd_c = yes_count(df_dash, "pwd_venue")
#             genset_c = yes_count(df_dash, "genset_availability")
#             ups_c = yes_count(df_dash, "ups_availability")

#             total_active = total_filtered if total_filtered > 0 else 1

#             r_values = [
#                 (cctv_c/total_active)*100,
#                 (pwd_c/total_active)*100,
#                 (ups_c/total_active)*100,
#                 (genset_c/total_active)*100
#             ]
#             r_labels = ['CCTV', 'PWD Ready', 'UPS', 'Genset']

#             r_values_loop = r_values + [r_values[0]]
#             r_labels_loop = r_labels + [r_labels[0]]

#             fig_radar = go.Figure()
#             fig_radar.add_trace(go.Scatterpolar(
#                 r=r_values_loop,
#                 theta=r_labels_loop,
#                 mode='lines+markers',
#                 fill='toself',
#                 fillcolor='rgba(129, 205, 236, 0.30)',
#                 line=dict(color='#236AA5', width=2.6),
#                 marker=dict(size=6, color='#236AA5', line=dict(color='#FFFFFF', width=1.1)),
#                 hoverinfo="theta+r",
#                 hovertemplate="<b>%{theta}</b><br>Readiness: %{r:.1f}%<extra></extra>"
#             ))
#             fig_radar.update_layout(
#                 polar=dict(
#                     bgcolor='rgba(255,255,255,0.45)',
#                     radialaxis=dict(
#                         visible=True,
#                         range=[0, 100],
#                         tickmode='linear',
#                         tick0=0,
#                         dtick=20,
#                         showline=True,
#                         linewidth=1,
#                         linecolor='#9AAFC3',
#                         gridcolor='#C6D5E5',
#                         gridwidth=1.2,
#                         tickfont=dict(size=10, color='#5F7285')
#                     ),
#                     angularaxis=dict(
#                         showline=True,
#                         linecolor='#AABDD0',
#                         gridcolor='#DCE7F2',
#                         gridwidth=1,
#                         tickfont=dict(size=12, color='#2D3748', family='sans-serif')
#                     )
#                 ),
#                 showlegend=False,
#                 margin=dict(t=20, b=5, l=28, r=28),
#                 paper_bgcolor="rgba(0,0,0,0)",
#                 plot_bgcolor="rgba(0,0,0,0)",
#                 height=305
#             )
#             st.plotly_chart(fig_radar, use_container_width=True)

#         with graph_col3:
#             st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>India Coverage Map</h3>", unsafe_allow_html=True)
#             map_df, india_geojson, _, _ = build_india_map_dataframe(df_dash)
#             if india_geojson is None:
#                 st.warning(f"GeoJSON not found at: {INDIA_GEOJSON_PATH}")
#             else:
#                 fig_map_preview = create_india_map_figure(
#                     map_df,
#                     india_geojson,
#                     height=250,
#                     show_data=False,
#                     zoom=2.35,
#                     center=INDIA_MAP_CENTER,
#                     opacity=0.9,
#                     line_color="#7C96B5",
#                     line_width=0.85,
#                 )
#                 if fig_map_preview is not None:
#                     id_to_state_map_preview = {
#                         str(row["id"]): str(row["state"])
#                         for _, row in map_df.iterrows()
#                     }
#                     selected_state_from_preview_click = None
#                     try:
#                         map_event = st.plotly_chart(
#                             fig_map_preview,
#                             use_container_width=True,
#                             key="india_map_preview_v5",
#                             on_select="rerun",
#                             selection_mode=("points",),
#                             config={"displayModeBar": False},
#                         )
#                         selected_state_from_preview_click = resolve_state_from_map_points(
#                             extract_selected_points(map_event),
#                             id_to_state_map_preview,
#                             ordered_names=map_df["state"].astype(str).tolist(),
#                         )
#                     except TypeError:
#                         st.plotly_chart(
#                             fig_map_preview,
#                             use_container_width=True,
#                             key="india_map_preview_static_v5",
#                             config={"displayModeBar": False},
#                         )

#                     if selected_state_from_preview_click:
#                         st.session_state.map_selected_state = selected_state_from_preview_click
#                         st.session_state.page = "india_map_detail"
#                         st.query_params["page"] = "india_map_detail"
#                         st.rerun()

#                     if st.button("Open Detailed Map", key="open_detailed_india_map", use_container_width=True):
#                         st.session_state.page = "india_map_detail"
#                         st.query_params["page"] = "india_map_detail"
#                         st.rerun()
             
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)


#         if not df_dash.empty:
#             st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Status Distribution</h3>", unsafe_allow_html=True)
#             status_col1, status_col2 = st.columns([1.5, 1.1], gap="medium")

#             c_data = df_dash.groupby(['region', 'status'], observed=False).size().reset_index(name='Count')
#             status_color_map = {"ACTIVE": "#81CDEC", "INACTIVE": "#CBD5E1", "BLACK LISTED": "#FECCCE"}
#             fig_stat = px.bar(c_data, x='region', y='Count', color='status', barmode='group', text='Count', color_discrete_map=status_color_map, category_orders={"region": REGION_ORDER})
#             fig_stat.update_traces(textposition='outside', textfont=dict(color='#334155', size=12), cliponaxis=False, marker_line_width=0)
#             fig_stat.update_layout(
#                 paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
#                 font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
#                 legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
#                 margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
#             )
#             with status_col1:
#                 st.plotly_chart(fig_stat, use_container_width=True)

#             pivot_status = df_dash.pivot_table(index='region', columns='status', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
#             status_row_order = [r for r in REGION_ORDER if r in pivot_status.index] + ["Grand Total"]
#             pivot_status = pivot_status.reindex(status_row_order)
            
#             headers_stat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_status.columns]
#             cells_stat = [pivot_status.index.tolist()] + [pivot_status[col].tolist() for col in pivot_status.columns]
            
#             fig_tab_stat = go.Figure(data=[go.Table(
#                 header=dict(values=headers_stat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
#                 cells=dict(values=cells_stat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_status)//2 + 1)][0][:len(pivot_status)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
#             )])
#             fig_tab_stat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")
            
#             with status_col2:
#                 st.plotly_chart(fig_tab_stat, use_container_width=True)


#             if 'category' in df_dash.columns:
#                 st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#                 st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Category Breakdown</h3>", unsafe_allow_html=True)
#                 cat_col1, cat_col2 = st.columns([1.5, 1.1], gap="medium")

#                 cat_data = df_dash.groupby(['region', 'category'], observed=False).size().reset_index(name='Count')
#                 color_map = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
#                 fig_cat = px.bar(cat_data, x='region', y='Count', color='category', barmode='group', text='Count', color_discrete_map=color_map, category_orders={"region": REGION_ORDER, "category": CAT_ORDER})
#                 fig_cat.update_traces(textposition='outside', textfont=dict(color='#334155', size=11), cliponaxis=False, marker_line_width=0)
#                 fig_cat.update_layout(
#                     paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
#                     font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
#                     legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
#                     margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
#                 )

#                 with cat_col1:
#                     st.plotly_chart(fig_cat, use_container_width=True)

#                 pivot_cat = df_dash.pivot_table(index='region', columns='category', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
#                 cat_col_order = [c for c in CAT_ORDER if c in pivot_cat.columns] + ["Grand Total"]
#                 pivot_cat = pivot_cat.reindex(index=status_row_order, columns=cat_col_order)
                
#                 headers_cat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_cat.columns]
#                 cells_cat = [pivot_cat.index.tolist()] + [pivot_cat[col].tolist() for col in pivot_cat.columns]
                
#                 fig_tab_cat = go.Figure(data=[go.Table(
#                     header=dict(values=headers_cat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
#                     cells=dict(values=cells_cat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_cat)//2 + 1)][0][:len(pivot_cat)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
#                 )])
#                 fig_tab_cat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")

#                 with cat_col2:
#                     st.plotly_chart(fig_tab_cat, use_container_width=True)

# # ================= INDIA MAP DETAIL PAGE =================
# elif st.session_state.page == "india_map_detail":
#     map_head_col, map_back_col = st.columns([4, 1], vertical_alignment="center")
#     with map_head_col:
#         st.markdown(
#             """
#             <div style="margin-bottom:8px;">
#                 <h2 style="margin:0; color:#1E293B; font-weight:800;">India Venue Coverage Map</h2>
#                 <p style="margin:4px 0 0 0; color:#64748B; font-size:13px;">Detailed state-wise view (click/hover for counts)</p>
#             </div>
#             """,
#             unsafe_allow_html=True,
#         )
#     with map_back_col:
#         if st.button("\u2190 Back to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.query_params["page"] = "dashboard"
#             st.rerun()

#     all_files = [
#         os.path.join(OUTPUT_PATH, f)
#         for f in os.listdir(OUTPUT_PATH)
#         if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")
#     ]

#     detail_df = pd.DataFrame()
#     if all_files:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         detail_df = pd.read_excel(latest_file_path, sheet_name=0)
#         detail_df.columns = detail_df.columns.str.lower().str.replace(" ", "_")
#         st.caption(f"Data Source: {os.path.basename(latest_file_path)}")
#     else:
#         st.warning("No processed reports found. Showing map boundaries only.")

#     map_df, india_geojson, mapped_states, mapped_venues = build_india_map_dataframe(detail_df)
#     if india_geojson is None:
#         st.error(f"India GeoJSON file not found: {INDIA_GEOJSON_PATH}")
#     else:
#         map_total_states = int(len(map_df))
#         data_total_venues = int(len(detail_df)) if not detail_df.empty else 0

#         mk1, mk2, mk3 = st.columns(3)
#         with mk1:
#             metric_card_compact("States in Map", map_total_states, accent="#7CAFD2")
#         with mk2:
#             metric_card_compact("States with Data", mapped_states, accent="#4B89B8")
#         with mk3:
#             metric_card_compact("Mapped Venues", mapped_venues if mapped_venues else data_total_venues, accent="#236AA5")

#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         fig_map_detail = create_india_map_figure(
#             map_df,
#             india_geojson,
#             height=760,
#             show_data=True,
#             zoom=2.45,
#             center=INDIA_MAP_CENTER,
#             opacity=0.92,
#             map_style="white-bg",
#             color_power=0.55,
#             line_color="#334155",
#             line_width=1.05,
#             restrict_to_india=True,
#         )
#         selected_state_from_click = None
#         id_to_state_map = {
#             str(row["id"]): str(row["state"])
#             for _, row in map_df.iterrows()
#         }

#         if fig_map_detail is not None:
#             try:
#                 detail_map_event = st.plotly_chart(
#                     fig_map_detail,
#                     use_container_width=True,
#                     key="india_map_detail_chart_v5",
#                     on_select="rerun",
#                     selection_mode=("points",),
#                     config={"displayModeBar": False, "scrollZoom": False},
#                 )
#                 selected_points = extract_selected_points(detail_map_event)
#                 selected_state_from_click = resolve_state_from_map_points(
#                     selected_points,
#                     id_to_state_map,
#                     ordered_names=map_df["state"].astype(str).tolist(),
#                 )
#                 selected_click_signature = get_last_point_signature(selected_points)
#             except TypeError:
#                 st.plotly_chart(
#                     fig_map_detail,
#                     use_container_width=True,
#                     key="india_map_detail_chart_static_v5",
#                     config={"displayModeBar": False, "scrollZoom": False},
#                 )
#                 selected_click_signature = None
#         else:
#             selected_click_signature = None

#         if selected_state_from_click and selected_click_signature:
#             prev_sig = st.session_state.get("_india_map_detail_last_click_sig")
#             if selected_click_signature != prev_sig:
#                 st.session_state["_india_map_detail_last_click_sig"] = selected_click_signature
#                 st.session_state.map_selected_state = selected_state_from_click
#         elif selected_click_signature is None:
#             st.session_state["_india_map_detail_last_click_sig"] = st.session_state.get("_india_map_detail_last_click_sig", None)

#         if selected_state_from_click and not selected_click_signature:
#             st.session_state.map_selected_state = selected_state_from_click

#         state_totals = (
#             map_df.groupby("state", as_index=False)["venue_count"]
#             .max()
#             .sort_values(["venue_count", "state"], ascending=[False, True], kind="stable")
#         )
#         available_states = state_totals["state"].astype(str).tolist()

#         if available_states:
#             if st.session_state.map_selected_state not in available_states:
#                 st.session_state.map_selected_state = available_states[0]

#             selected_state = st.selectbox(
#                 "Selected State",
#                 options=available_states,
#                 index=available_states.index(st.session_state.map_selected_state),
#                 key="india_detail_state_picker",
#             )
#             if selected_state != st.session_state.map_selected_state:
#                 st.session_state.map_selected_state = selected_state
#                 st.session_state["_india_map_detail_last_click_sig"] = None
#             selected_state = st.session_state.map_selected_state
#             st.caption(f"Current State: {selected_state}")
#         else:
#             selected_state = None

#         if selected_state and not detail_df.empty:
#             state_df = build_state_detail_payload(detail_df, selected_state)
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             st.markdown(
#                 f"<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>{html.escape(selected_state)} Detailed Snapshot</h3>",
#                 unsafe_allow_html=True,
#             )
#             if state_df.empty:
#                 st.info(f"No records available for {selected_state} in current output.")
#             else:
#                 status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([""] * len(state_df))
#                 venue_type_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([""] * len(state_df))
#                 total_state = int(len(state_df))
#                 active_state = int((status_series == "ACTIVE").sum())
#                 inactive_state = int((status_series == "INACTIVE").sum())
#                 black_state = int(status_series.str.contains("BLACK", na=False).sum())
#                 datc_state = int((venue_type_series == "DATC").sum())
#                 dotc_state = int((venue_type_series == "DOTC").sum())

#                 sk1, sk2, sk3, sk4, sk5, sk6 = st.columns(6)
#                 with sk1:
#                     metric_card_compact("State Total", total_state, accent="#236AA5")
#                 with sk2:
#                     metric_card_compact("DATC", datc_state, accent="#0EA5E9")
#                 with sk3:
#                     metric_card_compact("DOTC", dotc_state, accent="#6366F1")
#                 with sk4:
#                     metric_card_compact("Active", active_state, accent="#10B981")
#                 with sk5:
#                     metric_card_compact("Inactive", inactive_state, accent="#94A3B8")
#                 with sk6:
#                     metric_card_compact("Blacklisted", black_state, accent="#EF4444")

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 render_dotc_datc_summary(f"{selected_state} Summary", compute_dotc_datc_summary(state_df))

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 state_slug = resolve_state_slug(selected_state)
#                 state_geojson = load_state_geojson_by_slug(state_slug) if state_slug else None

#                 if st.session_state.map_selected_district_state != selected_state:
#                     st.session_state.map_selected_district = None
#                     st.session_state.map_selected_district_state = selected_state

#                 district_map_df = pd.DataFrame()
#                 district_id_map = {}

#                 if state_geojson is not None:
#                     district_map_df, district_id_map = build_district_map_dataframe(state_df, state_geojson)
#                     if not district_map_df.empty:
#                         dplot = district_map_df.copy()
#                         dplot["_color_metric"] = np.power(pd.to_numeric(dplot["venue_count"], errors="coerce").fillna(0).astype(float), 0.55)
#                         fig_district_map = px.choropleth(
#                             dplot,
#                             geojson=state_geojson,
#                             locations="id",
#                             featureidkey="id",
#                             color="_color_metric",
#                             hover_name="district",
#                             custom_data=["district", "id", "venue_count", "datc_count", "dotc_count", "active_count", "inactive_count", "black_count"],
#                             color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
#                             range_color=(0, max(1.0, float(dplot["_color_metric"].max()))),
#                         )
#                         fig_district_map.update_traces(
#                             marker_line_color="#334155",
#                             marker_line_width=1.0,
#                             hovertemplate=(
#                                 "<b>%{hovertext}</b><br>"
#                                 "Total: %{customdata[2]:,}<br>"
#                                 "DATC: %{customdata[3]:,}<br>"
#                                 "DOTC: %{customdata[4]:,}<br>"
#                                 "Active: %{customdata[5]:,}<br>"
#                                 "Inactive: %{customdata[6]:,}<br>"
#                                 "Blacklisted: %{customdata[7]:,}<extra></extra>"
#                             ),
#                         )
                        
#                         id_to_dname = dict(zip(dplot["id"], dplot["district"]))
#                         dlons, dlats, dtexts = [], [], []
#                         if state_geojson and "features" in state_geojson:
#                             for f in state_geojson["features"]:
#                                 fid = f.get("id")
#                                 if fid in id_to_dname:
#                                     cx, cy = _get_feature_centroid(f)
#                                     dlons.append(cx)
#                                     dlats.append(cy)
#                                     dtexts.append(id_to_dname[fid])
                        
#                         if dlons:
#                             fig_district_map.add_trace(go.Scattergeo(
#                                 lon=dlons,
#                                 lat=dlats,
#                                 text=dtexts,
#                                 mode="text",
#                                 textfont=dict(color="#1E293B", size=9, family="Inter, sans-serif"),
#                                 showlegend=False,
#                                 hoverinfo="skip"
#                             ))
                            
#                         fig_district_map.update_geos(fitbounds="locations", visible=False)
#                         fig_district_map.update_layout(
#                             clickmode="event+select",
#                             dragmode="pan",
#                             margin=dict(t=0, b=0, l=0, r=0),
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             height=560,
#                             coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.75),
#                         )

#                         selected_district_from_click = None
#                         try:
#                             district_map_event = st.plotly_chart(
#                                 fig_district_map,
#                                 use_container_width=True,
#                                 key=f"district_map_{state_slug}",
#                                 on_select="rerun",
#                                 selection_mode=("points",),
#                             )
#                             selected_district_from_click = resolve_state_from_map_points(
#                                 extract_selected_points(district_map_event),
#                                 district_id_map,
#                                 ordered_names=district_map_df["district"].astype(str).tolist(),
#                             )
#                         except TypeError:
#                             st.plotly_chart(fig_district_map, use_container_width=True, key=f"district_map_{state_slug}_static")

#                         if selected_district_from_click:
#                             st.session_state.map_selected_district = selected_district_from_click

#                         district_options = district_map_df["district"].astype(str).tolist()
#                         if district_options:
#                             if st.session_state.map_selected_district not in district_options:
#                                 district_options_sorted = (
#                                     district_map_df.sort_values("venue_count", ascending=False)["district"].astype(str).tolist()
#                                 )
#                                 st.session_state.map_selected_district = district_options_sorted[0]

#                             picked_district = st.selectbox(
#                                 f"Selected District in {selected_state}",
#                                 options=district_options,
#                                 index=district_options.index(st.session_state.map_selected_district),
#                                 key=f"district_picker_{state_slug}",
#                             )
#                             if picked_district != st.session_state.map_selected_district:
#                                 st.session_state.map_selected_district = picked_district

#                             selected_district = st.session_state.map_selected_district
#                             st.caption(f"Current District: {selected_district}")
#                             district_df = state_df[
#                                 state_df["district"].fillna("").astype(str).apply(normalize_district_name) == normalize_district_name(selected_district)
#                             ].copy()

#                             st.markdown(
#                                 f"<h4 style='color:#1E293B; margin-top:6px; margin-bottom:8px;'>{html.escape(selected_district)} District Stats</h4>",
#                                 unsafe_allow_html=True,
#                             )

#                             if district_df.empty:
#                                 st.info(f"No data available for {selected_district}.")
#                             else:
#                                 d_status_series = district_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in district_df.columns else pd.Series([""] * len(district_df))
#                                 d_vtype_series = district_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in district_df.columns else pd.Series([""] * len(district_df))
#                                 d_total = int(len(district_df))
#                                 d_datc = int((d_vtype_series == "DATC").sum())
#                                 d_dotc = int((d_vtype_series == "DOTC").sum())
#                                 d_active = int((d_status_series == "ACTIVE").sum())
#                                 d_inactive = int((d_status_series == "INACTIVE").sum())
#                                 d_black = int(d_status_series.str.contains("BLACK", na=False).sum())

#                                 dk1, dk2, dk3, dk4, dk5, dk6 = st.columns(6)
#                                 with dk1:
#                                     metric_card_compact("District Total", d_total, accent="#1D4ED8")
#                                 with dk2:
#                                     metric_card_compact("DATC", d_datc, accent="#0EA5E9")
#                                 with dk3:
#                                     metric_card_compact("DOTC", d_dotc, accent="#6366F1")
#                                 with dk4:
#                                     metric_card_compact("Active", d_active, accent="#10B981")
#                                 with dk5:
#                                     metric_card_compact("Inactive", d_inactive, accent="#94A3B8")
#                                 with dk6:
#                                     metric_card_compact("Blacklisted", d_black, accent="#EF4444")

#                                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                                 render_dotc_datc_summary(f"{selected_district} Summary", compute_dotc_datc_summary(district_df))

#                                 dc1, dc2 = st.columns([1.2, 1], gap="large")
#                                 with dc1:
#                                     d_cat_counts = (
#                                         district_df["category"].fillna("NA").astype(str).str.strip().replace("", "NA").value_counts().reset_index()
#                                         if "category" in district_df.columns else pd.DataFrame(columns=["index", "category"])
#                                     )
#                                     if not d_cat_counts.empty:
#                                         d_cat_counts.columns = ["Category", "Count"]
#                                         fig_d_cat = px.bar(
#                                             d_cat_counts,
#                                             x="Category",
#                                             y="Count",
#                                             text="Count",
#                                             color="Category",
#                                             color_discrete_map={"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"},
#                                         )
#                                         fig_d_cat.update_traces(textposition="outside", cliponaxis=False)
#                                         fig_d_cat.update_layout(
#                                             paper_bgcolor="rgba(0,0,0,0)",
#                                             plot_bgcolor="rgba(0,0,0,0)",
#                                             margin=dict(t=20, b=10, l=10, r=10),
#                                             xaxis=dict(title=""),
#                                             yaxis=dict(title="Venue Count", gridcolor="#E2E8F0"),
#                                             showlegend=False,
#                                             height=300,
#                                         )
#                                         st.plotly_chart(fig_d_cat, use_container_width=True, key=f"district_category_{state_slug}_{normalize_district_name(selected_district)}")
#                                     else:
#                                         st.info("Category data unavailable for this district.")
#                                 with dc2:
#                                     infra_fields = [
#                                         ("CCTV Ready", "cctv", "#3B82F6"),
#                                         ("PWD Ready", "pwd_venue", "#14B8A6"),
#                                         ("UPS Ready", "ups_availability", "#6366F1"),
#                                         ("Genset Ready", "genset_availability", "#0EA5E9"),
#                                     ]
#                                     for label, col_name, color_code in infra_fields:
#                                         yes_count = 0
#                                         if col_name in district_df.columns:
#                                             yes_count = int((district_df[col_name].fillna("").astype(str).str.strip().str.upper() == "YES").sum())
#                                         pct_val = (yes_count / d_total) * 100 if d_total > 0 else 0
#                                         metric_card_compact(label, f"{pct_val:.1f}%", accent=color_code)
#                         else:
#                             st.info("District list not available for this state map.")
#                 else:
#                     st.warning(f"District map file not available for {selected_state}.")

#                 st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                 sd_col1, sd_col2 = st.columns([1.25, 1], gap="large")
#                 with sd_col1:
#                     if "district" in state_df.columns:
#                         district_counts = (
#                             state_df["district"]
#                             .fillna("Unknown")
#                             .astype(str)
#                             .str.strip()
#                             .replace("", "Unknown")
#                             .value_counts()
#                             .head(15)
#                             .reset_index()
#                         )
#                         district_counts.columns = ["District", "Venues"]
#                         fig_dist = px.bar(
#                             district_counts,
#                             x="District",
#                             y="Venues",
#                             text="Venues",
#                             color_discrete_sequence=["#4299E1"],
#                         )
#                         fig_dist.update_traces(textposition="outside", cliponaxis=False)
#                         fig_dist.update_layout(
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             xaxis=dict(title="", tickangle=-30, showgrid=False),
#                             yaxis=dict(title="Venue Count", showgrid=True, gridcolor="#E2E8F0"),
#                             margin=dict(t=20, b=10, l=10, r=10),
#                             height=320,
#                             showlegend=False,
#                         )
#                         st.plotly_chart(fig_dist, use_container_width=True, key=f"district_breakdown_{normalize_state_name(selected_state)}")
#                     else:
#                         st.info("District-wise details are not available in this dataset.")
#                 with sd_col2:
#                     if "category" in state_df.columns:
#                         cat_counts = state_df["category"].fillna("NA").astype(str).str.strip().replace("", "NA").value_counts().reset_index()
#                         cat_counts.columns = ["Category", "Count"]
#                         fig_cat_state = px.pie(
#                             cat_counts,
#                             names="Category",
#                             values="Count",
#                             color="Category",
#                             color_discrete_map={"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"},
#                             hole=0.35,
#                         )
#                         fig_cat_state.update_layout(
#                             paper_bgcolor="rgba(0,0,0,0)",
#                             plot_bgcolor="rgba(0,0,0,0)",
#                             margin=dict(t=20, b=10, l=10, r=10),
#                             height=320,
#                             legend=dict(orientation="h", y=-0.12),
#                         )
#                         st.plotly_chart(fig_cat_state, use_container_width=True, key=f"category_breakdown_{normalize_state_name(selected_state)}")
#                     else:
#                         st.info("Category-wise details are not available in this dataset.")

#                 if normalize_state_name(selected_state) == "BIHAR":
#                     st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#                     st.markdown("##### Bihar District Reference (Attached PPT)")
#                     ref_images = load_ppt_reference_images(BIHAR_REFERENCE_PPTX_PATH, max_images=6)
#                     if ref_images:
#                         img_col1, img_col2 = st.columns(2, gap="medium")
#                         for idx, (img_name, img_bytes) in enumerate(ref_images):
#                             with (img_col1 if idx % 2 == 0 else img_col2):
#                                 st.image(img_bytes, caption=f"Reference {idx + 1}: {img_name}", use_container_width=True)
#                     else:
#                         st.info("Reference PPT found, but image previews were not extractable.")
#                     if os.path.exists(BIHAR_REFERENCE_PPTX_PATH):
#                         with open(BIHAR_REFERENCE_PPTX_PATH, "rb") as ref_file:
#                             st.download_button(
#                                 "Download Bihar District Reference PPT",
#                                 data=ref_file.read(),
#                                 file_name=os.path.basename(BIHAR_REFERENCE_PPTX_PATH),
#                                 mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
#                                 use_container_width=False,
#                             )

#         state_summary = state_totals[state_totals["venue_count"] > 0].head(12)
#         if not state_summary.empty:
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             with st.expander("Top States by Venue Count", expanded=False):
#                 st.dataframe(
#                     state_summary.rename(columns={"state": "State", "venue_count": "Venue Count"})[["State", "Venue Count"]],
#                     use_container_width=True,
#                     hide_index=True,
#                 )

# # ================= LOGIN PAGE =================
# elif st.session_state.page == "login":
#     # Inject large top spacing to vertically center the card mentally
#     st.markdown('<div class="gap-lg"></div><div class="gap-lg"></div>', unsafe_allow_html=True)
    
#     col_l, col_center, col_r = st.columns([1, 1.2, 1])
    
#     with col_center:
#         # High-End Security Icon & Header
#         st.markdown(f"""
#         <div style="text-align: center; margin-bottom: 30px;">
#             <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect>
#                 <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
#             </svg>
#             <h2 style="color:var(--primary-dark); font-weight:800; text-transform:uppercase; letter-spacing:1px; margin-top:16px;">Secure Executive Portal</h2>
#             <p style="color:#64748B; font-size:14px; margin-top:-8px;">Authorized Administrative Personnel Only</p>
#         </div>
#         """, unsafe_allow_html=True)
        
#         with st.form("login_form", border=True):
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             u = st.text_input("Username", placeholder="e.g. Admin")
#             st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#             p = st.text_input("Password", type="password", placeholder="\u2022\u2022\u2022\u2022\u2022\u2022\u2022\u2022")
#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
            
#             # Using primary type for maximum focus
#             if st.form_submit_button("Authenticate Access", type="primary", use_container_width=True):
#                 if u in USERS and USERS[u] == p:
#                     st.session_state.logged_in, st.session_state.user, st.session_state.page = True, u, "dashboard"
#                     st.rerun()
#                 else: 
#                     st.error("Authentication Failed: Invalid Credentials")
                
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()

# # ================= MASTER DATA VIEW =================
# elif st.session_state.page == "main":
#     allowed = [k for k, v in FILES.items() if st.session_state.user in v["users"]]
#     st.session_state.dataset_key = st.sidebar.radio("ACTIVE MASTER", allowed)
#     if st.query_params.get("dataset", None) != st.session_state.dataset_key:
#         st.query_params["dataset"] = st.session_state.dataset_key
#     current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
#     current_df = load_smart_data(current_path)

#     # --- CEO Level Header Layout ---
#     head_col1, head_col2 = st.columns([3, 1])
#     with head_col1:
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:14px; margin-bottom: 10px;">
#             <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <ellipse cx="12" cy="5" rx="9" ry="3"></ellipse>
#                 <path d="M21 12c0 1.66-4 3-9 3s-9-1.34-9-3"></path>
#                 <path d="M3 5v14c0 1.66 4 3 9 3s9-1.34 9-3V5"></path>
#             </svg>
#             <div>
#                 <h2 style="margin:0; font-weight:800; color:#1E293B;">Master Data Center</h2>
#                 <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Active Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
#             </div>
#         </div>
#         """, unsafe_allow_html=True)

#     with head_col2:
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()

#     if st.session_state.dataset_key == "Venue Inventory" and not current_df.empty:
#         status_series_main = (
#             current_df["status"].astype(str).str.strip().str.upper()
#             if "status" in current_df.columns else pd.Series([""] * len(current_df))
#         )
#         main_active = int((status_series_main == "ACTIVE").sum())
#         main_inactive = int((status_series_main == "INACTIVE").sum())
#         main_black = int(status_series_main.str.contains("BLACK", na=False).sum())
#         main_total = int(len(current_df))

#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         mk1, mk2, mk3, mk4 = st.columns(4, gap="small")
#         with mk1:
#             metric_card_compact("Active Venues", main_active, accent="#86d7a8")
#         with mk2:
#             metric_card_compact("Blacklisted", main_black, accent="#f4a3b5")
#         with mk3:
#             metric_card_compact("Inactive", main_inactive, accent="#b8c6d9")
#         with mk4:
#             metric_card_compact("Total Venues", main_total, accent="#81cdec")

#     st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
    
#     # --- Action Panel & Metrics ---
#     act_col1, act_col2, ext_col = st.columns([1.5, 3, 1.5])
#     with act_col1:
#         if st.button("\u2699 Open Advanced Editor", type="primary", use_container_width=True): 
#             st.session_state.page = "edit"
#             st.rerun()
            
#     with act_col2:
#         # Display dataset health quickly as premium pill badges
#         st.markdown(f"""
#         <div style="display:flex; gap:15px; padding-top:6px;">
#             <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Rows Scanned: {len(current_df):,}</span>
#             <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Columns Mapped: {len(current_df.columns)}</span>
#         </div>
#         """, unsafe_allow_html=True)
    
#     st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
#     st.dataframe(current_df, use_container_width=True, height=600)


# # ================= EDIT PAGE =================
# elif st.session_state.page == "edit":
#     st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
#     st.markdown(f"""
#     <div style="display:flex; align-items:center; gap:14px; margin-bottom: 25px;">
#         <svg width="36" height="36" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
#             <polyline points="16 18 22 12 16 6"></polyline>
#             <polyline points="8 6 2 12 8 18"></polyline>
#         </svg>
#         <div>
#             <h2 style="margin:0; font-weight:800; color:#1E293B;">Advanced Data Controller</h2>
#             <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Target Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
#         </div>
#     </div>
#     """, unsafe_allow_html=True)

#     current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
#     current_df = load_smart_data(current_path)

#     # --- TOP ACTIONS & FILTER ROW ---
#     st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#     col_search, col_save, col_clear, col_exit = st.columns([3.5, 1.5, 1.2, 1.2], gap="small")
    
#     with col_search:
#         search = st.text_input("Search", key="search_main_box", placeholder="\U0001F50D Search live database by keywords, target names...", label_visibility="collapsed")
#     with col_save: 
#         save_trigger = st.button("\U0001F4BE Commit Changes", type="primary", use_container_width=True)
#     with col_clear:
#         if st.button("\U0001F9F9 Clear", use_container_width=True):
#             st.session_state.search_main_box = ""
#             st.rerun()
#     with col_exit: 
#         if st.button("\u2190 Back", use_container_width=True):
#             st.session_state.page = "main"
#             st.rerun()
#     filtered = current_df.copy()
#     if search:
#         filtered = filtered[filtered.astype(str).apply(lambda x: x.str.contains(search, case=False).any(), axis=1)]

#     # --- BULK TOOLS (CLEANING & DUPLICATE CHECK) ---
#     if st.session_state.dataset_key != "Call Logs":
#         st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
#         with st.container(border=True):
#             st.subheader("Bulk Injection & Extraction Tools")
#             tool1, tool2 = st.columns([1, 2], gap="large")
            
#             with tool1:
#                 st.caption("Download Blank Template")
#                 blank_df = pd.DataFrame(columns=current_df.columns)
#                 buffer = io.BytesIO()
#                 with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#                     blank_df.to_excel(writer, index=False)
#                 st.download_button(label="\U0001F4E5 Download Format", data=buffer.getvalue(), 
#                                    file_name=f"Format_{FILES[st.session_state.dataset_key]['file']}", mime="application/vnd.ms-excel", use_container_width=True)
            
#             with tool2:
#                 st.caption("Bulk Upload Data (Clean & Update)")
#                 uploaded_file = st.file_uploader("Upload Excel", type=["xlsx"], label_visibility="collapsed")
                
#                 if uploaded_file:
#                     try:
#                         new_data = pd.read_excel(uploaded_file)

#                         d_master = find_dms_col(current_df)
#                         d_new = find_dms_col(new_data)

#                         if d_master and d_new:
#                             # Pre-clean the codes for matching
#                             new_data[d_new] = new_data[d_new].astype(str).str.replace(r'\s+', '', regex=True).str.upper()
#                             current_df[d_master] = current_df[d_master].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

#                             # Check length (informational)
#                             invalid_count = len(new_data[new_data[d_new].str.len() != 11])
#                             if invalid_count > 0:
#                                 st.warning(f"{invalid_count} codes in upload are not 11 digits.")

#                             if st.button("Confirm Bulk Update & Clean", use_container_width=True):
#                                 # 1. Set index to DMS Code for both dataframes
#                                 current_df.set_index(d_master, inplace=True)
#                                 new_data.set_index(d_new, inplace=True)

#                                 # 2. Update existing rows with new data
#                                 current_df.update(new_data)

#                                 # 3. Identify and append completely NEW rows
#                                 new_ids = new_data.index.difference(current_df.index)
#                                 new_rows = new_data.loc[new_ids]
                                
#                                 # 4. Combine and reset index back to original
#                                 final_df = pd.concat([current_df, new_rows]).reset_index()
                                
#                                 # Rename the index column back to the original Master column name if it changed
#                                 final_df.rename(columns={'index': d_master}, inplace=True)

#                                 # 5. Save to File
#                                 final_df.to_excel(current_path, index=False)
#                                 st.success(f"Done! {len(new_data)} rows processed (Updates + New entries).")
#                                 st.rerun()
#                         else:
#                             st.error(f"Missing ID Column! Master needs '{d_master}' and Upload needs a DMS/Venue Code column.")
                    
#                     except Exception as e:
#                         st.error(f"Error processing file: {e}")

#     # --- ML AUTO-TAGGING FOR CALL LOGS ---
#     if st.session_state.dataset_key == "Call Logs":
#         st.markdown("<hr style='margin: 30px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:10px; margin-bottom: 12px;">
#             <svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
#                 <rect x="2" y="2" width="20" height="8" rx="2" ry="2"></rect>
#                 <rect x="2" y="14" width="20" height="8" rx="2" ry="2"></rect>
#                 <line x1="6" y1="6" x2="6.01" y2="6"></line>
#                 <line x1="6" y1="18" x2="6.01" y2="18"></line>
#             </svg>
#             <h3 style="margin:0; color:#1E293B; font-size:18px;">AI Neural Auto-Classification Engine</h3>
#         </div>
#         """, unsafe_allow_html=True)
        
#         ml_col1, ml_col2 = st.columns(2, gap="large")
#         with ml_col1:
#             with st.container(border=True):
#                 st.subheader("1. Model Training Module")
#                 st.caption("Upload historical Call Logs data so the AI can learn human categorizations.")
    
#                 uploaded_train = st.file_uploader(
#                     "Upload Training Batch",
#                     type=["xlsx", "xls", "csv"],
#                     key="ml_train_uploader",
#                     label_visibility="collapsed",
#                     help="File must contain FINAL_ISSUE_REMARK_BY_PRAFULL (and other target columns if available).",
#                 )

#             if st.button("Train/Update ML Model", key="ml_train_btn", use_container_width=True):
#                 with st.spinner("Training Model... This may take a minute..."):
#                     import ml_engine

#                     if not uploaded_train:
#                         st.error("Please upload a training Excel file first.")
#                     else:
#                         train_ext = os.path.splitext(uploaded_train.name)[1].lower()
#                         if train_ext not in [".xlsx", ".xls", ".csv"]:
#                             train_ext = ".xlsx"
#                         train_path = os.path.join(BASE_PATH, f"temp_train_upload{train_ext}")
#                         with open(train_path, "wb") as f:
#                             f.write(uploaded_train.getbuffer())

#                         success, msg = ml_engine.train_ml_model(train_path)
#                         if success:
#                             st.success(f"{msg} (Source: uploaded file)")
#                         else:
#                             st.error(msg)
#         with ml_col2:
#             with st.container(border=True):
#                 st.subheader("2. Live Processing Module")
#                 st.caption("Upload newly downloaded raw Call Logs to classify and format them automatically.")
                
#                 uploaded_raw = st.file_uploader("Upload Raw Data Stream", type=["xlsx", "xls", "csv"], key="ml_uploader", label_visibility="collapsed")
#                 if uploaded_raw and st.button("\u26A1 Process with AI Engine", type="primary", use_container_width=True):
#                     with st.spinner("Processing..."):
#                         import ml_engine
#                         raw_ext = os.path.splitext(uploaded_raw.name)[1].lower()
#                         if raw_ext not in [".xlsx", ".xls", ".csv"]:
#                             raw_ext = ".xlsx"
#                         temp_path = os.path.join(BASE_PATH, f"temp_raw_upload{raw_ext}")
                        
#                         # --- AUTO-COMBINE Gateway/Core/Mock sheets if present ---
#                         try:
#                             if raw_ext in [".xlsx", ".xls"]:
#                                 xl = pd.ExcelFile(io.BytesIO(uploaded_raw.getvalue()))
#                                 sheet_names = xl.sheet_names
                                
#                                 # Check for Gateway/Core/Mock sheets
#                                 gwcm_sheets = [s for s in sheet_names if any(k in s.lower() for k in ["gateway", "core", "mock"])]
                                
#                                 if len(gwcm_sheets) > 1:
#                                     # Auto-combine multiple Gateway/Core/Mock sheets
#                                     st.info(f"Auto-combining {len(gwcm_sheets)} Gateway/Core/Mock sheets: {', '.join(gwcm_sheets)}")
#                                     frames = []
#                                     for s in gwcm_sheets:
#                                         df_part = xl.parse(s)
#                                         s_low = str(s).lower()
#                                         if "gateway" in s_low:
#                                             data_source = "GATEWAY"
#                                         elif "core" in s_low:
#                                             data_source = "CORE"
#                                         elif "mock" in s_low:
#                                             data_source = "MOCK"
#                                         else:
#                                             data_source = str(s).upper()
#                                         df_part.insert(0, "DATA_SOURCE", data_source)
#                                         df_part.insert(1, "SOURCE_SHEET", s)
#                                         frames.append(df_part)
                                    
#                                     combined_df = pd.concat(frames, ignore_index=True, sort=False)
#                                     # Save combined to single Excel file
#                                     with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
#                                         combined_df.to_excel(writer, sheet_name="Sheet1", index=False)
#                                     st.success(f"Combined {len(gwcm_sheets)} sheets: {len(combined_df):,} total rows")
#                                 else:
#                                     # Single sheet or no Gateway/Core/Mock sheets - save as is
#                                     with open(temp_path, "wb") as f:
#                                         f.write(uploaded_raw.getbuffer())
#                             else:
#                                 # CSV file - save as is
#                                 with open(temp_path, "wb") as f:
#                                     f.write(uploaded_raw.getbuffer())
#                         except Exception as e:
#                             st.warning(f"Could not auto-combine sheets: {e}")
#                             with open(temp_path, "wb") as f:
#                                 f.write(uploaded_raw.getbuffer())
                        
#                         # --- Process with ML ---
#                         res = ml_engine.process_call_logs(temp_path)
#                         if res.get("success"):
#                             btn1_file = res.get("output_file_btn1", res.get("output_file"))
#                             btn1_path = res.get("output_path_btn1", res.get("output_path"))
#                             btn2_file = res.get("output_file_btn2")
#                             btn2_path = res.get("output_path_btn2")

#                             if btn2_file:
#                                 st.success(
#                                     f"Processing Complete! Generated files: {btn1_file} and {btn2_file}"
#                                 )
#                             else:
#                                 st.success(f"Processing Complete! Saved to: {btn1_file}")
#                             ml_output = None
#                             try:
#                                 if btn1_path:
#                                     ml_output = pd.read_excel(btn1_path)
#                             except Exception:
#                                 pass
                            
#                             if "auto_approved" in res and "needs_review" in res and "auto_approved_threshold" in res:
#                                 st.write(
#                                     f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}"
#                                     f" | Auto-approved (\u2265{res['auto_approved_threshold']:.2f}): {res['auto_approved']}"
#                                     f" | Needs review: {res['needs_review']}"
#                                 )
#                             else:
#                                 st.write(f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}")
                            
#                             # Show preview if output was loaded successfully
#                             if ml_output is not None and not ml_output.empty:
#                                 st.info("Processed Data Preview (AI-tagged output)")
#                                 st.dataframe(ml_output.head(100), use_container_width=True)
                            
#                             dl_col1, dl_col2 = st.columns(2)
#                             with dl_col1:
#                                 if btn1_path and btn1_file:
#                                     with open(btn1_path, "rb") as f:
#                                         st.download_button(
#                                             label="Download Combined (New + Training)",
#                                             data=f,
#                                             file_name=btn1_file,
#                                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                                             use_container_width=True
#                                         )
#                             with dl_col2:
#                                 if btn2_path and btn2_file:
#                                     with open(btn2_path, "rb") as f:
#                                         st.download_button(
#                                             label="Download Combined (Training + New)",
#                                             data=f,
#                                             file_name=btn2_file,
#                                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                                             use_container_width=True
#                                         )
#                         else:
#                             st.error(res.get("msg", "Unknown error occurred."))

#     # --- DATA EDITOR ---
#     edited = st.data_editor(filtered, use_container_width=True, height=500, num_rows="dynamic", key="data_editor_tool")

#     if save_trigger:
#         # Manual Edit Cleaning
#         d_edit = find_dms_col(edited)
#         if d_edit:
#             edited[d_edit] = edited[d_edit].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

#         if search:
#             current_df.update(edited)
#             current_df.to_excel(current_path, index=False)
#         else:
#             edited.to_excel(current_path, index=False)
#         st.success("Changes Saved & Cleaned Successfully!")
#         st.rerun()
        

# # ================= PROCESS PAGE =================
# elif st.session_state.page == "process":
#     # --- CEO Level Header Layout ---
#     head_col1, head_col2 = st.columns([3, 1])
#     with head_col1:
#         st.markdown(f"""
#         <div style="display:flex; align-items:center; gap:14px; margin-bottom: 5px;">
#             <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
#                 <polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"></polygon>
#             </svg>
#             <div>
#                 <h2 style="margin:0; font-weight:800; color:#1E293B;">Central Analysis Engine</h2>
#                 <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Core Processing & Consolidation Node</p>
#             </div>
#         </div>
#         """, unsafe_allow_html=True)

#     with head_col2:
#         st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
#         if st.button("\u2190 Return to Dashboard", use_container_width=True):
#             st.session_state.page = "dashboard"
#             st.rerun()
    
#     st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)

#     with st.container(border=True):
#         st.markdown(f"""
#         <div style="margin-bottom: 20px;">
#             <h3 style="margin:0; color:#1E293B; font-size:18px;">Full Ecosystem Compilation</h3>
#             <p style="margin-top:4px; color:#64748B; font-size:14px;">Running the core engine will securely sequence Blacklisted Venues, FFA Data, Daily Observations, Market Reviews, and Machine-Learning Auto-Tagging into a singular unified Master Database.</p>
#         </div>
#         """, unsafe_allow_html=True)
        
#         col_run, col_space = st.columns([2, 3])
#         with col_run:
#             if st.button("\u26A1 Execute Core Analysis Consolidation", type="primary", use_container_width=True):
#                 status = st.empty()
#                 with st.spinner("Initializing Heavy Processing Sequence..."):
#                     try:
#                         # 1. LOAD
#                         vms = load_smart_data(os.path.join(BASE_PATH, "VMS Master.xlsx"))
#                         blacklisted = load_smart_data(os.path.join(BASE_PATH, "Blacklisted Venues.xlsx"))
#                         ffa = load_smart_data(os.path.join(BASE_PATH, "FFA Data.xlsx"))
#                         deo = load_smart_data(os.path.join(BASE_PATH, "Day End Observations.xlsx"))
#                         suspicious = load_smart_data(os.path.join(BASE_PATH, "Suspicious Activities.xlsx"))
#                         market_review = load_smart_data(os.path.join(BASE_PATH, "Google review and ratings.xlsx"))
#                         call_logs = load_smart_data(os.path.join(BASE_PATH, "Master Call Logs.xlsx"))
    
#                         # 2. ALIGN
#                         mapping = {"VENUE_CODE": "dms_code", "DMS_CODE": "dms_code", "DMS_Code": "dms_code", "DMS Code": "dms_code", "UPDATED_DMS_CODE": "dms_code", "DMS CODE": "dms_code", "Row Labels": "dms_code"}
#                         def align_and_lower(df_map):
#                             df_map.columns = df_map.columns.str.strip()
#                             for old, new in mapping.items():
#                                 if old in df_map.columns: df_map = df_map.rename(columns={old: new})
#                             df_map.columns = df_map.columns.str.lower().str.replace(" ", "_")
#                             return df_map
    
#                         vms, blacklisted, ffa = align_and_lower(vms), align_and_lower(blacklisted), align_and_lower(ffa)
#                         deo, suspicious, market_review, call_logs = align_and_lower(deo), align_and_lower(suspicious), align_and_lower(market_review), align_and_lower(call_logs)
    
#                         # 3. CLEANING
#                         def clean_dms(df_c, col):
#                             if col in df_c.columns:
#                                 df_c[col] = df_c[col].astype(str).str.replace(r'\.0$', '', regex=True).str.replace(r'\s+', '', regex=True).str.strip().str.upper()
#                             return df_c
#                         for d in [vms, blacklisted, ffa, deo, suspicious, market_review, call_logs]: clean_dms(d, "dms_code")
#                         vms = vms[vms["dms_code"].notna() & (~vms["dms_code"].isin(["NAN", "", "GRANDTOTAL"]))]
    
#                         # 4. MERGING
#                         rem_bl = 'remark' if 'remark' in blacklisted.columns else 'remarks'
#                         blacklisted["blacklist_info"] = blacklisted[rem_bl].fillna("").astype(str) + " by " + blacklisted["blacklisted_by"].fillna("").astype(str)
#                         merged = vms.merge(blacklisted[["dms_code", "blacklist_info"]], on="dms_code", how="left").fillna({"blacklist_info": ""})
#                         ffa_count = ffa.groupby("dms_code").size().reset_index(name="ffa_frequency")
#                         merged = merged.merge(ffa_count, on="dms_code", how="left").fillna({"ffa_frequency": 0})
    
#                         rem_deo = 'remark' if 'remark' in deo.columns else 'remarks'
#                         deo[rem_deo] = deo[rem_deo].fillna("").astype(str).str.strip()
#                         deo_summary = deo.groupby(["dms_code", rem_deo]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x[rem_deo], x["count"])), include_groups=False).reset_index(name="deo_remarks_summary")
#                         merged = merged.merge(deo_summary, on="dms_code", how="left").fillna({"deo_remarks_summary": ""})
    
#                         suspicious["issue_type"] = suspicious["issue_type"].fillna("").astype(str).str.strip()
#                         susp_summary = suspicious.groupby(["dms_code", "issue_type"]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x["issue_type"], x["count"])), include_groups=False).reset_index(name="suspicious_summary")
#                         merged = merged.merge(susp_summary, on="dms_code", how="left").fillna({"suspicious_summary": ""})

#                         # 4A. Call Logs Enrichment (moved from Call Logs AI module to Central Analysis Engine)
#                         def _first_non_blank(series: pd.Series) -> str:
#                             s = series.fillna("").astype(str).str.strip()
#                             s = s[~s.str.upper().isin(["", "NAN", "NONE", "NOT ASSIGNED", "NA", "-"])]
#                             return s.iloc[0] if not s.empty else ""

#                         if not call_logs.empty:
#                             if "project_name" in call_logs.columns:
#                                 call_logs["updated_project_name"] = clean_project_name(call_logs["project_name"])

#                             cl_dms_source = (
#                                 "updated_dms_code" if "updated_dms_code" in call_logs.columns else
#                                 "center_code" if "center_code" in call_logs.columns else
#                                 "dms_code" if "dms_code" in call_logs.columns else None
#                             )
#                             if cl_dms_source:
#                                 call_logs["updated_dms_code"] = (
#                                     call_logs[cl_dms_source]
#                                     .astype(str)
#                                     .str.replace(r"\.0$", "", regex=True)
#                                     .str.replace(r"\s+", "", regex=True)
#                                     .str.strip()
#                                     .str.upper()
#                                 )
#                                 if "dms_code" not in call_logs.columns:
#                                     call_logs["dms_code"] = call_logs["updated_dms_code"]
#                                 else:
#                                     call_logs["dms_code"] = call_logs["dms_code"].astype(str).str.strip()
#                                     blank_mask = call_logs["dms_code"].str.upper().isin(["", "NAN", "NONE"])
#                                     call_logs.loc[blank_mask, "dms_code"] = call_logs.loc[blank_mask, "updated_dms_code"]

#                             clean_dms(call_logs, "dms_code")
#                             cl_valid = call_logs[call_logs["dms_code"].notna()].copy()
#                             cl_valid["dms_code"] = cl_valid["dms_code"].astype(str).str.strip().str.upper()
#                             cl_valid = cl_valid[~cl_valid["dms_code"].isin(["", "NAN", "NONE", "GRANDTOTAL"])]

#                             if not cl_valid.empty and "dms_code" in vms.columns:
#                                 vms_cols_needed = ["dms_code", "name", "address", "landmark", "city", "district", "state", "pincode"]
#                                 vms_ref = vms.copy()
#                                 for vc in vms_cols_needed:
#                                     if vc not in vms_ref.columns:
#                                         vms_ref[vc] = ""
#                                 vms_ref = vms_ref[vms_cols_needed].drop_duplicates(subset="dms_code", keep="first")
#                                 vms_map_name = dict(zip(vms_ref["dms_code"], vms_ref.get("name", "")))
#                                 vms_map_addr = dict(zip(vms_ref["dms_code"], vms_ref.get("address", "")))
#                                 vms_map_land = dict(zip(vms_ref["dms_code"], vms_ref.get("landmark", "")))
#                                 vms_map_city = dict(zip(vms_ref["dms_code"], vms_ref.get("city", "")))
#                                 vms_map_dist = dict(zip(vms_ref["dms_code"], vms_ref.get("district", "")))
#                                 vms_map_state = dict(zip(vms_ref["dms_code"], vms_ref.get("state", "")))
#                                 vms_map_pin = dict(zip(vms_ref["dms_code"], vms_ref.get("pincode", "")))

#                                 mapped_name = cl_valid["dms_code"].map(vms_map_name).fillna("")
#                                 venue_fallback = (
#                                     cl_valid["venue_name"].fillna("").astype(str).str.strip().str.upper()
#                                     if "venue_name" in cl_valid.columns else pd.Series([""] * len(cl_valid), index=cl_valid.index)
#                                 )
#                                 cl_valid["updated_venue_name"] = mapped_name.where(mapped_name.astype(str).str.strip() != "", venue_fallback)

#                                 cl_valid["address"] = cl_valid["dms_code"].map(vms_map_addr).fillna(cl_valid.get("address", ""))
#                                 cl_valid["landmark"] = cl_valid["dms_code"].map(vms_map_land).fillna(cl_valid.get("landmark", ""))
#                                 cl_valid["city"] = cl_valid["dms_code"].map(vms_map_city).fillna(cl_valid.get("city", ""))
#                                 cl_valid["district"] = cl_valid["dms_code"].map(vms_map_dist).fillna(cl_valid.get("district", ""))
#                                 cl_valid["state"] = cl_valid["dms_code"].map(vms_map_state).fillna(cl_valid.get("state", ""))
#                                 cl_valid["pincode"] = cl_valid["dms_code"].map(vms_map_pin).fillna(cl_valid.get("pincode", ""))
#                                 cl_valid["concat_address"] = build_concat_address(cl_valid)

#                                 profile_aggs = {}
#                                 if "updated_project_name" in cl_valid.columns:
#                                     profile_aggs["call_logs_project_name"] = ("updated_project_name", _first_non_blank)
#                                 if "updated_venue_name" in cl_valid.columns:
#                                     profile_aggs["call_logs_updated_venue_name"] = ("updated_venue_name", _first_non_blank)
#                                 if "concat_address" in cl_valid.columns:
#                                     profile_aggs["call_logs_concat_address"] = ("concat_address", _first_non_blank)

#                                 if profile_aggs:
#                                     call_logs_profile = cl_valid.groupby("dms_code", as_index=False).agg(**profile_aggs)
#                                     merged = merged.merge(call_logs_profile, on="dms_code", how="left")
#                                     for col_name in profile_aggs.keys():
#                                         merged[col_name] = merged[col_name].fillna("")
    
#                         if "final_issue_remark_by_prafull" in call_logs.columns:
                            
#                             # Remove 'Other low impact issues' directly from remark column
#                             filtered_logs = call_logs[
#                                 ~call_logs["final_issue_remark_by_prafull"]
#                                 .fillna("")
#                                 .str.strip()
#                                 .str.lower()
#                                 .str.contains("other low impact issues", na=False)
#                             ]
                            
#                             cl_summary = (
#                                 filtered_logs
#                                 .groupby(["dms_code", "final_issue_remark_by_prafull"])
#                                 .size()
#                                 .reset_index(name="count")
#                                 .groupby("dms_code")
#                                 .apply(
#                                     lambda x: ", ".join(
#                                         f"{r}({c})" for r, c in zip(
#                                             x["final_issue_remark_by_prafull"], x["count"]
#                                         )
#                                     ),
#                                     include_groups=False
#                                 )
#                                 .reset_index(name="call_log_summary")
#                             )
    
#                             merged = merged.merge(cl_summary, on="dms_code", how="left").fillna({"call_log_summary": ""})
                            
                            
#                         market_review["google_ratings_num"] = pd.to_numeric(market_review["google_ratings"], errors="coerce")
#                         market_review["google_review_count_num"] = pd.to_numeric(market_review["google_review_count"], errors="coerce")
#                         m_cond = [market_review["google_review_count_num"] < 10, (market_review["google_ratings_num"] >= 4) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] <= 2) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] < 4) & (market_review["google_review_count_num"] >= 10)]
#                         m_choice = ["Doubtful (Reviews < 10)", "Good (Rating >=4 & Reviews >=10)", "Poor (Rating <=2 & Reviews >=10)", "Average (Rating <4 & Reviews >=10)"]
#                         market_review["market_remark"] = np.select(m_cond, m_choice, default="")
#                         merged = merged.merge(market_review[["dms_code", "market_remark"]], on="dms_code", how="left").fillna({"market_remark": ""})
    
    
#                         # 5. ANALYSIS RULES
#                         merged["final_analysis"] = "Good"; merged["reason"] = "Good"
#                         mask_comp = merged["blacklist_info"].str.lower().str.startswith("completely blacklisted", na=False)
#                         mask_cust = merged["blacklist_info"].str.lower().str.startswith("customer specific blacklisted", na=False)
                        
#                         # Assigning individually to avoid ValueError
#                         merged.loc[mask_comp, "final_analysis"] = "Completely Blacklisted"
#                         merged.loc[mask_comp, "reason"] = merged.loc[mask_comp, "blacklist_info"]
#                         merged.loc[mask_cust, "final_analysis"] = "Customer Specific Blacklisted"
#                         merged.loc[mask_cust, "reason"] = merged.loc[mask_cust, "blacklist_info"]
    
#                         nb = merged[~(mask_comp | mask_cust)].copy()
                        
#                         mask_ffa = nb["ffa_frequency"] >= 1
#                         nb.loc[mask_ffa, "final_analysis"] = np.where(nb.loc[mask_ffa, "ffa_frequency"] < 5, "Average", "Sensitive")
#                         nb.loc[mask_ffa, "reason"] = "Found in FFA | Count: " + nb.loc[mask_ffa, "ffa_frequency"].astype(int).astype(str)
    
#                         if "deo_remarks_summary" in nb.columns:
#                             mask_deo = (nb["deo_remarks_summary"] != "") & (nb["final_analysis"] == "Good")
#                             deo_counts = deo.groupby("dms_code").size()
#                             nb["deo_count"] = nb["dms_code"].map(deo_counts).fillna(0)
                            
#                             # Force Sensitive Venue if any comment exists
#                             nb.loc[mask_deo, "final_analysis"] = "Sensitive"
#                             nb.loc[mask_deo, "reason"] = nb.loc[mask_deo, "deo_remarks_summary"]
    
#                         if "suspicious_summary" in nb.columns:
#                             mask_susp = (nb["suspicious_summary"] != "") & (nb["final_analysis"] == "Good")
#                             susp_counts = suspicious.groupby("dms_code").size()
#                             nb["susp_count"] = nb["dms_code"].map(susp_counts).fillna(0)
                            
#                             # Force Sensitive Venue if any comment exists
#                             nb.loc[mask_susp, "final_analysis"] = "Sensitive"
#                             nb.loc[mask_susp, "reason"] = nb.loc[mask_susp, "suspicious_summary"]
    
#                         if "call_log_summary" in nb.columns:
#                             remark = nb["call_log_summary"].fillna("").str.strip()
                            
#                             mask_call = (remark != "") & (nb["final_analysis"] == "Good")
                            
#                             # Exclude anything containing 'Other low impact issues'
#                             mask_valid_call = ~remark.str.lower().str.contains("other low impact issues", na=False)
                            
#                             # Final mask
#                             mask_final = mask_call & mask_valid_call
    
#                             nb.loc[mask_final, "final_analysis"] = "Average"
#                             nb.loc[mask_final, "reason"] = "Call Logs | " + nb.loc[mask_final, "call_log_summary"]
                        
#                         if "market_remark" in nb.columns:
#                             # Normalize text
#                             remark = nb["market_remark"].fillna("").str.strip()
#                             call_remark = nb["call_log_summary"].fillna("").str.strip()
                            
#                             # Mask for rows where final_analysis is still Good
#                             mask_base = nb["final_analysis"] == "Good"
                            
#                             # [OK] Apply market logic ONLY if call_log_summary is blank OR only 'Other low impact issues'
#                             mask_call_valid = (
#                                 (call_remark == "") |
#                                 (call_remark.str.lower().str.contains("other low impact issues", na=False))
#                             )
                            
#                             # Final base mask
#                             mask_main = mask_base & mask_call_valid
                            
#                             # Check if remark contains doubtful/poor/average
#                             mask_issue = remark.str.lower().str.contains("doubtful|poor|average", na=False)
                            
#                             # --- Case 1: Issue found -> Average ---
#                             mask_avg = mask_main & mask_issue
#                             nb.loc[mask_avg, "final_analysis"] = "Average"
#                             nb.loc[mask_avg, "reason"] = "Market: " + nb.loc[mask_avg, "market_remark"]
                            
#                             # --- Case 2: Other remarks -> Good ---
#                             mask_good = mask_main & (remark != "") & (~mask_issue)
#                             nb.loc[mask_good, "final_analysis"] = "Good"
#                             nb.loc[mask_good, "reason"] = "Market: " + nb.loc[mask_good, "market_remark"]
                            
#                             # --- Case 3: Blank -> Good Venue ---
#                             mask_blank = mask_main & (remark == "")
#                             nb.loc[mask_blank, "final_analysis"] = "Good"
#                             nb.loc[mask_blank, "reason"] = "Good Venue"
                            
    
#                         final_result = pd.concat([merged[mask_comp | mask_cust], nb], ignore_index=True)
    
#                         # Rename output column: NAME -> VENUE_NAME
#                         final_result = final_result.rename(columns={"name": "venue_name"})
                        
#                         # --- 6. COLUMN WISE ANALYSIS ---
#                         analysis_results = []
#                         for col in final_result.columns:
#                             if col in ['deo_count', 'susp_count']: continue
#                             null_count = (final_result[col] == "").sum() + final_result[col].isna().sum()
#                             analysis_results.append({
#                                 "Column Name": col,
#                                 "Fill Rate (%)": round(((len(final_result) - null_count) / len(final_result)) * 100, 2),
#                                 "Unique Values": final_result[col].nunique(),
#                                 "Top Frequent": str(final_result[col].mode()[0]) if not final_result[col].empty else "N/A"
#                             })
#                         audit_df = pd.DataFrame(analysis_results)
    
#                         # Create Category column based on final_analysis
#                         category_map = {
#                             "Good": "A",
#                             "Average": "B",
#                             "Sensitive": "C",
#                             "Completely Blacklisted": "BL",
#                             "Customer Specific Blacklisted": "BL-C"
#                         }
    
#                         final_result["category"] = final_result["final_analysis"].map(category_map).fillna("")

#                         # DOTC default override: keep DATC untouched
#                         if "venue_type" in final_result.columns:
#                             dotc_mask = final_result["venue_type"].astype(str).str.strip().str.upper() == "DOTC"
#                             final_result.loc[dotc_mask, "final_analysis"] = "Good"
#                             final_result.loc[dotc_mask, "reason"] = "Good Venue"
#                             final_result.loc[dotc_mask, "category"] = "A"
                        
                        
#                         # Cleanup
#                         final_result = final_result.drop(columns=['deo_count', 'susp_count'], errors='ignore')
                        
    
#                         # 7. SAVE
#                         ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
#                         # Convert column names to uppercase for main file
#                         final_result.columns = [col.upper() for col in final_result.columns]
#                         audit_df.columns = [col.upper() for col in audit_df.columns]
    
#                         out_path = os.path.join(OUTPUT_PATH, f"VMS_Full_Analysis_{ts}.xlsx")
    
#                         # If available, also include the latest previous output in the new workbook
#                         prev_out_path = None
#                         try:
#                             prev_candidates = [
#                                 os.path.join(OUTPUT_PATH, f)
#                                 for f in os.listdir(OUTPUT_PATH)
#                                 if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and (not f.startswith("~$"))
#                             ]
#                             prev_candidates = [p for p in prev_candidates if os.path.abspath(p) != os.path.abspath(out_path)]
#                             if prev_candidates:
#                                 prev_out_path = max(prev_candidates, key=os.path.getmtime)
#                         except:
#                             prev_out_path = None
    
#                         with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
#                             final_result.to_excel(writer, sheet_name='Main_Analysis', index=False)
#                             audit_df.to_excel(writer, sheet_name='Column_Audit', index=False)
    
#                             run_info = pd.DataFrame([{
#                                 "CURRENT_OUTPUT_FILE": os.path.basename(out_path),
#                                 "PREVIOUS_OUTPUT_FILE": os.path.basename(prev_out_path) if prev_out_path else "",
#                                 "GENERATED_ON": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#                             }])
#                             run_info.to_excel(writer, sheet_name="Run_Info", index=False)
    
#                             if prev_out_path:
#                                 try:
#                                     old_main = pd.read_excel(prev_out_path, sheet_name="Main_Analysis")
#                                     old_main.to_excel(writer, sheet_name="Old_Main_Analysis", index=False)
#                                 except:
#                                     pass
#                                 try:
#                                     old_audit = pd.read_excel(prev_out_path, sheet_name="Column_Audit")
#                                     old_audit.to_excel(writer, sheet_name="Old_Column_Audit", index=False)
#                                 except:
#                                     pass
    
#                         # --- Inventory File ---
#                         inventory_cols = [
#                             "venue_type", "region", "state", "city", "district", "dms_code",
#                             "venue_name", "address", "pincode", "landmark", "complete_address",
#                             "venue_max_capacity", "pwd_venue", "status", "final_analysis", "reason", "category"
#                         ]
    
#                         inventory_df = final_result[[col.upper() for col in inventory_cols if col.upper() in final_result.columns]].copy()
    
#                         # Folder
#                         inventory_folder = os.path.join(OUTPUT_PATH, "Venue Inventory")
#                         os.makedirs(inventory_folder, exist_ok=True)
    
#                         inventory_path = os.path.join(inventory_folder, f"Venue Inventory as on {ts}.xlsx")
#                         inventory_df.to_excel(inventory_path, index=False)
    
#                         # --- APPLY FORMATTING FUNCTION ---
#                         def format_excel(file_path):
#                             wb = load_workbook(file_path)
                            
#                             header_font = Font(color="FFFFFF", bold=True)
#                             header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
#                             header_alignment = Alignment(horizontal="left")
    
#                             for sheet in wb.sheetnames:
#                                 ws = wb[sheet]
                                
#                                 for cell in ws[1]:  # First row (headers)
#                                     cell.font = header_font
#                                     cell.fill = header_fill
#                                     cell.alignment = header_alignment
    
#                             wb.save(file_path)
    
#                         # Apply formatting to both files
#                         format_excel(out_path)
#                         format_excel(inventory_path)
    
#                         # --- Status ---
#                         status.success(f"Generated Analysis Reports.")
#                         st.dataframe(final_result, use_container_width=True)
    
#                     except Exception as e: st.error(f"Error: {e}"); st.exception(e)


# # ================= VENUE DETAIL PAGE =================
# elif st.session_state.page == "venue_detail":
#     # Load data
#     all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH)
#                  if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]

#     if not all_files:
#         st.error("No analysis data found.")
#         if st.button("\u2190 Back to Dashboard"):
#             st.session_state.page = "dashboard"
#             st.rerun()
#     else:
#         latest_file_path = max(all_files, key=os.path.getmtime)
#         data = pd.read_excel(latest_file_path, sheet_name=0)
#         data.columns = data.columns.str.lower().str.replace(" ", "_")

#         venue_code = st.session_state.venue_detail_code
#         venue_row = data[data['dms_code'].astype(str).str.strip().str.upper() == str(venue_code).strip().upper()]

#         if venue_row.empty:
#             st.error(f"Venue with code **{venue_code}** not found.")
#             if st.button("\u2190 Back to Dashboard"):
#                 st.session_state.page = "dashboard"
#                 st.rerun()
#         else:
#             v = venue_row.iloc[0]

#             # Helper to safely get value
#             def _val(col, default="N/A"):
#                 val = v.get(col, default)
#                 if pd.isna(val) or str(val).strip().upper() in ["NAN", "NONE", ""]:
#                     return default
#                 res = str(val).strip()
#                 if res.endswith(".0") and res[:-2].replace('-', '').isdigit():
#                     res = res[:-2]
#                 return res

#             venue_name = _val("venue_name", _val("name", "Unknown Venue"))
#             dms_code = _val("dms_code")
#             category = _val("category")
#             final_analysis = _val("final_analysis")
#             status_val = _val("status")
#             region = _val("region")
#             city = _val("city")
#             state = _val("state")
#             district = _val("district")
#             venue_type = _val("venue_type")

#             # Category color
#             cat_colors = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
#             cat_color = cat_colors.get(category, "#DDE8F0")

#             # Analysis color
#             analysis_colors = {
#                 "Good": "#C8F6C2", "Average": "#FFF6A3", "Sensitive": "#FECCCE",
#                 "Completely Blacklisted": "#FDBED2", "Customer Specific Blacklisted": "#E9B6FA"
#             }
#             analysis_color = analysis_colors.get(final_analysis, "#DDE8F0")

#             # --- TOP BAR ---
#             top_left, top_right = st.columns([1, 4])
#             with top_left:
#                 if st.button("\u2190 Back to Dashboard", use_container_width=True):
#                     st.session_state.page = "dashboard"
#                     st.rerun()

#             # --- HERO HEADER ---
#             st.markdown(f"""
#             <div style="background: linear-gradient(135deg, white 0%, #EBF3F8 40%, #CBEAFC 100%);
#                         border-radius: 16px; padding: 20px 24px; margin: 8px 0 20px 0;
#                         box-shadow: 0 4px 16px rgba(0,0,0,0.05); position:relative; overflow:hidden;
#                         border: 1px solid #DDE8F0;">
#                 <div style="position:absolute; top:-20px; right:-20px; width:140px; height:140px;
#                             background:rgba(174,235,254,0.3); border-radius:50%;"></div>
#                 <div style="position:absolute; bottom:-30px; right:60px; width:90px; height:90px;
#                             background:rgba(174,235,254,0.2); border-radius:50%;"></div>
#                 <div>
#                     <div style="color:#2A3B4C; font-size:28px; font-weight:800; letter-spacing:-0.5px;
#                                 line-height:1.2;">{html.escape(venue_name)}</div>
#                     <div style="color:#566D82; font-size:14px; margin-top:4px; font-weight:600; letter-spacing:0.5px; text-transform:uppercase;">
#                         {html.escape(dms_code)} <span style="margin:0 6px; color:#839BB2;">|</span> {html.escape(venue_type)}
#                     </div>
#                 </div>
#                 <!-- Badges removed as per user request -->
#             </div>
#             """, unsafe_allow_html=True)

#             # --- VENUE PHOTO ---
#             venue_photo_url = _val("venue_photo_url", _val("image_url", ""))
            
#             # Check for local photo if URL is not provided
#             if not venue_photo_url or venue_photo_url in ["N/A", ""]:
#                 import glob
#                 import base64
#                 # Search for any image matching the exact DMS code (e.g. EST-AR-1161.jpg)
#                 photo_matches = glob.glob(os.path.join(PHOTOS_PATH, f"{dms_code}.*"))
#                 if photo_matches:
#                     # Filter for actual images
#                     img_files = [p for p in photo_matches if p.lower().endswith(('.png', '.jpg', '.jpeg', '.webp'))]
#                     if img_files:
#                         with open(img_files[0], "rb") as image_file:
#                             b64_str = base64.b64encode(image_file.read()).decode("utf-8")
#                             ext = os.path.splitext(img_files[0])[1][1:].lower()
#                             ext = "jpeg" if ext == "jpg" else ext
#                             venue_photo_url = f"data:image/{ext};base64,{b64_str}"
            
#             img_col, info_col, metrics_col = st.columns([1.2, 1.8, 1])

#             with img_col:
#                 photo_content = ""
#                 if venue_photo_url and venue_photo_url not in ["N/A", ""]:
#                     photo_content = f"""<div style="border-radius:12px; overflow:hidden; flex:1; min-height:0; background:var(--surface-muted); display:flex; justify-content:center; align-items:center;">
#     <img src="{venue_photo_url}" style="width:100%; height:100%; object-fit:cover;" />
# </div>"""
#                 else:
#                     photo_content = """<div style="border-radius:12px; flex:1; min-height:0; background:var(--surface-muted); border: 1px solid var(--border); display:flex; flex-direction:column; justify-content:center; align-items:center;">
#     <div style="width:48px; height:48px; border-radius:50%; background:var(--border-dark); margin-bottom:12px; display:flex; justify-content:center; align-items:center;">
#         <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="var(--surface)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"></rect><circle cx="8.5" cy="8.5" r="1.5"></circle><polyline points="21 15 16 10 5 21"></polyline></svg>
#     </div>
#     <span style="color:var(--text-muted); font-size:13px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Photo Not Available</span>
# </div>"""

#                 # Render entire block at 0 indentation to prevent Markdown falling into <pre> logic
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto; display:flex; flex-direction:column;">
#     <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Photo</div>
# {photo_content}
# </div>""", unsafe_allow_html=True)

#             with info_col:
#                 address = _val("address", "")
#                 landmark = _val("landmark", "")
#                 pincode = _val("pincode", "")
#                 complete_address = _val("complete_address", "")

#                 loc_items = [
#                     ("Region", region),
#                     ("State", state),
#                     ("City", city),
#                     ("District", district),
#                     ("Pincode", pincode),
#                     ("Landmark", landmark),
#                 ]

#                 loc_html = ""
#                 for label, val in loc_items:
#                     if val != "N/A":
#                         loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border);">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">{html.escape(val)}</span>
# </div>"""

#                 if complete_address != "N/A":
#                     loc_html += f"""<div style="margin-top:12px; padding:10px 14px; background:var(--surface-muted); border-radius:10px; border:1px solid var(--border);">
#     <div style="color:var(--text-muted); font-size:12px; font-weight:600; margin-bottom:4px;">FULL ADDRESS</div>
#     <div style="color:var(--text); font-size:13px; line-height:1.5;">{html.escape(complete_address)}</div>
# </div>"""

#                 google_coords = _val("google_cordinates", "")
#                 if google_coords != "N/A" and google_coords != "":
#                     loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border); margin-top:4px;">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">Google Location</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">
#         <a href="https://maps.google.com/?q={html.escape(google_coords)}" target="_blank" style="color:#2563EB; text-decoration:none;">{html.escape(google_coords)}</a>
#     </span>
# </div>"""

#                 added_on = _val("added_on-date", "")
#                 if added_on != "N/A" and added_on != "":
#                     loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border); margin-top:4px;">
#     <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">Added On</span>
#     <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">{html.escape(added_on)}</span>
# </div>"""

#                 header_text = "Location Profile"
#                 if added_on != "N/A" and added_on != "":
#                     header_text = f"Location Profile with added on date"
                    
#                 # Render entire profile block in a single markdown wrapper at 0 indentation
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto;">
#     <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">{header_text}</div>
# {loc_html}
# </div>""", unsafe_allow_html=True)

#             with metrics_col:
#                 def _mini_card(label, value, color="#2563EB"):
#                     return f"""<div style="flex:1; background:var(--surface-muted); border:1px solid var(--border); border-radius:8px; padding:6px; text-align:center; border-left:3px solid {color}; box-sizing:border-box; display:flex; flex-direction:column; justify-content:center; align-items:center;">
# <div style="font-size:17px; font-weight:800; color:{color}; line-height:1;">{html.escape(str(value))}</div>
# <div style="font-size:12px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.4px; margin-top:2px;">{label}</div>
# </div>"""

#                 cctv = _val("cctv")
#                 cctv_color = "#10B981" if cctv.upper() == "YES" else "#EF4444"
#                 pwd = _val("pwd_venue")
#                 pwd_color = "#10B981" if pwd.upper() == "YES" else "#EF4444"
#                 genset = _val("genset_availability")
#                 genset_color = "#10B981" if genset.upper() == "YES" else "#EF4444"
#                 ups = _val("ups_availability")
#                 ups_color = "#10B981" if ups.upper() == "YES" else "#EF4444"

#                 used_tested = _val("used_and_tested_venue")
#                 used_color = "#10B981" if used_tested.upper() == "YES" else "#EF4444"

#                 cards_html = (
#                     _mini_card("Max Capacity", _val("venue_max_capacity"), "#2563EB") +
#                     _mini_card("CCTV", cctv, cctv_color) +
#                     _mini_card("PWD Venue", pwd, pwd_color) +
#                     _mini_card("Genset", genset, genset_color) +
#                     _mini_card("UPS", ups, ups_color) +
#                     _mini_card("Used & Tested", used_tested.title(), used_color)
#                 )

#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:16px 14px; box-shadow:var(--shadow-sm); height:410px; box-sizing:border-box; display:flex; flex-direction:column;">
# <div style="font-size:12px; font-weight:700; color:var(--text-muted); margin-bottom:8px; text-transform:uppercase; letter-spacing:0.5px; flex-shrink:0;">Key Metrics</div>
# <div style="display:flex; flex-direction:column; flex:1; gap:5px; margin-top:0;">
# {cards_html}
# </div>
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             # --- VENUE INFO & CONTACT INFO ---
#             venue_info_col, contact_col = st.columns(2)

#             with venue_info_col:
#                 v_info_items = [
#                     ("Venue SPOC", _val("venue_spoc_name")),
#                     ("SPOC Contact", _val("spoc_contact_number")),
#                     ("SPOC Email", _val("spoc_email_id"))
#                 ]
#                 v_info_html = ""
#                 for label, val in v_info_items:
#                     if val != "N/A":
#                         v_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
# <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
# <span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
# </div>"""
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
# <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Information</div>
# {v_info_html}
# </div>""", unsafe_allow_html=True)

#             with contact_col:
#                 c_info_items = [
#                     ("Channel Manager", _val("channel_manager_name")),
#                     ("CM Contact", _val("channel_manager_contact_no.")),
#                     ("CM Email", _val("channel_manager_email_id")),
#                 ]
#                 c_info_html = ""
#                 for label, val in c_info_items:
#                     if val != "N/A":
#                         c_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
# <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
# <span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
# </div>"""
#                 st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
# <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">DEXIT SPOC CONTACT INFORMATION</div>
# {c_info_html}
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             is_dotc = str(venue_type).strip().upper() == "DOTC"
#             suppress_analysis_sections = is_dotc

#             # --- ANALYSIS & REMARKS ---
#             if not suppress_analysis_sections:
#                 st.markdown("""
#                 <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px;
#                             text-transform:uppercase; letter-spacing:0.5px;">Analysis & Remarks</div>
#                 """, unsafe_allow_html=True)

#             cat_val = _val("category", "N/A")
#             subcat_val = _val("subcategory", "N/A")

#             fields = []
#             if (not suppress_analysis_sections) and cat_val != "N/A" and cat_val != "" and cat_val != "0" and str(cat_val).lower() != "nan":
#                 fields.append(("Category", cat_val))
#             if subcat_val != "N/A" and subcat_val != "" and subcat_val != "0" and str(subcat_val).lower() != "nan":
#                 fields.append(("Subcategory", subcat_val))
#             if (not suppress_analysis_sections) and final_analysis != "N/A" and final_analysis != "" and final_analysis != "0" and final_analysis != "0.0":
#                 fields.append(("Venue Intel", final_analysis))

#             if fields:
#                 inner_html = ""
#                 for lbl, field_val in fields:
#                     inner_html += f"""
# <div style="margin-right:48px; margin-bottom:8px;">
#     <div style="font-size:11px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:4px;">{lbl}</div>
#     <div style="font-size:14px; color:#1A202C; line-height:1.6; font-weight:500;">{html.escape(str(field_val))}</div>
# </div>"""

#                 st.markdown(f"""
# <div style="background:#fff; border:1px solid #E2E8F0; border-left:4px solid {analysis_color};
#             border-radius:12px; padding:16px 18px 8px 18px; margin:8px 0;
#             box-shadow:0 1px 4px rgba(0,0,0,0.03); display:flex; flex-wrap:wrap;">
#     {inner_html}
# </div>""", unsafe_allow_html=True)

#             st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

#             # --- FINAL ISSUE REMARKS ---
#             final_issue = _val("call_log_summary", "")
#             if (not suppress_analysis_sections) and final_issue and final_issue != "N/A":
#                 st.markdown("""
#                 <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Major Call Log Issues Identified</div>
#                 """, unsafe_allow_html=True)

#                 st.markdown(f"""
#                 <div style="background:var(--surface); border:1px solid var(--border); border-left:4px solid #F59E0B; border-radius:12px; padding:14px 18px; margin:8px 0; box-shadow:var(--shadow-sm);">
#                     <div style="font-size:14px; color:var(--text); line-height:1.6; white-space:pre-wrap;">{html.escape(final_issue)}</div>
#                 </div>
#                 """, unsafe_allow_html=True)

#             st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)




import streamlit as st
import pandas as pd
import os
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import html
import json
import re
import zipfile
from datetime import datetime
from streamlit_autorefresh import st_autorefresh
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- AUTO UNZIP FOR GITHUB REPO FULL PROJECT SIZE FIX ---
def _ensure_files_unzipped():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith('.zip'):
                zip_path = os.path.join(root, file)
                # target will be original filename (e.g. ml_model.pkl from ml_model.zip)
                # We'll just extract everything inside the zip to the same directory
                try:
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        for entry in zip_ref.namelist():
                            target_path = os.path.join(root, entry)
                            if not os.path.exists(target_path):
                                zip_ref.extract(entry, root)
                except Exception as e:
                    pass

_ensure_files_unzipped()
# --------------------------------------------------


# ---------------- PAGE CONFIG ----------------
st.set_page_config(layout="wide", page_title="Venue Inventory Master", initial_sidebar_state="expanded")

# ---------------- STYLE ----------------
STYLE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ui", "styles.css")
try:
    with open(STYLE_PATH, "r", encoding="utf-8") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
except FileNotFoundError:
    # Styling is optional, don't break if CSS file not found
    pass
except Exception as e:
    # Log styling loading errors but don't crash
    print(f"Warning: Could not load styles from {STYLE_PATH}: {e}")


def metric_card(label, value, tooltip: str | None = None, pct: float | None = None, trend: str | None = None):
    safe_label = html.escape(str(label))
    safe_value = html.escape(str(value))

    details_html = ""
    graph_html = ""
    trend_html = ""

    if trend:
        t_str = str(trend).strip()
        if "+" in t_str:
            t_class = "trend-up"
        elif "-" in t_str:
            t_class = "trend-down"
        else:
            t_class = "trend-neutral"
        trend_html = f'<span class="{t_class}">{html.escape(t_str)}</span>'

    if pct is not None:
        safe_pct = min(max(pct, 0), 100)
        graph_html = f'<div class="kpi-graph"><div class="kpi-graph-fill" style="width: {safe_pct}%;"></div></div>'

    if tooltip:
        parts = str(tooltip).split("\n")
        formatted_parts = []
        for p in parts:
            if ":" in p:
                k, v = p.split(":", 1)
                formatted_parts.append(
                    "<span class='kpi-detail-item'>"
                    f"<span style='color:var(--text);font-weight:700'>{html.escape(k)}:</span>"
                    f"<span>{html.escape(v.strip())}</span>"
                    "</span>"
                )
            else:
                formatted_parts.append(f"<span class='kpi-detail-item'>{html.escape(p)}</span>")
        safe_details = ' <span style="margin:0 6px; color:var(--border-dark)">|</span> '.join(formatted_parts)

        details_html = f'<div class="kpi-subtext">{trend_html} {safe_details}</div>'
    elif trend_html:
        details_html = f'<div class="kpi-subtext">{trend_html}</div>'

    final_html = f"""<div class="kpi-static">
<div class="kpi-label">{safe_label}</div>
<div class="kpi-value">{safe_value}</div>
{details_html}
</div>"""
    st.markdown(final_html, unsafe_allow_html=True)

def metric_card_compact(label: str, value: int, accent: str = "#81CDEC"):
    """Compact KPI card for inline executive summaries."""
    safe_label = html.escape(str(label))
    safe_value = html.escape(f"{int(value):,}" if isinstance(value, (int, np.integer, float, np.floating)) else str(value))
    st.markdown(
        f"""
        <div class="kpi-compact" style="--kpi-accent:{accent};">
            <div class="kpi-compact-label">{safe_label}</div>
            <div class="kpi-compact-value">{safe_value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_brand_logo():
    """Render company logo inline — call inside a st.columns block for positioning."""
    import base64 as _b64
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
    if not os.path.exists(logo_path):
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.jpg")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as _lf:
            _logo_b64 = _b64.b64encode(_lf.read()).decode("utf-8")
        _ext = os.path.splitext(logo_path)[1][1:].lower()
        _mime = "jpeg" if _ext == "jpg" else _ext
        st.markdown(
            f"""
            <div style="display:flex; justify-content:flex-end; align-items:flex-start; margin:-14px 0 8px 0; padding:0;">
                <img src="data:image/{_mime};base64,{_logo_b64}"
                     alt="DEXIT Global"
                     style="width:205px; height:auto; display:block; margin:0; padding:0; object-fit:contain;" />
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div style="display:flex; justify-content:flex-end; margin:-4px 0 0 0; padding:0;
                        font-size:15px; font-weight:800; color:#1A202C; letter-spacing:0.5px;">VMS AUTOMATION</div>
            """,
            unsafe_allow_html=True,
        )

def sidebar_checkbox_filter(
    label: str,
    options: list,
    key_prefix: str,
    *,
    expanded: bool = False,
    columns: int = 1,
) -> list:
    """Render an executive-style checkbox filter block and return selected items."""
    clean_options = [opt for opt in options if str(opt).strip() != ""]
    if not clean_options:
        return []

    for idx in range(len(clean_options)):
        state_key = f"{key_prefix}_{idx}"
        if state_key not in st.session_state:
            st.session_state[state_key] = True

    selected_before = sum(
        1 for idx in range(len(clean_options)) if st.session_state.get(f"{key_prefix}_{idx}", True)
    )

    with st.sidebar.expander(f"{label}  ({selected_before}/{len(clean_options)})", expanded=expanded):
        search_query = st.text_input(
            f"Search {label}",
            key=f"{key_prefix}_search",
            placeholder="Search options...",
            label_visibility="collapsed",
        ).strip().lower()

        ctl_col1, ctl_col2 = st.columns(2)
        if ctl_col1.button("Select All", key=f"{key_prefix}_all", use_container_width=True):
            for idx in range(len(clean_options)):
                st.session_state[f"{key_prefix}_{idx}"] = True
        if ctl_col2.button("Clear All", key=f"{key_prefix}_none", use_container_width=True):
            for idx in range(len(clean_options)):
                st.session_state[f"{key_prefix}_{idx}"] = False

        visible_options = []
        for idx, opt in enumerate(clean_options):
            opt_str = str(opt)
            if not search_query or search_query in opt_str.lower():
                visible_options.append((idx, opt_str))

        if not visible_options:
            st.caption("No matching options.")
        else:
            check_cols = st.columns(max(1, columns))
            for pos, (idx, opt_text) in enumerate(visible_options):
                with check_cols[pos % max(1, columns)]:
                    st.checkbox(opt_text, key=f"{key_prefix}_{idx}")

    selected = [opt for idx, opt in enumerate(clean_options) if st.session_state.get(f"{key_prefix}_{idx}", True)]
    return selected

# Import photo manager
try:
    from photo_manager import PhotoManager, render_photo_manager_ui
except ImportError:
    PhotoManager = None
    render_photo_manager_ui = None

# ---------------- PATHS ----------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_PATH = os.path.join(SCRIPT_DIR, "Data")
OUTPUT_PATH = os.path.join(BASE_PATH, "Output")
PHOTOS_PATH = os.path.join(BASE_PATH, "Venue_Photos")
INDIA_GEOJSON_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "india.geojson")
STATE_LIST_PATH = os.path.join(SCRIPT_DIR, "india-maps-data-main", "state-list.json")
STATE_GEOJSON_DIR = os.path.join(SCRIPT_DIR, "india-maps-data-main", "geojson", "states")
BIHAR_REFERENCE_PPTX_PATH = os.path.join("C:\\Users", "DELL", "Downloads", "Bihar district maps.pptx")
INDIA_MAP_CENTER = {"lat": 21.92, "lon": 82.74}
INDIA_MAP_BOUNDS = {"west": 65.5, "east": 100.0, "south": 4.8, "north": 39.8}
if not os.path.exists(OUTPUT_PATH): os.makedirs(OUTPUT_PATH)
if not os.path.exists(PHOTOS_PATH): os.makedirs(PHOTOS_PATH)

# ---------------- CONFIGURATION ----------------
FILES = {
    "Venue Inventory": {"file": "VMS Master.xlsx", "users": ["Nishant", "Prafull", "Admin"]},
    "Blacklisted": {"file": "Blacklisted Venues.xlsx", "users": ["Prafull", "Admin"]},
    "FFA Data": {"file": "FFA Data.xlsx", "users": ["Anil", "Admin"]},
    "Suspicious": {"file": "Suspicious Activities.xlsx", "users": ["Prafull", "Admin"]},
    "Day End": {"file": "Day End Observations.xlsx", "users": ["Prafull", "Admin"]},
    "Market Reviews": {"file": "Google review and ratings.xlsx", "users": ["Prafull", "Admin"]},
    "Call Logs": {"file": "Master Call Logs.xlsx", "users": ["Prafull", "Admin"]},
    "Project Master": {"file": "Project Master.xlsx", "users": ["Prafull", "Admin"]}
}

USERS = {"Prafull": "Prafull@123", "Nishant": "Nishant@123", "Mayuresh": "Mayuresh@123", "Anil": "Anil@123", "Admin": "Admin@123"}

# --- SORTING DEFINITIONS ---
CAT_ORDER = ["A", "B", "C", "BL-C", "BL"]
REGION_ORDER = ["North One", "North Two", "East", "West", "South"]

# ---------------- SESSION STATE ----------------
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "user" not in st.session_state: st.session_state.user = None
if "search_main_box" not in st.session_state: st.session_state.search_main_box = ""
if "venue_search_query" not in st.session_state: st.session_state.venue_search_query = ""
if "manpower_search_code" not in st.session_state: st.session_state.manpower_search_code = ""
if "venue_detail_code" not in st.session_state: st.session_state.venue_detail_code = None
if "map_selected_state" not in st.session_state: st.session_state.map_selected_state = None
if "map_selected_district" not in st.session_state: st.session_state.map_selected_district = None
if "map_selected_district_state" not in st.session_state: st.session_state.map_selected_district_state = None

VALID_PAGES = {"dashboard", "login", "main", "edit", "process", "venue_detail", "india_map_detail", "manpower_dashboard", "project_wise_details", "city_datc_dotc"}
VALID_DATASETS = set(FILES.keys())

qp_page = st.query_params.get("page", None)
if isinstance(qp_page, list):
    qp_page = qp_page[0] if qp_page else None

qp_dataset = st.query_params.get("dataset", None)
if isinstance(qp_dataset, list):
    qp_dataset = qp_dataset[0] if qp_dataset else None

if "page" not in st.session_state:
    st.session_state.page = qp_page if qp_page in VALID_PAGES else "dashboard"
elif st.session_state.page not in VALID_PAGES:
    st.session_state.page = "dashboard"

if "dataset_key" not in st.session_state:
    st.session_state.dataset_key = qp_dataset if qp_dataset in VALID_DATASETS else "Venue Inventory"
elif st.session_state.dataset_key not in VALID_DATASETS:
    st.session_state.dataset_key = "Venue Inventory"
elif qp_dataset in VALID_DATASETS and qp_dataset != st.session_state.dataset_key:
    st.session_state.dataset_key = qp_dataset

if qp_page != st.session_state.page:
    st.query_params["page"] = st.session_state.page
if qp_dataset != st.session_state.dataset_key:
    st.query_params["dataset"] = st.session_state.dataset_key

if st.session_state.page == "dashboard":
    st.session_state["_auto_refresh_count"] = st_autorefresh(interval=60000, key="auto_refresh_main")
else:
    st.session_state["_auto_refresh_count"] = st.session_state.get("_auto_refresh_count", 0)

# ---------------- SMART DATA LOADER ----------------
def load_smart_data(file_path):
    if not os.path.exists(file_path): return pd.DataFrame()
    try:
        if file_path.lower().endswith(".xlsx"):
            preview = pd.read_excel(file_path, header=None, nrows=10)
            header_idx = 0
            for i, row in preview.iterrows():
                row_str = " ".join([str(x) for x in row.values])
                if any(key in row_str.upper() for key in ["ROW LABELS", "VENUE_TYPE", "CODE", "ACTIVE", "DMS_CODE", "STATUS"]):
                    header_idx = i; break
            return pd.read_excel(file_path, header=header_idx)
        return pd.read_csv(file_path)
    except: return pd.DataFrame()

# ---------------- HELPER FUNCTIONS FOR DATA PROCESSING ----------------
def clean_project_name(series: pd.Series) -> pd.Series:
    """Clean project names by removing special chars, normalizing spacing, and filtering blanks."""
    s = series.fillna("").astype(str).str.strip()
    s_upper = s.str.upper()
    is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
    s = s.mask(is_blank, "")
    s = s.str.replace(r"[^A-Za-z ]+", " ", regex=True)
    s = s.str.replace(r"\s+", " ", regex=True).str.strip()
    return s.str.upper()

def _blankify_text(series: pd.Series) -> pd.Series:
    """Convert blank-like values (-, NAN, NONE) to empty strings."""
    s = series.fillna("").astype(str).str.strip()
    s_upper = s.str.upper()
    is_blank = (s == "") | (s == "-") | (s_upper == "NAN") | (s_upper == "NONE")
    return s.mask(is_blank, "")

def _col_or_empty(df_in: pd.DataFrame, col_names: list) -> pd.Series:
    """Return column from dataframe or empty series if not present. Case-insensitive."""
    col_names_lower = [c.lower() for c in col_names]
    for c in df_in.columns:
        if str(c).lower() in col_names_lower:
            return df_in[c]
    return pd.Series([""] * len(df_in), index=df_in.index)

def build_concat_address(df_in: pd.DataFrame) -> pd.Series:
    """Build concatenated full address in format: name, address, landmark, city, district, state - pincode"""
    name_s = _blankify_text(_col_or_empty(df_in, ["name", "venue_name", "updated_venue_name"]))
    address_s = _blankify_text(_col_or_empty(df_in, ["address"]))
    landmark_s = _blankify_text(_col_or_empty(df_in, ["landmark"]))
    city_s = _blankify_text(_col_or_empty(df_in, ["city"]))
    district_s = _blankify_text(_col_or_empty(df_in, ["district"]))
    state_s = _blankify_text(_col_or_empty(df_in, ["state"]))

    parts_df = pd.concat([name_s, address_s, landmark_s, city_s, district_s, state_s], axis=1)
    joined_address = parts_df.apply(lambda row: ", ".join([v for v in row.tolist() if v]), axis=1)

    pincode_s = _blankify_text(_col_or_empty(df_in, ["pincode", "pin_code"]))
    pincode_s = pincode_s.str.replace(r"\.0$", "", regex=True)

    # Format: name, address, landmark, city, district, state - pincode
    result = np.where(
        joined_address != "",
        joined_address + np.where(pincode_s != "", " - " + pincode_s, ""),
        ""
    )
    return pd.Series(result, index=df_in.index)

def find_dms_col(df: pd.DataFrame):
    """Find the DMS/Venue code column in a dataframe."""
    return next((c for c in df.columns if c.lower().replace(" ", "_") in ['dms_code', 'venue_code']), None)

STATE_NAME_ALIASES = {
    "NCT OF DELHI": "DELHI",
    "NCT DELHI": "DELHI",
    "DELHI NCR": "DELHI",
    "ORISSA": "ODISHA",
    "UTTARANCHAL": "UTTARAKHAND",
    "JAMMU KASHMIR": "JAMMU AND KASHMIR",
    "JAMMU & KASHMIR": "JAMMU AND KASHMIR",
    "DADRA AND NAGAR HAVELI AND DAMAN AND DIU": "DADRA AND NAGAR HAVELI DAMAN AND DIU",
    "ANDAMAN NICOBAR": "ANDAMAN AND NICOBAR ISLANDS",
    "PONDICHERRY": "PUDUCHERRY",
}

FOREIGN_LOCATIONS = ["NEPAL", "UAE"]


@st.cache_data(show_spinner=False)
def load_india_geojson(path: str):
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            gj = json.load(f)
        if isinstance(gj, dict) and isinstance(gj.get("features"), list):
            return gj
    except Exception:
        return None
    return None


def normalize_state_name(value: str) -> str:
    name = str(value or "").upper()
    name = name.replace("&", " AND ")
    name = re.sub(r"[^A-Z0-9]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return STATE_NAME_ALIASES.get(name, name)


def build_india_map_dataframe(source_df: pd.DataFrame | None):
    geojson_data = load_india_geojson(INDIA_GEOJSON_PATH)
    if geojson_data is None:
        return pd.DataFrame(), None, 0, 0

    state_count_map: dict[str, int] = {}
    if source_df is not None and not source_df.empty and "state" in source_df.columns:
        state_series = source_df["state"].fillna("").astype(str).str.strip()
        state_series = state_series[state_series != ""]
        for state_name, count in state_series.value_counts().items():
            normalized = normalize_state_name(state_name)
            if normalized:
                state_count_map[normalized] = state_count_map.get(normalized, 0) + int(count)

    rows = []
    matched_states = 0
    mapped_venues = 0
    features = geojson_data.get("features", [])

    for idx, feature in enumerate(features):
        properties = feature.get("properties", {}) or {}
        state_name = (
            properties.get("st_nm")
            or properties.get("STATE")
            or properties.get("name")
            or f"State {idx + 1}"
        )
        district_code = properties.get("dt_code") or properties.get("district_code")
        state_code = properties.get("st_code") or properties.get("state_code")
        feature_id = str(
            feature.get("id")
            or district_code
            or state_code
            or properties.get("ID_1")
            or f"{normalize_state_name(state_name)}_{idx + 1}"
        )
        feature["id"] = feature_id

        count = int(state_count_map.get(normalize_state_name(state_name), 0))
        if count > 0:
            matched_states += 1
            mapped_venues += count

        rows.append(
            {
                "id": feature_id,
                "state": state_name,
                "venue_count": count,
            }
        )

    return pd.DataFrame(rows), geojson_data, matched_states, mapped_venues

def _get_feature_centroid(feature: dict) -> tuple[float, float]:
    geom = feature.get("geometry", {})
    if not geom:
        return 0.0, 0.0
    coords = geom.get("coordinates", [])
    flat_coords = []

    def extract_coords(c):
        if not c:
            return
        if isinstance(c[0], (int, float)):
            flat_coords.append(c)
        else:
            for item in c:
                extract_coords(item)
    extract_coords(coords)
    
    if not flat_coords:
        return 0.0, 0.0
    arr = np.array(flat_coords)
    return float(np.mean(arr[:, 0])), float(np.mean(arr[:, 1]))


def create_india_map_figure(
    map_df: pd.DataFrame,
    geojson_data: dict,
    *,
    height: int = 320,
    show_data: bool = True,
    zoom: float = 3.5,
    center: dict | None = None,
    opacity: float = 0.78,
    map_style: str = "white-bg",
    color_power: float = 1.0,
    line_color: str = "#94A3B8",
    line_width: float = 0.9,
    restrict_to_india: bool = True,
):
    if map_df.empty or not geojson_data:
        return None

    if show_data:
        plot_df = map_df.copy()
        counts = pd.to_numeric(plot_df.get("venue_count", 0), errors="coerce").fillna(0).astype(float)
        if color_power != 1.0:
            plot_df["_color_metric"] = np.power(counts, color_power)
        else:
            plot_df["_color_metric"] = counts
        max_metric = float(plot_df["_color_metric"].max()) if "_color_metric" in plot_df.columns else 0.0
        fig_map = px.choropleth(
            plot_df,
            geojson=geojson_data,
            locations="id",
            featureidkey="id",
            color="_color_metric",
            hover_name="state",
            custom_data=["state", "id", "venue_count"],
            hover_data={"venue_count": False, "id": False, "_color_metric": False},
            color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
            range_color=(0, max(1.0, max_metric)),
        )
    else:
        preview_df = map_df.copy()
        preview_df["map_fill"] = 1
        plot_df = preview_df
        fig_map = px.choropleth(
            preview_df,
            geojson=geojson_data,
            locations="id",
            featureidkey="id",
            color="map_fill",
            hover_name="state",
            custom_data=["state", "id", "venue_count"],
            hover_data={"map_fill": False, "id": False},
            color_continuous_scale=["#C9E1F2", "#C9E1F2"],
            range_color=(0, 1),
        )

    fig_map.update_traces(
        marker_line_color=line_color,
        marker_line_width=line_width,
        hovertemplate="<b>%{hovertext}</b><extra></extra>" if not show_data else "<b>%{hovertext}</b><br>Venues: %{customdata[2]:,}<extra></extra>",
    )
    
    id_to_name = dict(zip(plot_df["id"], plot_df["state"]))
    state_to_coords = {}
    if geojson_data and "features" in geojson_data:
        for f in geojson_data["features"]:
            fid = f.get("id")
            if fid in id_to_name:
                state_name = id_to_name[fid]
                geom = f.get("geometry", {})
                if geom:
                    if state_name not in state_to_coords:
                        state_to_coords[state_name] = []
                    state_to_coords[state_name].append(geom.get("coordinates", []))

    HIDE_LABELS = {
        "Dadra and Nagar Haveli and Daman and Diu", 
        "Lakshadweep",
    }
    LABEL_ALIASES = {
        "Andaman and Nicobar Islands": "A&N",
        "Jammu and Kashmir": "J&K",
        "Himachal Pradesh": "HP",
        "Uttarakhand": "UT",
        "Punjab": "PB",
        "West Bengal": "WB",
        "Jharkhand": "JH",
        "Arunachal Pradesh": "AR",
        "Chandigarh": "CH",
        "Delhi": "DL",
        "Goa": "GA",
        "Sikkim": "SK",
        "Tripura": "TR",
        "Mizoram": "MZ",
        "Manipur": "MN",
        "Nagaland": "NL",
        "Meghalaya": "ML",
        "Haryana": "HR",
        "Puducherry": "PY",
        "Uttar Pradesh": "UP",
        "Madhya Pradesh": "MP",
        "Andhra Pradesh": "AP",
        "Tamil Nadu": "TN",
        "Karnataka": "KA",
        "Kerala": "KL",
        "Telangana": "TG",
        "Rajasthan": "RJ",
        "Gujarat": "GJ",
        "Maharashtra": "MH",
        "Odisha": "OD",
        "Chhattisgarh": "CG",
        "Bihar": "BR",
        "Assam": "AS",
        "Ladakh": "LA"
    }
    
    lons, lats, texts = [], [], []
    for state_name, coords_list in state_to_coords.items():
        if state_name in HIDE_LABELS:
            continue
        cx, cy = _get_feature_centroid({"geometry": {"coordinates": coords_list}})
        if cx != 0.0 and cy != 0.0:
            lons.append(cx)
            lats.append(cy)
            texts.append(LABEL_ALIASES.get(state_name, state_name))

    if show_data and lons:
        fig_map.add_trace(go.Scattergeo(
            lon=lons,
            lat=lats,
            text=texts,
            mode="text",
            textfont=dict(color="#0F172A", size=8.5, family="Arial, sans-serif", weight="bold"),
            showlegend=False,
            hoverinfo="skip"
        ))
        
    fig_map.update_geos(fitbounds="locations", visible=False)
    
    fig_map.update_layout(
        clickmode="event+select",
        dragmode="pan",
        margin=dict(t=0, b=0, l=0, r=0),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        height=height,
        coloraxis_showscale=show_data,
        coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.78),
        geo=dict(bgcolor="rgba(0,0,0,0)")
    )
    return fig_map


def extract_selected_points(selection_event) -> list:
    if selection_event is None:
        return []
    if isinstance(selection_event, dict):
        return selection_event.get("selection", {}).get("points", []) or []
    selection_obj = getattr(selection_event, "selection", None)
    if selection_obj is None:
        return []
    return getattr(selection_obj, "points", []) or []


def get_last_point_signature(points: list) -> str | None:
    if not isinstance(points, (list, tuple)) or not points:
        return None
    point = points[-1]
    if point is None:
        return None
    customdata = _point_get(point, "customdata")
    if isinstance(customdata, (list, tuple)):
        custom_sig = "|".join("" if x is None else str(x) for x in customdata[:3])
    else:
        custom_sig = ""
    sig_parts = [
        str(_point_get(point, "location") or ""),
        str(_point_get(point, "id") or ""),
        str(_point_get(point, "hovertext") or ""),
        str(_point_get(point, "point_index") or _point_get(point, "pointIndex") or ""),
        custom_sig,
    ]
    signature = "||".join(sig_parts).strip()
    return signature if signature else None


def _point_get(point, key: str):
    if isinstance(point, dict):
        return point.get(key)
    return getattr(point, key, None)


@st.cache_data(show_spinner=False)
def load_ppt_reference_images(pptx_path: str, max_images: int = 6):
    if not os.path.exists(pptx_path):
        return []
    images = []
    try:
        with zipfile.ZipFile(pptx_path, "r") as zf:
            media_files = [
                n for n in zf.namelist()
                if n.startswith("ppt/media/") and n.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
            ]
            media_files = sorted(media_files)
            for media_path in media_files[:max_images]:
                try:
                    images.append((os.path.basename(media_path), zf.read(media_path)))
                except Exception:
                    continue
    except Exception:
        return []
    return images


def resolve_state_from_map_points(points: list, id_to_state_map: dict, ordered_names: list | None = None):
    if not points:
        return None
    point = points[-1] if isinstance(points, (list, tuple)) else None
    if point is None:
        return None

    # 0) customdata fallback using known payload order:
    #    [name, id, ...] for both state and district map traces.
    cd = _point_get(point, "customdata")
    if isinstance(cd, (list, tuple)):
        if len(cd) >= 2:
            cand_id = str(cd[1]).strip() if cd[1] is not None else ""
            if cand_id:
                resolved = id_to_state_map.get(cand_id)
                if resolved:
                    return resolved
        if len(cd) >= 1:
            cand_name = str(cd[0]).strip() if cd[0] is not None else ""
            if cand_name:
                for mapped_name in id_to_state_map.values():
                    if str(mapped_name).strip().lower() == cand_name.lower():
                        return mapped_name

        # legacy fallback for unexpected payload shape
        for item in cd:
            item_s = str(item).strip() if item is not None else ""
            if not item_s:
                continue
            resolved = id_to_state_map.get(item_s)
            if resolved:
                return resolved
            for mapped_name in id_to_state_map.values():
                if str(mapped_name).strip().lower() == item_s.lower():
                    return mapped_name

    # 1) direct id/location mapping (common for choropleth)
    for key in ("location", "id"):
        loc = _point_get(point, key)
        if loc is not None:
            resolved = id_to_state_map.get(str(loc))
            if resolved:
                return resolved

    # 2) index-based fallback (point_index / point_number)
    if ordered_names:
        for idx_key in ("point_index", "point_number", "pointIndex", "pointNumber"):
            idx_val = _point_get(point, idx_key)
            try:
                idx_int = int(idx_val)
            except Exception:
                idx_int = None
            if idx_int is not None and 0 <= idx_int < len(ordered_names):
                return ordered_names[idx_int]

    # 3) text-based fallback from hover payload
    for txt_key in ("hovertext", "label", "text", "name"):
        txt = _point_get(point, txt_key)
        if txt is None:
            continue
        txt_clean = str(txt).strip()
        if not txt_clean:
            continue
        # First try direct exact map value
        for mapped_name in id_to_state_map.values():
            if str(mapped_name).strip().lower() == txt_clean.lower():
                return mapped_name

    return None


def build_state_detail_payload(df: pd.DataFrame, state_name: str):
    if df is None or df.empty or "state" not in df.columns or not state_name:
        return pd.DataFrame()
    target = normalize_state_name(state_name)
    matched = df[df["state"].fillna("").astype(str).apply(normalize_state_name) == target].copy()
    return matched


def normalize_district_name(value: str) -> str:
    name = str(value or "").upper()
    name = re.sub(r"[^A-Z0-9]+", " ", name)
    return re.sub(r"\s+", " ", name).strip()


@st.cache_data(show_spinner=False)
def load_state_slug_mapping(path: str):
    mapping = {}
    if not os.path.exists(path):
        return mapping
    try:
        with open(path, "r", encoding="utf-8") as f:
            rows = json.load(f)
        for row in rows:
            state_name = row.get("name")
            slug = row.get("slug")
            if state_name and slug:
                mapping[normalize_state_name(state_name)] = slug
    except Exception:
        return {}
    return mapping


@st.cache_data(show_spinner=False)
def load_state_geojson_by_slug(slug: str):
    if not slug:
        return None
    path = os.path.join(STATE_GEOJSON_DIR, f"{slug}.geojson")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            gj = json.load(f)
        if isinstance(gj, dict) and isinstance(gj.get("features"), list):
            return gj
    except Exception:
        return None
    return None


def resolve_state_slug(state_name: str):
    if not state_name:
        return None
    state_map = load_state_slug_mapping(STATE_LIST_PATH)
    return state_map.get(normalize_state_name(state_name))


def _flatten_coordinates(coords):
    if isinstance(coords, (list, tuple)) and coords:
        if isinstance(coords[0], (int, float)) and len(coords) >= 2:
            yield float(coords[0]), float(coords[1])
        else:
            for item in coords:
                yield from _flatten_coordinates(item)


def get_geojson_center_zoom(geojson_data: dict):
    lons, lats = [], []
    for feature in geojson_data.get("features", []):
        geometry = feature.get("geometry") or {}
        coords = geometry.get("coordinates")
        if coords is None:
            continue
        for lon, lat in _flatten_coordinates(coords):
            lons.append(lon)
            lats.append(lat)

    if not lons or not lats:
        return {"lat": 22.7, "lon": 79.2}, 5.4

    min_lon, max_lon = min(lons), max(lons)
    min_lat, max_lat = min(lats), max(lats)
    span = max(max_lon - min_lon, max_lat - min_lat)

    if span > 14:
        zoom = 4.9
    elif span > 10:
        zoom = 5.3
    elif span > 6:
        zoom = 5.8
    elif span > 3.5:
        zoom = 6.3
    elif span > 2:
        zoom = 6.9
    else:
        zoom = 7.5

    center = {"lat": (min_lat + max_lat) / 2, "lon": (min_lon + max_lon) / 2}
    return center, zoom


def build_district_map_dataframe(state_df: pd.DataFrame, state_geojson: dict):
    if state_geojson is None:
        return pd.DataFrame(), {}

    district_count_map = {}
    datc_map = {}
    dotc_map = {}
    active_map = {}
    black_map = {}

    if state_df is not None and not state_df.empty:
        district_series = state_df["district"].fillna("").astype(str).str.strip() if "district" in state_df.columns else pd.Series([], dtype=str)
        venue_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([], dtype=str)
        status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([], dtype=str)

        for idx, district_name in district_series.items():
            norm = normalize_district_name(district_name)
            if not norm:
                continue
            district_count_map[norm] = district_count_map.get(norm, 0) + 1

            vtype = venue_series.get(idx, "")
            if vtype == "DATC":
                datc_map[norm] = datc_map.get(norm, 0) + 1
            elif vtype == "DOTC":
                dotc_map[norm] = dotc_map.get(norm, 0) + 1

            status_val = status_series.get(idx, "")
            if status_val == "ACTIVE":
                active_map[norm] = active_map.get(norm, 0) + 1
            elif "BLACK" in status_val:
                black_map[norm] = black_map.get(norm, 0) + 1

    rows = []
    id_to_district = {}
    for i, feature in enumerate(state_geojson.get("features", [])):
        props = feature.get("properties", {}) or {}
        district_name = props.get("district") or props.get("DISTRICT") or props.get("name") or f"District {i + 1}"
        norm = normalize_district_name(district_name)
        feature_id = str(feature.get("id") or props.get("dt_code") or props.get("district_code") or district_name)
        feature["id"] = feature_id
        id_to_district[feature_id] = district_name

        rows.append(
            {
                "id": feature_id,
                "district": district_name,
                "venue_count": int(district_count_map.get(norm, 0)),
                "datc_count": int(datc_map.get(norm, 0)),
                "dotc_count": int(dotc_map.get(norm, 0)),
                "active_count": int(active_map.get(norm, 0)),
                "black_count": int(black_map.get(norm, 0)),
            }
        )

    return pd.DataFrame(rows), id_to_district


def compute_dotc_datc_summary(df_in: pd.DataFrame):
    if df_in is None or df_in.empty:
        return {
            "dotc_centres": 0, "datc_centres": 0, "total_centres": 0,
            "dotc_seats": 0, "datc_seats": 0, "total_seats": 0,
        }

    vtype = df_in["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in df_in.columns else pd.Series([""] * len(df_in))
    seat_series = pd.to_numeric(df_in.get("venue_max_capacity", 0), errors="coerce").fillna(0)

    dotc_mask = vtype == "DOTC"
    datc_mask = vtype == "DATC"

    dotc_centres = int(dotc_mask.sum())
    datc_centres = int(datc_mask.sum())
    total_centres = int(len(df_in))

    dotc_seats = int(seat_series[dotc_mask].sum())
    datc_seats = int(seat_series[datc_mask].sum())
    total_seats = int(seat_series.sum())

    return {
        "dotc_centres": dotc_centres,
        "datc_centres": datc_centres,
        "total_centres": total_centres,
        "dotc_seats": dotc_seats,
        "datc_seats": datc_seats,
        "total_seats": total_seats,
    }


def render_dotc_datc_summary(title: str, summary: dict):
    st.markdown(
        f"""
        <div style="border:1px solid #A8BBC8; border-radius:10px; overflow:hidden; background:#F8FBFD;">
            <div style="padding:8px 12px; border-bottom:1px solid #A8BBC8; font-weight:800; color:#2A3B4C; text-transform:uppercase; letter-spacing:0.4px;">
                {html.escape(title)}
            </div>
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead>
                    <tr style="background:#EAF2F7; color:#334155;">
                        <th style="padding:7px; border:1px solid #D1DEE7; text-align:left;">Metric</th>
                        <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DOTC</th>
                        <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">DATC</th>
                        <th style="padding:7px; border:1px solid #D1DEE7; text-align:center;">Total</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Centre</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_centres", 0):,}</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_centres", 0):,}</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_centres", 0):,}</td>
                    </tr>
                    <tr>
                        <td style="padding:7px; border:1px solid #D1DEE7; font-weight:700;">Seat Capacity</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("dotc_seats", 0):,}</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("datc_seats", 0):,}</td>
                        <td style="padding:7px; border:1px solid #D1DEE7; text-align:center;">{summary.get("total_seats", 0):,}</td>
                    </tr>
                </tbody>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


# If protected pages are refreshed without an active session, route to login cleanly.
if st.session_state.page in {"main", "edit", "process"} and not st.session_state.logged_in:
    st.session_state.page = "login"
    st.query_params["page"] = "login"

# ================= DASHBOARD PAGE =================
if st.session_state.page == "dashboard":
    _title_col, _logo_col = st.columns([3, 1], vertical_alignment="top")
    with _title_col:
        st.title("Venue Analysis Dashboard")
    with _logo_col:
        render_brand_logo()
    st.markdown('<div style="height: 8px;"></div>', unsafe_allow_html=True)
    
    if st.session_state.logged_in:
        st.sidebar.success(f"User: {st.session_state.user}")
        if st.sidebar.button("Logout"):
            st.session_state.logged_in, st.session_state.user = False, None
            st.rerun()
    else:
        if st.sidebar.button("Admin Login"):
            st.session_state.page = "login"
            st.rerun()

    all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH) 
                 if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]
    
    if not all_files:
        st.warning("No processed reports found. Admin must run the Analysis Engine.")
    else:
        latest_file_path = max(all_files, key=os.path.getmtime)
        data = pd.read_excel(latest_file_path, sheet_name=0)
        data.columns = data.columns.str.lower().str.replace(" ", "_")
        
        # Calculate Inactive count for the card BEFORE global filtering
        inactive_overall_count = 0
        if "status" in data.columns:
            inactive_overall_count = int((data["status"].fillna("").astype(str).str.upper() == "INACTIVE").sum())
            # Apply Global Filter: Remove Inactive venues from the rest of the dashboard
            data = data[data["status"].fillna("").astype(str).str.upper() != "INACTIVE"].copy()
        
        # Clean State Data: Normalize names and filter out Foreign Locations
        if "state" in data.columns:
            data["state"] = data["state"].fillna("").astype(str).apply(normalize_state_name)
            data = data[~data["state"].isin(FOREIGN_LOCATIONS)]
        
        st.info(f"Dashboard Data: **{os.path.basename(latest_file_path)}**")

# --- VENUE SEARCH BAR ---
        st.markdown("""
        <div class="dashboard-hero-card">
            <div class="dashboard-hero-title">Venue Search</div>
            <p class="dashboard-hero-subtitle">
                Search by Venue Name, DMS Code, City, or any keyword to view complete venue profile and metrics
            </p>
        </div>
        """, unsafe_allow_html=True)

        vs_col1, vs_col2 = st.columns([4, 1], vertical_alignment="center")
        with vs_col1:
            venue_search_input = st.text_input(
                "Search Venue",
                value=st.session_state.venue_search_query,
                placeholder="e.g. AP COMPUTER POINT or EST-AR-1161 or ITANAGAR...",
                label_visibility="collapsed"
            )
        with vs_col2:
            search_btn = st.button("Search Venue", use_container_width=True, type="primary")

        # --- RESTRICTION LOGIC START ---
        if search_btn or (venue_search_input != st.session_state.venue_search_query and venue_search_input != ""):
            # Check authorization exactly like your sidebar buttons
            if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                st.session_state.venue_search_query = venue_search_input
            elif not st.session_state.logged_in:
                st.session_state.page = "login"
                st.rerun()
            else:
                st.error("Restricted to Admin/Prafull.")
                st.session_state.venue_search_query = "" 

        query = st.session_state.venue_search_query.strip()
        
        # Only run the search if the user is authorized
        if query and st.session_state.user in ["Admin", "Prafull"]:
            # --- START SEARCH PROCESSING ---
            search_mask = data.astype(str).apply(
                lambda col: col.str.contains(query, case=False, na=False)
            ).any(axis=1)
            results = data[search_mask]
            

        # Update session state if button is clicked or if they type and press enter
        if search_btn or venue_search_input != st.session_state.venue_search_query:
            st.session_state.venue_search_query = venue_search_input

        query = st.session_state.venue_search_query.strip()
        if query:
            # Search across all text columns
            search_mask = data.astype(str).apply(
                lambda col: col.str.contains(query, case=False, na=False)
            ).any(axis=1)
            results = data[search_mask]

            is_authorized = True

            if results.empty:
                st.warning(f"No venues found matching **\"{query}\"**. Try a different keyword.")
                if st.button("Clear Search", key="clear_search_empty"):
                    st.session_state.venue_search_query = ""
                    st.rerun()
            elif len(results) == 1 and search_btn and is_authorized:
                # Single result -> go directly to detail page (only auto-redirect if search button was explicitly clicked)
                st.session_state.venue_detail_code = str(results.iloc[0].get('dms_code', ''))
                st.session_state.page = "venue_detail"
                st.session_state.venue_search_query = "" # Clear search after found
                st.rerun()
            else:
                # Multiple results (or single result persistent view) -> show selection list
                head_col, clear_col = st.columns([5, 1], vertical_alignment="center")
                with head_col:
                    st.markdown(f"##### Found {len(results)} venue(s) matching **\"{query}\"**")
                with clear_col:
                    if st.button("Clear Search", key="clear_search_results", use_container_width=True):
                        st.session_state.venue_search_query = ""
                        st.rerun()

                for idx, row in results.iterrows():
                    dms = str(row.get('dms_code', 'N/A'))
                    name = str(row.get('venue_name', row.get('name', 'N/A')))
                    city = str(row.get('city', ''))
                    region = str(row.get('region', ''))
                    status = str(row.get('status', ''))
                    category = str(row.get('category', ''))
                    
                    cat_colors = {"A": "#10B981", "B": "#3B82F6", "C": "#F59E0B", "BL": "#EF4444", "BL-C": "#8B5CF6"}
                    cat_color = cat_colors.get(category, "#6B7280")
                    
                    # We split the row into two columns: one for the card, one for the action button
                    card_col, btn_col = st.columns([5.15, 0.85], vertical_alignment="center")

                    with card_col:
                        st.markdown(f"""
                        <div class="vms-card-wrapper">
                            <div class="vms-card-visual" style="border-left:4px solid {cat_color};">
                                <div>
                                    <div class="vms-card-title">{html.escape(name)}</div>
                                    <div class="vms-card-meta">
                                        <span>CODE: {html.escape(dms)}</span>
                                        <span class="vms-card-meta-sep">|</span>
                                        <span>CITY: {html.escape(city)}</span>
                                        <span class="vms-card-meta-sep">|</span>
                                        <span>REGION: {html.escape(region)}</span>
                                    </div>
                                </div>
                                <div class="vms-card-pills">
                                    <span class="vms-pill-category" style="background:{cat_color}22; color:{cat_color};">
                                        Cat {html.escape(category)}
                                    </span>
                                    <span class="vms-pill-status">
                                        {html.escape(status)}
                                    </span>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                    with btn_col:
                        def set_venue_callback(code):
                            st.session_state.venue_detail_code = code
                            st.session_state.page = "venue_detail"
                            st.session_state.venue_search_query = ""
                        
                        # Visible button shifted to the end (right side)
                        st.button("View Details", key=f"view_btn_{dms}_{idx}", use_container_width=True, on_click=set_venue_callback, args=(dms,))

                st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)

        filter_top_col1, filter_top_col2, filter_top_col3, filter_top_col4 = st.columns([1, 1, 1, 1], gap="medium")

        avail_top_regs = sorted(data['region'].dropna().unique().tolist()) if 'region' in data.columns else []
        top_region = filter_top_col1.selectbox(
            "Region",
            options=["All Regions"] + avail_top_regs,
            index=0,
            key="top_dash_region",
        ) if avail_top_regs else "All Regions"

        avail_top_states = sorted(data['state'].dropna().unique().tolist()) if 'state' in data.columns else []
        top_state = filter_top_col2.selectbox(
            "State",
            options=["All States"] + avail_top_states,
            index=0,
            key="top_dash_state",
        ) if avail_top_states else "All States"

        avail_top_districts = sorted(data['district'].dropna().unique().tolist()) if 'district' in data.columns else []
        top_district = filter_top_col3.selectbox(
            "District",
            options=["All Districts"] + avail_top_districts,
            index=0,
            key="top_dash_district",
        ) if avail_top_districts else "All Districts"

        avail_top_cities = sorted(data['city'].dropna().unique().tolist()) if 'city' in data.columns else []
        top_city = filter_top_col4.selectbox(
            "City",
            options=["All Cities"] + avail_top_cities,
            index=0,
            key="top_dash_city",
        ) if avail_top_cities else "All Cities"

        if 'region' in data.columns:
            data['region'] = data['region'].astype(str).str.title().str.strip()
            data['region'] = pd.Categorical(data['region'], categories=REGION_ORDER, ordered=True)
        if 'state' in data.columns:
            data['state'] = data['state'].fillna("").astype(str).apply(normalize_state_name)
        if 'district' in data.columns:
            data['district'] = data['district'].fillna("").astype(str).str.title().str.strip()
        if 'city' in data.columns:
            data['city'] = data['city'].fillna("").astype(str).str.title().str.strip()
        if 'category' in data.columns:
            data['category'] = pd.Categorical(data['category'], categories=CAT_ORDER, ordered=True)

        # Executive sidebar filter header
        st.sidebar.markdown(
            """
            <div class="filters-shell">
                <div class="filters-title-row">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
                        <polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"></polygon>
                    </svg>
                    <span class="filters-title-text">Dashboard Filters</span>
                </div>
                <p class="filters-subtitle">Refine live metrics by operational segments</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.sidebar.button("Reset All Filters", key="dash_reset_all", use_container_width=True):
            for k in list(st.session_state.keys()):
                if (
                    k.startswith("dash_vtype_")
                    or k.startswith("dash_region_")
                    or k.startswith("dash_state_")
                    or k.startswith("dash_district_")
                    or k.startswith("dash_city_")
                    or k.startswith("dash_status_")
                    or k.startswith("dash_abc_")
                ):
                    if k.endswith("_search"):
                        st.session_state[k] = ""
                    elif k.endswith("_all") or k.endswith("_none"):
                        continue
                    else:
                        st.session_state[k] = True
            st.rerun()
        st.sidebar.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        
        v_type_list = sorted([v for v in data['venue_type'].dropna().unique().tolist() if str(v).strip().upper() not in ["DOTC", "DATC"]])
        sel_vtype = sidebar_checkbox_filter("Venue Type", v_type_list, "dash_vtype", expanded=True, columns=1)

        avail_regs = sorted(data['region'].dropna().unique().tolist())
        sel_reg = sidebar_checkbox_filter("Region", avail_regs, "dash_region", expanded=False, columns=1)

        avail_states = sorted(data['state'].dropna().unique().tolist()) if 'state' in data.columns else []
        sel_state = sidebar_checkbox_filter("State", avail_states, "dash_state", expanded=False, columns=1) if avail_states else None

        avail_districts = sorted(data['district'].dropna().unique().tolist()) if 'district' in data.columns else []
        sel_district = sidebar_checkbox_filter("District", avail_districts, "dash_district", expanded=False, columns=1) if avail_districts else None

        avail_cities = sorted(data['city'].dropna().unique().tolist()) if 'city' in data.columns else []
        sel_city = sidebar_checkbox_filter("City", avail_cities, "dash_city", expanded=False, columns=1) if avail_cities else None

        stat_list = sorted(data['status'].dropna().unique().tolist())
        sel_stat = sidebar_checkbox_filter("Status", stat_list, "dash_status", expanded=False, columns=1)

        avail_cats = sorted(data['category'].dropna().unique().tolist()) if 'category' in data.columns else []
        if avail_cats:
            sel_cat = sidebar_checkbox_filter("ABC Category", avail_cats, "dash_abc", expanded=False, columns=2)
        else:
            sel_cat = []

        st.sidebar.markdown(
            f"""
            <div class="filters-summary-card">
                <div class="filters-summary-title">Current Selection</div>
                <div class="filters-summary-item">Venue Type: <b>{len(sel_vtype)}</b></div>
                <div class="filters-summary-item">Region: <b>{len(sel_reg)}</b></div>
                <div class="filters-summary-item">State: <b>{len(sel_state) if sel_state is not None else 'All'}</b></div>
                <div class="filters-summary-item">District: <b>{len(sel_district) if sel_district is not None else 'All'}</b></div>
                <div class="filters-summary-item">City: <b>{len(sel_city) if sel_city is not None else 'All'}</b></div>
                <div class="filters-summary-item">Status: <b>{len(sel_stat)}</b></div>
                <div class="filters-summary-item">ABC Category: <b>{len(sel_cat)}</b></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.sidebar.markdown(
            """
            <div class="sidebar-actions-shell">
                <div class="sidebar-actions-title">Quick Actions</div>
                <div class="sidebar-actions-subtitle">Navigate to admin tools instantly</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.sidebar.button("Manage Master Data", key="dash_manage_master_sidebar", use_container_width=True):
            st.session_state.page = "main" if st.session_state.logged_in else "login"
            st.rerun()
        if st.sidebar.button("Process Analysis Engine", key="dash_process_engine_sidebar", use_container_width=True):
            if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                st.session_state.page = "process"
            elif not st.session_state.logged_in:
                st.session_state.page = "login"
            else:
                st.error("Restricted to Admin/Prafull.")
            st.rerun()
            
        # if st.sidebar.button("Project-wise Details", key="dash_proj_manpower_sidebar", use_container_width=True):
            # if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                # st.session_state.manpower_search_code = ""
                # st.session_state.page = "project_wise_details"
            # elif not st.session_state.logged_in:
                # st.session_state.page = "login"
            # else:
                # st.error("Restricted to Admin/Prafull.")
            # st.rerun()

        if st.sidebar.button("Manpower Details", key="dash_manpower_dashboard_sidebar", use_container_width=True):
            if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                st.session_state.manpower_search_code = ""
                st.session_state.page = "manpower_dashboard"
            elif not st.session_state.logged_in:
                st.session_state.page = "login"
            else:
                st.error("Restricted to Admin/Prafull.")
            st.rerun()

        if st.sidebar.button("DATC & DOTC Inventory Summary", key="dash_city_datc_dotc_sidebar", use_container_width=True):
            if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                st.session_state.page = "city_datc_dotc"
            elif not st.session_state.logged_in:
                st.session_state.page = "login"
            else:
                st.error("Restricted to Admin/Prafull.")
            st.rerun()   

            
        mask = ((data['venue_type'].isin(sel_vtype)) | (data['venue_type'].str.strip().str.upper() == "DATC") | (data['venue_type'].str.strip().str.upper() == "DOTC")) & (data['status'].isin(sel_stat))

        region_filter = sel_reg
        if top_region != "All Regions":
            region_filter = [top_region]
        if 'region' in data.columns:
            mask &= data['region'].isin(region_filter)

        state_filter = None
        if top_state != "All States":
            state_filter = [top_state]
        elif sel_state is not None:
            state_filter = sel_state
        if state_filter is not None and 'state' in data.columns:
            mask &= data['state'].isin(state_filter)

        district_filter = None
        if top_district != "All Districts":
            district_filter = [top_district]
        elif sel_district is not None:
            district_filter = sel_district
        if district_filter is not None and 'district' in data.columns:
            mask &= data['district'].isin(district_filter)

        city_filter = None
        if top_city != "All Cities":
            city_filter = [top_city]
        elif sel_city is not None:
            city_filter = sel_city
        if city_filter is not None and 'city' in data.columns:
            mask &= data['city'].isin(city_filter)

        if 'category' in data.columns:
            mask = mask & (data['category'].isin(sel_cat))
        df_dash = data[mask].copy()

        def fmt_pct(part: int, total: int) -> str:
            if total <= 0:
                return "N/A"
            return f"{(part / total) * 100:.1f}%"

        def yes_count(df_in, col_name: str) -> int:
            if col_name not in df_in.columns:
                return 0
            s = df_in[col_name].astype(str).str.strip().str.upper()
            return int((s == "YES").sum())

        def numeric_sum(df_in, col_name: str) -> float:
            if col_name not in df_in.columns:
                return 0.0
            return float(pd.to_numeric(df_in[col_name], errors="coerce").fillna(0).sum())

        def numeric_nonnull(df_in, col_name: str) -> int:
            if col_name not in df_in.columns:
                return 0
            return int(pd.to_numeric(df_in[col_name], errors="coerce").notna().sum())

        total_filtered = int(len(df_dash))
        total_overall = int(len(data))

        active_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'ACTIVE'])) if 'status' in df_dash.columns else 0
        inactive_filtered = int(len(df_dash[df_dash['status'].str.upper() == 'INACTIVE'])) if 'status' in df_dash.columns else 0
        black_filtered = int(len(df_dash[df_dash['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in df_dash.columns else 0

        active_overall = int(len(data[data['status'].str.upper() == 'ACTIVE'])) if 'status' in data.columns else 0
        inactive_overall = int(len(data[data['status'].str.upper() == 'INACTIVE'])) if 'status' in data.columns else 0
        black_overall = int(len(data[data['status'].str.contains('BLACK', na=False, case=False)])) if 'status' in data.columns else 0

        if 'category' in data.columns:
            cat_series_overall = data['category'].astype(str).str.strip().str.upper()
            cat_counts_overall = cat_series_overall.value_counts()
        else:
            cat_counts_overall = {}

        def render_clickable_metric_card(col, card_id, label, value, selected_col, tooltip=None, pct=None):
            with col:
                metric_card(label, value, tooltip=tooltip, pct=pct)
                if st.button("View rows", key=f"dash_card_{card_id}", use_container_width=True):
                    st.session_state.dash_selected_card = card_id
                    st.session_state.dash_selected_card_label = label
                    st.session_state.dash_selected_card_column = selected_col

        # --- Metric Cards Row ---
        col_dotc1, col_dotc2, col_sep, col_gen1, col_gen2, col_gen3, col_sep_2, col_gen4 = st.columns([1, 1, 0.1, 1, 1, 1, 0.1, 1])

        _dotc_df_temp2 = df_dash[df_dash['venue_type'].str.strip().str.upper() == "DOTC"] if 'venue_type' in df_dash.columns else pd.DataFrame()
        _dotc_total_temp = int(len(_dotc_df_temp2)) if not _dotc_df_temp2.empty else 0
        pct = (_dotc_total_temp/total_overall)*100 if total_overall > 0 else 0
        render_clickable_metric_card(col_dotc1, "dotc_centres", "DOTC Centres", _dotc_total_temp, "venue_type", pct=pct)

        _dotc_df_temp = df_dash[df_dash['venue_type'].str.strip().str.upper() == "DOTC"] if 'venue_type' in df_dash.columns else pd.DataFrame()
        _dotc_seats_temp = int(pd.to_numeric(_dotc_df_temp['venue_max_capacity'], errors='coerce').fillna(0).sum()) if 'venue_max_capacity' in _dotc_df_temp.columns and not _dotc_df_temp.empty else 0
        pct = (_dotc_seats_temp/total_overall)*100 if total_overall > 0 else 0
        render_clickable_metric_card(col_dotc2, "dotc_capacity", "DOTC Seat Capacity", _dotc_seats_temp, "venue_max_capacity", pct=pct)

        with col_sep:
            st.markdown('<div style="border-left: 3px solid #A1C4FC; height: 86px; margin: 3px auto; opacity: 0.6;"></div>', unsafe_allow_html=True)

        _datc_active_df = df_dash[
            (df_dash['venue_type'].str.strip().str.upper() == "DATC") & 
            (df_dash['status'].str.upper() == "ACTIVE")
        ] if 'venue_type' in df_dash.columns and 'status' in df_dash.columns else pd.DataFrame()
        datc_active_count = len(_datc_active_df)
        _datc_total_df = df_dash[df_dash['venue_type'].str.strip().str.upper() == "DATC"] if 'venue_type' in df_dash.columns else pd.DataFrame()
        datc_total_count = len(_datc_total_df)
        pct = (datc_active_count / datc_total_count) * 100 if datc_total_count > 0 else 0
        render_clickable_metric_card(col_gen1, "active_datc", "Active DATC", datc_active_count, "status", pct=pct)

        pct = (black_filtered/total_filtered)*100 if total_filtered > 0 else 0
        render_clickable_metric_card(col_gen2, "blacklisted_venues", "Blacklisted Venues", black_filtered, "status", pct=pct)

        _datc_df_total = df_dash[df_dash['venue_type'].str.strip().str.upper() == "DATC"] if 'venue_type' in df_dash.columns else pd.DataFrame()
        datc_total_filtered = len(_datc_df_total)
        pct = (datc_total_filtered / total_overall) * 100 if total_overall > 0 else 0
        render_clickable_metric_card(col_gen3, "total_datc", "Total DATC Venues", datc_total_filtered, "venue_type", pct=pct)

        with col_sep_2:
            st.markdown('<div style="border-left: 3px solid #A1C4FC; height: 86px; margin: 3px auto; opacity: 0.6;"></div>', unsafe_allow_html=True)

        render_clickable_metric_card(col_gen4, "inactive_venues", "Inactive Venues", inactive_overall_count, "status", pct=0)


        # Insert a custom small gap (e.g., 10 pixels)
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)

        if st.session_state.get("dash_selected_card"):
            def get_selected_card_rows(df_in, card_key):
                if card_key == "dotc_centres":
                    return df_in[df_in['venue_type'].str.strip().str.upper() == "DOTC"] if 'venue_type' in df_in.columns else df_in.head(0)
                if card_key == "dotc_capacity":
                    return df_in[df_in['venue_type'].str.strip().str.upper() == "DOTC"] if 'venue_type' in df_in.columns else df_in.head(0)
                if card_key == "active_datc":
                    return df_in[(df_in['venue_type'].str.strip().str.upper() == "DATC") & (df_in['status'].str.strip().str.upper() == "ACTIVE")] if 'venue_type' in df_in.columns and 'status' in df_in.columns else df_in.head(0)
                if card_key == "blacklisted_venues":
                    return df_in[df_in['status'].str.contains('BLACK', case=False, na=False)] if 'status' in df_in.columns else df_in.head(0)
                if card_key == "total_datc":
                    return df_in[df_in['venue_type'].str.strip().str.upper() == "DATC"] if 'venue_type' in df_in.columns else df_in.head(0)
                if card_key == "inactive_venues":
                    return df_in[df_in['status'].str.strip().str.upper() == "INACTIVE"] if 'status' in df_in.columns else df_in.head(0)
                return df_in.head(0)

            selected_card = st.session_state.get("dash_selected_card")
            selected_label = st.session_state.get("dash_selected_card_label", selected_card)
            selected_col = st.session_state.get("dash_selected_card_column")
            selected_df = get_selected_card_rows(df_dash, selected_card)

            st.markdown(
                f"""
                <div style='border:1px solid var(--border); background:var(--surface); border-radius:12px; padding:14px 18px; margin-bottom:16px;'>
                    <div style='display:flex; justify-content:space-between; align-items:center; gap:10px;'>
                        <div style='font-weight:700; color:var(--text); font-size:14px;'>{html.escape(selected_label)} - {len(selected_df):,} row(s)</div>
                        <div style='color:var(--text-muted); font-size:12px;'>{html.escape(selected_col) if selected_col else ''}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            if selected_df.empty:
                st.info(f"No rows available for {selected_label} with the current filter set.")
            else:
                ordered_cols = list(selected_df.columns)
                if selected_col in ordered_cols:
                    ordered_cols = [selected_col] + [c for c in ordered_cols if c != selected_col]
                st.dataframe(selected_df[ordered_cols], use_container_width=True, height=420)

            if st.button("Clear card selection", key="dash_clear_card_selection", use_container_width=False):
                st.session_state.dash_selected_card = None
                st.session_state.dash_selected_card_label = None
                st.session_state.dash_selected_card_column = None

        # --- Interactive Executive Dashboards ---
        st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
        graph_col1, graph_col2, graph_col3 = st.columns([1, 1, 1], gap="medium")

        with graph_col1:
            st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Category Distribution</h3>", unsafe_allow_html=True)
            if 'category' in df_dash.columns:
                cat_series = df_dash['category'].astype(str).str.strip().str.upper()
                cat_counts = cat_series.value_counts()

                labels = ["A", "B", "C", "BL", "BL-C"]
                values = [cat_counts.get(l, 0) for l in labels]
                colors = ["#C8F6C2", "#AEEBFE", "#FFF6A3", "#FECCCE", "#E9B6FA"]

                chart_rows = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0]
                if chart_rows:
                    pie_labels = [r[0] for r in chart_rows]
                    pie_values = [r[1] for r in chart_rows]
                    pie_colors = [r[2] for r in chart_rows]
                    total_pie = sum(pie_values)
                    pie_text = [
                        f"<b>{lab}</b><br>{(val/total_pie)*100:.1f}%"
                        if total_pie and (val / total_pie) >= 0.06 else ""
                        for lab, val in zip(pie_labels, pie_values)
                    ]

                    fig_pie = go.Figure(data=[go.Pie(
                        labels=pie_labels,
                        values=pie_values,
                        sort=False,
                        direction="clockwise",
                        marker=dict(colors=pie_colors, line=dict(color='#FFFFFF', width=2)),
                        text=pie_text,
                        textinfo='text',
                        textposition='inside',
                        insidetextfont=dict(size=12, color='#243447', family='sans-serif'),
                        hovertemplate="<b>%{label}</b><br>Count: %{value:,}<br>Share: %{percent}<extra></extra>"
                    )])
                    fig_pie.update_layout(
                        showlegend=True,
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=-0.2,
                            xanchor="center",
                            x=0.5,
                            font=dict(size=11, color="#334155")
                        ),
                        margin=dict(t=8, b=26, l=8, r=8),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        height=300
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)

        with graph_col2:
            st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>Infrastructure Readiness Index</h3>", unsafe_allow_html=True)
            cctv_c = yes_count(df_dash, "cctv")
            pwd_c = yes_count(df_dash, "pwd_venue")
            genset_c = yes_count(df_dash, "genset_availability")
            ups_c = yes_count(df_dash, "ups_availability")

            total_active = total_filtered if total_filtered > 0 else 1

            r_values = [
                (cctv_c/total_active)*100,
                (pwd_c/total_active)*100,
                (ups_c/total_active)*100,
                (genset_c/total_active)*100
            ]
            r_labels = ['CCTV', 'PWD Ready', 'UPS', 'Genset']

            r_values_loop = r_values + [r_values[0]]
            r_labels_loop = r_labels + [r_labels[0]]

            fig_radar = go.Figure()
            fig_radar.add_trace(go.Scatterpolar(
                r=r_values_loop,
                theta=r_labels_loop,
                mode='lines+markers',
                fill='toself',
                fillcolor='rgba(129, 205, 236, 0.30)',
                line=dict(color='#236AA5', width=2.6),
                marker=dict(size=6, color='#236AA5', line=dict(color='#FFFFFF', width=1.1)),
                hoverinfo="theta+r",
                hovertemplate="<b>%{theta}</b><br>Readiness: %{r:.1f}%<extra></extra>"
            ))
            fig_radar.update_layout(
                polar=dict(
                    bgcolor='rgba(255,255,255,0.45)',
                    radialaxis=dict(
                        visible=True,
                        range=[0, 100],
                        tickmode='linear',
                        tick0=0,
                        dtick=20,
                        showline=True,
                        linewidth=1,
                        linecolor='#9AAFC3',
                        gridcolor='#C6D5E5',
                        gridwidth=1.2,
                        tickfont=dict(size=10, color='#5F7285')
                    ),
                    angularaxis=dict(
                        showline=True,
                        linecolor='#AABDD0',
                        gridcolor='#DCE7F2',
                        gridwidth=1,
                        tickfont=dict(size=12, color='#2D3748', family='sans-serif')
                    )
                ),
                showlegend=False,
                margin=dict(t=20, b=5, l=28, r=28),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                height=300
            )
            st.plotly_chart(fig_radar, use_container_width=True)

        with graph_col3:
            st.markdown("<h3 style='color:var(--primary-dark); font-size:16px; margin-bottom: 2px;'>India Coverage Map</h3>", unsafe_allow_html=True)
            map_df, india_geojson, _, _ = build_india_map_dataframe(df_dash)
            if india_geojson is None:
                st.warning(f"GeoJSON not found at: {INDIA_GEOJSON_PATH}")
            else:
                fig_map_preview = create_india_map_figure(
                    map_df,
                    india_geojson,
                    height=300,
                    show_data=False,
                    zoom=2.35,
                    center=INDIA_MAP_CENTER,
                    opacity=0.9,
                    line_color="#7C96B5",
                    line_width=0.85,
                )
                if fig_map_preview is not None:
                    id_to_state_map_preview = {
                        str(row["id"]): str(row["state"])
                        for _, row in map_df.iterrows()
                    }
                    selected_state_from_preview_click = None
                    try:
                        map_event = st.plotly_chart(
                            fig_map_preview,
                            use_container_width=True,
                            key="india_map_preview_v5",
                            on_select="rerun",
                            selection_mode=("points",),
                            config={"displayModeBar": False},
                        )
                        selected_state_from_preview_click = resolve_state_from_map_points(
                            extract_selected_points(map_event),
                            id_to_state_map_preview,
                            ordered_names=map_df["state"].astype(str).tolist(),
                        )
                    except TypeError:
                        st.plotly_chart(
                            fig_map_preview,
                            use_container_width=True,
                            key="india_map_preview_static_v5",
                            config={"displayModeBar": False},
                        )

                    if selected_state_from_preview_click:
                        st.session_state.map_selected_state = selected_state_from_preview_click
                        st.session_state.page = "india_map_detail"
                        st.query_params["page"] = "india_map_detail"
                        st.rerun()

                    if st.button("Open Detailed Map", key="open_detailed_india_map", use_container_width=True):
                        st.session_state.page = "india_map_detail"
                        st.query_params["page"] = "india_map_detail"
                        st.rerun()
             
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)


        if not df_dash.empty:
            st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Status Distribution</h3>", unsafe_allow_html=True)
            status_col1, status_col2 = st.columns([1.5, 1.1], gap="medium")

            c_data = df_dash.groupby(['region', 'status'], observed=False).size().reset_index(name='Count')
            status_color_map = {"ACTIVE": "#81CDEC", "INACTIVE": "#CBD5E1", "BLACK LISTED": "#FECCCE"}
            fig_stat = px.bar(c_data, x='region', y='Count', color='status', barmode='group', text='Count', color_discrete_map=status_color_map, category_orders={"region": REGION_ORDER})
            fig_stat.update_traces(textposition='outside', textfont=dict(color='#334155', size=12), cliponaxis=False, marker_line_width=0)
            fig_stat.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
            )
            with status_col1:
                st.plotly_chart(fig_stat, use_container_width=True)

            pivot_status = df_dash.pivot_table(index='region', columns='status', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
            status_row_order = [r for r in REGION_ORDER if r in pivot_status.index] + ["Grand Total"]
            pivot_status = pivot_status.reindex(status_row_order)
            
            headers_stat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_status.columns]
            cells_stat = [pivot_status.index.tolist()] + [pivot_status[col].tolist() for col in pivot_status.columns]
            
            fig_tab_stat = go.Figure(data=[go.Table(
                header=dict(values=headers_stat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
                cells=dict(values=cells_stat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_status)//2 + 1)][0][:len(pivot_status)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
            )])
            fig_tab_stat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")
            
            with status_col2:
                st.plotly_chart(fig_tab_stat, use_container_width=True)


            if 'category' in df_dash.columns:
                st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
                st.markdown("<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>Region-wise Category Breakdown</h3>", unsafe_allow_html=True)
                cat_col1, cat_col2 = st.columns([1.5, 1.1], gap="medium")

                cat_data = df_dash.groupby(['region', 'category'], observed=False).size().reset_index(name='Count')
                color_map = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
                fig_cat = px.bar(cat_data, x='region', y='Count', color='category', barmode='group', text='Count', color_discrete_map=color_map, category_orders={"region": REGION_ORDER, "category": CAT_ORDER})
                fig_cat.update_traces(textposition='outside', textfont=dict(color='#334155', size=11), cliponaxis=False, marker_line_width=0)
                fig_cat.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#2A3B4C"), xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zerolinecolor="#F1F5F9", title="Venue Count"),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                    margin=dict(t=20, b=10, l=10, r=10), height=350, bargap=0.15
                )

                with cat_col1:
                    st.plotly_chart(fig_cat, use_container_width=True)

                pivot_cat = df_dash.pivot_table(index='region', columns='category', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name="Grand Total", observed=False)
                cat_col_order = [c for c in CAT_ORDER if c in pivot_cat.columns] + ["Grand Total"]
                pivot_cat = pivot_cat.reindex(index=status_row_order, columns=cat_col_order)
                
                headers_cat = ['<b>REGION</b>'] + [f"<b>{str(c).upper()}</b>" for c in pivot_cat.columns]
                cells_cat = [pivot_cat.index.tolist()] + [pivot_cat[col].tolist() for col in pivot_cat.columns]
                
                fig_tab_cat = go.Figure(data=[go.Table(
                    header=dict(values=headers_cat, fill_color='#F8FAFC', align='left', font=dict(color='#475569', size=12), line_color='#E2E8F0', height=40),
                    cells=dict(values=cells_cat, fill_color=[['#FFFFFF', '#F8FAFC'] * (len(pivot_cat)//2 + 1)][0][:len(pivot_cat)], align='left', font=dict(color='#1E293B', size=13), line_color='#E2E8F0', height=35)
                )])
                fig_tab_cat.update_layout(margin=dict(t=10, b=10, l=0, r=0), height=350, paper_bgcolor="rgba(0,0,0,0)")

                with cat_col2:
                    st.plotly_chart(fig_tab_cat, use_container_width=True)

            st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)

            # --- DOTC & DATC VENTURE DISTRIBUTION (NEW SECTION) ---

            st.markdown("""
            <div style="background:var(--surface); border-radius:12px; padding:0.5px 15px; border-left:5px solid #8B5CF6; box-shadow:var(--shadow-sm); margin-bottom:24px;">
                <h2 style='color:var(--primary-dark); font-weight:700; font-size:15px; margin:0; line-height:1.1; padding:0;'>Operational Distribution (DOTC vs DATC)</h2>
                <p style='color:var(--text-muted); font-size:11px; margin:0; padding:0; line-height:1.2; font-weight:500;'>Detailed geographic breakdown of venue classifications</p>
            </div>
            """, unsafe_allow_html=True)

            # Filter and normalize
            df_type = df_dash.copy()
            if 'venue_type' in df_type.columns:
                df_type['v_type_norm'] = df_type['venue_type'].astype(str).str.strip().str.upper()
                df_type = df_type[df_type['v_type_norm'].isin(['DOTC', 'DATC'])]

                # Helper to render a distribution block (Chart + Table)
                def render_type_breakdown_block(df_subset, dimension, title_label, show_chart=True, chart_top_n=None):
                    st.markdown(f"<h3 style='color:var(--primary-dark); font-size:16px; font-weight:700; margin-top:20px; margin-bottom:12px;'>{title_label}</h3>", unsafe_allow_html=True)
                    
                    if df_subset.empty or dimension not in df_subset.columns:
                        st.info(f"No specific DOTC/DATC data available for {title_label}.")
                        return

                    # --- Build enriched pivot with Venue Count & Capacity ---
                    # Count pivot (DATC / DOTC counts)
                    pivot_count = df_subset.pivot_table(index=dimension, columns='v_type_norm', values='dms_code', aggfunc='count', fill_value=0, margins=True, margins_name='Grand Total', observed=False)
                    for col in ['DATC', 'DOTC']:
                        if col not in pivot_count.columns:
                            pivot_count[col] = 0
                    pivot_count = pivot_count[['DATC', 'DOTC']]
                    pivot_count['Total Venues'] = pivot_count['DATC'] + pivot_count['DOTC']

                # Capacity pivot
                    if 'venue_max_capacity' in df_subset.columns:
                        cap_series = pd.to_numeric(df_subset['venue_max_capacity'], errors='coerce').fillna(0)
                        df_cap = df_subset.copy()
                        df_cap['_cap_num'] = cap_series
                        
                        # NEW: Calculate capacity specifically for DATC and DOTC
                        cap_pivot = df_cap.pivot_table(index=dimension, columns='v_type_norm', values='_cap_num', aggfunc='sum', fill_value=0, observed=False)
                        for col in ['DATC', 'DOTC']:
                            if col not in cap_pivot.columns: cap_pivot[col] = 0
                        
                        total_cap = cap_series.sum()
                    else:
                        cap_pivot = pd.DataFrame(columns=['DATC', 'DOTC'])
                        total_cap = 0

                    # NEW: Assembly of columns to include DATC Cap and DOTC Cap
                    pivot_count['DATC Cap'] = pivot_count.index.map(lambda x: int(cap_pivot.loc[x, 'DATC']) if x in cap_pivot.index else (int(cap_pivot['DATC'].sum()) if x == 'Grand Total' else 0))
                    pivot_count['DOTC Cap'] = pivot_count.index.map(lambda x: int(cap_pivot.loc[x, 'DOTC']) if x in cap_pivot.index else (int(cap_pivot['DOTC'].sum()) if x == 'Grand Total' else 0))
                    
                    pivot_count['Total Capacity'] = pivot_count['DATC Cap'] + pivot_count['DOTC Cap']

                    if dimension == 'region':
                        row_order = [r for r in REGION_ORDER if r in pivot_count.index] + ["Grand Total"]
                        pivot_count = pivot_count.reindex(row_order)
                    else:
                        pivot_count = pivot_count.sort_values('Total Venues', ascending=False)

                    # --- Chart & Table Layout ---
                    if show_chart:
                        with st.container(border=True):
                            ch1, ch2 = st.columns(2, gap="large")
                            
                            with ch1:
                                st.markdown("<h4 style='color:#334155; font-size:14px; font-weight:700; margin-bottom:8px;'>Venue Count Distribution</h4>", unsafe_allow_html=True)
                                grp = df_subset.groupby([dimension, 'v_type_norm'], observed=False).size().reset_index(name='Count')
                                chart_df = grp.copy()
                                if chart_top_n:
                                    top_dims = df_subset[dimension].value_counts().head(chart_top_n).index.tolist()
                                    chart_df = chart_df[chart_df[dimension].isin(top_dims)]
    
                                fig = px.bar(chart_df, x=dimension, y='Count', color='v_type_norm', barmode='group', text='Count', 
                                             color_discrete_map={'DOTC': '#8B5CF6', 'DATC': '#3B82F6'},
                                             category_orders={dimension: REGION_ORDER if dimension == 'region' else None})
                                fig.update_traces(textposition='outside', textfont=dict(size=11), marker_line_width=0)
                                fig.update_layout(height=340, margin=dict(t=10, b=10, l=10, r=10), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                                  xaxis=dict(title=""), yaxis=dict(title="Count"),
                                                  legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""))
                                st.plotly_chart(fig, use_container_width=True)
                                
                            with ch2:
                                st.markdown("<h4 style='color:#334155; font-size:14px; font-weight:700; margin-bottom:8px;'>Total Seat Capacity</h4>", unsafe_allow_html=True)
                                if 'venue_max_capacity' in df_subset.columns:
                                    cap_grp = df_cap.groupby([dimension, 'v_type_norm'], observed=False)['_cap_num'].sum().reset_index(name='Capacity')
                                    if chart_top_n:
                                        cap_grp = cap_grp[cap_grp[dimension].isin(top_dims)]
                                    
                                    fig_cap = px.bar(cap_grp, x=dimension, y='Capacity', color='v_type_norm', barmode='group', text='Capacity', 
                                                 color_discrete_map={'DOTC': '#8B5CF6', 'DATC': '#3B82F6'},
                                                 category_orders={dimension: REGION_ORDER if dimension == 'region' else None})
                                    fig_cap.update_traces(textposition='outside', texttemplate='%{text:,.0f}', textfont=dict(size=11), marker_line_width=0)
                                    fig_cap.update_layout(height=340, margin=dict(t=10, b=10, l=10, r=10), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                                      xaxis=dict(title=""), yaxis=dict(title="Seat Capacity"),
                                                      showlegend=False)
                                    st.plotly_chart(fig_cap, use_container_width=True)
                                else:
                                    st.info("Seat capacity data not available.")
                                    
                            st.markdown("<hr style='margin: 15px 0; border-color: #E2E8F0;'>", unsafe_allow_html=True)
                            st.markdown("<h4 style='color:#334155; font-size:14px; font-weight:700; margin-bottom:12px;'>Data Summary Table</h4>", unsafe_allow_html=True)

                            header_map = {'DATC': 'DATC Venues', 'DOTC': 'DOTC Venues', 'Total Venues': 'Total Venues', 'Seat Capacity': 'Seat Capacity'}
                            dim_label = dimension.upper()
                            if dimension == 'region':
                                dim_label = 'REGION'
                            
                            headers = [f"<b>{dim_label}</b>"] + [f"<b>{header_map.get(c, c)}</b>" for c in pivot_count.columns]
                            cells = [pivot_count.index.tolist()] + [pivot_count[c].tolist() for c in pivot_count.columns]
                            
                            fig_tab = go.Figure(data=[go.Table(
                                columnwidth=[1.5, 1, 1, 1, 1],
                                header=dict(values=headers, fill_color='#F8FAFC', align='left', font=dict(size=12, color='#475569'), line_color='#E2E8F0', height=38),
                                cells=dict(values=cells, fill_color='#FFFFFF', align='left', font=dict(size=12, color='#1E293B'), line_color='#E2E8F0', height=34)
                            )])
                            table_height = min(400, 60 + (len(pivot_count) * 34))
                            fig_tab.update_layout(margin=dict(t=0, b=0, l=0, r=0), height=table_height, paper_bgcolor="rgba(0,0,0,0)")
                            st.plotly_chart(fig_tab, use_container_width=True)
                    else:
                        header_map = {'DATC': 'DATC Venues', 'DOTC': 'DOTC Venues', 'Total Venues': 'Total Venues', 'Seat Capacity': 'Seat Capacity'}
                        dim_label = dimension.upper()
                        headers = [f"<b>{dim_label}</b>"] + [f"<b>{header_map.get(c, c)}</b>" for c in pivot_count.columns]
                        cells = [pivot_count.index.tolist()] + [pivot_count[c].tolist() for c in pivot_count.columns]
                        
                        fig_tab = go.Figure(data=[go.Table(
                            header=dict(values=headers, fill_color='#F8FAFC', align='left', font=dict(size=11, color='#475569'), line_color='#E2E8F0', height=35),
                            cells=dict(values=cells, fill_color='#FFFFFF', align='left', font=dict(size=11, color='#1E293B'), line_color='#E2E8F0', height=30)
                        )])
                        fig_tab.update_layout(margin=dict(t=0, b=0, l=0, r=0), height=400, paper_bgcolor="rgba(0,0,0,0)")
                        st.plotly_chart(fig_tab, use_container_width=True)

                # Render Region-wise
                render_type_breakdown_block(df_type, 'region', "Region-wise Distribution")

            else:
                st.warning("'venue_type' column missing from data for DOTC/DATC breakdown.")


# ================= INDIA MAP DETAIL PAGE =================
elif st.session_state.page == "india_map_detail":
    map_head_col, map_back_col = st.columns([4, 1], vertical_alignment="center")
    with map_head_col:
        dataset_name = st.session_state.get('dataset_key')
        if dataset_name and dataset_name != "Venue Inventory":
            header_text = f"DEXIT | {dataset_name} - Pan-India Venue Coverage"
        else:
            header_text = "DEXIT | Master Venue Coverage Map"
            
        st.markdown(
            f"""
            <div style="margin-bottom:0px; padding-top:0px;">
                <h2 style="margin:0; color:#1E293B; font-weight:800;">{header_text}</h2>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with map_back_col:
        if st.button("\u2190 Back to Dashboard", use_container_width=True):
            st.session_state.page = "dashboard"
            st.query_params["page"] = "dashboard"
            st.rerun()

    all_files = [
        os.path.join(OUTPUT_PATH, f)
        for f in os.listdir(OUTPUT_PATH)
        if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")
    ]

    detail_df = pd.DataFrame()
    if all_files:
        latest_file_path = max(all_files, key=os.path.getmtime)
        detail_df = pd.read_excel(latest_file_path, sheet_name=0)
        detail_df.columns = detail_df.columns.str.lower().str.replace(" ", "_")
        
        # Globally filter out INACTIVE venues
        if "status" in detail_df.columns:
            detail_df = detail_df[detail_df["status"].fillna("").astype(str).str.upper() != "INACTIVE"].copy()
            
        # Clean State Data
        if "state" in detail_df.columns:
            detail_df["state"] = detail_df["state"].fillna("").astype(str).apply(normalize_state_name)
            detail_df = detail_df[~detail_df["state"].isin(FOREIGN_LOCATIONS)]
        
        st.caption(f"Data Source: {os.path.basename(latest_file_path)}")
    else:
        st.warning("No processed reports found. Showing map boundaries only.")

    map_df, india_geojson, mapped_states, mapped_venues = build_india_map_dataframe(detail_df)
    if india_geojson is None:
        st.error(f"India GeoJSON file not found: {INDIA_GEOJSON_PATH}")
    else:
        map_total_states = int(len(map_df))
        data_total_venues = int(len(detail_df)) if not detail_df.empty else 0

        fig_map_detail = create_india_map_figure(
            map_df,
            india_geojson,
            height=760,
            show_data=True,
            zoom=3.2,
            center=INDIA_MAP_CENTER,
            opacity=0.92,
            map_style="white-bg",
            color_power=0.55,
            line_color="#334155",
            line_width=1.05,
            restrict_to_india=True,
        )
        selected_state_from_click = None
        id_to_state_map = {
            str(row["id"]): str(row["state"])
            for _, row in map_df.iterrows()
        }

        if fig_map_detail is not None:
            try:
                detail_map_event = st.plotly_chart(
                    fig_map_detail,
                    use_container_width=True,
                    key="india_map_detail_chart_v5",
                    on_select="rerun",
                    selection_mode=("points",),
                    config={"displayModeBar": False, "scrollZoom": False},
                )
                selected_points = extract_selected_points(detail_map_event)
                selected_state_from_click = resolve_state_from_map_points(
                    selected_points,
                    id_to_state_map,
                    ordered_names=map_df["state"].astype(str).tolist(),
                )
                selected_click_signature = get_last_point_signature(selected_points)
            except TypeError:
                st.plotly_chart(
                    fig_map_detail,
                    use_container_width=True,
                    key="india_map_detail_chart_static_v5",
                    config={"displayModeBar": False, "scrollZoom": False},
                )
                selected_click_signature = None
        else:
            selected_click_signature = None

        if selected_state_from_click and selected_click_signature:
            prev_sig = st.session_state.get("_india_map_detail_last_click_sig")
            if selected_click_signature != prev_sig:
                st.session_state["_india_map_detail_last_click_sig"] = selected_click_signature
                st.session_state.map_selected_state = selected_state_from_click
        elif selected_click_signature is None:
            st.session_state["_india_map_detail_last_click_sig"] = st.session_state.get("_india_map_detail_last_click_sig", None)

        if selected_state_from_click and not selected_click_signature:
            st.session_state.map_selected_state = selected_state_from_click

        # --- PAN INDIA DOTC/DATC SUMMARY TABLE ---
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        st.markdown("### 📊 Pan India: DOTC & DATC Network Overall Summary", unsafe_allow_html=True)
        
        if not detail_df.empty and "venue_type" in detail_df.columns:
            pan_df = detail_df[detail_df["venue_type"].astype(str).str.strip().str.upper().isin(["DATC", "DOTC"])].copy()
            if not pan_df.empty:
                pan_df["v_type_norm"] = pan_df["venue_type"].astype(str).str.strip().str.upper()
                
                # Venue Count
                pan_count = pan_df.groupby("v_type_norm", as_index=False)["dms_code"].nunique()
                pan_count.rename(columns={"dms_code": "Venue Count", "v_type_norm": "Venue Type"}, inplace=True)
                
                # Seat Capacity
                if "venue_max_capacity" in pan_df.columns:
                    pan_cap = pan_df.assign(_cap=pd.to_numeric(pan_df["venue_max_capacity"], errors="coerce").fillna(0)).groupby("v_type_norm", as_index=False)["_cap"].sum()
                    pan_cap["_cap"] = pan_cap["_cap"].astype(int)
                    pan_cap.rename(columns={"_cap": "Total Seat Capacity", "v_type_norm": "Venue Type"}, inplace=True)
                else:
                    pan_cap = pd.DataFrame({"Venue Type": ["DATC", "DOTC"], "Total Seat Capacity": [0, 0]})
                    
                pan_summary = pd.merge(pd.DataFrame({"Venue Type": ["DATC", "DOTC"]}), pan_count, on="Venue Type", how="left").fillna(0)
                pan_summary = pd.merge(pan_summary, pan_cap, on="Venue Type", how="left").fillna(0)
                
                pan_summary["Venue Count"] = pan_summary["Venue Count"].astype(int)
                pan_summary["Total Seat Capacity"] = pan_summary["Total Seat Capacity"].astype(int)
                
                # Add total row
                total_row = pd.DataFrame({
                    "Venue Type": ["Grand Total"],
                    "Venue Count": [pan_summary["Venue Count"].sum()],
                    "Total Seat Capacity": [pan_summary["Total Seat Capacity"].sum()]
                })
                pan_summary = pd.concat([pan_summary, total_row], ignore_index=True)
                
                st.dataframe(pan_summary, use_container_width=True, hide_index=True)
            else:
                st.info("No DATC or DOTC venues found in the data.")
        else:
            st.info("Venue Type information is missing from the dataset.")
            
        st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

        state_totals = (
            map_df.groupby("state", as_index=False)["venue_count"]
            .max()
            .sort_values(["venue_count", "state"], ascending=[False, True], kind="stable")
        )
        available_states = state_totals["state"].astype(str).tolist()

        if available_states:
            # Check if we have a valid selected state in session, otherwise default to None (no selection)
            state_index = None
            if st.session_state.map_selected_state in available_states:
                state_index = available_states.index(st.session_state.map_selected_state)

            selected_state = st.selectbox(
                "Selected State",
                options=available_states,
                index=state_index,
                placeholder="Choose a State...",
                key="india_detail_state_picker",
            )
            if selected_state != st.session_state.map_selected_state:
                st.session_state.map_selected_state = selected_state
                st.session_state["_india_map_detail_last_click_sig"] = None
            selected_state = st.session_state.map_selected_state
            if selected_state:
                st.caption(f"Current State: {selected_state}")
        else:
            selected_state = None

        if selected_state and not detail_df.empty:
            state_df = build_state_detail_payload(detail_df, selected_state)
            # st.markdown(
            #     f"<h3 style='color:var(--primary-dark); font-weight:800; border-bottom:1px solid #E2E8F0; padding-bottom:8px;'>{html.escape(selected_state)} Detailed Snapshot</h3>",
            #     unsafe_allow_html=True,
            # )
            if state_df.empty:
                st.info(f"No records available for {selected_state} in current output.")
            else:
                # status_series = state_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in state_df.columns else pd.Series([""] * len(state_df))
                # venue_type_series = state_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in state_df.columns else pd.Series([""] * len(state_df))
                # total_state = int(len(state_df))
                # active_state = int((status_series == "ACTIVE").sum())
                # datc_state = int((venue_type_series == "DATC").sum())
                # dotc_state = int((venue_type_series == "DOTC").sum())

                # sk1, sk2, sk3, sk4 = st.columns(4)
                # with sk1:
                #     metric_card_compact("State Total", total_state, accent="#236AA5")
                # with sk2:
                #     metric_card_compact("DATC", datc_state, accent="#0EA5E9")
                # with sk3:
                #     metric_card_compact("DOTC", dotc_state, accent="#6366F1")
                # with sk4:
                #     metric_card_compact("Active", active_state, accent="#10B981")


                st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
                render_dotc_datc_summary(f"{selected_state} Summary", compute_dotc_datc_summary(state_df))

                st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
                state_slug = resolve_state_slug(selected_state)
                state_geojson = load_state_geojson_by_slug(state_slug) if state_slug else None

                if st.session_state.map_selected_district_state != selected_state:
                    st.session_state.map_selected_district = None
                    st.session_state.map_selected_district_state = selected_state

                district_map_df = pd.DataFrame()
                district_id_map = {}

                if state_geojson is not None:
                    district_map_df, district_id_map = build_district_map_dataframe(state_df, state_geojson)
                    if not district_map_df.empty:
                        dplot = district_map_df.copy()
                        dplot["_color_metric"] = np.power(pd.to_numeric(dplot["venue_count"], errors="coerce").fillna(0).astype(float), 0.55)
                        fig_district_map = px.choropleth(
                            dplot,
                            geojson=state_geojson,
                            locations="id",
                            featureidkey="id",
                            color="_color_metric",
                            hover_name="district",
                            custom_data=["district", "id", "venue_count", "datc_count", "dotc_count", "active_count", "black_count"],
                            color_continuous_scale=["#D8ECFA", "#8EC9F2", "#2D8CD6", "#0B4F8A"],
                            range_color=(0, max(1.0, float(dplot["_color_metric"].max()))),
                        )
                        fig_district_map.update_traces(
                            marker_line_color="#334155",
                            marker_line_width=1.0,
                            hovertemplate=(
                                "<b>%{hovertext}</b><br>"
                                "Total: %{customdata[2]:,}<br>"
                                "DATC: %{customdata[3]:,}<br>"
                                "DOTC: %{customdata[4]:,}<br>"
                                "Active: %{customdata[5]:,}<br>"
                                "Blacklisted: %{customdata[6]:,}<extra></extra>"
                            ),
                        )
                        
                        id_to_dname = dict(zip(dplot["id"], dplot["district"]))
                        dlons, dlats, dtexts = [], [], []
                        if state_geojson and "features" in state_geojson:
                            for f in state_geojson["features"]:
                                fid = f.get("id")
                                if fid in id_to_dname:
                                    cx, cy = _get_feature_centroid(f)
                                    dlons.append(cx)
                                    dlats.append(cy)
                                    dtexts.append(id_to_dname[fid])
                        
                        if dlons:
                            fig_district_map.add_trace(go.Scattergeo(
                                lon=dlons,
                                lat=dlats,
                                text=dtexts,
                                mode="text",
                                textfont=dict(color="#0F172A", size=11, family="Arial, sans-serif", weight="bold"),
                                textposition="middle center",
                                showlegend=False,
                                hoverinfo="skip"
                            ))
                            
                        fig_district_map.update_geos(fitbounds="locations", visible=False)
                        fig_district_map.update_layout(
                            clickmode="event+select",
                            dragmode="pan",
                            margin=dict(t=0, b=0, l=0, r=0),
                            paper_bgcolor="rgba(0,0,0,0)",
                            plot_bgcolor="rgba(0,0,0,0)",
                            height=560,
                            coloraxis_colorbar=dict(title="Venues", thickness=12, len=0.75),
                        )

                        selected_district_from_click = None
                        try:
                            district_map_event = st.plotly_chart(
                                fig_district_map,
                                use_container_width=True,
                                key=f"district_map_{state_slug}",
                                on_select="rerun",
                                selection_mode=("points",),
                            )
                            selected_district_from_click = resolve_state_from_map_points(
                                extract_selected_points(district_map_event),
                                district_id_map,
                                ordered_names=district_map_df["district"].astype(str).tolist(),
                            )
                        except TypeError:
                            st.plotly_chart(fig_district_map, use_container_width=True, key=f"district_map_{state_slug}_static")

                        if selected_district_from_click:
                            st.session_state.map_selected_district = selected_district_from_click

                        district_options = district_map_df["district"].astype(str).tolist()
                        if district_options:
                            # Use None for index to avoid default selection
                            district_idx = None
                            if st.session_state.map_selected_district in district_options:
                                district_idx = district_options.index(st.session_state.map_selected_district)

                            picked_district = st.selectbox(
                                f"Selected District in {selected_state}",
                                options=district_options,
                                index=district_idx,
                                placeholder="Choose a District...",
                                key=f"district_picker_{state_slug}",
                            )
                            if picked_district != st.session_state.map_selected_district:
                                st.session_state.map_selected_district = picked_district

                            selected_district = st.session_state.map_selected_district
                            if selected_district:
                                st.caption(f"Current District: {selected_district}")
                                # Robust column detection for district
                                d_col = next((c for c in state_df.columns if c.lower() in ["district", "district_name", "dist"]), "district")
                                
                                # Try clean uppercase match first
                                district_df = state_df[
                                    state_df[d_col].fillna("").astype(str).str.upper().str.strip() == str(selected_district).upper().strip()
                                ].copy()
                                
                                # Fallback to normalization match if direct match fails
                                if district_df.empty:
                                    district_df = state_df[
                                        state_df[d_col].fillna("").astype(str).apply(normalize_district_name) == normalize_district_name(selected_district)
                                    ].copy()


                                # st.markdown(
                                #     f"<h4 style='color:#1E293B; margin-top:6px; margin-bottom:8px;'>{html.escape(selected_district)} District Stats</h4>",
                                #     unsafe_allow_html=True,
                                # )

                                if district_df.empty:
                                    st.info(f"No data available for {selected_district}.")
                                else:
                                    d_status_series = district_df["status"].fillna("").astype(str).str.strip().str.upper() if "status" in district_df.columns else pd.Series([""] * len(district_df))
                                    d_vtype_series = district_df["venue_type"].fillna("").astype(str).str.strip().str.upper() if "venue_type" in district_df.columns else pd.Series([""] * len(district_df))
                                    d_total = int(len(district_df))
                                    d_datc = int((d_vtype_series == "DATC").sum())
                                    d_dotc = int((d_vtype_series == "DOTC").sum())
                                    d_active = int((d_status_series == "ACTIVE").sum())

                                    # dk1, dk2, dk3, dk4 = st.columns(4)
                                    # with dk1:
                                    #     metric_card_compact("District Total", d_total, accent="#1D4ED8")
                                    # with dk2:
                                    #     metric_card_compact("DATC", d_datc, accent="#0EA5E9")
                                    # with dk3:
                                    #     metric_card_compact("DOTC", d_dotc, accent="#6366F1")
                                    # with dk4:
                                    #     metric_card_compact("Active", d_active, accent="#10B981")

                                    st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
                                    render_dotc_datc_summary(f"{selected_district} Summary", compute_dotc_datc_summary(district_df))


                        else:
                            st.info("District list not available for this state map.")
                else:
                    st.warning(f"District map file not available for {selected_state}.")

# ================= LOGIN PAGE =================
elif st.session_state.page == "login":
    # Inject large top spacing to vertically center the card mentally
    st.markdown('<div class="gap-lg"></div><div class="gap-lg"></div>', unsafe_allow_html=True)
    
    col_l, col_center, col_r = st.columns([1, 1.2, 1])
    
    with col_center:
        # High-End Security Icon & Header
        st.markdown(f"""
        <div style="text-align: center; margin-bottom: 30px;">
            <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect>
                <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
            </svg>
            <h2 style="color:var(--primary-dark); font-weight:800; text-transform:uppercase; letter-spacing:1px; margin-top:16px;">Secure Executive Portal</h2>
            <p style="color:#64748B; font-size:14px; margin-top:-8px;">Authorized Administrative Personnel Only</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("login_form", border=True):
            st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
            u = st.text_input("Username", placeholder="e.g. Admin")
            st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
            p = st.text_input("Password", type="password", placeholder="\u2022\u2022\u2022\u2022\u2022\u2022\u2022\u2022")
            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
            
            # Using primary type for maximum focus
            if st.form_submit_button("Authenticate Access", type="primary", use_container_width=True):
                if u in USERS and USERS[u] == p:
                    st.session_state.logged_in, st.session_state.user, st.session_state.page = True, u, "dashboard"
                    st.rerun()
                else: 
                    st.error("Authentication Failed: Invalid Credentials")
                
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        if st.button("\u2190 Return to Dashboard", use_container_width=True):
            st.session_state.page = "dashboard"
            st.rerun()

# ================= MASTER DATA VIEW =================
elif st.session_state.page == "main":
    allowed = [k for k, v in FILES.items() if st.session_state.user in v["users"]]
    st.session_state.dataset_key = st.sidebar.radio("ACTIVE MASTER", allowed)
    if st.query_params.get("dataset", None) != st.session_state.dataset_key:
        st.query_params["dataset"] = st.session_state.dataset_key
    current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
    current_df = load_smart_data(current_path)

    # --- CEO Level Header Layout ---
    head_col1, head_col2 = st.columns([3, 1])
    with head_col1:
        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:14px; margin-bottom: 10px;">
            <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <ellipse cx="12" cy="5" rx="9" ry="3"></ellipse>
                <path d="M21 12c0 1.66-4 3-9 3s-9-1.34-9-3"></path>
                <path d="M3 5v14c0 1.66 4 3 9 3s9-1.34 9-3V5"></path>
            </svg>
            <div>
                <h2 style="margin:0; font-weight:800; color:#1E293B;">Master Data Center</h2>
                <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Active Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with head_col2:
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        if st.button("\u2190 Return to Dashboard", use_container_width=True):
            st.session_state.page = "dashboard"
            st.rerun()

    if st.session_state.dataset_key == "Venue Inventory" and not current_df.empty:
        status_series_main = (
            current_df["status"].astype(str).str.strip().str.upper()
            if "status" in current_df.columns else pd.Series([""] * len(current_df))
        )
        main_active = int((status_series_main == "ACTIVE").sum())
        main_black = int(status_series_main.str.contains("BLACK", na=False).sum())
        main_total = int(len(current_df))

        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        mk1, mk2, mk3 = st.columns(3, gap="small")
        with mk1:
            metric_card_compact("Active Venues", main_active, accent="#86d7a8")
        with mk2:
            metric_card_compact("Blacklisted", main_black, accent="#f4a3b5")
        with mk3:
            metric_card_compact("Total Venues", main_total, accent="#81cdec")

    st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
    
    # --- Action Panel & Metrics ---
    act_col1, act_col2, ext_col = st.columns([1.5, 3, 1.5])
    with act_col1:
        if st.button("\u2699 Open Advanced Editor", type="primary", use_container_width=True): 
            st.session_state.page = "edit"
            st.rerun()
            
    with act_col2:
        # Display dataset health quickly as premium pill badges
        st.markdown(f"""
        <div style="display:flex; gap:15px; padding-top:6px;">
            <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Rows Scanned: {len(current_df):,}</span>
            <span style="background:#F1F5F9; color:#334155; padding:6px 14px; border-radius:6px; font-weight:700; font-size:12px; border: 1px solid #E2E8F0;">Columns Mapped: {len(current_df.columns)}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
    st.dataframe(current_df, use_container_width=True, height=600)


# ================= EDIT PAGE =================
elif st.session_state.page == "edit":
    st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="display:flex; align-items:center; gap:14px; margin-bottom: 25px;">
        <svg width="36" height="36" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
            <polyline points="16 18 22 12 16 6"></polyline>
            <polyline points="8 6 2 12 8 18"></polyline>
        </svg>
        <div>
            <h2 style="margin:0; font-weight:800; color:#1E293B;">Advanced Data Controller</h2>
            <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Target Engine: <b style="color:var(--primary-dark);">{st.session_state.dataset_key}</b></p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    current_path = os.path.join(BASE_PATH, FILES[st.session_state.dataset_key]["file"])
    current_df = load_smart_data(current_path)

    # --- TOP ACTIONS & FILTER ROW ---
    st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
    col_search, col_save, col_clear, col_exit = st.columns([3.5, 1.5, 1.2, 1.2], gap="small")
    
    with col_search:
        search = st.text_input("Search", key="search_main_box", placeholder="\U0001F50D Search live database by keywords, target names...", label_visibility="collapsed")
    with col_save: 
        save_trigger = st.button("\U0001F4BE Commit Changes", type="primary", use_container_width=True)
    with col_clear:
        if st.button("\U0001F9F9 Clear", use_container_width=True):
            st.session_state.search_main_box = ""
            st.rerun()
    with col_exit: 
        if st.button("\u2190 Back", use_container_width=True):
            st.session_state.page = "main"
            st.rerun()
    filtered = current_df.copy()
    if search:
        filtered = filtered[filtered.astype(str).apply(lambda x: x.str.contains(search, case=False).any(), axis=1)]

    # --- BULK TOOLS (CLEANING & DUPLICATE CHECK) ---
    if st.session_state.dataset_key != "Call Logs":
        st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
        with st.container(border=True):
            st.subheader("Bulk Injection & Extraction Tools")
            tool1, tool2 = st.columns([1, 2], gap="large")
            
            with tool1:
                st.caption("Download Blank Template")
                blank_df = pd.DataFrame(columns=current_df.columns)
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    blank_df.to_excel(writer, index=False)
                st.download_button(label="\U0001F4E5 Download Format", data=buffer.getvalue(), 
                                   file_name=f"Format_{FILES[st.session_state.dataset_key]['file']}", mime="application/vnd.ms-excel", use_container_width=True)
            
            with tool2:
                st.caption("Bulk Upload Data (Clean & Update)")
                uploaded_file = st.file_uploader("Upload Excel", type=["xlsx"], label_visibility="collapsed")
                
                if uploaded_file:
                    try:
                        new_data = pd.read_excel(uploaded_file)

                        d_master = find_dms_col(current_df)
                        d_new = find_dms_col(new_data)

                        if d_master and d_new:
                            # Pre-clean the codes for matching
                            new_data[d_new] = new_data[d_new].astype(str).str.replace(r'\s+', '', regex=True).str.upper()
                            current_df[d_master] = current_df[d_master].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

                            # Check length (informational)
                            invalid_count = len(new_data[new_data[d_new].str.len() != 11])
                            if invalid_count > 0:
                                st.warning(f"{invalid_count} codes in upload are not 11 digits.")

                            if st.button("Confirm Bulk Update & Clean", use_container_width=True):
                                # 1. Set index to DMS Code for both dataframes
                                current_df.set_index(d_master, inplace=True)
                                new_data.set_index(d_new, inplace=True)

                                # 2. Update existing rows with new data
                                current_df.update(new_data)

                                # 3. Identify and append completely NEW rows
                                new_ids = new_data.index.difference(current_df.index)
                                new_rows = new_data.loc[new_ids]
                                
                                # 4. Combine and reset index back to original
                                final_df = pd.concat([current_df, new_rows]).reset_index()
                                
                                # Rename the index column back to the original Master column name if it changed
                                final_df.rename(columns={'index': d_master}, inplace=True)

                                # 5. Save to File
                                final_df.to_excel(current_path, index=False)
                                st.success(f"Done! {len(new_data)} rows processed (Updates + New entries).")
                                st.rerun()
                        else:
                            st.error(f"Missing ID Column! Master needs '{d_master}' and Upload needs a DMS/Venue Code column.")
                    
                    except Exception as e:
                        st.error(f"Error processing file: {e}")

    # --- ML AUTO-TAGGING FOR CALL LOGS ---
    if st.session_state.dataset_key == "Call Logs":
        st.markdown("<hr style='margin: 30px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:10px; margin-bottom: 12px;">
            <svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <rect x="2" y="2" width="20" height="8" rx="2" ry="2"></rect>
                <rect x="2" y="14" width="20" height="8" rx="2" ry="2"></rect>
                <line x1="6" y1="6" x2="6.01" y2="6"></line>
                <line x1="6" y1="18" x2="6.01" y2="18"></line>
            </svg>
            <h3 style="margin:0; color:#1E293B; font-size:18px;">AI Neural Auto-Classification Engine</h3>
        </div>
        """, unsafe_allow_html=True)
        
        ml_col1, ml_col2 = st.columns(2, gap="large")
        with ml_col1:
            with st.container(border=True):
                st.subheader("1. Model Training Module")
                st.caption("Upload historical Call Logs data so the AI can learn human categorizations.")
    
                uploaded_train = st.file_uploader(
                    "Upload Training Batch",
                    type=["xlsx", "xls", "csv"],
                    key="ml_train_uploader",
                    label_visibility="collapsed",
                    help="File must contain FINAL_ISSUE_REMARK_BY_PRAFULL (and other target columns if available).",
                )

            if st.button("Train/Update ML Model", key="ml_train_btn", use_container_width=True):
                with st.spinner("Training Model... This may take a minute..."):
                    import ml_engine

                    if not uploaded_train:
                        st.error("Please upload a training Excel file first.")
                    else:
                        train_ext = os.path.splitext(uploaded_train.name)[1].lower()
                        if train_ext not in [".xlsx", ".xls", ".csv"]:
                            train_ext = ".xlsx"
                        train_path = os.path.join(BASE_PATH, f"temp_train_upload{train_ext}")
                        with open(train_path, "wb") as f:
                            f.write(uploaded_train.getbuffer())

                        success, msg = ml_engine.train_ml_model(train_path)
                        if success:
                            st.success(f"{msg} (Source: uploaded file)")
                        else:
                            st.error(msg)
        with ml_col2:
            with st.container(border=True):
                st.subheader("2. Live Processing Module")
                st.caption("Upload newly downloaded raw Call Logs to classify and format them automatically.")
                
                uploaded_raw = st.file_uploader("Upload Raw Data Stream", type=["xlsx", "xls", "csv"], key="ml_uploader", label_visibility="collapsed")
                if uploaded_raw and st.button("\u26A1 Process with AI Engine", type="primary", use_container_width=True):
                    with st.spinner("Processing..."):
                        import ml_engine
                        raw_ext = os.path.splitext(uploaded_raw.name)[1].lower()
                        if raw_ext not in [".xlsx", ".xls", ".csv"]:
                            raw_ext = ".xlsx"
                        temp_path = os.path.join(BASE_PATH, f"temp_raw_upload{raw_ext}")
                        
                        # --- AUTO-COMBINE Gateway/Core/Mock sheets if present ---
                        try:
                            if raw_ext in [".xlsx", ".xls"]:
                                xl = pd.ExcelFile(io.BytesIO(uploaded_raw.getvalue()))
                                sheet_names = xl.sheet_names
                                
                                # Check for Gateway/Core/Mock sheets
                                gwcm_sheets = [s for s in sheet_names if any(k in s.lower() for k in ["gateway", "core", "mock"])]
                                
                                if len(gwcm_sheets) > 1:
                                    # Auto-combine multiple Gateway/Core/Mock sheets
                                    st.info(f"Auto-combining {len(gwcm_sheets)} Gateway/Core/Mock sheets: {', '.join(gwcm_sheets)}")
                                    frames = []
                                    for s in gwcm_sheets:
                                        df_part = xl.parse(s)
                                        s_low = str(s).lower()
                                        if "gateway" in s_low:
                                            data_source = "GATEWAY"
                                        elif "core" in s_low:
                                            data_source = "CORE"
                                        elif "mock" in s_low:
                                            data_source = "MOCK"
                                        else:
                                            data_source = str(s).upper()
                                        df_part.insert(0, "DATA_SOURCE", data_source)
                                        df_part.insert(1, "SOURCE_SHEET", s)
                                        frames.append(df_part)
                                    
                                    combined_df = pd.concat(frames, ignore_index=True, sort=False)
                                    # Save combined to single Excel file
                                    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                                        combined_df.to_excel(writer, sheet_name="Sheet1", index=False)
                                    st.success(f"Combined {len(gwcm_sheets)} sheets: {len(combined_df):,} total rows")
                                else:
                                    # Single sheet or no Gateway/Core/Mock sheets - save as is
                                    with open(temp_path, "wb") as f:
                                        f.write(uploaded_raw.getbuffer())
                            else:
                                # CSV file - save as is
                                with open(temp_path, "wb") as f:
                                    f.write(uploaded_raw.getbuffer())
                        except Exception as e:
                            st.warning(f"Could not auto-combine sheets: {e}")
                            with open(temp_path, "wb") as f:
                                f.write(uploaded_raw.getbuffer())
                        
                        # --- Process with ML ---
                        res = ml_engine.process_call_logs(temp_path)
                        if res.get("success"):
                            btn1_file = res.get("output_file_btn1", res.get("output_file"))
                            btn1_path = res.get("output_path_btn1", res.get("output_path"))
                            btn2_file = res.get("output_file_btn2")
                            btn2_path = res.get("output_path_btn2")

                            if btn2_file:
                                st.success(
                                    f"Processing Complete! Generated files: {btn1_file} and {btn2_file}"
                                )
                            else:
                                st.success(f"Processing Complete! Saved to: {btn1_file}")
                            ml_output = None
                            try:
                                if btn1_path:
                                    ml_output = pd.read_excel(btn1_path)
                            except Exception:
                                pass
                            
                            if "auto_approved" in res and "needs_review" in res and "auto_approved_threshold" in res:
                                st.write(
                                    f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}"
                                    f" | Auto-approved (\u2265{res['auto_approved_threshold']:.2f}): {res['auto_approved']}"
                                    f" | Needs review: {res['needs_review']}"
                                )
                            else:
                                st.write(f"Stats | Total: {res['total_rows']} | Cached Fast: {res['cached']} | ML Predicted: {res['ml_predicted']}")
                            
                            # Show preview if output was loaded successfully
                            if ml_output is not None and not ml_output.empty:
                                st.info("Processed Data Preview (AI-tagged output)")
                                st.dataframe(ml_output.head(100), use_container_width=True)
                            
                            dl_col1, dl_col2 = st.columns(2)
                            with dl_col1:
                                if btn1_path and btn1_file:
                                    with open(btn1_path, "rb") as f:
                                        st.download_button(
                                            label="Download Combined (New + Training)",
                                            data=f,
                                            file_name=btn1_file,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True
                                        )
                            with dl_col2:
                                if btn2_path and btn2_file:
                                    with open(btn2_path, "rb") as f:
                                        st.download_button(
                                            label="Download Combined (Training + New)",
                                            data=f,
                                            file_name=btn2_file,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True
                                        )
                        else:
                            st.error(res.get("msg", "Unknown error occurred."))

    # --- DATA EDITOR ---
    edited = st.data_editor(filtered, use_container_width=True, height=500, num_rows="dynamic", key="data_editor_tool")

    if save_trigger:
        # Manual Edit Cleaning
        d_edit = find_dms_col(edited)
        if d_edit:
            edited[d_edit] = edited[d_edit].astype(str).str.replace(r'\s+', '', regex=True).str.upper()

        if search:
            current_df.update(edited)
            current_df.to_excel(current_path, index=False)
        else:
            edited.to_excel(current_path, index=False)
        st.success("Changes Saved & Cleaned Successfully!")
        st.rerun()

    # --- PHOTO MANAGER (AUTOMATED PHOTO UPLOAD) ---
    if st.session_state.dataset_key == "Venue Inventory" and render_photo_manager_ui:
        st.markdown('<div class="gap-lg"></div>', unsafe_allow_html=True)
        st.markdown("---")
        st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
        render_photo_manager_ui(current_df, PHOTOS_PATH)

        
# ================= PROCESS PAGE =================
elif st.session_state.page == "process":
    # --- CEO Level Header Layout ---
    head_col1, head_col2 = st.columns([3, 1])
    with head_col1:
        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:14px; margin-bottom: 5px;">
            <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="var(--primary-dark)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"></polygon>
            </svg>
            <div>
                <h2 style="margin:0; font-weight:800; color:#1E293B;">Central Analysis Engine</h2>
                <p style="margin:0; color:#64748B; font-size:14px; text-transform:uppercase; letter-spacing:1px;">Core Processing & Consolidation Node</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with head_col2:
        st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
        if st.button("\u2190 Return to Dashboard", use_container_width=True):
            st.session_state.page = "dashboard"
            st.rerun()
    
    st.markdown("<hr style='margin: 10px 0 20px 0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown(f"""
        <div style="margin-bottom: 20px;">
            <h3 style="margin:0; color:#1E293B; font-size:18px;">Full Ecosystem Compilation</h3>
            <p style="margin-top:4px; color:#64748B; font-size:14px;">Running the core engine will securely sequence Blacklisted Venues, FFA Data, Daily Observations, Market Reviews, and Machine-Learning Auto-Tagging into a singular unified Master Database.</p>
        </div>
        """, unsafe_allow_html=True)
        
        col_run, col_space = st.columns([2, 3])
        with col_run:
            if st.button("\u26A1 Execute Core Analysis Consolidation", type="primary", use_container_width=True):
                status = st.empty()
                with st.spinner("Initializing Heavy Processing Sequence..."):
                    try:
                        # 1. LOAD
                        vms = load_smart_data(os.path.join(BASE_PATH, "VMS Master.xlsx"))
                        blacklisted = load_smart_data(os.path.join(BASE_PATH, "Blacklisted Venues.xlsx"))
                        ffa = load_smart_data(os.path.join(BASE_PATH, "FFA Data.xlsx"))
                        deo = load_smart_data(os.path.join(BASE_PATH, "Day End Observations.xlsx"))
                        suspicious = load_smart_data(os.path.join(BASE_PATH, "Suspicious Activities.xlsx"))
                        market_review = load_smart_data(os.path.join(BASE_PATH, "Google review and ratings.xlsx"))
                        call_logs = load_smart_data(os.path.join(BASE_PATH, "Master Call Logs.xlsx"))
    
                        # 2. ALIGN
                        mapping = {"VENUE_CODE": "dms_code", "DMS_CODE": "dms_code", "DMS_Code": "dms_code", "DMS Code": "dms_code", "UPDATED_DMS_CODE": "dms_code", "DMS CODE": "dms_code", "Row Labels": "dms_code"}
                        def align_and_lower(df_map):
                            df_map.columns = df_map.columns.str.strip()
                            for old, new in mapping.items():
                                if old in df_map.columns: df_map = df_map.rename(columns={old: new})
                            df_map.columns = df_map.columns.str.lower().str.replace(" ", "_")
                            return df_map
    
                        vms, blacklisted, ffa = align_and_lower(vms), align_and_lower(blacklisted), align_and_lower(ffa)
                        deo, suspicious, market_review, call_logs = align_and_lower(deo), align_and_lower(suspicious), align_and_lower(market_review), align_and_lower(call_logs)
    
                        # 3. CLEANING
                        def clean_dms(df_c, col):
                            if col in df_c.columns:
                                df_c[col] = df_c[col].astype(str).str.replace(r'\.0$', '', regex=True).str.replace(r'\s+', '', regex=True).str.strip().str.upper()
                            return df_c
                        for d in [vms, blacklisted, ffa, deo, suspicious, market_review, call_logs]: clean_dms(d, "dms_code")
                        vms = vms[vms["dms_code"].notna() & (~vms["dms_code"].isin(["NAN", "", "GRANDTOTAL"]))]
    
                        # 4. MERGING
                        rem_bl = 'remark' if 'remark' in blacklisted.columns else 'remarks'
                        blacklisted["blacklist_info"] = blacklisted[rem_bl].fillna("").astype(str) + " by " + blacklisted["blacklisted_by"].fillna("").astype(str)
                        merged = vms.merge(blacklisted[["dms_code", "blacklist_info"]], on="dms_code", how="left").fillna({"blacklist_info": ""})
                        ffa_count = ffa.groupby("dms_code").size().reset_index(name="ffa_frequency")
                        merged = merged.merge(ffa_count, on="dms_code", how="left").fillna({"ffa_frequency": 0})
    
                        rem_deo = 'remark' if 'remark' in deo.columns else 'remarks'
                        deo[rem_deo] = deo[rem_deo].fillna("").astype(str).str.strip()
                        deo_summary = deo.groupby(["dms_code", rem_deo]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x[rem_deo], x["count"])), include_groups=False).reset_index(name="deo_remarks_summary")
                        merged = merged.merge(deo_summary, on="dms_code", how="left").fillna({"deo_remarks_summary": ""})
    
                        suspicious["issue_type"] = suspicious["issue_type"].fillna("").astype(str).str.strip()
                        susp_summary = suspicious.groupby(["dms_code", "issue_type"]).size().reset_index(name="count").groupby("dms_code").apply(lambda x: ", ".join(f"{r}({c} instances)" for r, c in zip(x["issue_type"], x["count"])), include_groups=False).reset_index(name="suspicious_summary")
                        merged = merged.merge(susp_summary, on="dms_code", how="left").fillna({"suspicious_summary": ""})

                        # 4A. Call Logs Enrichment (moved from Call Logs AI module to Central Analysis Engine)
                        def _first_non_blank(series: pd.Series) -> str:
                            s = series.fillna("").astype(str).str.strip()
                            s = s[~s.str.upper().isin(["", "NAN", "NONE", "NOT ASSIGNED", "NA", "-"])]
                            return s.iloc[0] if not s.empty else ""

                        if not call_logs.empty:
                            if "project_name" in call_logs.columns:
                                call_logs["updated_project_name"] = clean_project_name(call_logs["project_name"])

                            cl_dms_source = (
                                "updated_dms_code" if "updated_dms_code" in call_logs.columns else
                                "center_code" if "center_code" in call_logs.columns else
                                "dms_code" if "dms_code" in call_logs.columns else None
                            )
                            if cl_dms_source:
                                call_logs["updated_dms_code"] = (
                                    call_logs[cl_dms_source]
                                    .astype(str)
                                    .str.replace(r"\.0$", "", regex=True)
                                    .str.replace(r"\s+", "", regex=True)
                                    .str.strip()
                                    .str.upper()
                                )
                                if "dms_code" not in call_logs.columns:
                                    call_logs["dms_code"] = call_logs["updated_dms_code"]
                                else:
                                    call_logs["dms_code"] = call_logs["dms_code"].astype(str).str.strip()
                                    blank_mask = call_logs["dms_code"].str.upper().isin(["", "NAN", "NONE"])
                                    call_logs.loc[blank_mask, "dms_code"] = call_logs.loc[blank_mask, "updated_dms_code"]

                            clean_dms(call_logs, "dms_code")
                            cl_valid = call_logs[call_logs["dms_code"].notna()].copy()
                            cl_valid["dms_code"] = cl_valid["dms_code"].astype(str).str.strip().str.upper()
                            cl_valid = cl_valid[~cl_valid["dms_code"].isin(["", "NAN", "NONE", "GRANDTOTAL"])]

                            if not cl_valid.empty and "dms_code" in vms.columns:
                                vms_cols_needed = ["dms_code", "name", "address", "landmark", "city", "district", "state", "pincode"]
                                vms_ref = vms.copy()
                                for vc in vms_cols_needed:
                                    if vc not in vms_ref.columns:
                                        vms_ref[vc] = ""
                                vms_ref = vms_ref[vms_cols_needed].drop_duplicates(subset="dms_code", keep="first")
                                vms_map_name = dict(zip(vms_ref["dms_code"], vms_ref.get("name", "")))
                                vms_map_addr = dict(zip(vms_ref["dms_code"], vms_ref.get("address", "")))
                                vms_map_land = dict(zip(vms_ref["dms_code"], vms_ref.get("landmark", "")))
                                vms_map_city = dict(zip(vms_ref["dms_code"], vms_ref.get("city", "")))
                                vms_map_dist = dict(zip(vms_ref["dms_code"], vms_ref.get("district", "")))
                                vms_map_state = dict(zip(vms_ref["dms_code"], vms_ref.get("state", "")))
                                vms_map_pin = dict(zip(vms_ref["dms_code"], vms_ref.get("pincode", "")))

                                mapped_name = cl_valid["dms_code"].map(vms_map_name).fillna("")
                                venue_fallback = (
                                    cl_valid["venue_name"].fillna("").astype(str).str.strip().str.upper()
                                    if "venue_name" in cl_valid.columns else pd.Series([""] * len(cl_valid), index=cl_valid.index)
                                )
                                cl_valid["updated_venue_name"] = mapped_name.where(mapped_name.astype(str).str.strip() != "", venue_fallback)

                                cl_valid["address"] = cl_valid["dms_code"].map(vms_map_addr).fillna(cl_valid.get("address", ""))
                                cl_valid["landmark"] = cl_valid["dms_code"].map(vms_map_land).fillna(cl_valid.get("landmark", ""))
                                cl_valid["city"] = cl_valid["dms_code"].map(vms_map_city).fillna(cl_valid.get("city", ""))
                                cl_valid["district"] = cl_valid["dms_code"].map(vms_map_dist).fillna(cl_valid.get("district", ""))
                                cl_valid["state"] = cl_valid["dms_code"].map(vms_map_state).fillna(cl_valid.get("state", ""))
                                cl_valid["pincode"] = cl_valid["dms_code"].map(vms_map_pin).fillna(cl_valid.get("pincode", ""))
                                cl_valid["concat_address"] = build_concat_address(cl_valid)

                                profile_aggs = {}
                                if "updated_project_name" in cl_valid.columns:
                                    profile_aggs["call_logs_project_name"] = ("updated_project_name", _first_non_blank)
                                if "updated_venue_name" in cl_valid.columns:
                                    profile_aggs["call_logs_updated_venue_name"] = ("updated_venue_name", _first_non_blank)
                                if "concat_address" in cl_valid.columns:
                                    profile_aggs["call_logs_concat_address"] = ("concat_address", _first_non_blank)

                                if profile_aggs:
                                    call_logs_profile = cl_valid.groupby("dms_code", as_index=False).agg(**profile_aggs)
                                    merged = merged.merge(call_logs_profile, on="dms_code", how="left")
                                    for col_name in profile_aggs.keys():
                                        merged[col_name] = merged[col_name].fillna("")
    
                        if "final_issue_remark_by_prafull" in call_logs.columns:
                            
                            # Remove 'Other low impact issues' directly from remark column
                            filtered_logs = call_logs[
                                ~call_logs["final_issue_remark_by_prafull"]
                                .fillna("")
                                .str.strip()
                                .str.lower()
                                .str.contains("other low impact issues", na=False)
                            ]
                            
                            cl_summary = (
                                filtered_logs
                                .groupby(["dms_code", "final_issue_remark_by_prafull"])
                                .size()
                                .reset_index(name="count")
                                .groupby("dms_code")
                                .apply(
                                    lambda x: ", ".join(
                                        f"{r}({c})" for r, c in zip(
                                            x["final_issue_remark_by_prafull"], x["count"]
                                        )
                                    ),
                                    include_groups=False
                                )
                                .reset_index(name="call_log_summary")
                            )
    
                            merged = merged.merge(cl_summary, on="dms_code", how="left").fillna({"call_log_summary": ""})
                            
                            
                        market_review["google_ratings_num"] = pd.to_numeric(market_review["google_ratings"], errors="coerce")
                        market_review["google_review_count_num"] = pd.to_numeric(market_review["google_review_count"], errors="coerce")
                        m_cond = [market_review["google_review_count_num"] < 10, (market_review["google_ratings_num"] >= 4) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] <= 2) & (market_review["google_review_count_num"] >= 10), (market_review["google_ratings_num"] < 4) & (market_review["google_review_count_num"] >= 10)]
                        m_choice = ["Doubtful (Reviews < 10)", "Good (Rating >=4 & Reviews >=10)", "Poor (Rating <=2 & Reviews >=10)", "Average (Rating <4 & Reviews >=10)"]
                        market_review["market_remark"] = np.select(m_cond, m_choice, default="")
                        merged = merged.merge(market_review[["dms_code", "market_remark"]], on="dms_code", how="left").fillna({"market_remark": ""})
    
    
                        # 5. ANALYSIS RULES
                        merged["final_analysis"] = "Good"; merged["reason"] = "Good"
                        mask_comp = merged["blacklist_info"].str.lower().str.startswith("completely blacklisted", na=False)
                        mask_cust = merged["blacklist_info"].str.lower().str.startswith("customer specific blacklisted", na=False)
                        
                        # Assigning individually to avoid ValueError
                        merged.loc[mask_comp, "final_analysis"] = "Completely Blacklisted"
                        merged.loc[mask_comp, "reason"] = merged.loc[mask_comp, "blacklist_info"]
                        merged.loc[mask_cust, "final_analysis"] = "Customer Specific Blacklisted"
                        merged.loc[mask_cust, "reason"] = merged.loc[mask_cust, "blacklist_info"]
    
                        nb = merged[~(mask_comp | mask_cust)].copy()
                        
                        mask_ffa = nb["ffa_frequency"] >= 1
                        nb.loc[mask_ffa, "final_analysis"] = np.where(nb.loc[mask_ffa, "ffa_frequency"] < 5, "Average", "Sensitive")
                        nb.loc[mask_ffa, "reason"] = "Found in FFA | Count: " + nb.loc[mask_ffa, "ffa_frequency"].astype(int).astype(str)
    
                        if "deo_remarks_summary" in nb.columns:
                            mask_deo = (nb["deo_remarks_summary"] != "") & (nb["final_analysis"] == "Good")
                            deo_counts = deo.groupby("dms_code").size()
                            nb["deo_count"] = nb["dms_code"].map(deo_counts).fillna(0)
                            
                            # Force Sensitive Venue if any comment exists
                            nb.loc[mask_deo, "final_analysis"] = "Sensitive"
                            nb.loc[mask_deo, "reason"] = nb.loc[mask_deo, "deo_remarks_summary"]
    
                        if "suspicious_summary" in nb.columns:
                            mask_susp = (nb["suspicious_summary"] != "") & (nb["final_analysis"] == "Good")
                            susp_counts = suspicious.groupby("dms_code").size()
                            nb["susp_count"] = nb["dms_code"].map(susp_counts).fillna(0)
                            
                            # Force Sensitive Venue if any comment exists
                            nb.loc[mask_susp, "final_analysis"] = "Sensitive"
                            nb.loc[mask_susp, "reason"] = nb.loc[mask_susp, "suspicious_summary"]
    
                        if "call_log_summary" in nb.columns:
                            remark = nb["call_log_summary"].fillna("").str.strip()
                            
                            mask_call = (remark != "") & (nb["final_analysis"] == "Good")
                            
                            # Exclude anything containing 'Other low impact issues'
                            mask_valid_call = ~remark.str.lower().str.contains("other low impact issues", na=False)
                            
                            # Final mask
                            mask_final = mask_call & mask_valid_call
    
                            nb.loc[mask_final, "final_analysis"] = "Average"
                            nb.loc[mask_final, "reason"] = "Call Logs | " + nb.loc[mask_final, "call_log_summary"]
                        
                        if "market_remark" in nb.columns:
                            # Normalize text
                            remark = nb["market_remark"].fillna("").str.strip()
                            call_remark = nb["call_log_summary"].fillna("").str.strip()
                            
                            # Mask for rows where final_analysis is still Good
                            mask_base = nb["final_analysis"] == "Good"
                            
                            # [OK] Apply market logic ONLY if call_log_summary is blank OR only 'Other low impact issues'
                            mask_call_valid = (
                                (call_remark == "") |
                                (call_remark.str.lower().str.contains("other low impact issues", na=False))
                            )
                            
                            # Final base mask
                            mask_main = mask_base & mask_call_valid
                            
                            # Check if remark contains doubtful/poor/average
                            mask_issue = remark.str.lower().str.contains("doubtful|poor|average", na=False)
                            
                            # --- Case 1: Issue found -> Average ---
                            mask_avg = mask_main & mask_issue
                            nb.loc[mask_avg, "final_analysis"] = "Average"
                            nb.loc[mask_avg, "reason"] = "Market: " + nb.loc[mask_avg, "market_remark"]
                            
                            # --- Case 2: Other remarks -> Good ---
                            mask_good = mask_main & (remark != "") & (~mask_issue)
                            nb.loc[mask_good, "final_analysis"] = "Good"
                            nb.loc[mask_good, "reason"] = "Market: " + nb.loc[mask_good, "market_remark"]
                            
                            # --- Case 3: Blank -> Good Venue ---
                            mask_blank = mask_main & (remark == "")
                            nb.loc[mask_blank, "final_analysis"] = "Good"
                            nb.loc[mask_blank, "reason"] = "Good Venue"
                            
    
                        final_result = pd.concat([merged[mask_comp | mask_cust], nb], ignore_index=True)
    
                        # Rename output column: NAME -> VENUE_NAME
                        final_result = final_result.rename(columns={"name": "venue_name"})
                        
                        # --- 6. COLUMN WISE ANALYSIS ---
                        analysis_results = []
                        for col in final_result.columns:
                            if col in ['deo_count', 'susp_count']: continue
                            null_count = (final_result[col] == "").sum() + final_result[col].isna().sum()
                            analysis_results.append({
                                "Column Name": col,
                                "Fill Rate (%)": round(((len(final_result) - null_count) / len(final_result)) * 100, 2),
                                "Unique Values": final_result[col].nunique(),
                                "Top Frequent": str(final_result[col].mode()[0]) if not final_result[col].empty else "N/A"
                            })
                        audit_df = pd.DataFrame(analysis_results)
    
                        # Create Category column based on final_analysis
                        category_map = {
                            "Good": "A",
                            "Average": "B",
                            "Sensitive": "C",
                            "Completely Blacklisted": "BL",
                            "Customer Specific Blacklisted": "BL-C"
                        }
    
                        final_result["category"] = final_result["final_analysis"].map(category_map).fillna("")

                        # DOTC default override: keep DATC untouched
                        if "venue_type" in final_result.columns:
                            dotc_mask = final_result["venue_type"].astype(str).str.strip().str.upper() == "DOTC"
                            final_result.loc[dotc_mask, "final_analysis"] = "Good"
                            final_result.loc[dotc_mask, "reason"] = "Good Venue"
                            final_result.loc[dotc_mask, "category"] = "A"
                        
                        
                        # Cleanup
                        final_result = final_result.drop(columns=['deo_count', 'susp_count'], errors='ignore')
                        
    
                        # 7. SAVE
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
                        # Convert column names to uppercase for main file
                        final_result.columns = [col.upper() for col in final_result.columns]
                        audit_df.columns = [col.upper() for col in audit_df.columns]
    
                        out_path = os.path.join(OUTPUT_PATH, f"VMS_Full_Analysis_{ts}.xlsx")
    
                        # If available, also include the latest previous output in the new workbook
                        prev_out_path = None
                        try:
                            prev_candidates = [
                                os.path.join(OUTPUT_PATH, f)
                                for f in os.listdir(OUTPUT_PATH)
                                if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and (not f.startswith("~$"))
                            ]
                            prev_candidates = [p for p in prev_candidates if os.path.abspath(p) != os.path.abspath(out_path)]
                            if prev_candidates:
                                prev_out_path = max(prev_candidates, key=os.path.getmtime)
                        except:
                            prev_out_path = None
    
                        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                            final_result.to_excel(writer, sheet_name='Main_Analysis', index=False)
                            audit_df.to_excel(writer, sheet_name='Column_Audit', index=False)
    
                            run_info = pd.DataFrame([{
                                "CURRENT_OUTPUT_FILE": os.path.basename(out_path),
                                "PREVIOUS_OUTPUT_FILE": os.path.basename(prev_out_path) if prev_out_path else "",
                                "GENERATED_ON": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            }])
                            run_info.to_excel(writer, sheet_name="Run_Info", index=False)
    
                            if prev_out_path:
                                try:
                                    old_main = pd.read_excel(prev_out_path, sheet_name="Main_Analysis")
                                    old_main.to_excel(writer, sheet_name="Old_Main_Analysis", index=False)
                                except:
                                    pass
                                try:
                                    old_audit = pd.read_excel(prev_out_path, sheet_name="Column_Audit")
                                    old_audit.to_excel(writer, sheet_name="Old_Column_Audit", index=False)
                                except:
                                    pass
    
                        # --- Inventory File ---
                        inventory_cols = [
                            "venue_type", "region", "state", "city", "district", "dms_code",
                            "venue_name", "address", "pincode", "landmark", "complete_address",
                            "venue_max_capacity", "pwd_venue", "status", "final_analysis", "reason", "category"
                        ]
    
                        inventory_df = final_result[[col.upper() for col in inventory_cols if col.upper() in final_result.columns]].copy()
    
                        # Folder
                        inventory_folder = os.path.join(OUTPUT_PATH, "Venue Inventory")
                        os.makedirs(inventory_folder, exist_ok=True)
    
                        inventory_path = os.path.join(inventory_folder, f"Venue Inventory as on {ts}.xlsx")
                        inventory_df.to_excel(inventory_path, index=False)
    
                        # --- APPLY FORMATTING FUNCTION ---
                        def format_excel(file_path):
                            wb = load_workbook(file_path)
                            
                            header_font = Font(color="FFFFFF", bold=True)
                            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                            header_alignment = Alignment(horizontal="left")
    
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                
                                for cell in ws[1]:  # First row (headers)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
    
                            wb.save(file_path)
    
                        # Apply formatting to both files
                        format_excel(out_path)
                        format_excel(inventory_path)
    
                        # --- Status ---
                        status.success(f"Generated Analysis Reports.")
                        st.dataframe(final_result, use_container_width=True)
    
                    except Exception as e: st.error(f"Error: {e}"); st.exception(e)


# ================= VENUE DETAIL PAGE =================
elif st.session_state.page == "venue_detail":
    # Load data
    all_files = [os.path.join(OUTPUT_PATH, f) for f in os.listdir(OUTPUT_PATH)
                 if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")]

    if not all_files:
        st.error("No analysis data found.")
        if st.button("\u2190 Back to Dashboard"):
            st.session_state.page = "dashboard"
            st.rerun()
    else:
        latest_file_path = max(all_files, key=os.path.getmtime)
        data = pd.read_excel(latest_file_path, sheet_name=0)
        data.columns = data.columns.str.lower().str.replace(" ", "_")

        # Globally filter out INACTIVE venues
        if "status" in data.columns:
            data = data[data["status"].fillna("").astype(str).str.upper() != "INACTIVE"].copy()

        venue_code = st.session_state.venue_detail_code
        venue_row = data[data['dms_code'].astype(str).str.strip().str.upper() == str(venue_code).strip().upper()]

        if venue_row.empty:
            st.error(f"Venue with code **{venue_code}** not found.")
            if st.button("\u2190 Back to Dashboard"):
                st.session_state.page = "dashboard"
                st.rerun()
        else:
            v = venue_row.iloc[0]

            # Helper to safely get value (Preserving your existing logic)
            def _val(col, default="N/A"):
                val = v.get(col, default)
                if pd.isna(val) or str(val).strip().upper() in ["NAN", "NONE", ""]:
                    return default
                res = str(val).strip()
                if res.endswith(".0") and res[:-2].replace('-', '').isdigit():
                    res = res[:-2]
                return res

            # --- SEARCH BAR ON DETAIL PAGE ---
            st.markdown("""
            <div style="background: var(--surface);
                        border-radius: 6px; padding: 6px 26px; margin: 6px 0 20px 0;
                        border: 1px solid var(--border); border-left: 5px solid var(--primary); 
                        box-shadow: var(--shadow-sm);">
                <div style="color:var(--text); font-size:18px; font-weight:700; letter-spacing:0.5px; margin-bottom:6px; text-transform:uppercase;">
                    Quick Search Other Venues
                </div>
            </div>
            """, unsafe_allow_html=True)

            ds_col1, ds_col2 = st.columns([4, 1], vertical_alignment="center")
            with ds_col1:
                detail_search_input = st.text_input(
                    "Search Venue",
                    value="",
                    placeholder="Search by name or code to jump to another venue...",
                    label_visibility="collapsed",
                    key="detail_page_search_input"
                )
            with ds_col2:
                detail_search_btn = st.button("Search", use_container_width=True, type="primary", key="detail_page_search_btn")

            # Logic to handle search from Detail Page
            if detail_search_btn and detail_search_input.strip() != "":
                if st.session_state.logged_in and st.session_state.user in ["Admin", "Prafull"]:
                    search_term = detail_search_input.strip()
                    
                    # 1. Check for exact DMS code match
                    exact_match = data[data['dms_code'].astype(str).str.upper() == search_term.upper()]
                    
                    if not exact_match.empty:
                        # Found exact code: Stay on this page but refresh with new code
                        st.session_state.venue_detail_code = str(exact_match.iloc[0].get('dms_code', ''))
                        st.session_state.page = "venue_detail"
                        st.rerun()
                    else:
                        # Not an exact code: treat as keyword, reset detail state, and go to Dashboard
                        st.session_state.venue_search_query = search_term
                        st.session_state.venue_detail_code = None  # Reset specific view
                        st.session_state.page = "dashboard"        # Move to result list
                        st.rerun()
                else:
                    st.error("Restricted to Admin/Prafull.")

            # Now continue using the _val helper for your existing Venue Detail UI variables
            venue_name = _val("venue_name", _val("name", "Unknown Venue"))
            dms_code = _val("dms_code")
            category = _val("category")
            final_analysis = _val("final_analysis")
            status_val = _val("status")
            region = _val("region")
            city = _val("city")
            state = _val("state")
            district = _val("district")
            venue_type = _val("venue_type")

            # Category color
            cat_colors = {"A": "#C8F6C2", "B": "#AEEBFE", "C": "#FFF6A3", "BL": "#FECCCE", "BL-C": "#E9B6FA"}
            cat_color = cat_colors.get(category, "#DDE8F0")

            is_dotc = str(venue_type).strip().upper() == "DOTC"
            suppress_analysis_sections = is_dotc

            # Analysis color
            analysis_colors = {
                "Good": "#C8F6C2", "Average": "#FFF6A3", "Sensitive": "#FECCCE",
                "Completely Blacklisted": "#FDBED2", "Customer Specific Blacklisted": "#E9B6FA"
            }
            analysis_color = analysis_colors.get(final_analysis, "#DDE8F0")

            # --- TOP BAR ---
            top_left, top_right = st.columns([1, 4])
            with top_left:
                if st.button("\u2190 Back to Dashboard", use_container_width=True):
                    st.session_state.page = "dashboard"
                    st.rerun()

            # --- HERO HEADER ---
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, white 0%, #EBF3F8 40%, #CBEAFC 100%);
                        border-radius: 16px; padding: 20px 24px; margin: 8px 0 20px 0;
                        box-shadow: 0 4px 16px rgba(0,0,0,0.05); position:relative; overflow:hidden;
                        border: 1px solid #DDE8F0;">
                <div style="position:absolute; top:-20px; right:-20px; width:140px; height:140px;
                            background:rgba(174,235,254,0.3); border-radius:50%;"></div>
                <div style="position:absolute; bottom:-30px; right:60px; width:90px; height:90px;
                            background:rgba(174,235,254,0.2); border-radius:50%;"></div>
                <div>
                    <div style="color:#2A3B4C; font-size:28px; font-weight:800; letter-spacing:-0.5px;
                                line-height:1.2;">{html.escape(venue_name)}</div>
                    <div style="color:#566D82; font-size:14px; margin-top:4px; font-weight:600; letter-spacing:0.5px; text-transform:uppercase;">
                        {html.escape(dms_code)} <span style="margin:0 6px; color:#839BB2;">|</span> {html.escape(venue_type)}
                    </div>
                </div>
                <!-- Badges removed as per user request -->
            </div>
            """, unsafe_allow_html=True)

            # --- VENUE PHOTO ---
            venue_photo_url = _val("venue_photo_url", _val("image_url", ""))
            
            # Check for local photos if URL is not provided
            photo_urls = []
            if not venue_photo_url or venue_photo_url in ["N/A", ""]:
                import glob
                import base64
                # Search for any image matching the DMS code pattern (e.g EST-AR-1161.jpg, EST-AR-1161_1.jpg, etc)
                photo_matches = glob.glob(os.path.join(PHOTOS_PATH, f"{dms_code}*.*"))
                if photo_matches:
                    # Filter for actual images and sort them
                    img_files = sorted([p for p in photo_matches if p.lower().endswith(('.png', '.jpg', '.jpeg', '.webp'))])
                    for img_file in img_files:
                        with open(img_file, "rb") as image_file:
                            b64_str = base64.b64encode(image_file.read()).decode("utf-8")
                            ext = os.path.splitext(img_file)[1][1:].lower()
                            ext = "jpeg" if ext == "jpg" else ext
                            photo_urls.append(f"data:image/{ext};base64,{b64_str}")
            else:
                photo_urls = [venue_photo_url]
            
            import streamlit.components.v1 as components

            img_col, info_col, metrics_col = st.columns([1.2, 1.8, 1])

            with img_col:
                if photo_urls:
                    total_photos = len(photo_urls)
                    # Build slides HTML
                    slides_html = ""
                    for idx, purl in enumerate(photo_urls):
                        slides_html += f'<div class="slide"><img src="{purl}" alt="Venue photo {idx+1}" /></div>\n'

                    # Arrows + dots + counter (only for multi-photo)
                    arrows_html = ""
                    dots_html = ""
                    counter_html = ""
                    scroll_js = ""

                    if total_photos > 1:
                        arrows_html = '''<button class="arrow left" id="arrowL" aria-label="Previous">&#10094;</button>
<button class="arrow right" id="arrowR" aria-label="Next">&#10095;</button>'''

                        dot_items = ""
                        for i in range(total_photos):
                            active = " active" if i == 0 else ""
                            dot_items += f'<button class="dot{active}" data-idx="{i}"></button>'
                        dots_html = f'<div class="dots">{dot_items}</div>'

                        counter_html = f'<div class="counter" id="counter">1 / {total_photos}</div>'

                        scroll_js = f'''<script>
(function(){{
  var track = document.getElementById('track');
  var arrowL = document.getElementById('arrowL');
  var arrowR = document.getElementById('arrowR');
  var counter = document.getElementById('counter');
  var dots = document.querySelectorAll('.dot');
  var total = {total_photos};

  function updateUI() {{
    var idx = Math.round(track.scrollLeft / track.offsetWidth);
    if (counter) counter.textContent = (idx+1) + ' / ' + total;
    dots.forEach(function(d,i) {{
      if (i===idx) d.classList.add('active'); else d.classList.remove('active');
    }});
  }}

  if(arrowL) arrowL.addEventListener('click', function(){{
    track.scrollBy({{left: -track.offsetWidth, behavior:'smooth'}});
  }});
  if(arrowR) arrowR.addEventListener('click', function(){{
    track.scrollBy({{left: track.offsetWidth, behavior:'smooth'}});
  }});
  dots.forEach(function(d){{
    d.addEventListener('click', function(){{
      var i = parseInt(d.getAttribute('data-idx'));
      track.scrollTo({{left: i * track.offsetWidth, behavior:'smooth'}});
    }});
  }});
  track.addEventListener('scroll', updateUI);
}})();
</script>'''

                    carousel_html = f'''<!DOCTYPE html>
<html><head><style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:transparent;overflow:hidden;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;}}
.card{{background:#fff;border:1px solid #E2E8F0;border-radius:16px;padding:20px 24px;height:410px;display:flex;flex-direction:column;box-shadow:inset 0 0 0 1px rgba(0,0,0,0.05);}}
.title{{font-size:13px;font-weight:700;color:#718096;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.5px;}}
.wrapper{{position:relative;border-radius:12px;overflow:hidden;background:#F8FAFC;flex:1;min-height:0;}}
.track{{display:flex;overflow-x:auto;scroll-snap-type:x mandatory;scroll-behavior:smooth;-webkit-overflow-scrolling:touch;width:100%;height:100%;scrollbar-width:none;-ms-overflow-style:none;}}
.track::-webkit-scrollbar{{display:none;}}
.slide{{flex:0 0 100%;scroll-snap-align:start;width:100%;height:100%;}}
.slide img{{width:100%;height:100%;object-fit:cover;display:block;}}
.arrow{{position:absolute;top:50%;transform:translateY(-50%);width:34px;height:34px;border-radius:50%;border:none;background:rgba(0,0,0,0.45);color:#fff;font-size:16px;font-weight:700;cursor:pointer;display:flex;align-items:center;justify-content:center;z-index:10;backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px);transition:background 0.2s;}}
.arrow:hover{{background:rgba(0,0,0,0.7);}}
.arrow.left{{left:8px;}}
.arrow.right{{right:8px;}}
.counter{{position:absolute;bottom:10px;left:50%;transform:translateX(-50%);background:rgba(0,0,0,0.5);color:#fff;font-size:11px;font-weight:700;padding:3px 12px;border-radius:20px;letter-spacing:0.5px;z-index:10;backdrop-filter:blur(4px);}}
.dots{{position:absolute;bottom:34px;left:50%;transform:translateX(-50%);display:flex;gap:6px;z-index:10;}}
.dot{{width:7px;height:7px;border-radius:50%;background:rgba(255,255,255,0.45);transition:all 0.3s;cursor:pointer;border:none;padding:0;}}
.dot.active{{background:#fff;transform:scale(1.25);box-shadow:0 0 4px rgba(255,255,255,0.5);}}
</style></head><body>
<div class="card">
<div class="title">Venue Photo</div>
<div class="wrapper">
<div class="track" id="track">
{slides_html}</div>
{arrows_html}
{dots_html}
{counter_html}
</div>
</div>
{scroll_js}
</body></html>'''

                    components.html(carousel_html, height=420, scrolling=False)
                else:
                    # No photos available – render placeholder
                    st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto; display:flex; flex-direction:column;">
    <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Photo</div>
<div style="border-radius:12px; flex:1; min-height:0; background:var(--surface-muted); border: 1px solid var(--border); display:flex; flex-direction:column; justify-content:center; align-items:center;">
    <div style="width:48px; height:48px; border-radius:50%; background:var(--border-dark); margin-bottom:12px; display:flex; justify-content:center; align-items:center;">
        <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="var(--surface)" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"></rect><circle cx="8.5" cy="8.5" r="1.5"></circle><polyline points="21 15 16 10 5 21"></polyline></svg>
    </div>
    <span style="color:var(--text-muted); font-size:13px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Photo Not Available</span>
</div>
</div>""", unsafe_allow_html=True)



            with info_col:
                address = _val("address", "")
                landmark = _val("landmark", "")
                pincode = _val("pincode", "")
                complete_address = _val("complete_address", "")

                loc_items = [
                    ("Region", region),
                    ("State", state),
                    ("City", city),
                    ("District", district),
                    ("Pincode", pincode),
                    ("Landmark", landmark),
                ]

                loc_html = ""
                for label, val in loc_items:
                    if val != "N/A":
                        loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border);">
    <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
    <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">{html.escape(val)}</span>
</div>"""

                if complete_address != "N/A":
                    loc_html += f"""<div style="margin-top:12px; padding:10px 14px; background:var(--surface-muted); border-radius:10px; border:1px solid var(--border);">
    <div style="color:var(--text-muted); font-size:12px; font-weight:600; margin-bottom:4px;">FULL ADDRESS</div>
    <div style="color:var(--text); font-size:13px; line-height:1.5;">{html.escape(complete_address)}</div>
</div>"""

                google_coords = _val("google_cordinates", "")
                if google_coords != "N/A" and google_coords != "":
                    loc_html += f"""<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px dashed var(--border); margin-top:4px;">
    <span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">Google Location</span>
    <span style="color:var(--text); font-size:13px; font-weight:600; text-align:right;">
        <a href="https://maps.google.com/?q={html.escape(google_coords)}" target="_blank" style="color:#2563EB; text-decoration:none;">{html.escape(google_coords)}</a>
    </span>
</div>"""

                header_text = "Location Profile"
                    
                # Render entire profile block in a single markdown wrapper at 0 indentation
                st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm); height:410px; overflow-y:auto;">
    <div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">{header_text}</div>
{loc_html}
</div>""", unsafe_allow_html=True)

            with metrics_col:
                def _mini_card(label, value, color="#2563EB"):
                    return f"""<div style="flex:1; background:var(--surface-muted); border:1px solid var(--border); border-radius:8px; padding:6px; text-align:center; border-left:3px solid {color}; box-sizing:border-box; display:flex; flex-direction:column; justify-content:center; align-items:center;">
<div style="font-size:17px; font-weight:800; color:{color}; line-height:1;">{html.escape(str(value))}</div>
<div style="font-size:12px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.4px; margin-top:2px;">{label}</div>
</div>"""

                cctv = _val("cctv")
                cctv_color = "#10B981" if cctv.upper() == "YES" else "#EF4444"
                pwd = _val("pwd_venue")
                pwd_color = "#10B981" if pwd.upper() == "YES" else "#EF4444"
                genset = _val("genset_availability")
                genset_color = "#10B981" if genset.upper() == "YES" else "#EF4444"
                ups = _val("ups_availability")
                ups_color = "#10B981" if ups.upper() == "YES" else "#EF4444"

                # --- Occupancy % for DOTC ---
                occ_html = ""
                if is_dotc:
                    try:
                        occ_df = pd.read_excel('Data/DOTC Occupancy.xlsx')
                        occ_row = occ_df[occ_df['DMS Code'].astype(str).str.strip().str.upper() == str(dms_code).strip().upper()]
                        if not occ_row.empty:
                            occ_val = occ_row.iloc[0].get('Seat Occupancy %', 0)
                            if pd.notna(occ_val):
                                formatted_occ = f"{float(occ_val) * 100:.1f}%"
                                occ_color = "#10B981" if float(occ_val) > 0.7 else "#F59E0B" if float(occ_val) > 0.4 else "#EF4444"
                                occ_html = _mini_card("Occupancy %", formatted_occ, occ_color)
                            else:
                                occ_html = _mini_card("Occupancy %", "N/A", "#64748B")
                        else:
                            occ_html = _mini_card("Occupancy %", "N/A", "#64748B")
                    except Exception as e:
                        occ_html = _mini_card("Occupancy %", "N/A", "#64748B")
                else:
                    used_tested = _val("used_and_tested_venue")
                    used_color = "#10B981" if used_tested.upper() == "YES" else "#EF4444"
                    occ_html = _mini_card("Used & Tested", used_tested.title(), used_color)

                cards_html = (
                    _mini_card("Max Capacity", _val("venue_max_capacity"), "#2563EB") +
                    _mini_card("CCTV", cctv, cctv_color) +
                    _mini_card("PWD Venue", pwd, pwd_color) +
                    _mini_card("Genset", genset, genset_color) +
                    _mini_card("UPS", ups, ups_color) +
                    occ_html
                )

                st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:16px 14px; box-shadow:var(--shadow-sm); height:410px; box-sizing:border-box; display:flex; flex-direction:column;">
<div style="font-size:12px; font-weight:700; color:var(--text-muted); margin-bottom:8px; text-transform:uppercase; letter-spacing:0.5px; flex-shrink:0;">Key Metrics</div>
<div style="display:flex; flex-direction:column; flex:1; gap:5px; margin-top:0;">
{cards_html}
</div>
</div>""", unsafe_allow_html=True)

            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

            # --- VENUE INFO & CONTACT INFO ---
            venue_info_col, contact_col = st.columns(2)

            with venue_info_col:
                v_info_items = [
                    ("Venue SPOC", _val("venue_spoc_name")),
                    ("SPOC Contact", _val("spoc_contact_number")),
                    ("SPOC Email", _val("spoc_email_id"))
                ]
                v_info_html = ""
                for label, val in v_info_items:
                    if val != "N/A":
                        v_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
<span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
<span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
</div>"""
                st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
<div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Venue Information</div>
{v_info_html}
</div>""", unsafe_allow_html=True)

            with contact_col:
                c_info_items = [
                    ("Channel Manager", _val("channel_manager_name")),
                    ("CM Contact", _val("channel_manager_contact_no.")),
                    ("CM Email", _val("channel_manager_email_id")),
                ]
                c_info_html = ""
                for label, val in c_info_items:
                    if val != "N/A":
                        c_info_html += f"""<div style="display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px dashed var(--border);">
<span style="color:var(--text-muted); font-size:13px; font-weight:500; text-transform:uppercase; letter-spacing:0.3px;">{label}</span>
<span style="color:var(--text); font-size:13px; font-weight:600;">{html.escape(val)}</span>
</div>"""
                st.markdown(f"""<div style="background:var(--surface); border:1px solid var(--border); border-radius:16px; padding:20px 24px; box-shadow:var(--shadow-sm);">
<div style="font-size:13px; font-weight:700; color:var(--text-muted); margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">DEXIT SPOC CONTACT INFORMATION</div>
{c_info_html}
</div>""", unsafe_allow_html=True)

            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

            # --- ANALYSIS & REMARKS ---
            if not suppress_analysis_sections:
                st.markdown("""
                <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px;
                            text-transform:uppercase; letter-spacing:0.5px;">Analysis & Remarks</div>
                """, unsafe_allow_html=True)

            cat_val = _val("category", "N/A")
            subcat_val = _val("subcategory", "N/A")

            fields = []
            if (not suppress_analysis_sections) and cat_val != "N/A" and cat_val != "" and cat_val != "0" and str(cat_val).lower() != "nan":
                fields.append(("Category", cat_val))
            if subcat_val != "N/A" and subcat_val != "" and subcat_val != "0" and str(subcat_val).lower() != "nan":
                fields.append(("Subcategory", subcat_val))
            if (not suppress_analysis_sections) and final_analysis != "N/A" and final_analysis != "" and final_analysis != "0" and final_analysis != "0.0":
                fields.append(("Venue Intel", final_analysis))

            if fields:
                inner_html = ""
                for lbl, field_val in fields:
                    inner_html += f"""
<div style="margin-right:48px; margin-bottom:8px;">
    <div style="font-size:11px; font-weight:700; color:#64748B; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:4px;">{lbl}</div>
    <div style="font-size:14px; color:#1A202C; line-height:1.6; font-weight:500;">{html.escape(str(field_val))}</div>
</div>"""

                st.markdown(f"""
<div style="background:#fff; border:1px solid #E2E8F0; border-left:4px solid {analysis_color};
            border-radius:12px; padding:16px 18px 8px 18px; margin:8px 0;
            box-shadow:0 1px 4px rgba(0,0,0,0.03); display:flex; flex-wrap:wrap;">
    {inner_html}
</div>""", unsafe_allow_html=True)
            
            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

            # --- FINAL ISSUE REMARKS ---
            final_issue = _val("call_log_summary", "")
            if (not suppress_analysis_sections) and final_issue and final_issue != "N/A":
                st.markdown("""
                <div style="font-size:13px; font-weight:700; color:#4A5568; margin-bottom:14px; text-transform:uppercase; letter-spacing:0.5px;">Major Call Log Issues Identified</div>
                """, unsafe_allow_html=True)

                st.markdown(f"""
                <div style="background:var(--surface); border:1px solid var(--border); border-left:4px solid #F59E0B; border-radius:12px; padding:14px 18px; margin:8px 0; box-shadow:var(--shadow-sm);">
                    <div style="font-size:14px; color:var(--text); line-height:1.6; white-space:pre-wrap;">{html.escape(final_issue)}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)
            # --- MANPOWER BUTTONS ---
            if st.button("Manpower Details", key=f"view_manpower_{dms_code}", use_container_width=True):
                st.session_state.manpower_search_code = dms_code
                st.session_state.page = "manpower_dashboard"
                st.rerun()

            # st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
            # if st.button("Project-wise Details", key=f"view_proj_manpower_{dms_code}", use_container_width=True):
                # st.session_state.manpower_search_code = dms_code
                # st.session_state.page = "project_wise_details"
                # st.rerun()



            st.markdown('<div class="gap-md"></div>', unsafe_allow_html=True)

elif st.session_state.page == "manpower_dashboard":
    map_head_col, map_back_col = st.columns([4, 1], vertical_alignment="center")
    with map_head_col:
        st.markdown(
            """
            <div>
                <h2>Manpower Dashboard</h2>
                <p>Deduplicated: Project → Drive → DMS Code → Max per metric → Sum</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with map_back_col:
        if st.button("← Back to Dashboard", use_container_width=True, key="manpower_back"):
            st.session_state.page = "dashboard"
            st.query_params["page"] = "dashboard"
            st.rerun()

    st.markdown('<div style="margin-bottom:12px;"></div>', unsafe_allow_html=True)

    f_col1, f_col2, f_col3 = st.columns([1.2, 1.3, 1.8])
    with f_col1:
        search_query = st.text_input("Search", value=st.session_state.manpower_search_code, placeholder="Type to search...", key="manpower_search_input").strip()
    with f_col2:
        emp_type_opts = ["DEXIT", "Outsourced"]
        selected_emp_types = st.segmented_control("Employee Type", options=emp_type_opts, default=emp_type_opts, selection_mode="multi", key="manpower_emp_type")
    with f_col3:
        range_mode = st.segmented_control("Date Filter", options=["All Data", "Custom Range"], default="All Data", key="manpower_date_mode")
        if range_mode == "Custom Range":
            rd_c1, rd_c2 = st.columns(2)
            with rd_c1:
                from_date = st.date_input("From Date", value=datetime.now().replace(day=1), key="manpower_date_from")
            with rd_c2:
                to_date = st.date_input("To Date", value=datetime.now(), key="manpower_date_to")

    st.session_state.manpower_search_code = search_query

    filepath = os.path.join("Data", "0 Final_Manpower_Data.xlsx")
    if os.path.exists(filepath):
        try:
            with st.spinner("Processing manpower data..."):
                df_manpower = pd.read_excel(filepath, sheet_name="Data")
                if 'Date' in df_manpower.columns:
                    df_manpower['Date'] = pd.to_datetime(df_manpower['Date'])

            st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)
            st.markdown("##### Advanced Filters")

            st.markdown(
                """
                <style>
                div[data-baseweb="select"], div[data-baseweb="base-input"] {
                    border: none !important;
                    background-color: transparent !important;
                }
                div[data-testid="stMultiSelect"] > div:nth-child(2),
                div[data-testid="stSelectbox"] > div:nth-child(2) {
                    border: 1px solid #9CA3AF !important;
                    border-radius: 8px !important;
                    min-height: 42px !important;
                    height: 42px !important;
                    display: flex !important;
                    align-items: center !important;
                    background-color: white !important;
                    box-sizing: border-box !important;
                    overflow: hidden !important;
                }
                div[data-testid="stMultiSelect"] > div:nth-child(2) > div:first-child {
                    flex: 1 !important;
                    overflow: hidden !important;
                    white-space: nowrap !important;
                    text-overflow: ellipsis !important;
                    padding: 0 10px !important;
                }
                div[data-baseweb="tag"] {
                    background-color: #EFF6FF !important;
                    background: #EFF6FF !important;
                    border: 1px solid #BFDBFE !important;
                    border-radius: 6px !important;
                    box-shadow: none !important;
                    height: 26px !important;
                    max-height: 26px !important;
                    margin: 2px 3px !important;
                    padding: 0 6px !important;
                    display: inline-flex !important;
                    align-items: center !important;
                    gap: 4px !important;
                }
                div[data-baseweb="tag"] > div,
                div[data-baseweb="tag"] > span > div {
                    background-color: transparent !important;
                    background: transparent !important;
                    border: none !important;
                    box-shadow: none !important;
                }
                div[data-testid="stMultiSelect"] input {
                    background: transparent !important;
                    border: none !important;
                    min-width: 2px !important;
                    width: 2px !important;
                    padding: 0 !important;
                    margin: 0 !important;
                    flex: 0 0 2px !important;
                }
                div[data-baseweb="tag"] span {
                    font-size: 12px !important;
                    line-height: 1 !important;
                    white-space: nowrap !important;
                    color: #1D4ED8 !important;
                    font-weight: 500 !important;
                    background: transparent !important;
                }
                div[data-baseweb="tag"] [data-baseweb="icon"],
                div[data-baseweb="tag"] svg {
                    color: #6B7280 !important;
                    background: transparent !important;
                    border: none !important;
                    width: 14px !important;
                    height: 14px !important;
                    flex-shrink: 0 !important;
                }
                div[role="button"], div[data-baseweb="select"] > div {
                    background: none !important;
                    border: none !important;
                }
                div[data-baseweb="select"] input,
                div[data-baseweb="select"] [data-baseweb="input"] {
                    padding-left: 10px !important;
                    text-align: left !important;
                    font-size: 14px !important;
                }
                div[data-testid="stSelectbox"] span {
                    font-size: 14px !important;
                    color: #374151 !important;
                }
                div[data-baseweb="select"] svg {
                    color: #6B7280 !important;
                    flex-shrink: 0 !important;
                }
                [data-testid="column"] {
                    width: 100% !important;
                }
                div[data-testid="stMultiSelect"] > div:nth-child(2):hover,
                div[data-testid="stSelectbox"] > div:nth-child(2):hover {
                    border-color: #6B7280 !important;
                }

                /* ── Clickable Card Styles ── */
                .mp-card-wrapper {
                    position: relative;
                    margin-bottom: 8px;
                }
                .mp-clickable-card {
                    background: white;
                    border: 1px solid #E5E7EB;
                    border-radius: 10px;
                    padding: 14px 10px 12px;
                    text-align: center;
                    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
                    transition: box-shadow 0.18s ease, transform 0.18s ease;
                    cursor: pointer;
                    position: relative;
                    z-index: 1;
                    pointer-events: none;
                }
                .mp-clickable-card:hover {
                    box-shadow: 0 6px 20px rgba(0,0,0,0.13);
                    transform: translateY(-2px);
                }
                .mp-card-label {
                    font-size: 10px;
                    color: #6B7280;
                    font-weight: 600;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                    margin-bottom: 6px;
                    line-height: 1.3;
                    min-height: 26px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }
                .mp-card-value {
                    font-size: 22px;
                    font-weight: 800;
                    line-height: 1;
                }
                /* Invisible full-coverage button overlay */
                .mp-card-wrapper [data-testid="stButton"] {
                    position: absolute !important;
                    top: 0 !important;
                    left: 0 !important;
                    width: 100% !important;
                    height: 100% !important;
                    z-index: 10 !important;
                }
                .mp-card-wrapper [data-testid="stButton"] button {
                    width: 100% !important;
                    height: 100% !important;
                    opacity: 0 !important;
                    cursor: pointer !important;
                    background: transparent !important;
                    border: none !important;
                    box-shadow: none !important;
                    padding: 0 !important;
                    margin: 0 !important;
                    position: absolute !important;
                    top: 0 !important;
                    left: 0 !important;
                }
                </style>
                """,
                unsafe_allow_html=True
            )

            # Auto-close multiselect dropdown after each selection
            st.markdown(
                """
                <script>
                (function() {
                    function closeOpenDropdowns() {
                        var active = document.activeElement;
                        if (active) {
                            var escEvent = new KeyboardEvent('keydown', {
                                key: 'Escape', code: 'Escape', keyCode: 27,
                                which: 27, bubbles: true, cancelable: true
                            });
                            active.dispatchEvent(escEvent);
                            active.blur();
                        }
                        var listboxes = document.querySelectorAll('[data-baseweb="popover"] ul[role="listbox"]');
                        listboxes.forEach(function(lb) {
                            var pop = lb.closest('[data-baseweb="popover"]');
                            if (pop) pop.style.display = 'none';
                            setTimeout(function() {
                                if (pop) pop.style.display = '';
                            }, 300);
                        });
                    }
                    function attachOptionListeners() {
                        document.removeEventListener('mousedown', handleOptionClick, true);
                        document.addEventListener('mousedown', handleOptionClick, true);
                    }
                    function handleOptionClick(e) {
                        var option = e.target.closest('[role="option"]');
                        if (!option) return;
                        var listbox = option.closest('[role="listbox"]');
                        if (!listbox) return;
                        setTimeout(closeOpenDropdowns, 80);
                    }
                    attachOptionListeners();
                    var bodyObserver = new MutationObserver(function() {
                        attachOptionListeners();
                    });
                    bodyObserver.observe(document.body, { childList: true, subtree: false });
                })();
                </script>
                """,
                unsafe_allow_html=True
            )

            # Auto-detect columns
            dms_target = next(
                (c for c in df_manpower.columns if c.upper().replace(" ", "_") in ['DMS_CODE', 'DMS CODE', 'CENTER_CODE']),
                df_manpower.columns[0]
            )
            emp_source_col = (
                'Employee Source' if 'Employee Source' in df_manpower.columns
                else 'Updated Employee Source' if 'Updated Employee Source' in df_manpower.columns
                else None
            )
            raw_map = {"DEXIT": "DexIT Global Employee", "Outsourced": "Outsource"}
            target_raw_values = [raw_map[t] for t in (selected_emp_types if selected_emp_types else [])]

            # --- Build base mask ---
            base_mask = pd.Series(True, index=df_manpower.index)
            if search_query:
                row_matches = pd.Series(False, index=df_manpower.index)
                for c in df_manpower.columns:
                    row_matches |= df_manpower[c].astype(str).str.contains(search_query, case=False, na=False)
                base_mask &= row_matches
            if selected_emp_types and emp_source_col:
                base_mask &= df_manpower[emp_source_col].astype(str).str.strip().isin(target_raw_values)
            if range_mode == "Custom Range":
                base_mask &= (df_manpower['Date'].dt.date >= from_date) & (df_manpower['Date'].dt.date <= to_date)

            df_base = df_manpower[base_mask].copy()

            # --- Detect location columns ---
            def _find_col(df, names):
                for c in df.columns:
                    if c.lower().strip().replace(' ', '_') in names:
                        return c
                return None

            region_col   = _find_col(df_manpower, ['region'])
            state_col    = _find_col(df_manpower, ['state', 'state_name'])
            district_col = _find_col(df_manpower, ['district', 'district_name'])
            city_col     = _find_col(df_manpower, ['city', 'city_name'])

            # --- Helper: apply advanced filters ---
            def apply_adv_filters(df_in, projects=None, emps=None, delay=None, issue=None,
                                  tenure=None, roles=None,
                                  regions=None, states=None, districts=None, cities=None):
                m = pd.Series(True, index=df_in.index)
                if projects:
                    m &= df_in['Project Name'].isin(projects)
                if emps:
                    m &= (df_in['Emp_ID'].astype(str) + " - " + df_in['Name']).isin(emps)
                if delay and delay != "All":
                    if isinstance(delay, list):
                        sub = pd.Series(False, index=df_in.index)
                        for d in delay:
                            if d in df_in.columns:
                                sub |= (df_in[d] > 0)
                        m &= sub
                    elif delay in df_in.columns:
                        m &= (df_in[delay] > 0)
                if issue and issue != "All":
                    issues_list = issue if isinstance(issue, list) else [issue]
                    isub = pd.Series(False, index=df_in.index)
                    if "FFA = Yes" in issues_list and 'FFA' in df_in.columns:
                        isub |= (df_in['FFA'] > 0)
                    if "Call Logs = Yes" in issues_list and 'Call Logs Issue Count' in df_in.columns:
                        isub |= (df_in['Call Logs Issue Count'] > 0)
                    m &= isub
                if tenure and 'Tenure' in df_in.columns:
                    m &= df_in['Tenure'].isin(tenure)
                if roles and 'Role' in df_in.columns:
                    m &= df_in['Role'].isin(roles)
                if regions and region_col and region_col in df_in.columns:
                    m &= df_in[region_col].isin(regions)
                if states and state_col and state_col in df_in.columns:
                    m &= df_in[state_col].isin(states)
                if districts and district_col and district_col in df_in.columns:
                    m &= df_in[district_col].isin(districts)
                if cities and city_col and city_col in df_in.columns:
                    m &= df_in[city_col].isin(cities)
                return df_in[m]

            # --- Session state init ---
            for _k, _def in [
                ("mp_sel_projects", []),
                ("mp_sel_emps", []),
                ("mp_sel_delay", []),
                ("mp_sel_issue", []),
                ("mp_sel_tenure", []),
                ("mp_sel_roles", []),
                ("mp_sel_region", []),
                ("mp_sel_state", []),
                ("mp_sel_district", []),
                ("mp_sel_city", []),
                ("mp_card_clicked", None),
            ]:
                if _k not in st.session_state:
                    st.session_state[_k] = _def

            # Reset button
            _reset_col, _ = st.columns([1, 5])
            with _reset_col:
                if st.button("Reset Filters", key="manpower_adv_reset", use_container_width=True):
                    for _k in ["mp_sel_projects","mp_sel_emps","mp_sel_delay","mp_sel_issue",
                               "mp_sel_tenure","mp_sel_roles","mp_sel_region","mp_sel_state",
                               "mp_sel_district","mp_sel_city"]:
                        st.session_state[_k] = []
                    st.session_state.mp_card_clicked = None
                    st.rerun()

            # Read current selections
            _cur_proj     = st.session_state.mp_sel_projects
            _cur_emps     = st.session_state.mp_sel_emps
            _cur_delay    = st.session_state.mp_sel_delay
            _cur_issue    = st.session_state.mp_sel_issue
            _cur_tenure   = st.session_state.mp_sel_tenure
            _cur_roles    = st.session_state.mp_sel_roles
            _cur_region   = st.session_state.mp_sel_region
            _cur_state    = st.session_state.mp_sel_state
            _cur_district = st.session_state.mp_sel_district
            _cur_city     = st.session_state.mp_sel_city

            # --- Compute available options (interdependent) ---
            _loc_kw = dict(regions=_cur_region, states=_cur_state, districts=_cur_district, cities=_cur_city)

            _df_for_proj  = apply_adv_filters(df_base, emps=_cur_emps, delay=_cur_delay, issue=_cur_issue, tenure=_cur_tenure, roles=_cur_roles, **_loc_kw)
            proj_list     = sorted(_df_for_proj['Project Name'].dropna().unique().tolist())
            _cur_proj     = [v for v in _cur_proj if v in proj_list]

            _df_for_emp   = apply_adv_filters(df_base, projects=_cur_proj, delay=_cur_delay, issue=_cur_issue, tenure=_cur_tenure, roles=_cur_roles, **_loc_kw)
            emp_list      = sorted((_df_for_emp['Emp_ID'].astype(str) + " - " + _df_for_emp['Name']).unique().tolist())
            _cur_emps     = [v for v in _cur_emps if v in emp_list]

            _df_for_delay   = apply_adv_filters(df_base, projects=_cur_proj, emps=_cur_emps, issue=_cur_issue, tenure=_cur_tenure, roles=_cur_roles, **_loc_kw)
            _all_delay_opts = ["Full Batch Delay", "Partially Batch Delay", "No Delay"]
            delay_opts      = [d for d in _all_delay_opts if d in _df_for_delay.columns and (_df_for_delay[d] > 0).any()]
            _cur_delay      = [v for v in _cur_delay if v in delay_opts]

            _df_for_issue = apply_adv_filters(df_base, projects=_cur_proj, emps=_cur_emps, delay=_cur_delay, tenure=_cur_tenure, roles=_cur_roles, **_loc_kw)
            issue_opts    = []
            if 'FFA' in _df_for_issue.columns and (_df_for_issue['FFA'] > 0).any():
                issue_opts.append("FFA = Yes")
            if 'Call Logs Issue Count' in _df_for_issue.columns and (_df_for_issue['Call Logs Issue Count'] > 0).any():
                issue_opts.append("Call Logs = Yes")
            _cur_issue = [v for v in _cur_issue if v in issue_opts]

            _df_for_tenure = apply_adv_filters(df_base, projects=_cur_proj, emps=_cur_emps, delay=_cur_delay, issue=_cur_issue, roles=_cur_roles, **_loc_kw)
            tenure_opts    = sorted(_df_for_tenure['Tenure'].dropna().unique().tolist()) if 'Tenure' in _df_for_tenure.columns else []
            _cur_tenure    = [v for v in _cur_tenure if v in tenure_opts]

            _df_for_role = apply_adv_filters(df_base, projects=_cur_proj, emps=_cur_emps, delay=_cur_delay, issue=_cur_issue, tenure=_cur_tenure, **_loc_kw)
            role_opts    = sorted(_df_for_role['Role'].dropna().unique().tolist()) if 'Role' in _df_for_role.columns else []
            _cur_roles   = [v for v in _cur_roles if v in role_opts]

            # --- Location filter options ---
            _base_kw = dict(projects=_cur_proj, emps=_cur_emps, delay=_cur_delay, issue=_cur_issue, tenure=_cur_tenure, roles=_cur_roles)

            def _loc_opts(col, exclude_key):
                if not col or col not in df_base.columns:
                    return []
                kw = {**_base_kw}
                if exclude_key == 'region':   kw.update(states=_cur_state, districts=_cur_district, cities=_cur_city)
                elif exclude_key == 'state':  kw.update(regions=_cur_region, districts=_cur_district, cities=_cur_city)
                elif exclude_key == 'district': kw.update(regions=_cur_region, states=_cur_state, cities=_cur_city)
                elif exclude_key == 'city':   kw.update(regions=_cur_region, states=_cur_state, districts=_cur_district)
                return sorted(apply_adv_filters(df_base, **kw)[col].dropna().unique().tolist())

            region_opts   = _loc_opts(region_col,   'region')
            state_opts    = _loc_opts(state_col,    'state')
            district_opts = _loc_opts(district_col, 'district')
            city_opts     = _loc_opts(city_col,     'city')

            _cur_region   = [v for v in _cur_region   if v in region_opts]
            _cur_state    = [v for v in _cur_state    if v in state_opts]
            _cur_district = [v for v in _cur_district if v in district_opts]
            _cur_city     = [v for v in _cur_city     if v in city_opts]

            # --- Render filter row 1 ---
            d_col1, d_col2, d_col3, d_col4, d_col5, d_col6 = st.columns([1, 1, 1, 1, 1, 1])
            with d_col1:
                sel_projects = st.multiselect("Project Name", options=proj_list, default=_cur_proj, key="mp_sel_projects", placeholder="All Projects")
            with d_col2:
                sel_emps = st.multiselect("Employee (ID - Name)", options=emp_list, default=_cur_emps, key="mp_sel_emps", placeholder="All Employees")
            with d_col3:
                sel_delay = st.multiselect("Batch Delay Status", options=delay_opts, default=_cur_delay, key="mp_sel_delay", placeholder="All")
                delay_filter = sel_delay[0] if len(sel_delay) == 1 else ("All" if not sel_delay else sel_delay)
            with d_col4:
                sel_issue = st.multiselect("Issue Filter", options=issue_opts, default=_cur_issue, key="mp_sel_issue", placeholder="All")
                issue_filter = sel_issue[0] if len(sel_issue) == 1 else ("All" if not sel_issue else sel_issue)
            with d_col5:
                sel_tenure = st.multiselect("Tenure", options=tenure_opts, default=_cur_tenure, key="mp_sel_tenure", placeholder="All Tenures")
            with d_col6:
                sel_roles = st.multiselect("Role", options=role_opts, default=_cur_roles, key="mp_sel_roles", placeholder="All Roles")

            # --- Render location filter row 2 ---
            _loc_filter_defs = [
                (region_col,   region_opts,   _cur_region,   "mp_sel_region",   "Region"),
                (state_col,    state_opts,    _cur_state,    "mp_sel_state",    "State"),
                (district_col, district_opts, _cur_district, "mp_sel_district", "District"),
                (city_col,     city_opts,     _cur_city,     "mp_sel_city",     "City"),
            ]
            _visible_loc = [(col, opts, cur, key, lbl) for col, opts, cur, key, lbl in _loc_filter_defs if col]
            if _visible_loc:
                l_cols = st.columns(len(_visible_loc))
                for lc, (col, opts, cur, key, lbl) in zip(l_cols, _visible_loc):
                    with lc:
                        st.multiselect(lbl, options=opts, default=cur, key=key, placeholder=f"All {lbl}s")

            sel_region   = st.session_state.mp_sel_region
            sel_state    = st.session_state.mp_sel_state
            sel_district = st.session_state.mp_sel_district
            sel_city     = st.session_state.mp_sel_city

            # --- Build final mask ---
            mask = base_mask.copy()
            if sel_projects:
                mask &= df_manpower['Project Name'].isin(sel_projects)
            if sel_emps:
                mask &= (df_manpower['Emp_ID'].astype(str) + " - " + df_manpower['Name']).isin(sel_emps)
            if sel_tenure and 'Tenure' in df_manpower.columns:
                mask &= df_manpower['Tenure'].isin(sel_tenure)
            if delay_filter != "All":
                if isinstance(delay_filter, list):
                    _dm = pd.Series(False, index=df_manpower.index)
                    for d in delay_filter:
                        if d in df_manpower.columns:
                            _dm |= (df_manpower[d] > 0)
                    mask &= _dm
                elif delay_filter in df_manpower.columns:
                    mask &= (df_manpower[delay_filter] > 0)
            if issue_filter != "All":
                issues_list = issue_filter if isinstance(issue_filter, list) else [issue_filter]
                _im = pd.Series(False, index=df_manpower.index)
                if "FFA = Yes" in issues_list and 'FFA' in df_manpower.columns:
                    _im |= (df_manpower['FFA'] > 0)
                if "Call Logs = Yes" in issues_list and 'Call Logs Issue Count' in df_manpower.columns:
                    _im |= (df_manpower['Call Logs Issue Count'] > 0)
                mask &= _im
            if sel_roles and 'Role' in df_manpower.columns:
                mask &= df_manpower['Role'].isin(sel_roles)
            if sel_region and region_col and region_col in df_manpower.columns:
                mask &= df_manpower[region_col].isin(sel_region)
            if sel_state and state_col and state_col in df_manpower.columns:
                mask &= df_manpower[state_col].isin(sel_state)
            if sel_district and district_col and district_col in df_manpower.columns:
                mask &= df_manpower[district_col].isin(sel_district)
            if sel_city and city_col and city_col in df_manpower.columns:
                mask &= df_manpower[city_col].isin(sel_city)

            df_filtered = df_manpower[mask].copy()

            if df_filtered.empty:
                st.warning("No data found for the selected filters.")
            else:
                METRIC_COLS = ['Call Logs Issue Count', 'FFA', 'Full Batch Delay', 'Partially Batch Delay', 'No Delay']
                for mc in METRIC_COLS:
                    if mc in df_filtered.columns:
                        df_filtered[mc] = pd.to_numeric(df_filtered[mc], errors='coerce').fillna(0)

                has_drive = 'Drive Name' in df_filtered.columns

                def build_deduped(df_in):
                    gkeys = ['Project Name']
                    if has_drive:
                        gkeys.append('Drive Name')
                    gkeys.append(dms_target)
                    agg = {m: 'max' for m in METRIC_COLS if m in df_in.columns}
                    return df_in.groupby(gkeys, dropna=False).agg(agg).reset_index()

                deduped_all = build_deduped(df_filtered)

                ov = {
                    'Project Count':   int(deduped_all['Project Name'].nunique()),
                    'Drive Count':     int(deduped_all['Drive Name'].nunique()) if has_drive else 0,
                    'Unique Manpower': int(df_filtered['Emp_ID'].nunique()) if 'Emp_ID' in df_filtered.columns else 0,
                }
                for m in METRIC_COLS:
                    ov[m] = int(deduped_all[m].sum()) if m in deduped_all.columns else 0

                st.markdown("### Overall Metrics")
                st.caption("Deduplicated: Project → Drive → DMS Code → Max per metric → Sum")

                total_batches = ov['Full Batch Delay'] + ov['Partially Batch Delay'] + ov['No Delay']

                _card_meta = [
                    ("Manpower",                    ov['Unique Manpower'],          "#64748B", None,                   "Emp_ID"),
                    ("Projects",                    ov['Project Count'],             "#3B82F6", None,                   "Project Name"),
                    ("Drives",                      ov['Drive Count'],               "#0EA5E9", None,                   "Drive Name"),
                    ("Call Logs",                   ov['Call Logs Issue Count'],     "#F59E0B", "Call Logs Issue Count", "Call Logs Issue Count"),
                    ("Fastest Finger Analysis (FFA)", ov['FFA'],                    "#10B981", "FFA",                   "FFA"),
                    ("Full Batch Delay",             ov['Full Batch Delay'],          "#EF4444", "Full Batch Delay",      "Full Batch Delay"),
                    ("Partial Batch Delay",          ov['Partially Batch Delay'],     "#8B5CF6", "Partially Batch Delay", "Partially Batch Delay"),
                    ("No Batch Delay",               ov['No Delay'],                  "#6366F1", "No Delay",              "No Delay"),
                    ("Total Batches",                total_batches,                   "#0F172A", None,                   "Project Name"),
                ]

                # ── Clickable Cards (no separate View button) ──────────────────
                k_cols = st.columns(9)
                for col, (lbl, val, clr, _fc, _btn_col) in zip(k_cols, _card_meta):
                    with col:
                        # Visible card HTML — pointer-events:none so clicks pass through
                        st.markdown(f"""
                            <div class="mp-card-wrapper">
                                <div class="mp-clickable-card" style="border-top: 4px solid {clr};">
                                    <div class="mp-card-label">{lbl}</div>
                                    <div class="mp-card-value" style="color:{clr};">{val:,}</div>
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                        # Invisible full-coverage button overlay
                        if st.button("‎", key=f"mp_card_btn_{lbl}", use_container_width=True, help=f"View {lbl} details"):
                            st.session_state.mp_card_clicked = (lbl, _fc, _btn_col)
                            st.rerun()

                # --- Card drill-down ---
                _clicked = st.session_state.mp_card_clicked
                if _clicked:
                    _card_lbl, _filter_col, _first_col = _clicked
                    _back_c, _title_c = st.columns([1, 5])
                    with _back_c:
                        if st.button("← Back to Summary", key="mp_card_back", use_container_width=True):
                            st.session_state.mp_card_clicked = None
                            st.rerun()
                    with _title_c:
                        st.markdown(f"#### 📋 Detail rows — **{_card_lbl}**")

                    _df_drill = df_filtered.copy()
                    if _filter_col and _filter_col in _df_drill.columns:
                        _df_drill = _df_drill[_df_drill[_filter_col] > 0]

                    _all_cols = list(_df_drill.columns)
                    if _first_col and _first_col in _all_cols:
                        _ordered = [_first_col] + [c for c in _all_cols if c != _first_col]
                    else:
                        _ordered = _all_cols

                    st.dataframe(_df_drill[_ordered], use_container_width=True, hide_index=True)
                    st.stop()

                # ── Summary tables ─────────────────────────────────────────────
                st.caption("Project × Drive × DMS Code | Max of each metric per row")
                pivot_display = deduped_all.copy()
                if 'Emp_ID' in df_filtered.columns:
                    gkeys_emp   = ['Project Name'] + (['Drive Name'] if has_drive else []) + [dms_target]
                    emp_grp     = df_filtered.groupby(gkeys_emp, dropna=False)['Emp_ID'].nunique().reset_index(name='Unique Manpower')
                    pivot_display = pivot_display.merge(emp_grp, on=gkeys_emp, how='left')

                grand = {'Project Name': 'GRAND TOTAL', dms_target: f"{deduped_all[dms_target].nunique()} centres"}
                if has_drive:
                    grand['Drive Name'] = f"{deduped_all['Drive Name'].nunique()} drives"
                for m in METRIC_COLS:
                    if m in deduped_all.columns:
                        grand[m] = int(deduped_all[m].sum())
                if 'Unique Manpower' in pivot_display.columns:
                    grand['Unique Manpower'] = int(df_filtered['Emp_ID'].nunique()) if 'Emp_ID' in df_filtered.columns else 0

                pivot_display = pd.concat([pivot_display, pd.DataFrame([grand])], ignore_index=True)
                st.dataframe(pivot_display, use_container_width=True, hide_index=True)

                st.markdown("### Project-wise Summary")

                def project_summary(df_proj):
                    gkeys = (['Drive Name', dms_target] if has_drive else [dms_target])
                    agg_c = {m: 'max' for m in METRIC_COLS if m in df_proj.columns}
                    d = df_proj.groupby(gkeys, dropna=False).agg(agg_c).reset_index()
                    res = {
                        'Drive Count':     int(df_proj['Drive Name'].nunique()) if has_drive else 0,
                        'Unique Manpower': int(df_proj['Emp_ID'].nunique()) if 'Emp_ID' in df_proj.columns else 0,
                    }
                    for m in METRIC_COLS:
                        res[m] = int(d[m].sum()) if m in d.columns else 0
                    return pd.Series(res)

                df_proj_agg = df_filtered.groupby('Project Name', dropna=False).apply(project_summary, include_groups=False).reset_index()
                proj_cols   = ['Project Name', 'Drive Count', 'Unique Manpower'] + [m for m in METRIC_COLS if m in df_proj_agg.columns]
                st.dataframe(df_proj_agg[proj_cols], use_container_width=True, hide_index=True)

                st.markdown("### Aggregated Manpower Table")
                st.caption("Per employee summary across all projects")

                def agg_emp(grp):
                    return pd.Series({
                        'Project Count': grp['Project Name'].nunique(),
                        **{m: int(grp[m].sum()) if m in grp.columns else 0 for m in METRIC_COLS}
                    })

                grp_emp_cols = ['Emp_ID', 'Name'] + (['Tenure'] if 'Tenure' in df_filtered.columns else [])
                df_emp_agg   = df_filtered.groupby(grp_emp_cols, dropna=False).apply(agg_emp, include_groups=False).reset_index()
                st.dataframe(df_emp_agg, use_container_width=True, hide_index=True)

                st.markdown("### Aggregated Manpower Table (with Roles)")
                st.caption("Same as above, further split by Role — each employee appears once per role")

                def agg_emp_roles(grp):
                    return pd.Series({
                        'Project Count': grp['Project Name'].nunique(),
                        **{m: int(grp[m].sum()) if m in grp.columns else 0 for m in METRIC_COLS}
                    })

                grp_role_cols = (
                    ['Emp_ID', 'Name']
                    + (['Tenure'] if 'Tenure' in df_filtered.columns else [])
                    + (['Role']   if 'Role'   in df_filtered.columns else [])
                )
                df_emp_role_agg = (
                    df_filtered
                    .groupby(grp_role_cols, dropna=False)
                    .apply(agg_emp_roles, include_groups=False)
                    .reset_index()
                )
                st.dataframe(df_emp_role_agg, use_container_width=True, hide_index=True)

                st.markdown("### DMS Code-wise Summary")

                def dms_summary(df_dms):
                    gkeys = (['Project Name', 'Drive Name'] if has_drive else ['Project Name'])
                    agg_c = {m: 'max' for m in METRIC_COLS if m in df_dms.columns}
                    d = df_dms.groupby(gkeys, dropna=False).agg(agg_c).reset_index()
                    return pd.Series({
                        'Drive Count':     int(df_dms['Drive Name'].nunique()) if has_drive else 0,
                        'Unique Manpower': int(df_dms['Emp_ID'].nunique()) if 'Emp_ID' in df_dms.columns else 0,
                        **{m: int(d[m].sum()) if m in d.columns else 0 for m in METRIC_COLS}
                    })

                df_dms_agg = df_filtered.groupby(dms_target, dropna=False).apply(dms_summary, include_groups=False).reset_index()
                dms_cols   = [dms_target, 'Drive Count', 'Unique Manpower'] + [m for m in METRIC_COLS if m in df_dms_agg.columns]
                st.dataframe(df_dms_agg[dms_cols], use_container_width=True, hide_index=True)

        except Exception as e:
            st.error(f"Error: {e}")
            st.exception(e)
    else:
        st.error("File not found: Data/0 Final_Manpower_Data.xlsx")
    

# # ================= PROJECT-WISE DASHBOARD PAGE =================
# elif st.session_state.page == "project_wise_details":
    # map_head_col, map_back_col = st.columns([4, 1], vertical_alignment="center")
    # with map_head_col:
        # st.markdown(
            # """
            # <div style="margin-bottom:8px;">
                # <h2 style="margin:0; color:#1E293B; font-weight:800;">Project-wise Details</h2>
                # <p style="margin:4px 0 0 0; color:#64748B; font-size:13px;">View aggregated metrics grouped by project for a specific DMS Code</p>
            # </div>
            # """,
            # unsafe_allow_html=True,
        # )
    # with map_back_col:
        # if st.button("\u2190 Back to Dashboard", use_container_width=True, key="proj_manpower_back"):
            # st.session_state.page = "dashboard"
            # st.query_params["page"] = "dashboard"
            # st.rerun()

    # st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
    
    # # Filter inputs
    # f_col1, f_col2, f_col3 = st.columns([1.2, 1.3, 1.8])
    # with f_col1:
        # search_query = st.text_input("Search", value=st.session_state.manpower_search_code, placeholder="Type to search...", key="proj_dms").strip()
    # with f_col2:
        # emp_type_opts = ["DEXIT", "Outsourced"]
        # selected_emp_types = st.segmented_control("Employee Type", options=emp_type_opts, default=emp_type_opts, selection_mode="multi", key="proj_emptype")
    # with f_col3:
        # range_mode = st.segmented_control("Date Filter", options=["All Data", "Custom Range"], default="All Data", key="proj_range_mode")
        # if range_mode == "Custom Range":
            # rd_c1, rd_c2 = st.columns(2)
            # with rd_c1:
                # from_date = st.date_input("From Date", value=datetime.now().replace(day=1), key="proj_from")
            # with rd_c2:
                # to_date = st.date_input("To Date", value=datetime.now(), key="proj_to")
        # else:
            # st.info("Showing all dates")
    
    # st.session_state.manpower_search_code = search_query

    # filepath = os.path.join("Data", "0 Final_Manpower_Data.xlsx")
    # if os.path.exists(filepath):
        # try:
            # with st.spinner("Loading project data..."):
                # df_manpower = pd.read_excel(filepath, sheet_name="Data")
                # if 'Date' in df_manpower.columns:
                    # df_manpower['Date'] = pd.to_datetime(df_manpower['Date'])
            
            # raw_map = {"DEXIT": "DexIT Global Employee", "Outsourced": "Outsource"}
            # target_raw_values = [raw_map[t] for t in (selected_emp_types if selected_emp_types else [])]

            # mask = pd.Series(True, index=df_manpower.index)
            # if search_query:
                # row_matches = pd.Series(False, index=df_manpower.index)
                # for c in df_manpower.columns:
                    # row_matches |= df_manpower[c].astype(str).str.contains(search_query, case=False, na=False)
                # mask &= row_matches
            
            # if selected_emp_types:
                # emp_src = (
                    # 'Employee Source' if 'Employee Source' in df_manpower.columns
                    # else 'Updated Employee Source' if 'Updated Employee Source' in df_manpower.columns
                    # else None
                # )
                # if emp_src:
                    # mask = mask & (df_manpower[emp_src].astype(str).str.strip().isin(target_raw_values))
            # else:
                # mask = pd.Series(False, index=df_manpower.index)

            # if range_mode == "Custom Range":
                # mask = mask & (df_manpower['Date'].dt.date >= from_date) & (df_manpower['Date'].dt.date <= to_date)

            # df_filtered = df_manpower[mask].copy()
            
            # if df_filtered.empty:
                # st.warning(f"No data found with the selected filters.")
            # else:
                # def calculate_unique_project_instances(df_in):
                    # if df_in.empty: return 0
                    # # Handle both full dataframe and single-project groups
                    # cols_to_use = ['Project Name', 'Date'] if 'Project Name' in df_in.columns else ['Date']
                    # sub = df_in[cols_to_use].drop_duplicates().sort_values(cols_to_use)
                    # if sub.empty: return 0
                    
                    # if 'Project Name' in sub.columns:
                        # sub['prev_date'] = sub.groupby('Project Name')['Date'].shift(1)
                    # else:
                        # sub['prev_date'] = sub['Date'].shift(1)
                        
                    # sub['gap'] = (sub['Date'] - sub['prev_date']).dt.days
                    # is_new = (sub['gap'].isna()) | (sub['gap'] >= 3)
                    # return int(is_new.sum())

                # def get_drive_level_sums_pw(df_in):
                    # totals = {'Drive Count': 0, 'Call Logs Issue Count': 0, 'FFA': 0, 'Full Batch Delay': 0, 'Partially Batch Delay': 0, 'No Delay': 0}
                    # if df_in.empty: return totals
                    # for proj_name, proj_df in df_in.groupby('Project Name', dropna=False):
                        # if proj_df.empty: continue
                        # u_dates = proj_df[['Date']].drop_duplicates().sort_values('Date')
                        # u_dates['prev'] = u_dates['Date'].shift(1)
                        # u_dates['is_new'] = (u_dates['Date'] - u_dates['prev']).dt.days >= 3
                        # u_dates['drive_id'] = u_dates['is_new'].cumsum()
                        # totals['Drive Count'] += int(u_dates['drive_id'].nunique())
                        # grp_with_id = proj_df.merge(u_dates[['Date', 'drive_id']], on='Date')
                        # drive_level = grp_with_id.groupby('drive_id').agg({
                            # 'Call Logs Issue Count': 'max', 'FFA': 'max', 'Full Batch Delay': 'max',
                            # 'Partially Batch Delay': 'max', 'No Delay': 'max'
                        # })
                        # for k in ['Call Logs Issue Count', 'FFA', 'Full Batch Delay', 'Partially Batch Delay', 'No Delay']:
                            # totals[k] += drive_level[k].sum()
                    # return totals

                # def agg_project(grp):
                    # if grp.empty: return pd.Series()
                    
                    # # 1. Identify unique Drives within this project
                    # # Sort by Date and find gaps >= 3 days
                    # u_dates = grp[['Date']].drop_duplicates().sort_values('Date')
                    # u_dates['prev'] = u_dates['Date'].shift(1)
                    # u_dates['is_new'] = (u_dates['Date'] - u_dates['prev']).dt.days >= 3
                    # u_dates['drive_id'] = u_dates['is_new'].cumsum()
                    
                    # # Merge drive_id back to the group
                    # return pd.Series({
                        # 'Drive Count': int(u_dates['drive_id'].nunique()),
                        # 'Employee Unique Count': grp['Emp_ID'].nunique(),
                        # 'Call Log Count': int(grp['Call Logs Issue Count'].sum()),
                        # 'FFA Count': float(grp['FFA'].sum()),
                        # 'Full Batch': int(grp['Full Batch Delay'].sum()),
                        # 'Partial': int(grp['Partially Batch Delay'].sum()),
                        # 'No Batch Delay': int(grp['No Delay'].sum())
                    # })
                
                # df_proj_agg = df_filtered.groupby('Project Name', dropna=False).apply(agg_project, include_groups=False).reset_index()

                # total_project_count = int(df_filtered['Project Name'].nunique()) if 'Project Name' in df_filtered.columns else 0
                # total_drive_count = int(df_proj_agg['Drive Count'].sum())
                # total_ffa = float(df_proj_agg['FFA Count'].sum())
                # total_call_logs = int(df_proj_agg['Call Log Count'].sum())
                # total_full_delay = int(df_proj_agg['Full Batch'].sum())
                # total_part_delay = int(df_proj_agg['Partial'].sum())
                # total_no_delay = int(df_proj_agg['No Batch Delay'].sum())
                # total_manpower = int(df_filtered['Emp_ID'].nunique()) if 'Emp_ID' in df_filtered.columns else 0
                
                # st.markdown("### Overall Metrics")
                # k1, k2, k3, k4, k5, k6, k7, k8 = st.columns(8)
                # with k1:
                    # metric_card_compact("Project Count", total_project_count, accent="#3B82F6")
                # with k2:
                    # metric_card_compact("Drive Count", total_drive_count, accent="#0EA5E9")
                # with k3:
                    # metric_card_compact("Call Logs Issue", total_call_logs, accent="#F59E0B")
                # with k4:
                    # metric_card_compact("FFA", total_ffa, accent="#10B981")
                # with k5:
                    # metric_card_compact("Full Delay", total_full_delay, accent="#EF4444")
                # with k6:
                    # metric_card_compact("Part Delay", total_part_delay, accent="#8B5CF6")
                # with k7:
                    # metric_card_compact("No Delay", total_no_delay, accent="#6366F1")
                # with k8:
                    # metric_card_compact("Unique Manpower", total_manpower, accent="#64748B")
                
                # st.markdown("### Project-wise Aggregation")
                # # Ensure correct column order as requested
                # cols = ['Project Name', 'Drive Count', 'Employee Unique Count', 'Call Log Count', 'FFA Count', 'Full Batch', 'Partial', 'No Batch Delay']
                # st.dataframe(df_proj_agg[cols], use_container_width=True, hide_index=True)
                
        # except Exception as e:
            # st.error(f"Error loading project data: {e}")
    # else:
        # st.error(f"Manpower data file not found at {filepath}")

# ================= DATC & DOTC INVENTORY SUMMARY PAGE =================
elif st.session_state.page == "city_datc_dotc":

    # --- Header ---
    head_col, back_col = st.columns([4, 1], vertical_alignment="center")
    with head_col:
        st.markdown(
            """
            <div style="margin-bottom:0px; padding-top:0px;">
                <h2 style="margin:0; color:#1E293B; font-weight:800;">DATC &amp; DOTC Inventory Summary</h2>
                <p style="margin:4px 0 0 0; color:#64748B; font-size:13px;">Venue breakdown by City, District, State — with Active/Inactive status split</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with back_col:
        if st.button("← Back to Dashboard", use_container_width=True, key="city_datc_back"):
            st.session_state.page = "dashboard"
            st.query_params["page"] = "dashboard"
            st.rerun()


    # --- Load data ---
    all_output_files = [
        os.path.join(OUTPUT_PATH, f)
        for f in os.listdir(OUTPUT_PATH)
        if f.startswith("VMS_Full_Analysis") and f.endswith(".xlsx") and not f.startswith("~$")
    ] if os.path.exists(OUTPUT_PATH) else []

    if not all_output_files:
        st.warning("No processed reports found. Admin must run the Analysis Engine first.")
        st.stop()

    latest_path = max(all_output_files, key=os.path.getmtime)
    city_data = pd.read_excel(latest_path, sheet_name=0)
    city_data.columns = city_data.columns.str.lower().str.replace(" ", "_")
    
    # Globally filter out INACTIVE venues
    if "status" in city_data.columns:
        city_data = city_data[city_data["status"].fillna("").astype(str).str.upper() != "INACTIVE"].copy()

    # Clean State Data
    if "state" in city_data.columns:
        city_data["state"] = city_data["state"].fillna("").astype(str).apply(normalize_state_name)
        city_data = city_data[~city_data["state"].isin(FOREIGN_LOCATIONS)]

    # Normalize key columns
    for col in ["venue_type", "status", "state", "region", "city", "district", "category"]:
        if col in city_data.columns:
            city_data[col] = city_data[col].fillna("").astype(str).str.strip()

    if "venue_type" not in city_data.columns:
        st.error("'venue_type' column not found in data.")
        st.stop()

    city_data["v_type_norm"] = city_data["venue_type"].str.upper()
    df_filtered_cd = city_data[city_data["v_type_norm"].isin(["DATC", "DOTC"])].copy()

    if df_filtered_cd.empty:
        st.warning("No DATC or DOTC venues found in the data.")
        st.stop()

    st.markdown('<div class="gap-sm"></div>', unsafe_allow_html=True)
    
    # Filter inputs
    f_col1, f_col2, f_col3 = st.columns([1.2, 1.3, 1.8])
    with f_col1:
        search_query = st.text_input(" Search", placeholder="e.g. EST-AS-1420 or City name...", key="city_search_q").strip()
    with f_col2:
        emp_type_opts = ["DEXIT", "Outsourced"]
        selected_emp_types = st.segmented_control("Employee Type", options=emp_type_opts, default=emp_type_opts, selection_mode="multi", key="city_emptype")
    with f_col3:
        range_mode = st.segmented_control("Date Filter", options=["All Data", "Custom Range"], default="All Data", key="city_range_mode")
        if range_mode == "Custom Range":
            rd_c1, rd_c2 = st.columns(2)
            with rd_c1:
                from_date = st.date_input("From Date", value=datetime.now().replace(day=1), key="city_from")
            with rd_c2:
                to_date = st.date_input("To Date", value=datetime.now(), key="city_to")
        else:
            st.info("Showing all dates")

    # --- Sidebar Filters ---
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🔽 Filters")

    regions_avail = sorted([r for r in df_filtered_cd["region"].unique() if r]) if "region" in df_filtered_cd.columns else []
    states_avail  = sorted([s for s in df_filtered_cd["state"].unique()  if s]) if "state"  in df_filtered_cd.columns else []

    sel_regions = sidebar_checkbox_filter("Region", regions_avail, "cd_region", expanded=False, columns=1)
    sel_states  = sidebar_checkbox_filter("State", states_avail, "cd_state", expanded=False, columns=1)
    sel_types   = sidebar_checkbox_filter("Venue Type", ["DATC", "DOTC"], "cd_vtype", expanded=True, columns=1)
    
    status_avail = sorted([s for s in df_filtered_cd["status"].unique() if s]) if "status" in df_filtered_cd.columns else []
    sel_status  = sidebar_checkbox_filter("Status", status_avail, "cd_status", expanded=False, columns=1)
    top_n = st.sidebar.slider("Top N Cities in Chart", min_value=5, max_value=50, value=20, step=5, key="cd_top_n")

    # Apply filters
    mask_cd = pd.Series(True, index=df_filtered_cd.index)
    if search_query:
        row_matches = pd.Series(False, index=df_filtered_cd.index)
        for c in df_filtered_cd.columns:
            row_matches |= df_filtered_cd[c].astype(str).str.contains(search_query, case=False, na=False)
        mask_cd &= row_matches
    
    if sel_regions and "region" in df_filtered_cd.columns:
        mask_cd &= df_filtered_cd["region"].isin(sel_regions)
    if sel_states and "state" in df_filtered_cd.columns:
        mask_cd &= df_filtered_cd["state"].isin(sel_states)
    if sel_types:
        mask_cd &= df_filtered_cd["v_type_norm"].isin(sel_types)
    if sel_status and "status" in df_filtered_cd.columns:
        mask_cd &= df_filtered_cd["status"].isin(sel_status)

    df_cd = df_filtered_cd[mask_cd].copy()

    if df_cd.empty:
        st.warning("No data matches the selected filters.")
        st.stop()

    # --- KPI Summary ---
    def render_top_metrics_cd(df_in):
        t_venues = len(df_in)
        t_datc   = int((df_in["v_type_norm"] == "DATC").sum())
        t_dotc   = int((df_in["v_type_norm"] == "DOTC").sum())
        st.markdown('<div style="height:16px;"></div>', unsafe_allow_html=True)
        mk1, mk2, mk3 = st.columns(3)
        with mk1: metric_card_compact("Total Venues", t_venues, accent="#3B82F6")
        with mk2: metric_card_compact("DATC",         t_datc,   accent="#8B5CF6")
        with mk3: metric_card_compact("DOTC",         t_dotc,   accent="#06B6D4")
        st.markdown('<div style="height:20px;"></div>', unsafe_allow_html=True)

    # ── Helper: build pivot ──────────────────────────────────────────────────
    def build_city_pivot(df_in, dimension):
        if dimension not in df_in.columns:
            return pd.DataFrame()
        piv = df_in.pivot_table(
            index=dimension, columns="v_type_norm",
            values="dms_code", aggfunc="count", fill_value=0, observed=False
        )
        for col in ["DATC", "DOTC"]:
            if col not in piv.columns:
                piv[col] = 0
        piv = piv[["DATC", "DOTC"]]
        piv["Total"] = piv["DATC"] + piv["DOTC"]

        # ── Seat Capacity columns (Total + Active) ──────────────────────────
        if "venue_max_capacity" in df_in.columns:
            cap = pd.to_numeric(df_in["venue_max_capacity"], errors="coerce").fillna(0)
            df_cap = df_in.copy()
            df_cap["_cap"] = cap

            # Total seat capacity (all venues)
            cap_by_dim = df_cap.groupby(dimension, observed=False)["_cap"].sum()
            piv["Seat Capacity"] = piv.index.map(lambda x: int(cap_by_dim.get(x, 0)))

            if "status" in df_in.columns:
                status_norm = df_cap["status"].str.upper()

                # Active venues count
                active_df   = df_cap[status_norm == "ACTIVE"]
                active_cnt   = active_df.groupby(dimension,   observed=False)["dms_code"].count()
                piv["Active"]   = piv.index.map(lambda x: int(active_cnt.get(x, 0)))

                # Active seat capacities (correctly matched)
                active_cap   = active_df.groupby(dimension,   observed=False)["_cap"].sum()
                piv["Active Capacity"]   = piv.index.map(lambda x: int(active_cap.get(x, 0)))
            else:
                piv["Active"]            = 0
                piv["Active Capacity"]   = 0
        else:
            piv["Seat Capacity"]     = 0
            if "status" in df_in.columns:
                status_norm  = df_in["status"].str.upper()
                active_by    = df_in[status_norm == "ACTIVE"].groupby(dimension, observed=False)["dms_code"].count()
                piv["Active"]   = piv.index.map(lambda x: int(active_by.get(x, 0)))
            else:
                piv["Active"]   = 0
            piv["Active Capacity"]   = 0

        piv = piv.sort_values("Total", ascending=False)

        # Grand Total row
        grand = piv.sum(numeric_only=True)
        grand.name = "Grand Total"
        piv = pd.concat([piv, grand.to_frame().T])
        return piv

    # ── Helper: render pivot as plotly table ─────────────────────────────────
    def render_pivot_table(piv, dim_label):
        if piv.empty:
            st.info("No data available.")
            return
        all_cols = piv.columns.tolist()

        # Column header display names
        col_display = {
            "DATC": "DATC", "DOTC": "DOTC", "Total": "Total",
            "Seat Capacity": "Seat<br>Capacity",
            "Active": "Active",
            "Active Capacity": "Active<br>Capacity",
        }
        # Color-code capacity columns
        col_header_colors = {
            "Active": "#166534",         # dark green header
            "Active Capacity": "#14532D",
            "Seat Capacity": "#1E3A5F",
        }
        header_fill_colors = [
            col_header_colors.get(c, "#1E3A5F") for c in ["_dim"] + all_cols
        ]

        headers   = [f"<b>{dim_label}</b>"] + [f"<b>{col_display.get(c, c)}</b>" for c in all_cols]
        cells_data = [piv.index.astype(str).tolist()] + [piv[c].tolist() for c in all_cols]
        n_rows = len(piv)

        # Row background colors
        row_colors = []
        for i in range(n_rows):
            if i == n_rows - 1:
                row_colors.append("#EFF6FF")   # Grand Total highlight
            elif i % 2 == 0:
                row_colors.append("#FFFFFF")
            else:
                row_colors.append("#F8FAFC")

        # Cell font colors — make capacity cols stand out
        def cell_font_colors(col):
            if col == "Active Capacity":
                return ["#15803D" if i < n_rows - 1 else "#052e16" for i in range(n_rows)]
            if col == "Active":
                return ["#166534" if i < n_rows - 1 else "#052e16" for i in range(n_rows)]
            return ["#1E293B"] * n_rows

        dim_font = ["#1E293B"] * n_rows
        cells_font_colors = [dim_font] + [cell_font_colors(c) for c in all_cols]

        # Column widths — wider for capacity/name cols
        col_widths = []
        for c in ["_dim"] + all_cols:
            if c == "_dim":
                col_widths.append(2.0)
            elif "Capacity" in c:
                col_widths.append(1.4)
            elif c == "Active":
                col_widths.append(1.0)
            else:
                col_widths.append(0.9)

        fig_t = go.Figure(data=[go.Table(
            columnwidth=col_widths,
            header=dict(
                values=headers,
                fill_color=header_fill_colors,
                align="center",
                font=dict(size=11, color="white"),
                line_color="#0F2644",
                height=40
            ),
            cells=dict(
                values=cells_data,
                fill_color=[row_colors] * len(headers),
                align="center",
                font=dict(size=11, color=cells_font_colors),
                line_color="#E2E8F0",
                height=30
            )
        )])
        fig_t.update_layout(
            margin=dict(t=0, b=0, l=0, r=0),
            height=max(380, min(n_rows * 32 + 60, 650)),
            paper_bgcolor="rgba(0,0,0,0)"
        )
        st.plotly_chart(fig_t, use_container_width=True)

    # ── Tabs ─────────────────────────────────────────────────────────────────
    tab_city, tab_district, tab_state, tab_region = st.tabs([
        "🏙️ City-wise", "🗺️ District-wise", "📍 State-wise", "🌏 Region-wise"
    ])

    # ────────── TAB 1: CITY-WISE ──────────────────────────────────────────────
    with tab_city:
        if "city" not in df_cd.columns:
            st.warning("'city' column not found.")
        else:
            df_city = df_cd[df_cd["city"] != ""].copy()
            render_top_metrics_cd(df_city)
            st.markdown("### 🏙️ City-wise DATC & DOTC Breakdown")

            piv_city = build_city_pivot(df_cd, "city")

            # Chart (top N cities excluding Grand Total) — full width
            chart_df_city = df_cd.groupby(["city", "v_type_norm"], observed=False).size().reset_index(name="Count")
            top_cities = df_cd["city"].value_counts().head(top_n).index.tolist()
            chart_df_city = chart_df_city[chart_df_city["city"].isin(top_cities)]

            st.markdown(f"#### Top {top_n} Cities — DATC vs DOTC")
            fig_city = px.bar(
                chart_df_city, x="city", y="Count", color="v_type_norm",
                barmode="group", text="Count",
                color_discrete_map={"DOTC": "#8B5CF6", "DATC": "#3B82F6"},
                labels={"v_type_norm": "Type", "city": "City"}
            )
            fig_city.update_traces(textposition="outside", textfont=dict(size=10), marker_line_width=0)
            fig_city.update_layout(
                height=420, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(tickangle=-35, title=""), yaxis=dict(title="Count"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                margin=dict(t=20, b=60, l=10, r=10)
            )
            st.plotly_chart(fig_city, use_container_width=True)

            # Pivot table — full width so all 8 columns are visible
            st.markdown("#### Pivot Table (All Cities)")
            render_pivot_table(piv_city, "CITY")


    # ────────── TAB 2: DISTRICT-WISE ─────────────────────────────────────────
    with tab_district:
        if "district" not in df_cd.columns:
            st.warning("'district' column not found.")
        else:
            df_dist = df_cd[df_cd["district"] != ""].copy()
            render_top_metrics_cd(df_dist)
            st.markdown("### 🗺️ District-wise DATC & DOTC Breakdown")
            piv_dist = build_city_pivot(df_cd, "district")

            top_districts = df_cd["district"].value_counts().head(top_n).index.tolist()
            chart_df_dist = df_cd[df_cd["district"].isin(top_districts)].groupby(
                ["district", "v_type_norm"], observed=False
            ).size().reset_index(name="Count")

            st.markdown(f"#### Top {top_n} Districts — DATC vs DOTC")
            fig_dist = px.bar(
                chart_df_dist, x="district", y="Count", color="v_type_norm",
                barmode="group", text="Count",
                color_discrete_map={"DOTC": "#8B5CF6", "DATC": "#3B82F6"},
                labels={"v_type_norm": "Type", "district": "District"}
            )
            fig_dist.update_traces(textposition="outside", textfont=dict(size=10), marker_line_width=0)
            fig_dist.update_layout(
                height=420, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(tickangle=-35, title=""), yaxis=dict(title="Count"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                margin=dict(t=20, b=60, l=10, r=10)
            )
            st.plotly_chart(fig_dist, use_container_width=True)

            st.markdown("#### Pivot Table (All Districts)")
            render_pivot_table(piv_dist, "DISTRICT")

    # ────────── TAB 3: STATE-WISE ─────────────────────────────────────────────
    with tab_state:
        if "state" not in df_cd.columns:
            st.warning("'state' column not found.")
        else:
            df_state = df_cd[df_cd["state"] != ""].copy()
            render_top_metrics_cd(df_state)
            st.markdown("### 📍 State-wise DATC & DOTC Breakdown")
            piv_state = build_city_pivot(df_cd, "state")

            chart_df_state = df_cd.groupby(["state", "v_type_norm"], observed=False).size().reset_index(name="Count")

            st.markdown("#### State-wise — DATC vs DOTC")
            fig_state = px.bar(
                chart_df_state, x="state", y="Count", color="v_type_norm",
                barmode="group", text="Count",
                color_discrete_map={"DOTC": "#8B5CF6", "DATC": "#3B82F6"},
                labels={"v_type_norm": "Type", "state": "State"}
            )
            fig_state.update_traces(textposition="outside", textfont=dict(size=10), marker_line_width=0)
            fig_state.update_layout(
                height=420, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(tickangle=-45, title=""), yaxis=dict(title="Count"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                margin=dict(t=20, b=80, l=10, r=10)
            )
            st.plotly_chart(fig_state, use_container_width=True)

            st.markdown("#### Pivot Table (All States)")
            render_pivot_table(piv_state, "STATE")

    # ────────── TAB 4: REGION-WISE ───────────────────────────────────────────
    with tab_region:
        if "region" not in df_cd.columns:
            st.warning("'region' column not found.")
        else:
            df_region = df_cd[df_cd["region"] != ""].copy()
            render_top_metrics_cd(df_region)
            st.markdown("### 🌏 Region-wise DATC & DOTC Breakdown")
            piv_region = build_city_pivot(df_cd, "region")

            chart_df_region = df_cd.groupby(["region", "v_type_norm"], observed=False).size().reset_index(name="Count")
            region_order = [r for r in REGION_ORDER if r in chart_df_region["region"].unique()]

            st.markdown("#### Region-wise — DATC vs DOTC")
            fig_region = px.bar(
                chart_df_region, x="region", y="Count", color="v_type_norm",
                barmode="group", text="Count",
                color_discrete_map={"DOTC": "#8B5CF6", "DATC": "#3B82F6"},
                category_orders={"region": region_order},
                labels={"v_type_norm": "Type", "region": "Region"}
            )
            fig_region.update_traces(textposition="outside", textfont=dict(size=11), marker_line_width=0)
            fig_region.update_layout(
                height=400, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(title=""), yaxis=dict(title="Count"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
                margin=dict(t=20, b=20, l=10, r=10)
            )
            st.plotly_chart(fig_region, use_container_width=True)

            st.markdown("#### Pivot Table (All Regions)")
            render_pivot_table(piv_region, "REGION")

    # ────────── EXPORT TO EXCEL ──────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Export Data")
    exp_col1, exp_col2, exp_col3 = st.columns(3)

    def build_export_bytes(df_in):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Sheet 1: Raw filtered data
            export_cols = [c for c in ["dms_code", "venue_type", "city", "district", "state", "region",
                                        "status", "category", "venue_max_capacity"] if c in df_in.columns]
            df_in[export_cols].to_excel(writer, sheet_name="Raw Data", index=False)
            # Sheet 2: City pivot
            if "city" in df_in.columns:
                build_city_pivot(df_in, "city").to_excel(writer, sheet_name="City-wise")
            # Sheet 3: District pivot
            if "district" in df_in.columns:
                build_city_pivot(df_in, "district").to_excel(writer, sheet_name="District-wise")
            # Sheet 4: State pivot
            if "state" in df_in.columns:
                build_city_pivot(df_in, "state").to_excel(writer, sheet_name="State-wise")
            # Sheet 5: Region pivot
            if "region" in df_in.columns:
                build_city_pivot(df_in, "region").to_excel(writer, sheet_name="Region-wise")
        return output.getvalue()

    with exp_col1:
        try:
            excel_bytes = build_export_bytes(df_cd)
            st.download_button(
                label="⬇️ Download Full Report (Excel)",
                data=excel_bytes,
                file_name=f"City_DATC_DOTC_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
        except Exception as e:
            st.error(f"Export error: {e}")

    with exp_col2:
        try:
            csv_bytes = df_cd[[c for c in ["dms_code", "venue_type", "city", "district",
                                             "state", "region", "status", "venue_max_capacity"]
                                if c in df_cd.columns]].to_csv(index=False).encode("utf-8")
            st.download_button(
                label="⬇️ Download Raw CSV",
                data=csv_bytes,
                file_name=f"City_DATC_DOTC_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"CSV export error: {e}")
